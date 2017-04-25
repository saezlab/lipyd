#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
#  This file is part of the `emese` python module
#
#  Copyright (c) 2015-2017 - EMBL-EBI
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#
#  This code is not for public use.
#  Please do not redistribute.
#  For permission please contact me.
#
#  Website: http://www.ebi.ac.uk/~denes
#

from __future__ import print_function
from future.utils import iteritems
from past.builtins import xrange, range, reduce

import os
import sys
import imp
import traceback
import copy
import time
import datetime

import collections
import itertools

try:
    import cPickle as pickle
except:
    import pickle

import warnings
import timeit

import re

import xlrd
import openpyxl
import xlsxwriter
import lxml.etree
import xmltodict

import zlib
import base64
import struct

import numpy as np
import pandas as pd
import scipy as sp
from scipy import stats
import scipy.cluster.hierarchy as hc
import sklearn.decomposition
import fastcluster
import matplotlib as mpl
import matplotlib.backends.backend_pdf
import matplotlib.pyplot as plt
import seaborn as sns
#import plotly.offline as pl
#import plotly.graph_objs as go
#import plotly.tools
try:
    import altair
except:
    sys.stdout.write('No module `altair` available.\n')

try:
    import rpy2.robjects.packages as rpackages
    rvcd = rpackages.importr('vcdExtra')
    rbase = rpackages.importr('base')
    rutils = rpackages.importr('utils')
    rstats = rpackages.importr('stats')
    rococo = rpackages.importr('rococo')
except:
    sys.stdout.write('Could not import rpy2 or some of the R packages.\n')

# from this module:
import emese.mass as mass
import emese.ms2 as ms2
import emese.fragments as fragments
import emese.mz as mzmod
import emese.progress as progress
import emese._curl as _curl
from emese.common import *

import rlcompleter, readline
readline.parse_and_bind('tab:complete')

warnings.filterwarnings('error')
warnings.filterwarnings('default')
warnings.simplefilter('always')


# ## Decorator classes: ## #

class get_hits(object):
    
    def __init__(self, fun):
        self.fun = fun
    
    def __call__(self, data, **kwargs):
        result = empty_dict(data)
        for protein in data.keys():
            print(protein)
            for pn, tbl in iteritems(data[protein]):
                result[protein][pn] = self.fun(tbl, **kwargs)
                if result[protein][pn] is None:
                    sys.stdout.write('No profile info for %s, %s, %s()\n' \
                        % (protein, pn, self.fun.__name__))
        return result

class count_hits(object):
    
    def __init__(self, fun):
        self.fun = fun
    
    def __call__(self, **kwargs):
        hits = self.obj.empty_dict(self.obj.data)
        phits = self.obj.empty_dict(self.obj.data)
        for protein in self.obj.data.keys():
            for pn, tbl in iteritems(self.obj.data[protein]):
                try:
                    hits[protein][pn] = self.fun(tbl, **kwargs)
                    phits[protein][pn] = self.fun(tbl, **kwargs) / \
                        float(len(tbl[kwargs['name']])) * 100
                except:
                    pass
        return hits, phits
    
    def __get__(self, instance, owner):
        self.obj = instance
        self.cls = owner
        
        return self.__call__

class combine_filters(object):
    
    def __init__(self, fun):
        self.fun = fun
    
    def __call__(self, *args, valids_only = False, **kwargs):
        data = self.obj.valids if valids_only else self.obj.data
        for protein in data.keys():
            for pn, tbl in iteritems(data[protein]):
                try:
                    self.fun(self.obj, tbl, *args, **kwargs)
                except KeyError:
                    sys.stdout.write('\t:: WARNING: error in '
                                     '`@combine_filters` decorator :((\n')
    
    def __get__(self, instance, owner):
        self.obj = instance
        self.cls = owner
        
        return self.__call__

class combine_filters_valids(combine_filters):
    
    def __call__(self, *args, **kwargs):
        super(combine_filters_valids, self).__call__(*args, valids_only = True, **kwargs)

# ## Main class: ## #

class Screening(object):
    """
    Currently this is the main class in this module.
    Accepts dozens of parameters as keyword arguments.
    To see all these check the ``defaults`` dict in the code of the
    ``__init__()`` method.
    """
    
    def __init__(self, **kwargs):
        
        self.defaults = {
            # The absolute root directory.
            # This should not be necessary, why is it here?
            'path_root': '/',
            # The basedir for every files and directories in the followings.
            'basedir': os.getcwd(),
            # If None will be the same as ``basedir``.
            'data_basedir': None,
            # List of the shared directories containing most of the data.
            # Due to their large size these are kept on separate partition.
            # Lists are joined by `os.path.join`.
            # E.g. two shared folders can be defined as
            # `[['my', 'huge', 'partition'], [`another`, `huge`, `disk`]]`.
            'datadirs': [['share']],
            # the overview table of Enric's screening
            # we read the protein containing and the
            # highest fractions from here
            'fractionsf': 'LTPsceenprogres_v07d.xlsx',
            # list of potentially problematic fractions
            'wrongfracsf': 'wrong_fractions.csv',
            # File with offsets of fractions from SEC in ml.
            # This we used at Antonella, but at Enric, these values
            # are different for each protein, and contained by the
            # individual SEC profile files.
            'ppfracf': 'fractions.csv',
            # Directory with the SEC absorbance profiles for each protein.
            'ppsecdir': 'SEC_profiles',
            # The first section in the SEC profile filenames is the protein
            # name or the second:
            'sec_filenames_protein_first': True,
            # This is a typo in some of the PEAK table headers.
            # If this cause trouble just turn it off.
            'fix_fraction_name_ab_typo': True,
            # Directory with the SDS PAGE protein quantities for some
            # proteins.
            'gelprofdir': 'gel_profiles',
            # The directory containing the standards in mzML format.
            # These are used for calculation of recalibration.
            'stddir': 'Standards_mzML format',
            # Directory with manually processed files. These are originally
            # output of this module, then have been manually processed, and
            # can be read again and compared/further analysed.
            'manualdir': 'Processed_files',
            # ending of processed files
            'manualend': 'final.xlsx',
            # For recalibration we read the dates of runs from here.
            'seqfile': 'Sequence_list_LTP_screen_2015.csv',
            # This is an output file to export the absorbance based
            # protein quantities for all proteins and fractions.
            'pptablef': 'proteins_by_fraction.csv',
            # Defines abbreviations of each lipid names and the keywords
            # to identify these in SwissLipids and LipidMaps.
            'lipnamesf': 'lipid_names_v2.csv',
            # Literature curated data about known binding properties of LTPs.
            'bindpropf': 'binding_properties.csv',
            # Lipid classes properties and database IDsb
            'lipipropf': 'lipid_properties.csv',
            # The file with recalibration values from Marco.
            'recalfile': 'Recalibration_values_LTP.csv',
            # If the recalibration performed by this module, we read the
            # expected values of the standards from this file.
            'metabsf': 'Metabolites.xlsx',
            # Simple list with lipid binding domains, names and UniProt
            # IDs of all lipid transfer proteins.
            'ltplistf': 'ltplist.csv',
            # Tolerate numpy warnings, or raise them as errors.
            # Useful for debugging.
            'tolerate_numpy_warnings': True,
            # URLs of external databases.
            # SwissLipids: calculated masses of hundreds of thousands
            # of lipid species.
            'swisslipids_url': 'http://www.swisslipids.org/php/'\
                'export.php?action=get&file=lipids.csv',
            # Experimentally verified masses of few tens of thousands lipids.
            'lipidmaps_url': 'http://www.lipidmaps.org/resources/downloads/'\
                'LMSDFDownload28Jun15.tar.gz',
            # The filename to use after extracting the archice above.
            'lipidmaps_fname': 'LMSDFDownload28Jun15/'\
                'LMSDFDownload28Jun15FinalAll.sdf',
            # ComPPI: protein subcellular localization data.
            'comppi_url': 'http://comppi.linkgroup.hu/downloads',
            # Gene Ontology.
            'goa_url': 'ftp://ftp.ebi.ac.uk/pub/databases/GO/goa/%s/'\
                'goa_%s.gaf.gz',
            # Gene Ontology.
            'quickgo_url': 'http://www.ebi.ac.uk/QuickGO/GAnnotation?format=tsv&'\
                'limit=-1%s&termUse=%s&tax=%u&col=proteinID,goID,goName,aspect',
            # PubChem webservice URL to query molecule properties
            # see details here:
            # https://pubchem.ncbi.nlm.nih.gov/help.html#Glossary
            'pubchem_url': ('https://pubchem.ncbi.nlm.nih.gov/'
                'rest/pug/compound/cid/%s/property/'
                'TPSA,' # polar surface area in square angstroms
                'XLogP,' # water octanol partitioning coefficient
                'Complexity,' # Bertz/Hendrickson/Ihlenfeldt formula
                'HBondDonorCount,' # O, N, P, S with hydrogene
                'HBondAcceptorCount,' # O, N, P, S with lone pair
                'HeavyAtomCount,' # non-H atom count
                'RotatableBondCount/XML'), # rotatable bondsgg
            # Manually curated data by Charlotte about localization
            # and membrane composition.
            'localizationf': 'subcellular_localisation_and_binding.xlsx',
            'membranesf': 'membranes_lipid_composition.xlsx',
            # Original MS2 fragment lists manually compiled by Marco.
            # One for positive and one for negative mode.
            'pfragmentsfile': 'lipid_fragments_positive_mode_v10d2%s.txt',
            'nfragmentsfile': 'lipid_fragments_negative_mode_v10d2.txt',
            # This is a file to output protein profiles to.
            # Serves only the purpose of checking for errors.
            'pptable_file': 'protein_profiles.txt',
            # Cache files. At certain points this module saves and reloads
            # data from pickles, so it don't need to reprocess everything
            # every time. Although if something changed, or something
            # goes wrong, you might need to delete these to ensure
            # everything processed by the most recent code.
            'featurescache': 'features.pickle',
            'pprofcache': 'pprofiles_raw.pickle',
            'abscache': 'absorbances.pickle',
            'flimcache': 'fraclims.pickle',
            # The directory with all MS2 MGF files. If not set, the MGF files
            # will be searched under the directory of each protein.
            'ms2dir': 'MGFfiles',
            # Directory with manually processed `golden standards`
            # from Marco.
            'marco_dir': 'marco',
            # table with manually set protein ratios among
            # many other columns
            'manual_ppratios_xls': 'Proteins_Overview_05.xlsx',
            # Columns to read from the manual ppratios XLS file.
            # These were different at Antonella and Enric, so we
            # need to set them here.
            #'manual_ppratios_xls_cols': [2, 4, 8, 9], # 03b
            'manual_ppratios_xls_cols': [3, 6, 10, 11], # 05
            'auxcache': 'save.pickle',
            'stdcachefile': 'calibrations.pickle',
            'validscache': 'valids.pickle',
            # Write a log about MS2 processing to this file.
            'ms2log': 'ms2identities.log',
            # use the items only of these levels from SwissLipids.
            # Useful to avoid flooding data with redundant subspecies.
            'swl_levels': ['Species'],
            # the source of recalibration: either from Marco, or calculated
            # by this software by processing raw data from the standards.
            'recal_source': 'marco',
            # at Antonella, these proteins come before our fractions, or
            # maybe in A9, so the further fractions we can consider void
            # and we can use their average to reduce the background of the
            # absorbances of the same fractions of other proteins.
            'background_proteins': set(['BNIPL', 'OSBP', 'SEC14L1']),
            # the minimum total intensities (average areas) of MS1 features
            # in negative and positive mode. Below these values features are
            # not highlighted and not included on the `best` sheet in the
            # output tables unless they have MS2 data.
            'aa_threshold': {
                'neg': 30000.0,
                'pos': 150000.0
            },
            'fr_offsets': [0.0],          # at Enric this is [0.0],
                                          # at Antonella [0.010, 0.045]
            'abs_cols': [1], # this can be [1, 3, 5] if we want to read
                             # also the 260 and 215 nm UV profiles...
            'fraclims_from_sec': True, # read the fraction limits from
                                       # the same files as SEC profiles
                                       # or simply from `fractions.csv`
            'pp_do_correction': False, # bypass corrections
            'pcont_fracs_from_abs': False, # determine protein containing
                                           # fractions from the absorbances
                                           # or read from separate file
            # allow features to have missing values in first or last
            # protein containing fractions
            'permit_profile_end_nan': True,
            # This is the maximum extent we can broaden the band
            # when selecting the features with intensity ratios
            # closest to the protein ratio. This is used as a
            # ratio-like limit, e.g. if the number here is 0.25,
            # and the protein ratio is 1.0, then the lower limit
            # will be 1.0 * 0.25 = 0.25 and the upper
            # 1.0 / 0.25 * 1.0 = 4.0.
            'peak_ratio_score_max_bandwidth': 0.25,
            'peak_ratio_score_optimal_bandwidth': 0.5,
            # This is the preferred number of features to
            # calculate the mean and SD in protein ratio score.
            # It means we try to select this number of features
            # with intensity ratio closest to the protein ratio,
            # and use their mean and SD to estimate the fit of 
            # other features. The lower the number the closer
            # features will be selected, but it should be enough
            # large to make the SD a meaningful metric. Similarly
            # we can set a minumum population, below this it
            # really does not make sense to calculate an SD.
            'peak_ratio_score_optimal_population': 10,
            'peak_ratio_score_still_good_population': 7,
            'peak_ratio_score_minimal_population': 4,
            # Use the adaptive method at the peak ratio score
            # calculation. This means to iteratively broaden the
            # band at selecting the features with intensity ratios
            # closest to the protein ratio, either until we have
            # the optimal population of features, or we reach the
            # maximum bandwith. If no features found this way,
            # a warning message will be displayed.
            'adaptive_peak_ratio_score': False,
            # Read externally determined protein peak ratios from file
            # these were provided by Marco and used for Antonella`s
            # data analysis
            'use_manual_ppratios': False,
            # omg, I don't remember what it is
            'use_last_ratio': False,
            # for calculation of peak ratio score, take all intensity ratios
            # within this range of tolerance around the protein ratio.
            # E.g if the protein ratio is 0.6, then the lower threshold
            # will be 0.6 * 0.5 = 0.3, and the upper 1 / 0.5 * 0.6 = 1.2;
            # Then the standard deviation of all intensity ratios within
            # this range is calculated from both positive and negative
            # features, and for each feature and for each pairs of fractions
            # the difference between its intensity ratio divided by the SD
            # results the `peak ratio score`. This is calculated for all
            # pairs of protein containing fractions, and we take their
            # mean if more than 2 fractions are available.
            'peak_ratio_range': 0.25,
            # The threshold below the peak ratio scores considered good.
            # We set a cut-off e.g. for highlighting with green in the
            # output tables, but otherwise it is a continuous measure
            # of goodness.
            'peak_ratio_score_threshold': 1.0,
            # constant fraction layout in Antonella's screening
            # not used at Enric, is kind of "deprecated"
            'fixed_fraction_layout': False,
            'fracs': ['A9', 'A10', 'A11', 'A12', 'B1'],
            'all_fracs': ['A5', 'A6', 'A7', 'A8', 'A9',
                          'A10', 'A11', 'A12', 'B1'],
            # constant fraction layout in Antonella's screening
            # not used at Enric, is kind of "deprecated"
            # uppercase version with leading zeros
            'fracsU': ['A09', 'A10', 'A11', 'A12', 'B01'],
            'pp_minratio': 3,
            # at Antonella's screening a fraction which is considered
            # void, so adjusting the zero of the absorbance profiles
            # to this fraction
            # see method ``pp_baseline_correction()``
            'basefrac': 'A5',
            # the tolerance when looking up MS1 m/z values in databases
            # (Daltons)
            'ms1_tolerance': 0.01,
            # the tolerance when looking up MS2 fragment masses (Daltons)
            'ms2_tolerance': 0.05,
            # the tolerance at identifying features in standards (Daltons)
            'std_tolerance': 0.02,
            # MS2 precursors must have their charges determined
            'ms2_precursor_charge': None,
            # Use only fragments from Marco's lists (note: the series)
            # are still programmatically generated and their m/z values
            # calculated automatically), or use an extended fragment list
            # constructed by Denes (note: this might contain more false
            # positives, i.e. fragments which do not occure, or are iso-
            # baric with others, resulting unnecessary noise; but some-
            # times these annotations still might help to have a clue
            # what is there in cases where otherwise you would see only
            # unknown fragments)
            'only_marcos_fragments': True,
            # Expect only certain adducts at certain lipid categories
            # or assume any lipid may form any type of adduct.
            'adducts_constraints': False,
            # In output tables construct the lipid names programmatically
            # or copy the ones from the database.
            'marco_lipnames_from_db': True,
            # Use the average area values from the `PEAK` software
            # output, or recalculate them here using the intensity
            # values of features across all fractions
            'use_original_average_area': True,
            # Overwrite UV absorbance based protein quantities with
            # those manually acquired from SDS PAGE when these are
            # available
            'use_gel_profiles': True,
            # Use slope_filter (by Enric) for the final
            # selection of profiles
            'slope_profile_selection': True,
            # In slope_filter, if there are 2 highest fractions
            # with approximately equal protein content,
            # we use this range of tolerance to filter
            # the intensity ratios between them. E.g if this number is
            # 0.75, ratios between 0.75 and 1.33 will be accepted.
            'slope_equal_fractions_ratio': 0.75,
            # When we have only the descending parts of the profiles
            # in the measured fractions, we set this additional
            # constraint to avoid include everything descending
            # and select those which still fit better the protein
            'slope_descending_slope_diff_tolerance': 0.8,
            # Do not use the fractions marked as wrong
            # at comparing protein and intensity profiles
            'filter_wrong_fractions': True,
            # The MS2 retention time values must be within the RT
            # range of the feature detected in MS1, otherwise
            # will be dropped
            'ms2_rt_within_range': False,
            # Consider the MS2 scans from only those fractions
            # containing the protein, or from all available fractions
            'ms2_only_protein_fractions' : False,
            # Above this threshold we consider the MS2 spectrum to not
            # belong to the protein and highlight with red in the
            # output tables.
            'deltart_threshold': 0.5,
            # Don't know what it is for
            'uniprots': None,
            # Method names to convert between adduct and exact masses
            'ad2ex': {
                1: {
                    'pos': {
                        '[M+H]+': 'remove_h',
                        '[M+NH4]+': 'remove_nh4',
                        '[M+Na]+': 'remove_na',
                    },
                    'neg': {
                        '[M-H]-': 'add_h',
                        '[M+HCOO]-': 'remove_fo'
                    }
                },
                2: {
                    'pos': {},
                    'neg': {
                        '[M-2H]2-': 'add_2h'
                    }
                },
                3: {
                    'pos': {},
                    'neg': {
                        '[M-3H]3-': 'add_3h'
                    }
                }
            },
            # method names to convert between exact and adduct masses
            'ex2ad': {
                1: {
                    'pos': {
                        '[M+H]+': 'add_h',
                        '[M+NH4]+': 'add_nh4',
                        '[M+Na]+': 'add_na'
                    },
                    'neg': {
                        '[M-H]-': 'remove_h',
                        '[M+HCOO]-': 'add_fo'
                    }
                },
                2: {
                    'pos': {},
                    'neg': {
                        '[M-2H]2-': 'remove_2h'
                    }
                },
                3: {
                    'pos': {},
                    'neg': {
                        '[M-3H]3-': 'remove_3h'
                    }
                }
            },
            # metrics to use for determining similarity of protein
            # and intensity profiles; these were attempts, and not
            # used any more, are "deprecated"
            'metrics': [
                ('Kendall\'s tau', 'ktv', False),
                ('Spearman corr.', 'spv', False),
                ('Pearson corr.', 'pev', False),
                ('Euclidean dist.', 'euv', True),
                ('Robust corr.', 'rcv', False),
                ('Goodman-Kruskal\'s gamma', 'gkv', False),
                ('Difference', 'dfv', True)
            ]
        }
        
        self.in_basedir = ['fractionsf', 'ppfracf', 'seqfile',
            'pptablef', 'lipnamesf', 'bindpropf', 'metabsf',
            'pfragmentsfile', 'nfragmentsfile', 'featurescache',
            'auxcache', 'stdcachefile', 'validscache', 'marco_dir',
            'abscache', 'pptable_file', 'recalfile', 'manual_ppratios_xls',
            'manualdir', 'ltplistf', 'flimcache', 'ppsecdir', 'gelprofdir']
        
        for attr, val in iteritems(self.defaults):
            if attr in kwargs:
                setattr(self, attr, kwargs[attr])
            else:
                setattr(self, attr, val)
        
        self.pfragmentsfile = \
            self.pfragmentsfile % ('m' if self.only_marcos_fragments else '')
        
        self.basedir = os.path.join(*([self.path_root] + self.basedir)) \
            if type(self.basedir) is list \
            else self.basedir
        
        for name in self.in_basedir:
            if hasattr(self, name):
                path = getattr(self, name)
                path = path if type(path) in charTypes else os.path.join(*path)
                setattr(self, name, os.path.join(self.basedir, path))
        
        self.data_basedir = self.basedir \
            if self.data_basedir is None else self.data_basedir
        
        self.datadirs = list(map(lambda p:
            os.path.join(*([self.data_basedir] + p)) if type(p) is list else p,
            self.datadirs
        ))
        
        self.stddir = os.path.join(self.datadirs[0], self.stddir)
        
        self.paths_exist()
        
        if type(self.swl_levels) in charTypes:
            self.swl_levels = set([self.swl_levels])
        elif type(self.swl_levels) is list:
            self.swl_levels = set(self.swl_levels)
        
        self.ms1_tlr = self.ms1_tolerance
        self.ms2_tlr = self.ms2_tolerance
        self.std_tlr = self.std_tolerance
        
        self.nAdducts = None
        self.pAdducts = None
        self.exacts = None
        self.proteins_drifts = None
        self.pp_zeroed = False
        self.wrong_fracs = {}
        
        self.aaa_threshold = {
            'neg': 30000.0,
            'pos': 150000.0
        }
        
        self.recount1 = re.compile(r'\(([Odt]?)-?([0-9]{1,2}):([0-9]{1,2})\)')
        self.recount2 = re.compile(r'\(([Odt]?)-?([0-9]{1,2}):([0-9]{1,2})/'
                                   r'([Odt]?)-?([0-9]{1,2}):([0-9]{1,2})/?'
                                   r'([Odt]?)-?([0-9]{0,2}):?([0-9]{0,2})\)')
        self.recount3 = re.compile(r'\(([Odt]?)-?([0-9]{1,2}):([0-9]{1,2})/?'
                                   r'([Odt]?)-?([0-9]{0,2}):?([0-9]{0,2})/?'
                                   r'([Odt]?)-?([0-9]{0,2}):?([0-9]{0,2})\)')
        self.readd = re.compile(r'(\[M[-\+][-\)\)\+A-Za-z0-9]*\][0-9]?[\-+])')
        self.refa  = re.compile(r'([dl]?)([0-9]+:[0-9]{1,2})\(?([,0-9EZ]+)?\)?')
        
        if os.path.exists('fonts.css'):
            with open('fonts.css', 'r') as f:
                fonts = f.read()
        
        self.html_table_template = """<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="utf-8">
                <title>%s</title>
            """ \
            + """
                <style type="text/css">
                    %s
                    html {
                        font-family: di;
                        color: #333333;
                        padding-top: 40px;
                    }
                    table {
                        border-width: 0px;
                    }
                    table, tr, th, td {
                        border: 1px solid #777777;
                        border-collapse: collapse;
                    }
                    th, .rowname {
                        font-weight: bold;
                    }
                    .positive {
                        background-color: #E25C49;
                        color: #FFFFFF;
                    }
                    .negative, .ms2tbl tr th {
                        background-color: #49969A;
                        color: #FFFFFF;
                    }
                    .both {
                        background: linear-gradient(120deg, #E25C49, #49969A);
                        background: -moz-linear-gradient(120deg, #E25C49, #49969A);
                        /*background: -webkit-linear-gradient(120deg, #E25C49, #49969A);
                        background: -o-linear-gradient(120deg, #E25C49, #49969A);*/
                        color: #FFFFFF;
                    }
                    .matching {
                        background-color: #03928C;
                        color: #FFFFFF;
                    }
                    .nothing {
                        background-color: #FFFFFF;
                        color: #000000;
                        font-weight: normal;
                    }
                    .clickable {
                        cursor: pointer;
                    }
                    td, th {
                        border: 1px solid #CCCCCC;
                    }
                    #ms2main {
                        position: fixed;
                        display: none;
                        top: 0px;
                    }
                    h1, h2, h3, strong, th {
                        font-weight: normal;
                    }
                    #ms2spectra {
                        display: none;
                        width: auto;
                        max-height: 60vh;
                        overflow-y: scroll;
                        overflow-x: hidden;
                        font-size: small;
                        background-color: #FFFFFF;
                        color: #333333;
                        padding: 5px 20px 5px 5px;
                        cursor: default;
                        position: relative;
                    }
                    .close {
                        cursor: pointer;
                        display: inline-block;
                        position: relative;
                        float: right;
                    }
                    .ms2cell {
                        cursor: pointer;
                    }
                    #shade {
                        position: fixed;
                        top: 0;
                        left: 0;
                        background: rgba(1,1,1,0.6);
                        z-index: 5;
                        width: 100%%%%;
                        height: 100%%%%;
                        display: none;
                    }
                    .after5 {
                        display: none;
                    }
                    .ms2hdr {
                        font-size: larger;
                        color: #FFFFFF;
                        background-color: #333333;
                        padding: 3px;
                    }
                    #ms2button {
                        position: fixed;
                        top: 0px;
                        left: 0px;
                        display: inline-block;
                        padding: 7px;
                        font-size: larger;
                        background-color: #6EA945;
                        color: #FFFFFF;
                        cursor: pointer;
                        max-height: 70vh;
                    }
                    .scansbutton {
                        display: inline-block;
                        color: #333333;
                        background-color: #FFFFFF;
                        padding: 3px;
                        margin: 3px;
                        cursor: pointer;
                    }
                    #heighthelper {
                        height: 100vh;
                        overflow: auto;
                        max-height: 60vh;
                    }
                    .scantbl {
                        width: 100%%%%;
                        margin-bottom: 5px;
                    }
                    .scantbl table tr td,th:nth-child(2){
                        width: 100px;
                    }
                    .scantbl table tr td,th:nth-child(1){
                        width: 100px;
                    }
                    .scantbl table tr td,th:nth-child(4){
                        width: 100px;
                    }
                    .ms2hdr {
                        width: 650px;
                    }
                    .primary {
                        display: none;
                    }
                    .secondary {
                        display: none;
                    }
                    .noprotein {
                        display: none;
                    }
                    .ms2tblcontainer {
                        margin-bottom: 5px;
                    }
                </style>
                <script type="text/javascript">
                    function showTooltip(e) {
                        e = e || window.event;
                        var targ = e.target || e.srcElement;
                        if (targ.nodeType == 3) targ = targ.parentNode;
                        text = targ.getAttribute('title')
                        if (text !== undefined) alert(text);
                    }
                </script>
                <script type="text/javascript" src="https://code.jquery.com/jquery-2.2.4.min.js"></script>
                <script type="text/javascript">
                    function SortScans(){
                        $("div[id^=ms2c_]").sort(function(a, b){
                            return ~~a.id.split("_")[1] < ~~b.id.split("_")[1]
                        }).appendTo("#ms2spectra");
                    }
                    $(document).on("click.ms2cell", ".ms2cell", $(this).attr("ms2spectra"), function(e){
                        $("#ms2spectra").append(
                            decodeURIComponent(
                                escape(
                                    window.atob(
                                        $(e.target).attr("ms2spectra")
                                    )
                                )
                            )
                        );
                        $("#ms2spectra").fadeIn();
                        SortScans();
                    });
                    $(document).on("click.ms2button", "#ms2button", function(e){
                        if(e.target.id == "ms2button"){
                            $("#ms2spectra").toggle();
                        }
                    });
                    $(document).on("click.morefrags", ".morefrags", function(e){
                        e.stopPropagation();
                        $(".after5").toggle();
                    });
                    $(document).on("click.remove", ".remove", function(e){
                        e.stopPropagation();
                        $(this).closest(".ms2tblcontainer").remove();
                        if(! $(".ms2tblcontainer").length ){
                            $("#ms2spectra").fadeOut();
                        }
                    });
                    $(document).on("click.scansp", ".morescans1", function(e){
                        e.stopPropagation();
                        $(this).closest(".ms2tblcontainer").children(".primary").toggle();
                    });
                    $(document).on("click.scansp", ".morescans2", function(e){
                        e.stopPropagation();
                        $(this).closest(".ms2tblcontainer").children(".secondary").toggle();
                    });
                    $(document).on("click.scansp", ".morescans3", function(e){
                        e.stopPropagation();
                        $(this).closest(".ms2tblcontainer").children(".noprotein").toggle();
                    });
                </script>
            </head>""" % (fonts) \
        + """
            <body>
                <div id="ms2button">MS2 spectra<br />
                    <div id="ms2spectra"></div>
                </div>
                <h1>%s</h1>
                <table id="features" border="0">
                    %s
                </table>
            </body>
            </html>"""
        
        self.colors = [
            '#B81466', # blue maguerite
            '#C441B3', # slate blue
            '#D6708B', # pale violet red
            '#69B3D6', # viking
            '#49969A', # chetwode blue
            '#608784', # gothic
            '#03928C', # java
            '#CF5836', # mandy
            '#B3443D', # blush
            '#65B9B9', # seagull
            '#009BB9', # pelorous
            '#D88776', # my pink
            '#755987', # kimberly
            '#6F3940', # sanguine brown
            '#A09255', # teak
            '#7C997B', # avocado
            '#582E5E', # minsk
            '#224A7A', # fun blue
            '#FE2E08', # orange red
            '#2D16E2', # medium blue
            '#CBE216', # fuego
            '#14B866', # medium sea green
            '#987B99'  # london hue
        ]
        
        self.df_header = [
            'protein',
            'ionm',
            'id',
            'mz',
            'mzcorr',
            'intensity',
            'cls',
            'headgroup1',
            'rtmean',
            'rtms2',
            'rtlow',
            'rtup',
            'headgroup',
            'lyso',
            'pref',
            'carb',
            'unsat',
            'fa1p',
            'fa1c',
            'fa1u',
            'fa2p',
            'fa2c',
            'fa2u',
            'fa3p',
            'fa3c',
            'fa3u',
            'fullhgroup',
            'mhgroup',
            'uhgroup',
            'hgcc',
            'cc',
            'hgfa',
            'ccfa',
            'screen'
        ]
        
        if not self.tolerate_numpy_warnings:
            self.numpy_warnings_as_errors()
    
    def paths_exist(self):
        paths = list(map(lambda name:
            getattr(self, name),
            self.in_basedir
        )) + self.datadirs
        for path in paths:
            if not os.path.exists(path) and path[-6:] != 'pickle':
                sys.stdout.write('\t:: Missing input file/path: %s\n' % path)
        sys.stdout.flush()
    
    def reload(self, children = False):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist=[modname.split('.')[0]])
        imp.reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
        
        if children and hasattr(self, 'valids'):
            
            for protein, d in iteritems(self.valids):
                
                for tbl in d.values():
                    
                    if 'ms2f' in tbl:
                        
                        for ms2f in tbl['ms2f'].values():
                            
                            ms2f.reload(children = True)
    
    @staticmethod
    def numpy_warnings_as_errors():
        np.seterr(all = 'raise')
        warnings.filterwarnings('error')
        warnings.showwarning = warn_with_traceback
    
    @staticmethod
    def numpy_warnings_as_warnings():
        np.seterr(all = 'warn')
        warnings.filterwarnings('always')
    
    #
    # read standards data
    #

    def which_day(self, protein, mode = 'pos'):
        """
        Which day the fractions for a given LTP have been run?
        """
        if not hasattr(self, 'seq'):
            self.read_seq()
        return map(lambda date_fr:
            date_fr[0],
            filter(lambda date_fr:
                (protein, mode) in map(lambda sample:
                    (sample[0], sample[1]),
                    date_fr[1]
                ),
                iteritems(self.seq)
            )
        )

    def standards_theoretic_masses(self):
        """
        Reads the exact masses and most abundant ion masses of
        the lipid standards.
        """
        result = {}
        tbl = self.read_xls(self.metabsf, 9)
        hdr = tbl[0]
        for pn in [('pos', 1, 6), ('neg', 8, 16)]:
            hdr = tbl[pn[1] - 1]
            result[pn[0]] = dict(
                map(lambda line:
                    (
                        line[0],
                        {
                            'exact': float(line[1]),
                            'ion': float(line[2]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[3])):
                                self.to_float(line[3]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[4])):
                                self.to_float(line[4]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[5])):
                                self.to_float(line[5])
                        }
                    ),
                    tbl[pn[1]:pn[2]]
                )
            )
        for pn, dd in iteritems(result):
            for lstd, d in iteritems(dd):
                for k in d.keys():
                    if k.isdigit():
                        d['diff_%s'%k] = d['ion'] - d[k] \
                            if type(d[k]) is float else np.nan
        self.stdmasses = result

    #
    # obtaining lipid data from SwissLipids
    #

    def get_id(self, dct, value):
        if value not in dct:
            i = 0 if len(dct) == 0 else max(dct.values()) + 1
            dct[value] = i
        return dct[value]

    def get_lipidmaps(self):
        lipidmaps = []
        fields = ['lm_id', 'systematic_name', 'synonyms', 'category',
            'main_class', 'exact_mass', 'formula', 'inchi_key', 'pubchem_sid',
            'pubchem_cid', 'common_name']
        record = {}
        expect = False
        c = _curl.Curl(self.lipidmaps_url, large = True,
            silent = False, files_needed = [self.lipidmaps_fname])
        lmapsf = c.result
        
        for l in list(lmapsf.values())[0]:
            
            l = l.decode('utf-8')
            if expect:
                record[expect] = l.strip() \
                    if expect != 'exact_mass' else float(l.strip())
                expect = False
            elif l[0] == '>':
                label = l.strip()[3:-1].lower()
                if label in fields:
                    expect = label
            elif l[0] == '$':
                lipidmaps.append([record[label] if label in record else None \
                    for label in fields])
                record = {}
        
        c.close()
        
        return lipidmaps

    def lipidmaps_adducts(self):
        _nAdducts = []
        _pAdducts = []
        if not hasattr(self, 'lipidmaps'):
            self.get_lipidmaps()
        for l in self.lipidmaps:
            if l[5] is not None:
                _pAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+H]+', mzmod.Mz(l[5]).add_h()])
                _pAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+NH4]+', mzmod.Mz(l[5]).add_nh4()])
                _nAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M-H]-', mzmod.Mz(l[5]).remove_h()])
                _nAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+HCOO]-', mzmod.Mz(l[5]).add_fo()])
        _pAdducts = np.array(sorted(_pAdducts, key = lambda x: x[-1]),
            dtype = np.object)
        _nAdducts = np.array(sorted(_nAdducts, key = lambda x: x[-1]),
            dtype = np.object)
        self.pAdducts = _pAdducts \
            if self.pAdducts is None \
            else np.vstack((pAdducts, _pAdducts))
        self.pAdducts = self.pAdducts[self.pAdducts[:,-1].argsort()]
        self.nAdducts = _nAdducts \
            if self.nAdducts is None \
            else np.vstack((nAdducts, _nAdducts))
        self.nAdducts = self.nAdducts[self.nAdducts[:,-1].argsort()]

    def lipidmaps_exact(self):
        
        _exacts = []
        lipidmaps = self.get_lipidmaps()
        
        prg = progress.Progress(len(lipidmaps), 'Processing LipidMaps', 10)
        
        for l in lipidmaps:
            prg.step()
            if l[5] is not None:
                _exacts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]),
                        str(l[2])]), l[6], '', l[5]])
        
        prg.terminate()
        
        _exacts = np.array(
            sorted(_exacts, key = lambda x: x[-1]),
            dtype = np.object)
        
        self.exacts = _exacts \
            if self.exacts is None \
            else np.vstack((self.exacts, _exacts))
        
        self.exacts = self.exacts[self.exacts[:,-1].argsort()]
    
    def export_lipidmaps(self, fname = 'lipidmaps.tab', one_name = False):
        with open(fname, 'w') as f:
            hdr = ['ID']
            if one_name:
                hdr.append('Name')
            else:
                hdr.extend(['Name1', 'Name2', 'Name3'])
            hdr.extend(['Constitution', 'MonoisotopicMass'])
            f.write('%s\n' % '\t'.join(hdr))
            for l in self.exacts:
                if l[0][0] == 'L' and l[5] is not None and l[5] > 0.0:
                    names = l[2].split('|')
                    if one_name:
                        for name in names:
                            if name != 'None':
                                names = [name]
                                break
                        names = [names[0]]
                    formula = '' if l[3] is None else l[3]
                    names = list(map(lambda x: '' if x == 'None' else x, names))
                    seq = [l[0]] + names + [l[3], '%.06f' % l[5]]
                    try:
                        f.write('%s\n' % \
                                '\t'.join([l[0]] + names + [formula, '%.06f' % l[5]])
                            )
                    except:
                        print(seq)
    
    def export_swisslipids(self, fname = 'swisslipids.tab'):
        with open(fname, 'w') as f:
            f.write('%s\t%s\t%s\t%s\n' % \
                ('ID', 'Name', 'Constitution', 'MonoisotopicMass'))
            for l in self.exacts:
                if l[0][0] == 'S' and l[1] == 'Species' and l[5] is not None:
                    f.write('%s\t%s\t%s\t%.06f\n' % \
                        (l[0], l[2], l[3], l[5]))
    
    def get_swisslipids(self, adducts = None,
        exact_mass = False):
        """
        Downloads the total SwissLipids lipids dataset.
        Returns numpy array with all queried adducts and optionally exact
        masses, calculates the formate adduct m/z's on demand.
        Please note, for a number of lipids some adduct m/z's are missing
        from SL, while the exact masses are given for all the lipids. Also
        storing masses of many adducts of the same lipid in an array results
        extensive memory usage (couple of hundreds MB).
        """
        adduct_method = {
            '[M+OAc]-': 'add_ac',
            '[M+H]+': 'add_h',
            '[M-H]-': 'remove_h',
            '[M+HCOO]-': 'add_fo',
            '[M+NH4]+': 'add_nh4'
        }
        if type(adducts) is list:
            adducts = set(adducts)
        readd = re.compile(r'.*(\[.*)')
        c = _curl.Curl(self.swisslipids_url, silent = False,
            compr = 'gz', large = True)
        swl = c.result
        swl.fileobj.seek(-4, 2)
        ucsize = struct.unpack('I', swl.fileobj.read(4))[0]
        swl.fileobj.seek(0)
        hdr = str(swl.readline()).split('\t')
        positives = []
        negatives = []
        exact_masses = []
        prg = progress.Progress(ucsize, 'Processing SwissLipids', 101)
        for l in swl:
            prg.step(len(l))
            l = str(l).split('\t')
            if len(l) > 22:
                for i in xrange(13, 23):
                    if len(l[i]) > 0:
                        add = '' if i == 13 else readd.findall(hdr[i])[0]
                        if not adducts or add in adducts or i == 13 and \
                            exact_mass or formiate and i == 13:
                            mz = self.to_float(l[i])
                            if i == 13:
                                if formiate:
                                    mzfo = mzmod.Mz(mz).add_fo()
                                    foline = [l[0], l[1], l[2],
                                        l[10], '[M+HCOO]-', mzfo]
                                    negatives.append(foline)
                                if exact_mass:
                                    exact_masses.append([l[0], l[1], l[2],
                                        l[10], add, mz])
                            elif hdr[i].endswith('+'):
                                positives.append([l[0], l[1], l[2],
                                    l[10], add, mz])
                            elif hdr[i].endswith('-'):
                                negatives.append([l[0], l[1], l[2],
                                    l[10], add, mz])
        prg.terminate()
        swl.close()
        positives = sorted(positives, key = lambda x: x[-1])
        positives = np.array(positives, dtype = np.object)
        negatives = sorted(negatives, key = lambda x: x[-1])
        negatives = np.array(negatives, dtype = np.object)
        return positives, negatives

    def get_swisslipids_exact(self):
        """
        Downloads the total SwissLipids lipids dataset.
        Returns numpy array with only exact mass values.
        """
        readd = re.compile(r'.*(\[.*)')
        c = _curl.Curl(self.swisslipids_url, silent = False,
            compr = 'gz', large = True)
        swl = c.result
        swl.fileobj.seek(-4, 2)
        ucsize = struct.unpack('I', swl.fileobj.read(4))[0]
        swl.fileobj.seek(0)
        hdr = swl.readline().decode('utf-8').split('\t')
        _exacts = []
        prg = progress.Progress(ucsize, 'Processing SwissLipids', 101)
        
        for l in swl:
            
            prg.step(len(l))
            l = l.decode('utf-8').split('\t')
            if len(l) > 22:
                mz = self.to_float(l[14])
                add = ''
                _exacts.append([l[0], l[1], l[2], l[10], add, mz or 0.0])
        
        prg.terminate()
        swl.fileobj.close()
        swl.close()
        
        del swl
        _exacts = sorted(_exacts, key = lambda x: x[-1])
        _exacts = np.array(_exacts, dtype = np.object)
        self.exacts = _exacts \
            if self.exacts is None \
            else np.vstack((self.exacts, _exacts))
        self.exacts = self.exacts[self.exacts[:,-1].argsort()]
        
    
    def add_nonidet(self):
        """
        Adds mass values of Nonidet P-40 to the mass database.
        """
        self.nonidet_mzs = [528.37402, 484.34353, 572.39251,
                            616.36121, 660.44410]
        nonidet = []
        for mz in self.nonidet_mzs:
            nonidet.append(['SLM:999999999', 'Species', 'Nonidet P-40', '', '', mz])
        nonidet = np.array(nonidet, dtype = np.object)
        self.exacts = nonidet \
            if self.exacts is None \
            else np.vstack((self.exacts, nonidet))
        self.exacts = self.exacts[self.exacts[:,-1].argsort()]

    def database_set(self, exacts, names, levels = ['Species']):
        levels = [levels] if type(levels) in [str, unicode] else levels
        levels = set([levels]) if type(levels) is list else levels
        names = [names] if type(names) in [str, unicode] else names
        idx = []
        for i, e in enumerate(exacts):
            if e[1] in levels:
                for n in names:
                    if n in e[2]:
                        idx.append(i)
                        break
        return exacts[idx,:]
    
    def process_db_keywords(self, kwdstr):
        return \
            list(
                map(
                    lambda kwdset:
                        {
                            'neg':
                                list(
                                    map(
                                        lambda kwd:
                                            kwd[1:],
                                        list(
                                            filter(
                                                lambda kwd:
                                                    len(kwd) and \
                                                    kwd[0] == '!',
                                                kwdset.split(';')
                                            )
                                        )
                                    )
                                ),
                            'pos':
                                list(
                                    filter(
                                        lambda kwd:
                                            len(kwd) and kwd[0] != '!',
                                        kwdset.split(';')
                                    )
                                )
                        },
                    kwdstr.split('|')
                )
            )
    
    def binders_of(self, hg):
        return \
            list(
                map(
                    lambda i:
                        i[0],
                    filter(
                        lambda i:
                            hg in i[1] and i[0] in self.fractions_upper,
                        iteritems(self.bindprop)
                    )
                )
            )

    #
    # reading the SEC absorption values and calculating the protein 
    # profiles in the fractions
    #
    
    def _GLTPD1_profile_correction_eq(self, propname):
        eq = (getattr(self, propname)['GLTPD1']['A11'] + \
            getattr(self, propname)['GLTPD1']['A12']) / 2.0
        getattr(self, propname)['GLTPD1']['A11'] = eq
        getattr(self, propname)['GLTPD1']['A12'] = eq
    
    def _GLTP1_profile_correction_inv(self, propname):
        a11 = getattr(self, propname)['GLTPD1']['A11']
        a12 = getattr(self, propname)['GLTPD1']['A12']
        getattr(self, propname)['GLTPD1']['A11'] = a12
        getattr(self, propname)['GLTPD1']['A12'] = a11
    
    def pp(self, **kwargs):
        self.secfracs = {}
        for offset, label in zip([0.0] + self.fr_offsets, ['', 'L', 'U']):
            self._pp(offset = offset, label = label, **kwargs)
    
    def raw_sec_absorbances(self, cache = True, fraclim = True):
        """
        Reads the SEC absorbance curves into numpy arrays.
        """
        refrac = re.compile(r'[0-9]*([A-Z][0-9]{1,2})')
        
        self.pp_zeroed = False
        if cache and os.path.exists(self.abscache) and \
            (not fraclim or os.path.exists(self.flimcache)):
            
            with open(self.abscache, 'rb') as fp:
                self.absorb = pickle.load(fp)
            
            if fraclim:
                with open(self.flimcache, 'rb') as fp:
                    self.fraclim = pickle.load(fp)
            
            return None
        
        reprotein = re.compile(r'.*?[\s_-]?([A-Za-z0-9]{3,})[\s_]?([A-Za-z0-9]+?)\.[a-z]{3}')
        self.absorb = {}
        secdir = os.path.join(self.basedir, self.ppsecdir)
        fnames = os.listdir(secdir)
        
        if not hasattr(self, 'fraclim'):
            self.fraclim = {}
        
        for fname in fnames:
            
            if self.sec_filenames_protein_first:
                protein = reprotein.findall(fname)[0][0]
            else:
                protein = reprotein.findall(fname)[0][1]
            if fname[-3:] == 'xls' or fname[-4:] == 'xlsx':
                
                try:
                    tbl = self.read_xls(os.path.join(secdir, fname))[3:]
                except xlrd.biffh.XLRDError:
                    sys.stdout.write('Error reading XLS:\n\t%s\n' % \
                        os.path.join(self.secdir, fname))
                    continue
                
                self.absorb[protein.upper()] = \
                    np.array(
                        list(
                            map(
                                lambda l:
                                    list(
                                        map(
                                            lambda num:
                                                self.to_float(num),
                                            l[:6]
                                        )
                                    ),
                                tbl
                            )
                        )
                    )
            else:
                with open(os.path.join(secdir, fname), 'r') as f:
                    
                    sec = \
                        list(
                            filter(
                                len,
                                map(
                                    lambda l:
                                        l.split(),
                                    f.read().split('\n')[3:]
                                )
                            )
                        )
                    
                    self.absorb[protein.upper()] = \
                        np.array(
                            list(
                                map(
                                    lambda l:
                                        [
                                            float(l[0].strip()),
                                            float(l[1].strip())
                                        ],
                                    sec
                                )
                            )
                        )
                    
                    if fraclim:
                        
                        fraclims = \
                            list(
                                map(
                                    lambda l:
                                        (
                                            l[3].replace('"', ''),
                                            float(l[2].strip())
                                        ),
                                    filter(
                                        lambda l:
                                            len(l) > 3,
                                        sec
                                    )
                                )
                            )
                        
                        self.fraclim[protein.upper()] = \
                            list(
                                map(
                                    lambda i:
                                        (
                                            i[1][1], # the lower boundary
                                            fraclims[i[0] + 1][1], # the upper
                                            refrac.match(i[1][0]).groups()[0] # the fraction label
                                        ),
                                    enumerate(fraclims[:-1]) # last one is the Waste
                                )
                            )
        
        with open(self.abscache, 'wb') as fp:
            pickle.dump(self.absorb, fp)
        
        if fraclim:
            with open(self.flimcache, 'wb') as fp:
                pickle.dump(self.fraclim, fp)
    
    def pp2(self):
        """
        Runs the complete workflow of all protein and
        intensity profile related calculations.
        This includes:
        - reading the UV absorbances
        - taking the mean absorbance for each fraction
        with multiple offsets if necessary
        - doing corrections for baseline and background
        - getting a mean from different offsets
        - optionally determining the protein containing
        fractions from the absorbances
        - reading manually defined protein ratios if
        ``use_manual_ppratios`` is ``True``
        - determining measured fractions for each protein
        - listing those proteins measured only in one
        fraction
        - setting the protein content values of all non
        protein containing fractions to zero
        
        """
        
        self.raw_sec_absorbances()
        self.absorbances_by_fractions()
        if self.pp_do_correction:
            self.pp_baseline_correction()
            self.pp_background()
            self.pp_background_correction()
        else:
            self.abs_by_frac_c = self.abs_by_frac
        self.mean_pp()
        if self.pcont_fracs_from_abs:
            self.protein_containing_fractions_from_absorbances()
        if not hasattr(self, 'fractions_upper'):
            self.fractions_upper = self.fractions
        self.fractions_marco()
        self.set_measured()
        self.set_onepfraction()
        self.zero_controls()
    
    def absorbances_by_fractions(self):
        """
        Calculates mean absorbances by fractions at
        all offsets in `fr_offsets`.
        """
        result = {}
        def get_segment(protein, c, lo, hi):
            a = self.absorb[protein]
            return np.nanmean(a[np.where(np.logical_and(a[:,c - 1] < hi,
                                                     a[:,c - 1] >= lo)),c])
        
        if not hasattr(self, 'fraclim') or not self.fraclim:
            self.read_fraction_limits()
        
        for protein in self.absorb.keys():
            result[protein] = {}
            # whether we have uniform limits,
            # or different for each protein
            fraclim = self.fraclim \
                if type(self.fraclim) is list \
                else self.fraclim[protein]
            
            for lim in fraclim:
                result[protein][lim[2]] = {}
                for c in self.abs_cols:
                    result[protein][lim[2]][c] = []
                    
                    if not self.pp_do_correction:
                        abs_min = np.nanmin(self.absorb[protein][:,c])
                    else:
                        abs_min = 0.0
                    
                    for o in self.fr_offsets:
                        result[protein][lim[2]][c].append(
                            get_segment(protein, c, lim[0] + o, lim[1] + o) - \
                                abs_min
                        )
        self.abs_by_frac = result
        
        self.gel_protein_quantification()
    
    def gel_protein_quantification(self):
        """
        For proteins with various issues with UV absorbances,
        this method reads quantifications based on manual
        processing of SDS PAGE images. The UV absorbance values
        in attribute `abs_by_frac` will be replaced.
        """
        if not self.use_gel_profiles:
            return None
        
        fnames = []
        self.missing_gel = {}
        
        if os.path.exists(self.gelprofdir):
            
            fnames = os.listdir(self.gelprofdir)
        
        if not len(fnames):
            sys.stdout.write('\t:: Warning: using SDS PAGE based protein '\
                'profiles turned on,\n'\
                '\t   but directory `%s` either '\
                'does not exist or is empty.\n' % (self.gelprofdir))
        
        for fname in fnames:
            
            with open(os.path.join(self.gelprofdir, fname), 'r') as fp:
                
                protein = fname.split('_')[0].upper()
                profile = {}
                
                for line in fp:
                    
                    if not len(line.strip()):
                        continue
                    
                    line = line.strip().split('\t')
                    try:
                        fr, val = line[0], float(line[1])
                    except:
                        print(protein, line)
                        continue
                    
                    profile[fr] = val
                
                missing_sec = (set(profile.keys()) -
                               set(self.abs_by_frac[protein].keys()))
                
                missing_gel = (set(self.abs_by_frac[protein].keys()) -
                               set(profile.keys()))
                
                if len(missing_gel):
                    
                    sys.stdout.write('\t\t -- These fractions at protein '\
                        '`%s` are not available from SDS PAGE: %s.\n'\
                        '\t\t    Setting them to zero.\n' % (
                            protein,
                            ', '.join(self.sort_fracs_str(missing_gel))
                        ))
                
                if len(missing_sec):
                    
                    sys.stdout.write('\t\t -- These fractions at protein '\
                        '`%s` present on SDS PAGE but missing from the '\
                        'SEC profile file: %s\n. '\
                        'Discarding them.\n' % (
                            protein,
                            ', '.join(self.sort_fracs_str(missing_sec))
                        ))
                
                for fr, val in iteritems(profile):
                    
                    # sometimes we have more columns
                    # if we read the 215 & 260 nm absorbances
                    # although we don't use them
                    # but let's keep the column count
                    # by default we use 280 nm absorbances
                    plen = len(list(
                        self.abs_by_frac[protein][fr].values())[0])
                    
                    for k in self.abs_by_frac[protein][fr].keys():
                        
                        self.abs_by_frac[protein][fr][k] = [val] * plen
                
                for fr in missing_gel:
                    
                    for k in self.abs_by_frac[protein][fr].keys():
                        
                        self.abs_by_frac[protein][fr][k] = [0.0] * plen
                
                self.missing_gel[protein] = missing_gel
    
    def check_missing_gel(self):
        """
        Check whether any of the protein containing fractions has been
        zeroed due to missing from the SDS PAGE based quantification.
        Prints warning if this happened.
        """
        if self.use_gel_profiles:
            
            for protein, missing in iteritems(self.missing_gel):
                
                with_protein = set(self.protein_containing_fractions(protein))
                zeroed = missing & with_protein
                
                if zeroed:
                    
                    sys.stdout.write('\t:: Warning: at protein `%s` fraction'\
                        '(s) have been set to zero because they are missing '\
                        'from the SDS PAGE.\n\t   But these are marked as pr'\
                        'otein containing fractions: %s\n' % (
                            protein,
                            ', '.join(self.sort_fracs_str(zeroed))
                        ))
    
    def pp_baseline_correction(self, base = None):
        """
        Subtracts the value of the `base` frabction from the values
        of each fraction. The `base` considered to be void.
        """
        result = dict(map(lambda p: (p, {}), self.abs_by_frac.keys()))
        
        base = self.basefrac if base is None else base
        
        for protein, d in iteritems(self.abs_by_frac):
            
            for fr, d2 in iteritems(d):
                
                result[protein][fr] = (
                    dict(
                        map(
                            lambda c:
                                (
                                    c[0],
                                    list(
                                        map(
                                            lambda ab:
                                                ab[1] - d[base][c[0]][ab[0]],
                                            enumerate(c[1])
                                        )
                                    )
                                ),
                            iteritems(d2)
                        )
                    )
                )
        
        self.abs_by_frac_b = result
    
    def pp_background_correction(self):
        
        self.abs_by_frac_c = \
            dict(
                map(
                    lambda prot:
                        (
                            prot[0],
                            dict(
                                map(
                                    lambda fr:
                                        (
                                            fr[0],
                                            dict(
                                                map(
                                                    lambda c:
                                                        (
                                                            c[0],
                                                            list(
                                                                map(
                                                                    lambda o:
                                    o[1] - self.profile_background[fr[0]][c[0]][o[0]],
                                                                    enumerate(c[1])
                                                                )
                                                            )
                                                        ),
                                                    iteritems(fr[1])
                                                )
                                            )
                                        ),
                                    iteritems(prot[1])
                                )
                            )
                        ),
                    iteritems(self.abs_by_frac_b)
                )
            )
    
    def pp_background(self):
        self.profile_background = \
            dict(
                map(
                    lambda fr:
                        (
                            fr,
                            dict(
                                map(
                                    lambda c:
                                        (
                                            c,
                                            list(
                                                map(
                                                    lambda o:
                                                        np.mean(
                                                            list(
                                                                map(
                                                                    lambda pra:
                                                                        pra[1][fr][c][o],
                                                                    filter(
                                                                        lambda pra:
                                                                            pra[0] in self.background_proteins,
                                                                        iteritems(self.abs_by_frac_b)
                                                                    )
                                                                )
                                                            )
                                                        ),
                                                    [0, 1]
                                                )
                                            )
                                        ),
                                    self.abs_cols
                                )
                            )
                        ),
                    self.all_fracs
                )
            )
    
    def mean_pp(self):
        
        self.pprofs = {}
        self.pprofsL =  {}
        self.pprofsU =  {}
        def get_mean(attr, get_val):
            getattr(self, attr)[protein] = \
                dict(
                    map(
                        lambda fr:
                            (
                                fr[0],
                                np.nanmean(
                                    list(
                                        map(
                                            lambda vals:
                                                get_val(vals),
                                            fr[1].values()
                                        )
                                    )
                                )
                            ),
                        iteritems(d)
                    )
                )
        
        for protein, d in iteritems(self.abs_by_frac_c):
            get_mean('pprofs', np.nanmean)
            get_mean('pprofsL', lambda x: x[0])
            get_mean('pprofsU', lambda x: x[-1])
    
    def read_fraction_limits(self):
        with open(self.ppfracf, 'r') as f:
            self.fraclim = \
                list(
                    map(
                        lambda i:
                            (self.to_float(i[0]),
                             self.to_float(i[1]),
                             i[2].upper()),
                        filter(
                            lambda l:
                                len(l) == 3,
                            list(
                                map(
                                    lambda l:
                                        l.split(';'),
                                    f.read().split('\n')
                                )
                            )
                        )
                    )
                )
    
    def _pp(self, offset = 0.0, label = '', cache = True,
        correct_GLTPD1 = True, GLTPD_correction = 'eq'):
        """
        For each protein, for each fraction, calculates the mean of
        absorptions of all the measurements belonging to one fraction.
        """
        cachefile = '%s%s.pickle' % (self.pprofcache.split('.')[0], label)
        propname = 'pprofs%s' % label
        if cache and os.path.exists(cachefile):
            with open(cachefile, 'rb') as fp:
                self.pprofs = pickle.load(fp)
            
            if correct_GLTPD1:
                getattr(self,
                    '_GLTPD1_profile_correction_%s' % GLTPD_correction)(propname)
            return None
        
        reprotein = re.compile(r'.*[\s_-]([A-Za-z0-9]{3,})\.xls')
        result = {}
        secdir = os.path.join(self.basedir, self.ppsecdir)
        fnames = os.listdir(secdir)
        
        with open(self.ppfracf, 'r') as f:
            frac = \
                list(
                    map(
                        lambda i:
                            (self.to_float(i[0]) + offset,
                            self.to_float(i[1]) + offset,
                            i[2].upper()),
                        filter(
                            lambda l:
                                len(l) == 3,
                            map(
                                lambda l:
                                    l.split(';'),
                                f.read().split('\n')
                            )
                        )
                    )
                )
        
        self.secfracs['%u' % int(offset * 1000)] = frac
        
        for fname in fnames:
            
            protein = reprotein.findall(fname)[0]
            frac_abs = dict((i[2], []) for i in frac)
            gfrac = (i for i in frac)
            fr = next(gfrac)
            
            try:
                tbl = self.read_xls(os.path.join(secdir, fname))[3:]
            except xlrd.biffh.XLRDError:
                sys.stdout.write('Error reading XLS:\n\t%s\n' % \
                    os.path.join(self.basedir, fname))
                continue
            
            minabs = min(0.0, min(self.to_float(l[5]) for l in tbl))
            
            for l in tbl:
                # l[4]: volume (ml)
                ml = self.to_float(l[4])
                if ml < fr[0]:
                    continue
                if ml >= fr[1]:
                    try:
                        fr = next(gfrac)
                    except StopIteration:
                        break
                # l[5] mAU UV3 215nm
                frac_abs[fr[2]].append(self.to_float(l[5]) - minabs)
            
            result[protein] = dict((fnum, np.mean(a)) \
                for fnum, a in iteritems(frac_abs))
        
        with open(cachefile, 'wb') as fp:
            pickle.dump(result, fp)
        
        setattr(self, propname, result)
        if correct_GLTPD1:
            getattr(self,
                '_GLTPD1_profile_correction_%s' % GLTPD_correction)(propname)
    
    def protein_profile_correction(self, propname):
        for protein_name, prof in iteritems(getattr(self, propname)):
            basefrac = prof[self.basefrac]
            for frac in self.fracs:
                prof[frac] -= basefrac
            cmin = min(map(lambda fr: prof[fr], self.fracs))
            for frac in self.fracs:
                prof[frac] -= cmin
    
    def zero_controls(self):
        """
        For protein profiles with offsets 0, 15 and 45
        it calls protein_profile_correction() to
        make the baseline correction, saves a copy
        to pprof_original<offset> property, and
        after sets the values for non LTP conatining
        fractions to zero.
        """
        if self.pp_zeroed:
            sys.stdout.write('\t:: Looks like controls already set to zero, '\
                'please set `pp_zeroed` to False to override.\n')
            sys.stdout.flush()
            return None
        for offset, label in zip([0.0] + self.fr_offsets, ['', 'L', 'U']):
            propname = 'pprofs%s' % label
            if self.pp_do_correction:
                self.protein_profile_correction(propname)
            setattr(self, 'pprofs_original%s' % label,
                copy.deepcopy(getattr(self, propname)))
            for protein_name, sample in iteritems(self.fractions):
                for fr, val in iteritems(sample):
                    if val == 0 or val is None:
                        getattr(self, propname)[protein_name.upper()][fr] = 0.0
        self.pp_zeroed = True
    
    def protein_containing_fractions_from_absorbances(self, abscol = 1):
        """
        Sets the protein containing fractions (`fractions`) based on
        the UV absorbances, considering protein containing those whit a
        value higher than the maximum / minratio.
        """
        self.fractions = dict(map(lambda protein: (protein, {}),
                                  self.abs_by_frac_c.keys()))
        
        for protein, ab in iteritems(self.abs_by_frac_c):
            pr = sorted(list(ab.values())[0].keys()) # the UV profile IDs,
                                                     # e.g. 3 = 260 nm
            abs_max = \
                max(
                    map(
                        lambda a:
                            np.mean(a[abscol]),
                        ab.values()
                    )
                )
            
            self.fractions[protein] = \
                dict(
                    map(
                        lambda f:
                            (
                                f[0],
                                int(np.mean(f[1][abscol]) > \
                                    abs_max / self.pp_minratio)
                            ),
                        iteritems(ab)
                    )
                )
        
        for protein, frs in iteritems(self.fractions):
            if protein in self._fractions:
                for fr, val in iteritems(frs):
                    if fr not in self._fractions[protein]:
                        self.fractions[protein][fr] = 0
                    else:
                        self.fractions[protein][fr] = 1
    
    def fractions_table(self, attr, fname = 'fractions_absorption.csv'):
        fracs = ['a5', 'a6', 'a7', 'a8', 'a9', 'a10', 'a11', 'a12', 'b1']
        values = getattr(self, attr)
        with open(fname, 'w') as f:
            f.write('%s\n' % '\t'.join([''] + fracs))
            for protein in sorted(values.keys()):
                f.write('%s\n' % \
                    '\t'.join([protein] + list(map(lambda fr:
                                           '%.08f' % values[protein][fr],
                                        fracs)))
                )
    
    def protein_containing_fractions_table(self, fname = 'proteins_in_fractions.csv'):
        with open(fname, 'w') as f:
            f.write('%s\n' % '\t'.join([''] + self.fracs))
            for protein in sorted(self.fractions):
                f.write('%s\n' % \
                        '\t'.join([protein] + list(map(lambda fr:
                                str(self.fractions[protein][fr + 1]),
                            xrange(5))))
                )
    
    def protein_containing_fractions(self, protein, except_wrong = False):
        """
        Returns a list of those fractions containing the protein,
        sorted by the order of measurement (their position on the plate).
        
        :param str protein: Name of the protein.
        :param bool except_wrong: Do not include fractions
                                  marked as potentially wrong.
        """
        
        return (
            np.array(
                list(
                    filter(
                        lambda fr:
                            self.fractions[protein][fr] == 1,
                        self.all_fractions(protein,
                                           except_wrong = except_wrong)
                    )
                )
            )
        )
    
    def all_fractions(self, protein, except_wrong = False):
        """
        Returns a list of all fractions for one protein,
        sorted by the order of measurement (their position on the plate).
        
        :param str protein: Name of the protein.
        :param bool except_wrong: Do not include fractions
                                  marked as potentially wrong.
        
        """
        
        if except_wrong:
            
            wfracs = self.wrong_fractions(protein, except_wrong)
            
            return list(filter(lambda fr: fr not in wfracs,
                               self.pfracs[protein][:-1]))
        
        # removing the SEC buffer control
        return self.pfracs[protein][:-1]
    
    def wrong_fractions(self, protein, mode):
        """
        Returns the fractions marked as potentially wrong,
        empty set if there is none.
        """
        
        if (protein in self.wrong_fracs and
            mode in self.wrong_fracs[protein]):
            
            return set(self.wrong_fracs[protein][mode])
        
        return set([])
    
    def fraction_indices(self, protein, except_wrong = False):
        """
        Returns a dict with column indices for all fractions
        and a boolean value whether if it contains protein.
        
        :param str protein: Name of the protein.
        :param bool except_wrong: Do not include fractions
                                  marked as potentially wrong.
        """
        
        return (
            dict(
                map(
                    lambda fr:
                        (
                            fr[1],
                            (fr[0], bool(self.fractions[protein][fr[1]]))
                        ),
                    enumerate(self.all_fractions(protein,
                                                 except_wrong = except_wrong))
                )
            )
        )
    
    def pcont_indices(self, protein, except_wrong = False):
        """
        Returns the indices of protein containing fraction columns.
        
        :param str protein: Name of the protein.
        :param bool except_wrong: Do not include fractions
                                  marked as potentially wrong.
        """
        
        return (
            np.array(
                sorted(
                    map(
                        lambda fr:
                            fr[1][0],
                        filter(
                            lambda fr:
                                fr[1][1],
                            iteritems(self.fraction_indices(protein))
                        )
                    )
                )
            )
        )
    
    def pcont_columns(self, protein, mode, attr = 'fe',
                      except_wrong = False):
        """
        Returns the protein containing fraction columns
        from the features x fractions array `attr`.
        """
        
        return self.valids[protein][mode][attr][:,
            self.pcont_indices(protein, except_wrong = except_wrong)]
    
    def fractions_by_protein_amount(self, except_wrong = False):
        """
        Orders the fractions by the amount of protein.
        Considers fraction offsets if those are assumed,
        and creates dicts `fracs_orderL` and fracs_orderU`
        (these are identical if no offset assumed).
        """
        
        def order(protein, with_protein, pprofs):
            return \
                sorted(
                    filter(
                        lambda fr:
                            fr[0] in with_protein,
                        iteritems(pprofs[protein])
                    ),
                    key = lambda fr: fr[1],
                    reverse = True
                )
        self.fracs_orderL = {}
        self.fracs_orderU = {}
        
        for protein in self.pprofsL.keys():
            
            if protein not in self.pfracs:
                continue
            
            with_protein = (
                self.protein_containing_fractions(protein,
                                                  except_wrong = except_wrong)
            )
            self.fracs_orderL[protein] = \
                order(protein, with_protein, self.pprofsL)
            self.fracs_orderU[protein] = \
                order(protein, with_protein, self.pprofsU)
    
    def primary_fractions(self):
        """
        Selects the fraction with highest protein amount (primary),
        all the others considered secondary. There might be more
        than one primary fractions if at different fraction offsets
        different fractions have the highest amount of protein.
        Creates dict `fracs_order`.
        """
        
        self.fracs_order = dict(
            map(
                lambda p:
                    (p, {'prim': set([]), 'sec': set([])}),
                self.fractions_upper.keys()
            )
        )
        
        for protein, d in iteritems(self.fracs_order):
            
            if protein not in self.pfracs:
                continue
            
            d['prim'].add(self.fracs_orderL[protein][0][0])
            d['prim'].add(self.fracs_orderU[protein][0][0])
            for fr, c in self.fracs_orderL[protein]:
                if fr not in d['prim']:
                    d['sec'].add(fr)
            for fr, c in self.fracs_orderU[protein]:
                if fr not in d['prim']:
                    d['sec'].add(fr)
    
    def protein_peak_ratios(self):
        """
        Calculates the expected minimum and maximum values for
        protein peak ratios.
        The result contains empty dicts for proteins with only
        one fraction, one ratio for those with 2 fractions, and
        2 ratios for those with 3 fractions.
        Creates dict `ppratios`.
        """
        self.ppratios = dict((protein, {}) \
            for protein in self.fractions_upper.keys())

        for protein, sample in iteritems(self.fractions_upper):
            
            if protein not in self.pfracs:
                continue
            
            ratios = {}
            ref = None
            pprofL = self.pprofsL[protein]
            pprofU = self.pprofsU[protein]
            for i, frac in enumerate(self.fracs):
                if sample[i + 1] == 1:
                    if ref is None:
                        ref = frac
                    else:
                        ratio1 = pprofL[ref] / pprofL[frac]
                        ratio2 = pprofU[ref] / pprofU[frac]
                        ratios[(ref, frac)] = tuple(sorted([ratio1, ratio2]))
            self.ppratios[protein] = ratios
    
    def read_manual_ppratios(self, verbose = False, log = None):
        refracs = re.compile(r'.*([AB])([0-9]{1,2})-([AB])([0-9]{1,2}).*')
        nondigit = re.compile(r'[^\d\.-]+')
        tbl = self.read_xls(self.manual_ppratios_xls)[1:]
        ix = self.manual_ppratios_xls_cols
        ppratios = {}
        first = {}
        fractions = {}
        if verbose:
            sys.stdout.write('\tReading manual ppratios\n')
        if log is not None:
            logf = open(log, 'w')
            logf.write('%s\n' % ('\t'.join(['diff', 'protein', 'fractions', 'old_lower', 'old_upper', 'new_lower', 'new_upper'])))
        for l in tbl:
            protein = l[ix[0]].split('=')[0].strip()
            ppratios[protein] = {}
            frm = refracs.findall(l[ix[1]])
            fractions[protein] = set([])
            if frm is not None:
                for i, (fr1l, fr1n, fr2l, fr2n) in enumerate(frm):
                    fr1 = '%s%u' % (fr1l.lower(), int(fr1n))
                    fr2 = '%s%u' % (fr2l.lower(), int(fr2n))
                    fractions[protein].add(fr1)
                    fractions[protein].add(fr2)
                    lower = float(nondigit.sub('', l[ix[2]].split('/')[i]))
                    upper = l[ix[3]].split('/')[i]
                    if upper.strip().lower() == 'inf':
                        upper = np.inf
                    else:
                        upper = float(nondigit.sub('', upper))
                    ppratios[protein][(fr1, fr2)] = (lower, upper)
                    old_lower = self.ppratios[protein][(fr1, fr2)][0]
                    old_upper = self.ppratios[protein][(fr1, fr2)][1]
                    if verbose:
                        sys.stdout.write('\t\tSetting %s:%s at %s to '\
                            '%.02f-%.02f (was %.02f-%.02f before).\n' % \
                            (fr1, fr2, protein,
                             lower, upper,
                             old_lower, old_upper)
                        )
                    if log is not None:
                        logf.write('%s\n' % ('\t'.join([
                            '#' if old_lower == lower and old_upper == upper else '!',
                            protein,
                            '%s:%s' % (fr1, fr2),
                            '%.03f' % old_lower,
                            '%.03f' % old_upper,
                            '%.03f' % lower,
                            '%.03f' % upper
                        ])))
                    if i == 0:
                        first[protein] = (fr1, fr2)
            else:
                if verbose:
                    sys.stdout.write('\t\tCould not read '\
                        'fraction IDs for %s\n' % protein)
            if len(protein) and not len(frm) and log is not None:
                fr = ':'.join(self.protein_containing_fractions(protein))
                sys.stdout.write('\t\tProtein %s present '\
                    'only in fraction %s\n' % (protein, fr))
                logf.write('%s\n' % \
                    '\t'.join(['#', protein, fr] + ['None'] * 5))
        if log is not None:
            logf.close()
        self.ppratios_manual = ppratios
        self.first_ratio_manual = first
        self.fractions_manual = fractions
    
    def ppratios_replace_manual(self):
        epsilon = 0.0000000000001
        for protein, pprs in iteritems(self.ppratios_manual):
            for fr, ratios in iteritems(pprs):
                self.ppratios[protein][fr] = ratios
                frr = (fr[1], fr[0])
                ratiosr = (1.0 / (ratios[0] + epsilon),
                           1.0 / (ratios[1] + epsilon))
                self.ppratios[protein][frr] = ratiosr
            if protein in self.first_ratio_manual:
                fr1i = self.fracs.index(self.first_ratio_manual[protein][0]) + 1
                fr2i = self.fracs.index(self.first_ratio_manual[protein][1]) + 1
                if self.fractions_upper[protein][fr1i] == 1 and \
                    self.fractions_upper[protein][fr2i] == 1:
                    self.first_ratio[protein] = self.first_ratio_manual[protein]
    
    def fractions_marco(self):
        """
        Reads manually defined protein peak ratios and protein containing
        fractions.
        """
        
        self.fractions_original = \
            copy.deepcopy(self.fractions_upper
                          if hasattr(self, 'fractions_upper')
                          else self.fractions)
        
        if not self.use_manual_ppratios:
            return None
        
        refrac = re.compile(r'([AB])([0-9]{1,2})')
        tbl = self.read_xls(self.manual_ppratios_xls)[1:]
        ix = self.manual_ppratios_xls_cols
        for l in tbl:
            protein = l[ix[0]].split('=')[0].strip()
            if not len(protein):
                continue
            fracs = set(map(
                lambda fr:
                    '%s%u' % (fr[0].lower(), int(fr[1])),
                refrac.findall(l[ix[1]])
            ))
            for i, fr in enumerate(self.fracs):
                if fr in fracs:
                    if self.fractions_upper[protein][i + 1] != 1:
                        sys.stdout.write('\t:: Setting fraction %s at %s to 1, '\
                            'this was %s before\n' % \
                            (fr, protein, self.fractions_upper[protein][i + 1]))
                        self.fractions_upper[protein][i + 1] = 1
                elif self.fractions_upper[protein][i + 1] == 1:
                    sys.stdout.write('\t:: Setting fraction %s at %s to 0, '\
                        'this was %s before\n' % \
                        (fr, protein, self.fractions_upper[protein][i + 1]))
                    self.fractions_upper[protein][i + 1] = 0
    
    def protein_peak_ratios2(self):
        """
        Calculates the expected minimum and maximum values for
        protein peak ratios.
        The result contains empty dicts for proteins with only
        one fraction, one ratio for those with 2 fractions, and
        2 ratios for those with 3 fractions.
        """
        
        def ratio_tuple(frac1, frac2, ifracs):
            return tuple(
                        sorted([frac1, frac2],
                            key = lambda fr: ifracs[fr]
                        )
                    )
        
        self.ppratios = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        
        self.first_ratioL = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        self.last_ratioL = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        self.first_ratioU = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        self.last_ratioU = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        self.first_ratio = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        self.last_ratio = dict((protein, {}) \
            for protein in self.fractions_upper.keys())
        
        def get_ratio(protein, ref, frac, o):
            return \
                np.mean(
                    list(
                        map(
                            lambda w:
                                self.abs_by_frac_c[protein][ref][w][o] / \
                                    self.abs_by_frac_c[protein][frac][w][o],
                            self.abs_cols
                        )
                    )
                )
        
        for protein, sample in iteritems(self.fractions_upper):
            
            if protein not in self.pfracs:
                continue
            
            fracs = self.protein_containing_fractions(protein)
            ifracs = dict(map(lambda fr: (fr[1], fr[0]), enumerate(fracs)))
            
            ratios = {}
            for i, frac1 in enumerate(fracs):
                
                for j, frac2 in enumerate(fracs):
                    
                    if frac1 != frac2:
                        
                        ratios[(frac1, frac2)] = \
                            tuple(
                                map(
                                    lambda r:
                                        get_ratio(
                                            protein,
                                            frac1, frac2, r),
                                    xrange(len(self.fr_offsets))
                                )
                            )
            
            self.ppratios[protein] = ratios
            
            if len(self.fracs_orderL) > 1:
                
                highestL = self.fracs_orderL[protein][0][0]
                secondL = self.fracs_orderL[protein][1][0]
                lowestL = self.fracs_orderL[protein][-1][0]
                highestU = self.fracs_orderU[protein][0][0]
                secondU = self.fracs_orderU[protein][1][0]
                lowestU = self.fracs_orderU[protein][-1][0]
                
                self.first_ratioL[protein] = \
                    ratio_tuple(highestL, secondL, ifracs)
                self.last_ratioL[protein] = \
                    ratio_tuple(highestL, lowestL, ifracs)
                self.first_ratioU[protein] = \
                    ratio_tuple(highestU, secondU, ifracs)
                self.last_ratioU[protein] = \
                    ratio_tuple(highestU, lowestU, ifracs)
                
                self.first_ratio[protein] = \
                    tuple(sorted(self.first_ratioL[protein],
                                 key = lambda fr: (fr[0], int(fr[1:]))))
                self.last_ratio[protein] = \
                    tuple(sorted(self.last_ratioL[protein],
                                 key = lambda fr: (fr[0], int(fr[1:]))))
                
            else:
                self.first_ratioL[protein] = None
                self.last_ratioL[protein]  = None
                self.first_ratioU[protein] = None
                self.last_ratioU[protein]  = None
                self.first_ratio[protein] = None
                self.last_ratio[protein]  = None
    
    def intensity_peak_ratios(self):
        """
        Calculates the intensity peak ratios which are to be
        compared with the protein peak ratios.
        For each protein for each mode a new array named `ipr`
        will be created with one row for each feature and
        0, 1 or 2 columns depending on the number of protein
        containing fractions.
        """
        for protein, d in iteritems(self.valids):
            
            fracs = self.protein_containing_fractions(protein)
            ifracs = self.fraction_indices(protein)
            
            for mode, tbl in iteritems(d):
                
                ratios = []
                indices = {}
                
                for fe in tbl['fe']:
                    ratio = []
                    
                    for i in xrange(len(fracs) - 1):
                        
                        for j in xrange(i + 1, len(fracs)):
                            
                            frac1, frac2 = fracs[i], fracs[j]
                            
                            f1 = fe[ifracs[frac1][0]]
                            f2 = fe[ifracs[frac2][0]]
                            r = 0.0 if f1 == 0.0 \
                                else np.inf if f2 == 0.0 \
                                else f1 / f2
                            
                            ratio.append(r)
                            
                            indices[(frac1, frac2)] = len(ratio) - 1
                    
                    ratios.append(ratio)
                
                tbl['ipr'] = np.array(ratios)
                tbl['ipri'] = indices
                
                if self.first_ratio[protein] is not None and \
                    self.first_ratio[protein] in tbl['ipri']:
                    fi = tbl['ipri'][self.first_ratio[protein]]
                    hi = tbl['ipri'][self.last_ratio[protein]]
                    tbl['iprf'] = tbl['ipr'][:,fi]
                    tbl['iprh'] = tbl['ipr'][:,hi]
                else:
                    tbl['iprf'] = None
                    tbl['iprh'] = None
    
    def ratios_in_range(self):
        """
        Creates a new boolean attribute `ppr`, which tells if the intensity
        peak ratios are in the range expected based on the fractions offset,
        or if we assume no offset, then a custom interval.
        """
        
        ppr1 = self.last_ratio if self.use_last_ratio else self.first_ratio
        iprvar = 'iprh' if self.use_last_ratio else 'iprf'
        
        for protein, d in iteritems(self.valids):
            
            fracs = self.protein_containing_fractions(protein)
            
            for mode, tbl in iteritems(d):
                # only if we have more than one fraction:
                
                if len(self.ppratios[protein]) and tbl[iprvar] is not None:
                    
                    in_range = []
                    
                    for i in xrange(len(fracs) - 1):
                        
                        for j in xrange(i + 1, len(fracs)):
                            
                            frac1, frac2 = fracs[i], fracs[j]
                            fkey = (frac1, frac2)
                            ppr = self.ppratios[protein][fkey]
                            idx = tbl['ipri'][fkey]
                            
                            # no fraction offsets:
                            if len(ppr) == 1:
                                ppr = (ppr[0] - ppr[0] * self.peak_ratio_range,
                                    ppr[0] + ppr[0] * self.peak_ratio_range)
                            
                            in_range.append(
                                np.logical_and(
                                    tbl['ipr'][:,idx] >= ppr[0],
                                    tbl['ipr'][:,idx] <= ppr[1]
                                )
                            )
                    
                    tbl['prra'] = np.column_stack(in_range)
                    fi = tbl['ipri'][self.first_ratio[protein]]
                    hi = tbl['ipri'][self.last_ratio[protein]]
                    tbl['prrf'] = tbl['prra'][:,fi]
                    tbl['prrh'] = tbl['prra'][:,hi]
                    tbl['prrv'] = np.sum(tbl['prra'], axis = 1)
                    tbl['prr'] = tbl['prrf']
                
                else:
                    tbl['prra'] = None
                    tbl['prr'] = None
                    tbl['prrf'] = None
                    tbl['prrv'] = None
                    tbl['prrh'] = None
    
    def write_pptable(self):
        """
        Writes protein profiles in a table, so we don't need to read
        all the 60 XLS files every time.
        """
        with open(self.pptablef, 'w') as f:
            
            allfr = \
                sorted(
                    set(
                        itertools.chain(
                            *map(
                                lambda d: d.keys(),
                                self.pprofs.values()
                            )
                        )
                    ),
                    key = lambda x: (x[0], int(x[1:]))
                )
            
            f.write('\t%s%s' % ('\t'.join(allfr), '\n'))
            
            f.write(
                '\n'.join(
                    map(
                        lambda p:
                            '%s\t%s' % (
                                p[0],
                                '\t'.join(
                                    map(
                                        lambda f:
                                            '%.20f' % (p[1][f] if f in p[1] else np.nan),
                                        allfr
                                    )
                                )
                            ),
                        iteritems(self.pprofs)
                    )
                )
            )
    
    def read_pptable(self):
        """
        Reads protein profiles from table.
        
        This method has not been updated for new fraction layout!
        """
        with open(self.pptablef, 'r') as f:
            header = f.readline().strip().split('\t')
            return dict((ll[0], dict(zip(header, self.float_lst(ll[1:])))) \
                for ll in (l.split('\t') for l in f.read().split('\n')))

    def read_xls(self, xls_file, sheet = 0, csv_file = None,
        return_table = True):
        """
        Generic function to read MS Excel XLS file, and convert one sheet
        to CSV, or return as a list of lists
        """
        table = []
        try:
            book = xlrd.open_workbook(xls_file, on_demand = True)
            try:
                if type(sheet) is int:
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)
            table = [[unicode(c.value) \
                for c in sheet.row(i)] \
                for i in xrange(sheet.nrows)]
        except IOError:
            sys.stdout.write('No such file: %s\n' % xls_file)
            sys.stdout.flush()
        except:
            try:
                book = openpyxl.load_workbook(filename = xls_file,
                    read_only = True)
            except:
                sys.stdout.write('\tCould not open xls: %s\n' % xls_file)
                if not os.path.exists(xls_file):
                    sys.stdout.write('\tFile does not exist.\n')
                sys.stdout.flush()
            try:
                if type(sheet) is int:
                    sheet = book.worksheets[sheet]
                else:
                    sheet = book[sheet]
            except:
                sheet = book.worksheets[0]
            cells = sheet.get_squared_range(1, 1,
                sheet.max_column, sheet.max_row)
            table = map(lambda row:
                map(lambda c:
                    unicode(c.value),
                    row
                ),
                cells
            )
        if csv_file:
            with open(csv_file, 'w') as csv:
                csv.write('\n'.join(['\t'.join(r) for r in table]))
        if not return_table:
            table = None
        if 'book' in locals() and hasattr(book, 'release_resources'):
            book.release_resources()
        return table
    
    def mz2oi(self, protein, mode, mz):
        protein = self.protein_name_upper2mixed(protein)
        try:
            tbl = self.data[protein[mode]][mode]
        except KeyError:
            print(protein, mode)
        ui = tbl['ann'][:,2].searchsorted(mz)
        du = 999.0 if ui == tbl['ann'].shape[0] else tbl['ann'][ui,2] - mz
        dl = 999.0 if ui == 0 else mz - tbl['ann'][ui - 1,2]
        d = min(du, dl)
        if d < 0.0001:
            oi = ui if du < dl else ui - 1
        else:
            oi = None
        return oi
    

    def read_seq(self):
        refrac = re.compile(r'[AB][9120]{1,2}')
        result = {}
        with open(self.seqfile, 'r') as f:
            for l in f:
                l = l.replace('"', '').replace('2606', '0626')\
                    .replace('2627', '0627')
                if len(l) and l[:6] != 'Bracke' and l[:6] != 'Sample':
                    l = l.split(',')[1]
                    mode = 'neg' if 'neg' in l \
                        else 'pos' if 'pos' in l \
                        else None
                    seq = l.split(mode)[-1] if mode is not None else None
                    if seq is not None and len(seq) and seq[0] == '_':
                        seq = seq[1:]
                    if seq == 'A9': seq = 'A09'
                    if seq == 'B1': seq = 'B01'
                    l = l.split('_')
                    date = l[0]
                    if 'extracted' in l:
                        if 'SEC' in l:
                            protein = '#BUF'
                        elif 'Std1' in l:
                            protein = '#STD'
                    elif refrac.match(l[-1]) or refrac.match(l[-2]):
                        protein = l[4]
                        # fix typo in filename
                        if protein == 'ORP9STARD15':
                            protein = protein[4:]
                    else:
                        protein = None
                    if date not in result:
                        result[date] = []
                    result[date].append((protein, mode, seq))
        del result['150330']
        del result['150331']
        self.seq = result
    
    def recalibration(self):
        if self.recal_source == 'marco':
            self.drifts_from_marco()
        else:
            self.standards_filenames()
            self.read_seq()
            self.standards_theoretic_masses()
            self.read_standards()
            self.drifts_by_standard()
            self.drifts_by_date()
            self.drifts2proteins()
        self.recalibrate()
    
    def standards_filenames(self):
        fnames = os.listdir(self.stddir)
        result = {}
        for fname in fnames:
            if 'Seq' not in fname:
                _fname = os.path.join(self.stddir, fname)
                lFname = fname[:-5].split('_')
                date = lFname[0]
                mode = 'pos' if 'pos' in lFname else 'neg'
                seq = fname.split(mode)[-1][:-5]
                if len(seq) and seq[0] == '_': seq = seq[1:]
                if date not in result:
                    result[date] = {}
                result[date][('#STD', mode, seq)] = _fname
        self.stdfiles = result

    def is_recalibrated(self):
        return np.any(['recalibrated' in tbl and tbl['recalibrated']\
            for d in self.valids.values() for tbl in d.values()])

    def recalibrate(self, missing = False):
        if self.is_recalibrated():
            if not missing:
                sys.stdout.write('\t:: Looks recalibration '\
                    'already has been done.\n'\
                    '\t   Set `recalibrated` '\
                    'values to `False` and call this method\n'\
                    '\t   again if you really want to repeat it.\n\n')
                sys.stdout.flush()
                return None
            else:
                sys.stdout.write('\t:: Recalibration already done.\n'\
                    '\t   Looking for missing LTP/mode cases, and '\
                    'doing only those.\n'\
                    '\t   If you want to recalibrate everything '\
                    'again,\n'\
                    '\t   reload the original data.\n\n')
                sys.stdout.flush()
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                if 'recalibrated' not in tbl or \
                (not tbl['recalibrated'] and missing):
                    if protein in self.proteins_drifts:
                        if mode in self.proteins_drifts[protein]:
                            ppm = np.median(
                                self.proteins_drifts[protein][mode].values())
                            ratio = self.ppm2ratio(ppm)
                            tbl['mz'] = tbl['mz'] * ratio
                            tbl['recalibrated'] = True
                        else:
                            tbl['recalibrated'] = False
                            sys.stdout.write('\t:: No drift value '\
                                'for %s in mode %s :(\n' % (protein, mode))
                    else:
                        tbl['recalibrated'] = False
                        sys.stdout.write('\t:: No drift values'\
                            ' for %s-%s :(\n' % (protein, mode))

    def read_standards(self, accuracy = 5, cache = True):
        if os.path.exists(self.stdcachefile) and cache:
            self.std_measured = pickle.load(open(self.stdcachefile, 'rb'))
            return None
        result = dict(map(lambda date:
            (date, {}),
            self.stdfiles.keys()
        ))
        for date, fractions in iteritems(self.stdfiles):
            for sample, fname in iteritems(fractions):
                if sample[0] == '#STD':
                    time, scans, centroids = self.read_mzml(fname)
                    peaks = self.find_peaks(scans, centroids, accuracy)
                    peaks = self.filter_peaks(peaks)
                    stdmeasured = self.standards_lookup(peaks, sample[1])
                    result[date][sample] = stdmeasured
        sys.stdout.write('\n\t:: Results saved to file %s\n' % \
            self.stdcachefile)
        sys.stdout.flush()
        pickle.dump(result, open(self.stdcachefile, 'wb'))
        self.std_measured = result

    def drifts_by_standard(self, write_table = 'drifts_ppm.tab'):
        drifts = dict(map(lambda date_fr:
            (date_fr[0], dict(map(lambda lipids:
                (lipids[0], dict(map(lambda lipid:
                    (lipid, None),
                    lipids[1].keys()))
                ),
                iteritems(fr[1])))
            ),
            iteritems(self.std_measured))
        )
        hdr = ['date', 'run', 'mode', 'lipid', 'theoretical',
            'measured', 'ratio', 'ppm']
        tab = [hdr]
        for date in sorted(self.std_measured.keys()):
            fractions = self.std_measured[date]
            for sample in sorted(fractions.keys()):
                lipids = fractions[sample]
                for lipid in sorted(lipids.keys()):
                    mmz = lipids[lipid]
                    mode = sample[1]
                    tmz = self.stdmasses[mode][lipid]['ion']
                    ratio = None if mmz is None else tmz / mmz
                    ppm = None if ratio is None else self.ratio2ppm(ratio)
                    drifts[date][sample][lipid] = ppm
                    tab.append([
                        '20%s-%s-%s' % (date[:2], date[2:4], date[4:6]),
                        sample[2],
                        'positive' if sample[1] == 'pos' else 'negative',
                        lipid,
                        '%.06f' % tmz,
                        '%.06f' % mmz if mmz is not None else 'n/a',
                        '%.09f' % ratio if ratio is not None else 'n/a',
                        '%.03f' % ppm if ppm is not None else 'n/a'
                    ])
        if write_table and type(write_table) in charTypes:
            tab = '\n'.join(map(lambda line:
                '\t'.join(line),
                tab
            ))
            with open(write_table, 'w') as f:
                f.write(tab)
        self.drifts = drifts

    def drifts_table(self, outfile = 'drifts_ppm.tab'):
        hdr = ['date', 'run', 'mode', 'lipid', 'ppm']
        tab = [hdr]
        for date, fractions in iteritems(self.drifts):
            for sample, lipids in iteritems(fractions):
                for lipid, ppm in iteritems(lipids):
                    tab.append([
                        '20%s-%s-%s' % (date[:2], date[2:4], date[4:6]),
                        sample[2],
                        'positive' if sample[1] == 'pos' else 'negative',
                        lipid,
                        '%.03f' % ppm if ppm is not None else 'n/a'
                    ])
        tab = '\n'.join(map(lambda line:
            '\t'.join(line),
            tab
        ))
        with open(outfile, 'w') as f:
            f.write(tab)

    def drifts_by_date(self):
        drifts2 = dict(map(lambda date_fr:
            (date_fr[0], dict(map(lambda sample:
                (sample, None),
                date_fr[1].keys()))
            ),
            iteritems(self.drifts))
        )
        for date, fractions in iteritems(self.drifts):
            for sample, lipids in iteritems(fractions):
                drifts2[date][sample] = \
                np.nanmedian(
                    self.remove_outliers(
                        np.array(
                            filter(lambda x:
                                x is not None,
                                lipids.values()
                            )
                        )
                    )
                )
                if np.isnan(drifts2[date][sample]):
                    sys.stdout.write('\t:: No values in %s %s\n' % \
                        (date, str(sample)))
        self.drifts2 = drifts2

    def drifts2proteins(self, write_table = 'LTPs_drifts.tab'):
        sys.stdout.write('\t:: Setting recalibration data via cache, from previously processed MzMLs\n')
        sys.stdout.flush()
        # before doing anything, fix some inconsistencies:
        stard10fractions = ['A09', 'A10', 'A11', 'A12']
        for i, fr in zip(
                map(lambda sample:
                    sample[0],
                    filter(lambda sample:
                        sample[1][0] == 'STARD10',
                        enumerate(self.seq['150310'])
                    )
                ),
                stard10fractions * 2
            ):
            self.seq['150310'][i] = \
                (self.seq['150310'][i][0], self.seq['150310'][i][1], fr)
        #
        proteins_drifts = {}
        hdr = ['LTP', 'mode', 'fraction', 'ratio', 'ppm']
        tab = [hdr]
        def standard_indices(mode, se):
            return np.array(
                map(lambda s:
                    s[0],
                    filter(lambda s:
                        s[1][0] == '#STD' and s[1][1] == mode,
                        enumerate(se)
                    )
                )
            )
        def add_row(tab, protein, mode, fr, d):
            tab.append([
                protein,
                'positive' if mode == 'pos' else 'negative',
                fr,
                '%.09f' % self.ppm2ratio(d),
                '%.06f' % d
            ])
            return tab
        for date in sorted(self.seq.keys()):
            se = self.seq[date]
            stdi = {}
            lastbuf = {'pos': None, 'neg': None}
            prev = None
            stdi['neg'] = standard_indices('neg', se)
            stdi['pos'] = standard_indices('pos', se)
            for i, sample in enumerate(se):
                typ = sample[0]
                if typ is not None and typ != '#STD':
                    mode = sample[1]
                    ui = stdi[mode].searchsorted(i)
                    if ui == 0:
                        std1i = stdi[mode][0]
                    else:
                        std1i = stdi[mode][ui - 1]
                    if ui >= len(stdi[mode]):
                        std2i = stdi[mode][-1]
                    else:
                        std2i = stdi[mode][ui]
                    if np.isnan(self.drifts2[date][se[std1i]]):
                        std1i = std2i
                    if std1i == std2i or \
                        np.isnan(self.drifts2[date][se[std2i]]):
                        std2i = None
                    if std2i is None:
                        d = self.drifts2[date][se[std1i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    elif std1i is not None and \
                        np.isnan(self.drifts2[date][se[std1i]]) and std2i is not None:
                        d = self.drifts2[date][se[std2i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    elif std2i is not None and \
                        np.isnan(self.drifts2[date][se[std2i]]) and std1i is not None:
                        d = self.drifts2[date][se[std1i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    else:
                        o1 = i - std1i
                        o2 = std2i - i
                        d = (self.drifts2[date][se[std1i]] * o2 + \
                            self.drifts2[date][se[std2i]] * o1) / (o1 + o2)
                    if typ == '#BUF':
                        lastbuf[mode] = d
                    else:
                        if typ not in proteins_drifts:
                            proteins_drifts[typ] = {}
                        if mode not in proteins_drifts[typ]:
                            proteins_drifts[typ][mode] = {}
                        if prev != typ:
                            proteins_drifts[typ][mode]['ctrl'] = lastbuf[mode]
                            tab = add_row(tab, typ, mode, 'CTL', lastbuf[mode])
                        proteins_drifts[typ][mode][sample[2]] = d
                        tab = add_row(tab, typ, mode, sample[2], d)
                    prev = typ
        if write_table and type(write_table) in charTypes:
            with open(write_table, 'w') as f:
                f.write('\n'.join(
                    map(lambda line:
                        '\t'.join(line),
                        tab)
                    )
                )
            sys.stdout.write('\n\t:: Table written to file %s\n\n'%write_table)
        self.proteins_drifts =  proteins_drifts
    
    def drifts_from_marco(self):
        sys.stdout.write('\t:: Reading recalibration data from Marco`s table\n')
        sys.stdout.flush()
        self.proteins_drifts = {}
        with open(self.recalfile, 'r') as f:
            for _ in xrange(4):
                null = f.readline()
            for l in f:
                l = l.strip().split(',')
                self.proteins_drifts[l[0]] = {'pos': {}, 'neg': {}}
                neg_ppm = self.to_float(l[1])
                pos_ppm = self.to_float(l[2])
                for frac in ['ctrl'] + self.fracsU:
                    self.proteins_drifts[l[0]]['pos'][frac] = pos_ppm
                    self.proteins_drifts[l[0]]['neg'][frac] = neg_ppm

    def remove_outliers(self, data, m = 2):
        return data[np.where(abs(data - np.nanmean(data)) < \
            m * np.nanstd(data))]

    def ratio2ppm(self, ratio):
        return (1.0 - ratio) * 1.0e06

    def ppm2ratio(self, ppm):
        return 1.0 - ppm / 1.0e06

    def standards_lookup(self, peaks, mode):
        measured = {}
        for lipid, values in iteritems(self.stdmasses[mode]):
            _mz = values['ion']
            ui = peaks[:,0].searchsorted(_mz)
            ud = 999.0 if ui >= peaks.shape[0] else peaks[ui, 0] - _mz
            ld = 999.0 if ui == 0 else _mz - peaks[ui - 1, 0]
            ci = ui if ud < ld else ui - 1
            mz = peaks[ci,0] if abs(peaks[ci,0] - _mz) <= self.std_tlr else None
            measured[lipid] = mz
        return measured

    def _process_binary_array(self, binaryarray, length = None):
        prefix = '{http://psi.hupo.org/ms/mzml}'
        cvparam = '%scvParam' % prefix
        _binary = '%sbinary' % prefix
        for par in binaryarray.findall(cvparam):
            if par.attrib['name'][-5:] == 'float':
                typ = par.attrib['name'][:2]
            if par.attrib['name'][-5:] == 'ssion':
                comp = par.attrib['name'][:4] == 'zlib'
            if par.attrib['name'][-5:] == 'array':
                name = par.attrib['name'].split()[0]
        binary = binaryarray.find(_binary)
        decoded = base64.decodestring(binary.text)
        if comp:
            decoded = zlib.decompress(decoded)
        # length is stated in the spectrum tag, but knowing the data type
        # we can also calculate:
        length = len(decoded) * 8 / int(typ) if length is None else length
        # float or double:
        _typ = 'd' if typ == '64' else 'f'
        # mzml binary data must be always little endian:
        dt = np.dtype('<%s'%_typ)
        arr = np.frombuffer(decoded, dtype = dt)
        # this does the same:
        # arr = struct.unpack('<%u%s' % (length, _typ), decoded)
        # arr = np.array(arr, dtype = np.float32)
        return name, arr

    def _centroid(self, raw):
        """
        Perform a Gauss fit to centroid the peaks for the property
        Code from http://pymzml.github.io/_modules/spec.html
        """
        peaks = []
        if 'm/z' in raw and 'intensity' in raw:
            intensity = raw['intensity']
            mz = raw['m/z']
            for i, ins in enumerate(intensity[:-1]):
                if i < 2:
                    continue
                if 0.0 < intensity[i - 1] < ins > intensity[i + 1] > 0.0:
                    x1 = mz[i - 1]
                    x2 = mz[i]
                    x3 = mz[i + 1]
                    y1 = intensity[i - 1]
                    y2 = intensity[i]
                    y3 = intensity[i + 1]
                    
                    if x2 - x1 > (x3 - x2) * 10 or (x2 - x1) * 10 < x3 - x2:
                        continue
                    if y3 == y1:
                        lower = 2
                        upper = 2
                        while (not 0 < y1 < y2 > y3 > 0) and \
                            y1 == y3 and upper < 10:
                            if i - lower < 0:
                                i_lower = 0
                            else:
                                i_lower = i - lower
                            if i + upper >= len(mz):
                                i_upper = len(mz) - 1
                            else:
                                i_upper = i + upper
                            x1 = mz[i_lower]
                            x3 = mz[i_upper]
                            y1 = intensity[i_lower]
                            y3 = intensity[i_upper]
                            if lower % 2 == 0:
                                upper += 1
                            else:
                                lower += 1
                    if not (0 < y1 < y2 > y3 > 0) or y1 == y3:
                        continue
                    try:
                        logratio = np.log(y2 / y1) / np.log(y3 / y1)
                        mu = (logratio * (x1**2 - x3**2) - x1**2 + x2**2) / \
                            (2 * (x2 - x1) - 2 * logratio* (x3 - x1))
                        csquared = (x2**2 - x1**2 - 2 * x2 * mu + \
                                2 * x1 * mu) / \
                            (2 * np.log(y1 / y2))
                        a = y1 * np.exp((x1 - mu)**2 / (2 * csquared))
                    except:
                        continue
                    peaks.append([mu, a])
        return np.array(peaks, dtype = np.float32)

    def find_peaks(self, scans, centroids, accuracy = 5, dope = False):
        """
        Finds peaks detected in consecutive scans.
        Returns list of dicts with centroid m/z,
        cumulative intensity and RT range for
        each peak.
        Returns array with 
        m/z, rt_min, rt_max, area, scan count
        as coulumns.
        
        accuracy : int, float
            Instrument accuracy in ppm.
            The difference over we consider 2
            m/z values to be distinct.
        
        # values order: (rt, mz, int)
        """
        accuracy = 1000000.0 / accuracy
        for scan, c in iteritems(centroids):
            centroids[scan]['centroids'] = \
                c['centroids'][c['centroids'][:,0].argsort(),:]
        peaks = []
        consecutive = {}
        def get_id():
            i = 0
            while True:
                yield i
                i += 1
        _peakid = get_id()
        prg = progress.Progress(len(scans), 'Processing scans', 1)
        for scan in scans:
            prg.step()
            if scan in centroids:
                _scan = centroids[scan]['centroids']
                rt = centroids[scan]['rt']
                added = set([])
                to_remove = set([])
                for peakid, cons in iteritems(consecutive):
                    ld = ud = 999.0
                    cons_mz = np.array(map(lambda x: x[1], cons))
                    cons_in = np.array(map(lambda x: x[2], cons))
                    cons_centroid_mz = np.sum(cons_mz * cons_in) / \
                        np.sum(cons_in)
                    ui = _scan[:,0].searchsorted(cons_centroid_mz)
                    if ui < _scan.shape[0]:
                        ud = _scan[ui, 0] - cons_centroid_mz
                    if ui > 0:
                        ld = cons_centroid_mz - _scan[ui - 1, 0]
                    ci = ui if ud < ld else ui - 1
                    if dope and _dope(cons_centroid_mz):
                        print('\n\t:: This mz looks DOPE: %f, closest in'\
                            ' this scan (%u): %f; accuracy: %f' % \
                            (cons_centroid_mz, scan, _scan[ci, 0],
                                cons_centroid_mz / accuracy))
                    if abs(_scan[ci, 0] - cons_centroid_mz) <= \
                        cons_centroid_mz / accuracy:
                        # same m/z detected in another consecutive scan:
                        consecutive[peakid].append(
                            (rt, _scan[ci, 0], _scan[ci, 1]))
                        if dope and _dope(_scan[ci, 0]):
                            print('\n\t:: DOPE found in next scan, '\
                                'appended to consecutive series: %u, %f' % \
                                    (scan, _scan[ci,0]))
                        # this index won't initiate a new consecutive series:
                        added.add(ci)
                    else:
                        # this consecutive series interrupted here,
                        # so move it to the peaks stack:
                        cons_rt = np.array(map(lambda x: x[0], cons))
                        peaks.append([
                            cons_centroid_mz,
                            min(cons_rt),
                            max(cons_rt),
                            cons_in.sum(),
                            len(cons)
                        ])
                        if dope and _dope(cons_centroid_mz):
                            print('\n\t:: DOPE series interrupted, moved to '\
                                'peaks stack: %f, %f' % \
                                (cons_centroid_mz, cons_in.sum()))
                        to_remove.add(peakid)
                for peakid in to_remove & set(consecutive.keys()):
                    del consecutive[peakid]
                for i, (mz, ins) in enumerate(_scan):
                    if i not in added:
                        peakid = _peakid.next()
                        consecutive[peakid] = [(rt, mz, ins)]
                        if dope and _dope(mz):
                            print('\n\t:: New DOPE found, '\
                                'new consecutive series started'\
                                ': scan %u, m/z = %f, peakid = %u' % \
                                (scan, mz, peakid))
        prg.terminate()
        peaks = np.array(peaks, dtype = np.float32)
        peaks = peaks[peaks[:,0].argsort(),:]
        return peaks

    def _dope(self, mz):
        return mz < 742.5331 and mz > 742.530

    def filter_peaks(self, peaks, min_scans = 3, min_area = 10000):
        peaks = peaks[np.where(
            np.logical_and(
                peaks[:,4] >= min_scans,
                peaks[:,3] >= min_area
            )
        )]
        peaks = peaks[peaks[:, 0].argsort(),:]
        return peaks

    def read_mzml(self, fname):
        prefix = '{http://psi.hupo.org/ms/mzml}'
        run = '%srun' % prefix
        spectrum = '%sspectrum' % prefix
        cvparam = '%scvParam' % prefix
        binarraylist = '%sbinaryDataArrayList' % prefix
        binarray = '%sbinaryDataArray' % prefix
        scanlist = '%sscanList' % prefix
        scan = '%sscan' % prefix
        mslevel = 'ms level'
        basepeakmz = 'base peak m/z'
        basepeakin = 'base peak intensity'
        starttime = 'scan start time'
        scans = []
        centroids = {}
        time = datetime.datetime.utcfromtimestamp(0)
        with open(fname, 'r') as f:
            sys.stdout.write('\t:: Reading file `%s`\n'%fname)
            mzml = lxml.etree.iterparse(f, events = ('end',))
            ms1 = False
            try:
                for ev, elem in mzml:
                    if elem.tag == run:
                        time = elem.attrib['startTimeStamp']
                        time = datetime.datetime.strptime(time,
                            '%Y-%m-%dT%H:%M:%SZ')
                    if elem.tag == spectrum:
                        ms1 = False
                        length = elem.attrib['defaultArrayLength']
                        scanid = int(elem.attrib['id'].split('=')[-1])
                        for cvp in elem.findall(cvparam):
                            if cvp.attrib['name'] == mslevel:
                                level = cvp.attrib['value']
                                if level != '1':
                                    break
                                else:
                                    ms1 = True
                            if cvp.attrib['name'] == basepeakmz:
                                mz = float(cvp.attrib['value'])
                            if cvp.attrib['name'] == basepeakin:
                                intensity = float(cvp.attrib['value'])
                        if ms1:
                            _scan = elem.find(scanlist).find(scan)
                            for cvp in _scan.findall(cvparam):
                                if cvp.attrib['name'] == starttime:
                                    rt = float(cvp.attrib['value'])
                            raw = {}
                            for bda in elem.find(binarraylist).\
                                findall(binarray):
                                name, arr = self._process_binary_array(bda, length)
                                raw[name] = arr
                            _centroids = self._centroid(raw)
                            scans.append(scanid)
                            centroids[scanid] = \
                                {'rt': rt, 'centroids': _centroids}
            except lxml.etree.XMLSyntaxError as error:
                sys.stdout.write('\n\tWARNING: XML processing error: %s\n' % \
                    str(error))
                sys.stdout.flush()
        return time, scans, centroids

    def dope(self, c):
        """
        Tells whether if a series of m/z's might be DOPE.
        """
        return np.any(np.logical_and(c[:,0] > 742.52, c[:,0] < 742.54))

    #
    # scanning the directory tree and collecting the MS data csv files
    #

    def get_filenames(self):
        """
        Files are placed in 2 root directories, in specific subdirectories,
        positive and negative MS mode in separate files.
        This function collects all the file paths and returns them 
        in a dict of dicts. Keys are LTP names, and 'pos'/'neg'.
        """
        redirname = re.compile(r'(^[0-9a-zA-Z]+)[ _]')
        fnames = {}
        datadlsts = \
            list(
                map(lambda d:
                    list(
                        filter(lambda dd:
                            os.path.isdir(os.path.join(d, dd)),
                            os.listdir(d)
                        )
                    ),
                    self.datadirs
                )
            )
        
        if sum(map(len, datadlsts)) == 0:
            sys.stdout.write('\t:: Please mount the shared folder!\n')
            return fnames
        
        for path, proteindl in zip(self.datadirs, datadlsts):
            
            for proteind in proteindl:
                
                protein = redirname.findall(proteind)
                pos = 'pos' if 'pos' in proteind \
                    else 'neg' if 'neg' in proteind else None
                if pos is not None and len(protein) > 0:
                    protein = protein[0]
                    fpath = [proteind, 'features']
                    for f in os.listdir(os.path.join(path, *fpath)):
                        if 'LABELFREE' in f or 'features' in f:
                            fpath.append(f)
                            break
                    for f in os.listdir(os.path.join(path, *fpath)):
                        if f.endswith('.csv'):
                            fpath.append(f)
                            break
                    if protein not in fnames:
                        fnames[protein] = {}
                    if pos not in fnames[protein] or proteind.endswith('update'):
                        fnames[protein][pos] = os.path.join(path, *fpath)
        self.datafiles = fnames

    def read_file_np(self, fname, read_vars = ['Normalized Area']):
        """
        Reads one MS file, returns numpy array,
        with void mask.
        Column order:
        quality, m/z, significance,
        rt-min, rt-max, charge, rtmean, avg.area,
        control, a9, a10, a11, a12, b1
        """
        # typos in the headers what need to be fixed
        typos = {
            'Sample 5': '',
            'Sample 6': '',
            '_ ': '_'
        }
        rehdr = re.compile(r'([_0-9a-zA-Z]+)[_\s]([/0-9a-zA-Z\s]+)')
        refra = re.compile(r'.*_[0-9]??([a-zA-Z]?)([0-9]{1,2}).*')
        retyp = re.compile(r'(' + '|'.join(typos.keys()) + r')')
        # order of fractions (fractions)
        prows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        # variable names
        vname = {
            'm/z': 0,
            'RT mean': 1,
            'Normalized Area': 2
        }
        # the dict which will be returned contains the
        # variables to be read, and the annotations
        data = dict(map(lambda k: (k, []), vname.keys()))
        data['annot'] = []
        #
        rtmkey = ('RT', 'mean')
        rtmcol = None
        with open(fname, 'r') as f:
            hdr = retyp.sub(lambda x:
                    typos[x.group()],
                f.readline().replace('"', '')
            ).split(',')[1:]
            for i, h in enumerate(hdr):
                this_hdr = rehdr.match(h)
                if this_hdr is not None and this_hdr.groups(0) == rtmkey:
                    rtmcol = i
            if rtmcol is None:
                sys.stdout.write('\t:: Could not find RT mean'
                                    ' column in file %s\n' % fname)
                sys.stdout.flush()
            cols = [tuple([i] + [x.strip() \
                    for x in list(rehdr.match(h).groups(0))]) \
                for i, h in enumerate(hdr[6:-7])]
            
            fracs = []
            prow_prev, pcol_prev = None, None
            for c in cols:
                if 'ctrl' in c[1] or 'ctlr' in c[1] or 'secbuffer' in c[1]:
                    prow, pcol = ('X', '0')
                else:
                    if self.fix_fraction_name_ab_typo:
                        c1field = c[1].split('_')
                        c1field[-1] = c1field[-1].replace('ab', 'b')
                        if len(c1field) > 2:
                            c1field[-2] = c1field[-2].replace('ab', 'b')
                        c1field = '_'.join(c1field)
                    else:
                        c1field = c[1]
                    
                    prow, pcol = refra.match(c1field).groups()
                    
                    if not len(prow):
                        if pcol_prev is not None:
                            if int(pcol_prev) >= int(pcol) - 1:
                                prow = prow_prev
                            else:
                                prow = prows[prows.index(prow_prev) + 1]
                    
                    prow_prev, pcol_prev = prow, pcol
                
                fracs.append((prow.upper(), pcol))
            
            # col nums for each variable, for each fraction
            scols = dict([(var, dict([(i, None) for i in fracs])) \
                for var in vname.keys()])
            
            # testing if order of columns is always
            # m/z, RT, norm area; all files passed
            for i, c in enumerate(cols):
                if c[0] % 3 != vname[c[2]]:
                    sys.stdout.write('erroneous column order: col %u '\
                        'with header %s\n\tin file %s\n' % (c[0], c[2], fname))
                fnum = fracs[i]
                # assigning columns to variable->fraction slots
                # col offsets from 6
                scols[c[2]][fnum] = c
            
            # now sorting fractions:
            fracs = sorted(set(fracs), key = lambda f: (f[0], int(f[1])))
            
            for l in f:
                # removing first column
                l = [i.strip() for i in l.replace('"', '').split(',')][1:]
                # reading annotations
                annot = \
                    [
                        self.to_float(l[0]),
                        self.to_float(l[1]),
                        self.to_float(l[2])
                    ] + \
                    [self.to_float(i.strip()) for i in l[3].split('-')] + \
                    [
                        self.to_float(l[4]),
                        self.to_float(l[rtmcol]),
                        self.to_float(l[5])
                    ]
                
                data['annot'].append(annot)
                
                ldata = {}
                for fr in fracs:
                    # number of columns depends on which variables we need
                    for var in read_vars:
                        if var not in ldata:
                            ldata[var] = []
                        
                        cnum = scols[var][fr][0]
                        
                        if cnum is None:
                            ldata[var].append(np.nan)
                        else:
                            # col offset is 6 !
                            ldata[var].append(self.to_float(l[cnum + 6]))
                
                for var in read_vars:
                    data[var].append(ldata[var])
        
        data = dict(map(lambda d: (d[0], np.array(d[1], dtype = np.float64)), iteritems(data)))
        mzsort = np.argsort(data['annot'][:,2])
        datalen = data['annot'].shape[0]
        data = \
            dict(
                map(
                    lambda d:
                        (d[0], d[1][mzsort]),
                    filter(
                        lambda d:
                            d[1].shape[0] == datalen,
                        iteritems(data)
                    )
                )
            )
        
        return data, fracs
    
    @classmethod
    def sort_fracs_str(cls, frs):
        """
        Sorts fraction labels in a standard way, i.e. first by
        their letter code, after their number as numeric value.
        Accepts list of strings, returns list of strings.
        """
        return (
            list(
                map(
                    lambda fr:
                        '%s%u' % fr[1],
                    cls.sort_fracs_tuple(
                        map(lambda fr:
                                (fr[0], fr[1:]),
                            frs
                        )
                    )
                )
            )
        )
    
    @staticmethod
    def sort_fracs_tuple(frs):
        """
        Sorts fraction labels in a standard way, i.e. first by
        their letter code, after their number as numeric value.
        Accepts list of tuples, returns list of tuples.
        Numbers can be either string or integer.
        """
        frs = map(lambda fr: (fr[0], int(fr[1])), frs)
        return sorted(enumerate(frs),
                      key = lambda x: (x[1][0], x[1][1]))
    
    def read_data(self):
        """
        Iterates through dict of dicts with file paths, reads
        each file, and makes array view and masks to make it easy
        to handle missing data, and access m/z values of fractions 
        and controls, and other values.
        """
        
        def get_indices(sfracs, cfracs, fun):
            return \
                np.array(
                    list(
                        map(
                            lambda fr:
                                (
                                    fr != 'X0' and
                                    fr in cfracs and
                                    fun(cfracs[fr])
                                ),
                            sfracs
                        )
                    )
                )
        
        pfracs = {}
        data = {}
        prg = progress.Progress(len(self.datafiles) * 2,
                                'Reading features (MS1 data)', 1)
        
        for _protein, pos_neg in iteritems(self.datafiles):
            
            protein = _protein.upper()
            if protein not in data:
                data[protein] = {}
            
            if hasattr(self, 'fractions'):
                cfracs = self.fractions[protein]
            else:
                cfracs = None
            
            for p, fname in iteritems(pos_neg):
                
                prg.step()
                
                #try:
                data[protein][p] = {}
                # this returns an array and the fraction sequence:
                pdata, fracs = self.read_file_np(fname)
                #
                ifracs, sfracs = zip(*self.sort_fracs_tuple(fracs))
                sfracs = list(map(lambda fr: '%s%s' % fr, sfracs))
                pfracs[protein] = np.array(sfracs)
                # rearrange columns according to fraction sequence:
                data[protein][p]['raw'] = \
                    pdata['Normalized Area'][:,np.array(ifracs)]
                #
                # another thing we need are the annotations:
                data[protein][p]['ann'] = pdata['annot']
                #
                # making a view with the intensities:
                # (this would not be necessary any more,
                # it was when these were in the same array
                # as annotations)
                data[protein][p]['int'] = data[protein][p]['raw']
                #
                
                if cfracs is not None:
                    # making a view with measured:
                    imes = get_indices(sfracs, cfracs, lambda x: x is not None)
                    
                    data[protein][p]['mes'] = data[protein][p]['int'][:,imes]
                    #
                    # making a view with the controls:
                    ictr = get_indices(sfracs, cfracs, lambda x: x == 0)
                    
                    data[protein][p]['ctr'] = data[protein][p]['int'][:,ictr]
                    #
                    # fractions with lipids:
                    
                    ilip = get_indices(sfracs, cfracs, lambda x: x == 1)
                    
                    data[protein][p]['lip'] = data[protein][p]['int'][:,ilip]
                    #
                    # all fractions except blank control:
                    if sfracs[-1] == 'X0':
                        imes[-1] = False
                    
                    data[protein][p]['smp'] = data[protein][p]['int'][:,imes]
                
                # average area:
                data[protein][p]['aa'] = data[protein][p]['ann'][:,7]
        
        prg.terminate()
        
        if not hasattr(self, 'data'):
            self.data = {}
        if not hasattr(self, 'pfracs'):
            self.pfracs = {}
        
        self.data.update(data)
        self.pfracs.update(pfracs)
    
    #
    # Generic helper functions
    #

    def to_float(self, num):
        """
        Extracts float from string, or returns None.
        """
        if type(num) is float or type(num) is np.float64:
            return num
        num = num.strip()
        renum = re.compile(r'([-]?[0-9]*[\.]?[0-9]+[eE]?[-\+]?[0-9]*)')
        fnum = renum.match(num)
        if fnum:
            return float(fnum.groups(0)[0])
        else:
            if type(num) is str:
                if num.lower() == 'inf':
                    return np.inf
                if num.lower() == '-inf':
                    return -np.inf
            else:
                return None

    def to_int(self, num):
        """
        Extracts int from string or returns None.
        """
        renum = re.compile(r'([-]?[0-9]+[\.]?[0-9]*)')
        num = renum.match(num.strip())
        if num:
            return int(num.groups(0)[0])
        else:
            return num

    def float_lst(self, l):
        """
        Converts elements of a list to floats.
        """
        return [self.to_float(x) for x in l]

    """
    Filtering functions
    """

    def quality_filter(self, threshold = 0.2):
        for protein, d in iteritems(self.data):
            for pn, tbl in iteritems(d):
                tbl['qly'] = np.array(tbl['ann'][:,0] >= threshold)

    def charge_filter(self, charge = 1):
        for protein, d in iteritems(self.data):
            for pn, tbl in iteritems(d):
                tbl['crg'] = np.array(tbl['ann'][:,5] == charge)
    
    def rt1_filter(self, rtmin = 1.0):
        for protein, d in iteritems(self.data):
            for pn, tbl in iteritems(d):
                tbl['rtm'] = tbl['ann'][:,6]
                tbl['rt1'] = np.array(tbl['rtm'] > rtmin)

    def area_filter(self, area = 10000.0):
        area = float(area)
        for protein, d in iteritems(self.data):
            for mode, tbl in iteritems(d):
                if self.use_original_average_area:
                    try:
                        with np.errstate(divide = 'ignore', invalid = 'ignore'):
                            tbl['are'] = \
                                np.mean(np.nan_to_num(tbl['lip']), 1) >= area
                    except:
                        print(protein, mode)
                else:
                    tbl['are'] = tbl['aa'] >= area

    def peaksize_filter(self, peakmin = 2.0, peakmax = 5.0, area = 10000):
        # minimum and maximum of all intensities over all proteins:
        
        prg = progress.Progress(len(self.data) * 2, 'Peak size filter', 1)
        
        for protein, d in iteritems(self.data):
            
            for mode, tbl in iteritems(d):
                
                prg.step()
                mini = np.min(
                    np.nanmax(
                        tbl['int'][
                            np.mean(
                                np.nan_to_num(
                                    tbl['lip']),
                                1
                            ) > 10000,
                        :], 1
                    )
                )
                
                maxi = np.max(np.nan_to_num(tbl['int']))
                tbl['peaksize'] = np.max(np.nan_to_num(tbl['lip']), 1) / \
                    (np.max(np.nan_to_num(tbl['ctr']), 1) + 0.001)
                tbl['pslim02'] = (((np.max(np.nan_to_num(tbl['lip']), 1) -
                                    mini) / (maxi - mini)) *
                        (2.0 - 2.0) + peakmin)
                tbl['pslim05'] = (((np.max(np.nan_to_num(tbl['lip']), 1) -
                                    mini)) / (maxi - mini) *
                        (5.0 - 2.0) + peakmin)
                tbl['pslim10'] = (((np.max(np.nan_to_num(tbl['lip']), 1) -
                                    mini) / (maxi - mini)) *
                        (10.0 - 2.0) + peakmin)
                tbl['pslim510'] = (((np.max(np.nan_to_num(tbl['lip']), 1) -
                                     mini) / (maxi - mini)) *
                        (10.0 - 5.0) + 5.0)
                
                tbl['pks'] = (
                    # the peaksize:
                    np.max(np.nan_to_num(tbl['lip']), 1) / \
                        (np.max(np.nan_to_num(tbl['ctr']), 1) + 0.001)
                    # must be larger:
                    >
                    # than the min peaksize as a function of intensity:
                    (((np.max(np.nan_to_num(tbl['lip']), 1) - mini) / (maxi - mini)) *
                        (peakmax - peakmin) + peakmin)
                )
        
        prg.terminate()

    def cprofile_filter(self):
        profile_filter(prfx = 'c')

    def profile_filter(self, prfx = ''):
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        notf = []
        cols = 'smp' if prfx == 'c' else 'lip'
        prg = progress.Progress(len(data) * 2, 'Profile filter', 1,
            percent = False)
        for protein, d in iteritems(self.data):
            for pn, tbl in iteritems(d):
                prg.step()
                if protein.upper() not in self.pprofs:
                    notf.append(protein)
                    continue
                # i != 0 : we always drop the blank control
                ppr = np.array([self.pprofs[protein.upper()][frs[i]] \
                    for i, fr in enumerate(self.fractions[protein]) \
                        if fr == 1 and i != 0])
                ppr = self.norm_profile(ppr).astype(np.float64)
                prr = stats.rankdata(ppr)
                pranks = sorted((i for i, x in enumerate(prr)),
                    key = lambda x: x, reverse = True)
                if tbl[cols][0,:].count() > 1:
                    flatp = ppr[pranks[0]] - ppr[pranks[1]] < \
                        ppr[pranks[0]] * 0.1
                # 0 if have only one fraction in sample
                tbl['%sprf'%prfx] = np.apply_along_axis(
                    lambda x: self.diff_profiles(ppr, norm_profile(x)),
                        axis = 1, arr = tbl[cols])
                # all True if have only one fraction in sample
                tbl['%srpr'%prfx] = np.array([True] * tbl[cols].shape[0]) \
                    if tbl[cols].shape[1] == 1 else \
                    np.apply_along_axis(
                    lambda x: self.comp_profiles(ppr, 
                        self.norm_profile(x), prr, flatp),
                        axis = 1, arr = tbl[cols])
        prg.terminate()
        sys.stdout.write('No protein profiles found for %s\n\n' % ', '.join(notf))
        sys.stdout.flush()

    def diff_profiles(self, p1, p2):
        # profiles are numpy arrays
        # of equal length
        if len(p1) > 1:
            return np.nansum(np.abs(p1 - p2))
        else:
            return 1.0 * len(p2)

    def _comp_profiles(self, p1, p2):
        p1r = stats.rankdata(p1)
        pranks = sorted((i for i, x in enumerate(p1r)), 
            key = lambda x: x, reverse = True)
        flatp = p1r[pranks[0]] - p1r[pranks[1]] < \
            p1r[pranks[0]] * 0.1
        return self.comp_profiles(p1, p2, p1r, flatp), 0.0

    def comp_profiles(self, p1, p2, p1r = False, flatp = False):
        highest = False
        flat = False
        hollow = False
        p2r = stats.rankdata(p2)
        if type(p1r) is not np.ndarray:
            p1r = stats.rankdata(p1)
        highest = np.any(np.logical_and(p1r == 1, p2r == 1))
        p2ranks = sorted((i for i, x in enumerate(p2r)), 
            key = lambda x: x, reverse = True)
        if len(p1r) == 3:
            hollow = p1r[1] == 2 or p2r[1] == 2
        if not highest:
            flat = flatp and sum(p2) > 0.0 and \
                p2[p2ranks[0]] - p2[p2ranks[1]] < \
                    p2[p2ranks][0] * 0.1
        return not hollow and (highest or flat)

    def norm_profile(self, profile):
        return (profile - np.nanmin(profile.astype(np.float64))) / \
            (np.nanmax(profile) - np.nanmin(profile))

    def norm_profiles(self, tbl):
        if type(tbl) == np.ma.core.MaskedArray:
            tbl = tbl.data
        with np.errstate(divide = 'ignore', invalid = 'ignore'):
            return (tbl - np.nanmin(tbl, axis = 1, keepdims = True)) / \
                (np.nanmax(tbl, axis = 1, keepdims = True) - \
                    np.nanmin(tbl, axis = 1, keepdims = True))

    def ubiquity_filter_old(self, only_valid = True):
        prg = progress.Progress(len(data)**2, 'Ubiquity filter', 1)
        proteins = sorted(self.data.keys())
        for pn in ['pos', 'neg']:
            for i, protein1 in enumerate(proteins):
                for protein2 in proteins[i+1:]:
                    protein1s = set([])
                    protein2s = set([])
                    prg.step()
                    if pn in data[protein1] and pn in data[protein2]:
                        protein1t = data[protein1][pn]
                        protein2t = data[protein2][pn]
                        if 'ubi' not in protein1t:
                            protein1t['ubi'] = \
                                np.zeros([protein1t['raw'].shape[0], 2], \
                                dtype = np.int8)
                        if 'ubi' not in protein2t:
                            protein2t['ubi'] = \
                                np.zeros([protein2t['raw'].shape[0], 2], \
                                dtype = np.int8)
                        for i1, mz1 in np.ndenumerate(protein1t['raw'][:,2]):
                            i2u = protein2t['raw'][:,2].searchsorted(mz1)
                            if i2u < protein2t['raw'].shape[0] and \
                                protein2t['raw'][i2u,2] - mz1 <= self.ms1_tlr:
                                if i1 not in protein1s:
                                    protein1t['ubi'][i1] += 1
                                    protein1s.add(i1)
                                if i2u not in protein2s:
                                    protein2t['ubi'][i2u] += 1
                                    protein2s.add(i2u)
                            if mz1 - protein2t['raw'][i2u - 1,2] <= self.ms1_tlr:
                                if i1 not in protein1s:
                                    protein1t['ubi'][i1] += 1
                                    protein1s.add(i1)
                                if i2u - 1 not in protein2s:
                                    protein2t['ubi'][i2u - 1] += 1
                                    protein2s.add(i2u - 1)
        prg.terminate()

    def ubiquity_filter(self, only_valid = True):
        prg = progress.Progress(len(self.valids)**2, 'Ubiquity filter', 1)
        proteins = sorted(self.valids.keys())
        for pn in ['pos', 'neg']:
            for i, protein1 in enumerate(proteins):
                for protein2 in proteins[i+1:]:
                    protein1s = set([])
                    protein2s = set([])
                    prg.step()
                    if pn in self.valids[protein1] and pn in self.valids[protein2]:
                        protein1t = self.valids[protein1][pn]
                        protein2t = self.valids[protein2][pn]
                        if 'ubi' not in protein1t:
                            protein1t['ubi'] = np.zeros(protein1t['mz'].shape[0],
                                dtype = np.int8)
                        if 'ubi' not in protein2t:
                            protein2t['ubi'] = np.zeros(protein2t['mz'].shape[0],
                                dtype = np.int8)
                        for i1, mz1 in np.ndenumerate(protein1t['mz']):
                            if protein1t['cpv'][i1] or not only_valid:
                                i2u = protein2t['mz'].searchsorted(mz1)
                                u = 0
                                while True:
                                    if i2u + u < protein2t['mz'].shape[0] and \
                                        protein2t['mz'][i2u + u] - mz1 <= \
                                        self.ms1_tlr:
                                        if protein2t['cpv'][i2u + u] or \
                                            not only_valid:
                                            if i1 not in protein1s:
                                                protein1t['ubi'][i1] += 1
                                                protein1s.add(i1)
                                            if i2u + u not in protein2s:
                                                protein2t['ubi'][i2u + u] += 1
                                                protein2s.add(i2u + u)
                                        u += 1
                                    else:
                                        break
                                l = 1
                                while True:
                                    if i2u - l >= 0 and \
                                        mz1 - protein2t['mz'][i2u - l] <= \
                                        self.ms1_tlr:
                                        if protein2t['cpv'][i2u - l] or \
                                            not only_valid:
                                            if i1 not in protein1s:
                                                protein1t['ubi'][i1] += 1
                                                protein1s.add(i1)
                                            if i2u - l not in protein2s:
                                                protein2t['ubi'][i2u - l] += 1
                                                protein2s.add(i2u - l)
                                        l += 1
                                    else:
                                        break
        prg.terminate()
    
    def remove_filter(self, attr_name):
        """
        Deletes the result of a previously done filtering.
        """
        for protein, d in iteritems(self.data):
            for pn, tbl in iteritems(d):
                if attr_name in tbl:
                    del tbl[attr_name]
    
    def rprofile_filter(self):
        """
        Dummy function for eval_filter()
        """
        pass
    
    @get_hits
    def val_ubi_prf_rpr_hits(self, tbl, ubiquity = 7, treshold = 0.15,
        tresholdB = 0.25, profile_best = False):
        """
        Column order:
        quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1,
        profile_score, control_profile_score, rank_profile_boolean,
        ubiquity_score, ubiquity_score,
        original_index
        """
        prf = 'cprf' if tbl['lip'][1,:].count() == 1 else 'prf'
        _treshold = (treshold, tresholdB)
        if prf not in tbl:
            return None
        # dict of profile matching score values and their indices
        selected = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'], tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['rpr']
                ),
                tbl['ubi'] <= ubiquity
            ),
            np.logical_or(
                np.logical_not(
                    np.isnan(tbl['prf'])
                ),
                np.logical_not(
                    np.isnan(tbl['cprf'])
                )
            )
        )
        prf_values = dict(zip(np.where(selected)[0],
            zip(tbl['prf'][selected], tbl['cprf'][selected])))
        if profile_best:
            # select the best scoring records
            prf_values = sorted(prf_values.items(), key = lambda x: x[1])
            try:
                score_treshold = prf_values[min(profile_best - 1,
                    len(prf_values) - 1)][1]
            except IndexError:
                print(prf_values)
                return None
            # comments?
            indices = np.array(sorted(map(lambda x: x[0], \
                filter(lambda x: x[1] <= score_treshold, prf_values))))
        else:
            # or select the records with profile
            # match less then or equal to treshold
            indices = np.array(sorted((i for i, v in iteritems(prf_values) \
                if v <= _treshold)))
        return (tbl['raw'][indices,:], tbl['prf'][indices],
            tbl['cprf'][indices],
            tbl['rpr'][indices], tbl['ubi'][indices], tbl['uby'][indices] \
                if 'uby' in tbl else np.zeros(len(indices)), indices)

    @get_hits
    def pass_through(self, tbl, ubiquity = 7, treshold = 0.15,
        tresholdB = 0.25, profile_best = False):
        """
        Column order:
        quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1,
        profile_score, control_profile_score, rank_profile_boolean,
        ubiquity_score, ubiquity_score
        """
        prf = 'cprf' if tbl['lip'][1,:].count() == 1 else 'prf'
        _treshold = (treshold, tresholdB)
        if prf not in tbl:
            return None
        # dict of profile matching score values and their indices
        selected = np.logical_or(
            np.logical_not(
                np.isnan(tbl['prf'])
            ),
            np.logical_not(
                np.isnan(tbl['cprf'])
            )
        )
        prf_values = dict(zip(np.where(selected)[0],
            zip(tbl['prf'][selected], tbl['cprf'][selected])))
        if profile_best:
            # select the best scoring records
            prf_values = sorted(prf_values.items(), key = lambda x: x[1])
            try:
                score_treshold = prf_values[min(profile_best - 1,
                    len(prf_values) - 1)][1]
            except IndexError:
                print(prf_values)
                return None
            # comments?
            indices = np.array(sorted(map(lambda x: x[0], \
                filter(lambda x: x[1] <= score_treshold, prf_values))))
        else:
            # or select the records with profile
            # match less then or equal to treshold
            indices = np.array(sorted((i for i, v in iteritems(prf_values) \
                if v <= _treshold)))
        print('Selected: ', np.nansum(tbl['cprf'][indices]))
        print('Total: ', np.nansum(tbl['cprf']))
        return (tbl['raw'][indices,:], tbl['prf'][indices],
            tbl['cprf'][indices],
            tbl['rpr'][indices], tbl['ubi'][indices], tbl['uby'][indices] \
                if 'uby' in tbl else np.zeros(len(indices)))
    
    @combine_filters_valids
    def logical_and(self, tbl, one, two, result):
        tbl[result] = np.logical_and(tbl[one], tbl[two])
    
    @combine_filters_valids
    def logical_or(self, tbl, one, two, result):
        tbl[result] = np.logical_or(tbl[one], tbl[two])
    
    @combine_filters_valids
    def logical_not(self, tbl, attr, result):
        tbl[result] = np.logical_not(tbl[attr])
    
    @combine_filters
    def validity_filter(self, tbl):
        tbl['vld'] = \
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            tbl['qly'],
                            tbl['crg']
                        ),
                        tbl['are']
                    ),
                    tbl['pks']
                ),
                tbl['rt1']
            )

    @combine_filters
    def val_ubi_filter(self, tbl, ubiquity = 7):
        tbl['vub'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['ubi'][:,0] <= ubiquity
        )
    
    @combine_filters
    def val_prf_filter(self, tbl, treshold = 0.15):
        tbl['vpr'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['prf'] <= treshold
        )
    
    @combine_filters
    def val_rpr_filter(self, tbl):
        tbl['vrp'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['rpr']
        )
    
    @combine_filters
    def val_ubi_prf_filter(self, tbl, ubiquity = 7, treshold = 0.15):
        tbl['vup'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            tbl['qly'],
                            tbl['crg']
                        ),
                        tbl['are']
                    ),
                    tbl['pks']
                ),
                tbl['prf'] <= treshold
            ), tbl['ubi'][:,0] <= ubiquity
        )

    @combine_filters
    def val_ubi_prf_rprf_filter(self, tbl, ubiquity = 7, treshold = 0.15):
        tbl['vur'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'],
                                tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['rpr']
                ),
                tbl['prf'] <= treshold
            ),
            tbl['ubi'][:,0] <= ubiquity
        )
    
    def apply_filters(self,
        filters = ['quality', 'charge', 'area', 'peaksize', 'rt1'],
        param = {}):
        for f in filters:
            p = param[f] if f in param else {}
            fun = getattr(self, '%s_filter' % f)
            fun(**p)
    
    def empty_dict(self):
        return dict((protein, {'pos': None, 'neg': None}) \
            for protein in self.data.keys())
    
    def eval_filter(self, filtr, param = {}, runtime = True,
        repeat = 3, number = 10, hit = lambda x: x):
        t = None
        fun = globals()['%s_filter' % filtr]
        filters = {
                'quality': 'qly',
                'charge': 'crg',
                'peaksize': 'pks',
                'ubiquity': 'ubi',
                'area': 'are',
                'profile': 'prf',
                'rprofile': 'rpr',
                'cprofile': 'cprf',
                'validity': 'vld',
                'val_ubi': 'vub',
                'val_prf': 'vpr',
                'val_ubi_prf': 'vup',
                'val_ubi_prf_rprf': 'vur',
                'val_rpr': 'vrp'
            }
        name = filters[filtr]
        if runtime:
            t = min(timeit.Timer(lambda: fun(self.data, **param)).\
                repeat(repeat = repeat, number = number)) / float(number)
        @count_hits
        def counter(self, tbl, name = None):
            return np.nansum(hit(tbl[name]))
        hits, phits = counter(self.data, name = name)
        nums = np.array(list(j for i in  hits.values() for j in i.values()),
            dtype = np.float)
        pcts = np.array(list(j for i in  phits.values() for j in i.values()),
            dtype = np.float)
        return t, hits, phits, (np.nanmin(nums), np.nanmax(nums)), \
            (np.nanmin(pcts), np.nanmax(pcts))

    def combined_hits(self, profile = 0.15,
        ubiquity = 20, verbose = True):
        @count_hits
        def counter(self, tbl):
            return np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'],
                                tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['prf'] <= profile
                ),
                tbl['ubi'][:,0] <= ubiquity
            ).sum()
        hits = counter(self.data)
        if verbose:
            sys.stdout.write('\n\tprotein\t\t+\t-\n\t%s\n'%('='*30))
            for protein in sorted(hits.keys()):
                sys.stdout.write('\t%s\t'%protein)
                for pn in ['pos', 'neg']:
                    num = hits[protein][pn]
                    sys.stdout.write('\t%s' % \
                        (str(num) if num is not None else 'n/a'))
                sys.stdout.write('\n')
        return hits

    """
    Save and reload functions
    """

    def save_data(self, fname = None):
        """
        Saves the raw data into pickle.
        """
        fname = os.path.join(self.basedir, self.featurescache) \
            if fname is None else fname
        sys.stdout.write('\t:: Saving data to %s ...\n' % fname)
        sys.stdout.flush()
        with open(fname, 'wb') as fp:
            pickle.dump(self.data, fp)
        
        sys.stdout.write('\t:: Data has been saved to %s.\n' % fname)
        sys.stdout.flush()

    def load_data(self, fname = None):
        fname = os.path.join(self.basedir, self.featurescache) \
            if fname is None else fname
        sys.stdout.write('\t:: Loading data from %s ...\n' % fname)
        sys.stdout.flush()
        with open(fname, 'rb') as fp:
            self.data = pickle.load(fp)

    def save(self, fname = None):
        """
        Actually this does not save anything important,
        it was useful only some time before.
        """
        fname = os.path.join(self.basedir, self.auxcache) \
            if fname is None else fname
        sys.stdout.write('\t:: Saving auxiliary data to %s ...\n' % fname)
        sys.stdout.flush()
        with open(fname, 'wb') as fp:
            pickle.dump((self.datafiles, self.fractions), fp)
        
        sys.stdout.write('\t:: Some annotations has been saved to %s.\n' % fname)
        sys.stdout.flush()

    def load(self, fname = None):
        fname = os.path.join(self.basedir, self.auxcache) \
            if fname is None else fname
        sys.stdout.write('\t:: Loading auxiliary data '\
            'from %s ...\n' % fname)
        sys.stdout.flush()
        with open(fname, 'rb') as fp:
            self.datafiles, self.fractions = \
                pickle.load(fp)

    """
    END: save & reload
    """

    """
    Lipid databases lookup functions
    """

    def find_lipids(self, hits, lipnames):
        """
        Column order:
        
        in:
        [0] quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1
        [1] profile_score
        [2] control_profile_score
        [3] rank_profile_boolean
        [4] ubiquity_score
        [5] ubiquity_score
        
        out:
        protein_name, m/z, 
        profile_score, control_profile_score,
        rank_profile_boolean, ubiquity_score,
        ubiquity_score, swisslipids_ac, level,
        lipid_name, lipid_formula, adduct, 
        adduct_m/z
        """
        # levels: 'Structural subspecies', 'Isomeric subspecies',
        # 'Species', 'Molecular subspecies'
        lipids = dict((protein.upper(), {}) for protein in hits.keys())
        for protein, d in iteritems(hits):
            for pn, tbl in iteritems(d):
                adducts = self.pAdducts if pn == 'pos' else self.nAdducts
                result = []
                if tbl[0] is not None:
                    for i in xrange(tbl[0].shape[0]):
                        swlipids = self.adduct_lookup(tbl[0][i,1], adducts)
                        if swlipids is not None:
                            for lip in swlipids:
                                hg, p_add, n_add, ether = \
                                    self.headgroup_from_lipid_name(lip)
                                fa = self.fattyacid_from_lipid_name(lip)
                                result.append(np.concatenate(
                                    # tbl[1] and tbl[2] are the profile 
                                    # and cprofile scores
                                    (np.array([protein.upper(), tbl[0][i,1],
                                        tbl[1][i], tbl[2][i],
                                        tbl[3][i], tbl[4][i],
                                        tbl[5][i]],
                                        dtype = np.object),
                                    lip,
                                    np.array(hg, fa, dtype = np.object)),
                                    axis = 0))
                    lipids[protein.upper()][pn] = np.vstack(
                        sorted(result, key = lambda x: x[1]))
        self.lipids = lipids

    def headgroup_from_lipid_name(self, lip):
        """
        For one database record attempts to identify the lipid class
        by looking up keywords.
        Calls greek name identification, greek fatty acid names are
        identified as 'FA'.
        """
        
        db = 'lmp' if lip[0][0] == 'L' else 'swl'
        for shortname, spec in iteritems(self.lipnames):
            for kwset in spec[db]:
                matched = [kw in lip[2] for kw in kwset['pos']]
                if sum(matched) == len(kwset['pos']) and \
                    sum(matched) > 0:
                    matched = [kw in lip[2] for kw in kwset['neg']]
                    if sum(matched) == 0:
                        ether = '(O-' in lip[2]
                        return shortname, spec['pos_adduct'], spec['neg_adduct'], ether
        
        return self.process_fa_name(lip)
    
    def process_fa_name(self, lip):
        """
        Identifies fatty acids based on their greek name.
        """
        
        if lip[2] in fa_greek:
            return 'FA', None, None, False
        
        return None, None, None, None
    
    def fattyacid_cc(self, names):
        """
        Returns carbon count and unsaturation from greek name.
        """
        
        for name in names:
            if name in fa_greek:
                return list(fa_greek[name])

    def fattyacid_from_lipid_name(self, lip, _sum = True):
        """
        Returns lipid carbon count and unsaturation for almost
        any name of common species in the databases.
        Handles greek names, e.g. `octadecenoate` --> [18,1]
        """
        
        _fa = None
        
        if lip[0][0] == 'L':
            names = lip[2].split('|')
            for i in [0, 2, 1]:
                fa = self.refa.findall(names[i])
                if len(fa) > 0:
                    _fa = fa
                    break
        
        elif lip[0][0] == 'S':
            fa = self.refa.findall(lip[2])
            if len(fa) > 0:
                _fa = fa
        
        if _fa is not None and _sum:
            _fa = list(map(sum, zip(*[[int(i) for i in f[1].split(':')] \
                for f in _fa])))
        
        if _fa is None:
            _fa = self.fattyacid_cc(lip[2].split('|'))
        
        return _fa

    def find_lipids_exact(self, verbose = False,
        outfile = None, proteins = None, charge = 1):
        """
        Looks up lipids by m/z among database entries in
        `exacts`, and stores the result in dict under key
        `lip`, where keys are the original indices (`i`).
        """
        
        if verbose:
            outfile = sys.stdout if outfile is None else open(outfile, 'w')
        for protein, d in iteritems(self.valids):
            if proteins is None or protein in proteins:
                for pn, tbl in iteritems(d):
                    tbl['lip'] = {}
                    for i in xrange(tbl['mz'].shape[0]):
                        tbl['lip'][tbl['i'][i]] = self.adduct_lookup_exact(
                            tbl['mz'][i], pn,
                            verbose, outfile, charge = 1
                        )
        if hasattr(outfile, 'close') and outfile != sys.stdout:
            outfile.close()

    def adduct_lookup(self, mz, adducts):
        """
        Looks up m/z values in the table containing the reference database
        already converted to a pool of all possible adducts.
        (Does not convert the m/z to other adducts.)
        """
        result = []
        iu = adducts[:,-1].searchsorted(mz)
        if adducts.shape[0] > iu:
            u = 0
            while True:
                if adducts[iu + u,-1] - mz <= self.ms1_tlr:
                    if adducts[iu + u,1] in self.swl_levels:
                        result.append(adducts[iu + u,:])
                    u += 1
                else:
                    break
        if iu > 0:
            l = 1
            while True:
                if iu - l >= 0 and mz - adducts[iu - l,-1] <= self.ms1_tlr:
                    if adducts[iu - l,1] in self.swl_levels:
                        result.append(adducts[iu - l,:])
                    l += 1
                else:
                    break
        return None if len(result) == 0 else np.vstack(result)

    def adduct_lookup_exact(self, mz, mode, verbose = False,
        outfile = None, charge = 1):
        """
        Looks up m/z values in the table containing the reference database
        casting the m/z to specific adducts.
        """
        if verbose and outfile is None:
            outfile = sys.stdout
        result = []
        adducts = self.ad2ex[charge][mode]
        for addName, addFun in iteritems(adducts):
            addMz = getattr(mzmod.Mz(mz), addFun)()
            if verbose:
                outfile.write('\t:: Searching for %s adducts.\n\t '\
                    '-- Adduct mass (measured): '\
                    '%.08f, exact mass (calculated): %.08f.\n' % \
                    (addName, mz, addMz))
            iu = self.exacts[:,-1].searchsorted(addMz)
            if self.exacts.shape[0] > iu:
                u = 0
                while True:
                    if self.exacts[iu + u,-1] - addMz <= self.ms1_tlr:
                        if self.exacts[iu + u,1] in self.swl_levels:
                            lip = self.exacts[iu + u,:]
                            hg, p_add, n_add, ether = \
                                self.headgroup_from_lipid_name(lip)
                            fa = self.fattyacid_from_lipid_name(lip)
                            lip = np.concatenate((lip,
                                np.array([addMz, hg, fa, ether],
                                    dtype = np.object)),
                                axis = 0)
                            lip[4] = addName
                            if verbose:
                                outfile.write('\t -- Found: %s\n' % \
                                    str(list(lip)))
                            result.append(lip)
                        u += 1
                    else:
                        break
            if iu > 0:
                l = 1
                while True:
                    if iu - l >= 0 and addMz - self.exacts[iu - l,-1] <= \
                        self.ms1_tlr:
                        if self.exacts[iu - l,1] in self.swl_levels:
                            lip = self.exacts[iu - l,:]
                            # headgroup and fatty acid guess from lipid name
                            # happens here!
                            hg, p_add, n_add, ether = \
                                self.headgroup_from_lipid_name(lip)
                            fa = self.fattyacid_from_lipid_name(lip)
                            lip = np.concatenate((lip, 
                                np.array([addMz, hg, fa, ether],
                                    dtype = np.object)),
                                axis = 0)
                            lip[4] = addName
                            if verbose:
                                outfile.write('\t -- Found: %s\n' % \
                                    str(list(lip)))
                            result.append(lip)
                        l += 1
                    else:
                        break
        return None if len(result) == 0 else np.vstack(result)

    def negative_positive(self, add_col = 12, mz_col = 1, swl_col = 8):
        """
        Column order:
        
        in:
        ltp_name, m/z,
        profile_score, control_profile_score,
        rank_profile_boolean, ubiquity_score, ubiquity_score,
        original_index,
        swisslipids_ac, level, lipid_name, lipid_formula, adduct, adduct_m/z
        
        out:
        [0] pos_m/z, pos_profile_score,
        [2] pos_control_profile_score, pos_rank_profile_boolean,
        [4] pos_ubiquity_score, pos_ubiquity_score,
        [6] pos_original_index, pos_swisslipids_ac, pos_level,
        [9] pos_lipid_name, pos_lipid_formula, pos_adduct, pos_adduct_m/z
        [13] neg_m/z, neg_profile_score, neg_control_profile_score,
        [16] neg_rank_profile_boolean,
        [17] neg_ubiquity_score, neg_ubiquity_score, neg_original_index,
        [18] neg_swisslipids_ac, neg_level,
        [22] neg_lipid_name, neg_lipid_formula, neg_adduct, neg_adduct_m/z
        """
        result = dict((protein.upper(), []) for protein in self.lipids.keys())
        prg = progress.Progress(len(result), 'Matching positive & negative',
            1, percent = False)
        for protein, tbl in iteritems(self.lipids):
            prg.step()
            if 'neg' in tbl and 'pos' in tbl:
                for neg in tbl['neg']:
                    adds = [neg[add_col]] \
                        if neg[add_col] != 'Unknown' else ['[M-H]-', '[M+HCOO]-']
                    for add in adds:
                        if add == '[M-H]-':
                            poshmz = mzmod.Mz(Mz(neg[mz_col]).add_h()).add_h()
                            posnh3mz = mzmod.Mz(mzmod.Mz(neg[mz_col]).add_h()).add_nh4()
                        elif add == '[M+HCOO]-':
                            poshmz = mzmod.Mz(mzmod.Mz(neg[mz_col]).remove_fo()).add_h()
                            posnh3mz = mzmod.Mz(mzmod.Mz(neg[mz_col]).\
                                remove_fo()).add_nh4()
                        else:
                            continue
                        iu = tbl['pos'][:,mz_col].searchsorted(poshmz)
                        u = 0
                        if iu < len(tbl['pos']):
                            while iu + u < len(tbl['pos']):
                                if tbl['pos'][iu + u,mz_col] - poshmz <= \
                                    self.ms1_tlr:
                                    if tbl['pos'][iu + u,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu + u,add_col] \
                                            == '[M+H]+' or \
                                        tbl['pos'][iu + u,add_col] \
                                            == 'Unknown':
                                        result[protein].append(np.concatenate(
                                            (tbl['pos'][iu + u,1:], neg[1:]),
                                                axis = 0
                                        ))
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if poshmz - tbl['pos'][iu - l,mz_col] <= \
                                    self.ms1_tlr:
                                    if tbl['pos'][iu - l,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu - l,add_col] \
                                            == '[M+H]+' or \
                                        tbl['pos'][iu - l,add_col] \
                                            == 'Unknown':
                                        result[protein].append(np.concatenate(
                                            (tbl['pos'][iu - l,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    l += 1
                                else:
                                    break
                        iu = tbl['pos'][:,1].searchsorted(posnh3mz)
                        u = 0
                        if iu < len(tbl['pos']):
                            while iu + u < len(tbl['pos']):
                                if tbl['pos'][iu + u,mz_col] - \
                                    posnh3mz <= self.ms1_tlr:
                                    if tbl['pos'][iu + u,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu + u,add_col] \
                                            == '[M+NH4]+' or \
                                        tbl['pos'][iu + u,add_col] \
                                            == 'Unknown':
                                        result[protein].append(np.concatenate(
                                            (tbl['pos'][iu + u,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if posnh3mz - tbl['pos'][iu - l,mz_col] \
                                    <= self.ms1_tlr:
                                    if tbl['pos'][iu - l,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu - l,add_col] \
                                            == '[M+NH4]+' or \
                                        tbl['pos'][iu - l,add_col] \
                                            == 'Unknown':
                                        result[protein].append(np.concatenate(
                                            (tbl['pos'][iu - l,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    l += 1
                                else:
                                    break
                if len(result[protein]) > 0:
                    result[protein] = np.vstack(result[protein])
        prg.terminate()
        return result

    def negative_positive2(self):
        """
        Results in dicts ['pos']['neg'] and ['neg']['pos'].
        
        Values in each array: assumed positive adduct, assumed negative adduct,
            measured positive m/z, measured negative m/z
        """
        ad2ex = self.ad2ex[1]
        ex2ad = self.ex2ad[1]
        self.sort_alll('mz')
        for protein, tbl in iteritems(self.valids):
            tbl['pos']['neg'] = dict((i, {}) for i in tbl['pos']['i'])
            tbl['pos']['neg_lip'] = dict((i, {}) for i in tbl['pos']['i'])
            tbl['neg']['pos'] = dict((i, {}) for i in tbl['neg']['i'])
            tbl['neg']['pos_lip'] = dict((i, {}) for i in tbl['neg']['i'])
        prg = progress.Progress(len(self.valids),
            'Matching positive & negative',
            1, percent = False)
        for protein, tbl in iteritems(self.valids):
            prg.step()
            for pi, poi in enumerate(tbl['pos']['i']):
                measured_pos_mz = tbl['pos']['mz'][pi]
                for pos_add, pos_add_fun in iteritems(ad2ex['pos']):
                    calculated_exact = \
                        getattr(mzmod.Mz(measured_pos_mz), pos_add_fun)()
                    for neg_add, neg_add_fun in iteritems(ex2ad['neg']):
                        calculated_neg_mz = getattr(mzmod.Mz(calculated_exact),
                            neg_add_fun)()
                        iu = tbl['neg']['mz'].searchsorted(calculated_neg_mz)
                        u = 0
                        if iu < len(tbl['neg']['i']):
                            while iu + u < len(tbl['neg']['i']):
                                if tbl['neg']['mz'][iu + u] - \
                                    calculated_neg_mz <= self.ms1_tlr:
                                    noi = tbl['neg']['i'][iu + u]
                                    match = np.array([pos_add, neg_add,
                                        measured_pos_mz,
                                        tbl['neg']['mz'][iu + u]],
                                        dtype = np.object)
                                    tbl['pos']['neg'][poi][noi] = match
                                    tbl['neg']['pos'][noi][poi] = match
                                    self.negative_positive_lipids(tbl, poi,
                                        tbl['neg']['i'][iu + u],
                                        pos_add, neg_add)
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if calculated_neg_mz - \
                                    tbl['neg']['mz'][iu - l] <= self.ms1_tlr:
                                    noi = tbl['neg']['i'][iu - l]
                                    match = np.array([pos_add, neg_add,
                                        measured_pos_mz,
                                        tbl['neg']['mz'][iu - l]],
                                        dtype = np.object)
                                    tbl['pos']['neg'][poi][noi] = match
                                    tbl['neg']['pos'][noi][poi] = match
                                    self.negative_positive_lipids(tbl, poi,
                                        tbl['neg']['i'][iu - l],
                                        pos_add, neg_add)
                                    l += 1
                                else:
                                    break
        prg.terminate()

    def negative_positive_lipids(self, tbl, poi, noi, pos_add, neg_add):
        """
        Result columns:
        database id positive, database id negative,
        lipid name positive, lipid name negative,
        adduct positive, adduct negative,
        database m/z positive, database m/z negative,
        measured m/z positive, measured m/z negative,
        headgroup positive, headgroup negative,
        fatty acids positive, fatty acids negative,
        dominant adduct positive, dominant adduct negative
        """
        result = []
        if tbl['pos']['lip'][poi] is not None and \
            tbl['neg']['lip'][noi] is not None:
            for plip in tbl['pos']['lip'][poi]:
                if plip[4] == pos_add:
                    for nlip in tbl['neg']['lip'][noi]:
                        if plip[7] is not None and nlip[7] is not None and \
                            nlip[4] == neg_add and plip[7] == nlip[7] and \
                            plip[8] == nlip[8]:
                                result.append(np.array(
                                [plip[0], nlip[0], plip[2], nlip[2], plip[4],
                                nlip[4], plip[5], nlip[5], plip[6], nlip[6],
                                plip[7], nlip[7], plip[8], nlip[8],
                                self.lipnames[plip[7]]['pos_adduct'],
                                self.lipnames[nlip[7]]['neg_adduct']],
                                dtype = np.object))
        result = np.vstack(result) if len(result) > 0 else None
        tbl['pos']['neg_lip'][poi][noi] = result
        tbl['neg']['pos_lip'][noi][poi] = result
    
    def average_area_5(self):
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                tbl['aaa'] = np.nansum(tbl['fe'], 1) / tbl['fe'].shape[0]
    
    def ms1(self):
        self.fractions_marco()
        self.fractions_by_protein_amount()
        self.primary_fractions()
        self.lookup_nans()
        self.average_area_5()
        self.protein_peak_ratios2()
        if self.use_manual_ppratios:
            self.read_manual_ppratios()
            self.ppratios_replace_manual()
        self.intensity_peak_ratios()
        self.ratios_in_range()
        self.peak_ratio_score()
        self.slope_filter()
        # this we do after slope filter as we collect
        # the peak ratio scores from the highest
        # fraction pairs defined in `slope_filter`
        self.combine_peak_ratio_scores()
        self.peak_ratio_score_bool()
        self.export_scores()
        self.lipid_lookup_exact()
        self.ms1_headgroups()
        self.negative_positive2()
        self.headgroups_negative_positive('ms1')
        # self.marco_standards()
    
    """
    END: lipid databases lookup
    """

    """
    Functions for MS2
    """
    def ms2(self):
        """
        Calls all methods in the MS2 pipeline. This involves
        
        - reading and autogenerating the fragment lists
        
        - identifying the MS2 mgf files
        
        - mapping mgf file offsets to scans
        
        - identifying detected fragments using the lists
        
        - identifying the features based on charecteristic
           fragments
        """
        self.ms2_init()
        self.ms2_map()
        self.ms2_main()
        self.ms2_headgroups()
        self.headgroups_by_fattya()
        self.identity_combined()
        self.ms2_scans_identify()
        self.ms2_headgroups2()
        self.consensus_identity()
        self.ms2_rt()
        self.good_features()
    
    def ms2_init(self):
        """
        Provides the generic data for MS2: looks up mgf files
        and reads/generates fragment lists.
        """
        
        self.ms2_metabolites()
        self.ms2_filenames()
    
    def ms2_onebyone(self, callback = 'std_layout_tables_xlsx',
                     proteins = None, identify = True, **kwargs):
        """
        Does the same as `ms2()`, but after doing it for one
        entity, it calls a callback, typically to export all
        important outcome, and after deletes the objects to
        free up memory. This is useful when working with so
        large data that it would require many gigs of memory
        otherwise.
        """
        
        self.ms2_metabolites()
        self.ms2_filenames()
        
        proteins = proteins if proteins else self.valids.keys()
        
        prg = progress.Progress(len(proteins),
                                'Reading & analysing MS2',
                                1, percent = False)
        
        for protein in proteins:
            
            prg.step()
            
            self.ms2_oneprotein(protein, identify = identify)
            
            if hasattr(callback, '__call__'):
                
                callback(self, proteins = [protein], **kwargs)
            
            else:
                
                getattr(self, callback)(proteins = [protein], **kwargs)
            
            self.ms2_cleanup(proteins = [protein])
        
        self.good_features()
        
        prg.terminate()
    
    def ms2_oneprotein(self, protein, identify = True):
        """
        Runs the whole MS2 workflow for only one protein.
        """
        
        if (
            not hasattr(self, 'ms2files') or
            not hasattr(self, 'nFragments') or
            not hasattr(self, 'pFragments')):
            
            self.ms2_init()
        
        self.ms2_map(proteins = [protein])
        self.ms2_main(proteins = [protein])
        self.ms2_headgroups(proteins = [protein])
        self.headgroups_by_fattya(proteins = [protein])
        self.identity_combined(proteins = [protein])
        self.ms2_scans_identify(proteins = [protein], identify = identify)
        self.ms2_headgroups2(proteins = [protein])
        self.consensus_identity(proteins = [protein])
        self.ms2_rt(proteins = [protein])
    
    def identify(self):
        
        self.headgroups_by_fattya()
        self.headgroups_negative_positive('ms2')
        self.identity_combined()
        self.identity_combined_ms2()
        self.feature_identity_table()
    
    def ms2_metabolites(self):
        """
        Calls the methods for reading and auto-generating fragment
        ion lists for both negative and positive modes.
        """
        
        self.pFragments, self.pHgfrags, self.pHeadgroups = \
            self.ms2_read_metabolites(self.pfragmentsfile)
        self.nFragments, self.nHgfrags, self.nHeadgroups = \
            self.ms2_read_metabolites(self.nfragmentsfile)
    
    def export_auto_fraglist(self, infile, outfile = None):
        """
        Exports the auto-generated fragment lists into file.
        """
        
        outfile = 'lipid_fragments_%s_mode_ext.txt' % \
            ('negative' if 'negative' in infile else 'positive') \
            if outfile is None else outfile
        lst = self.ms2_read_metabolites(infile, extra_fragments = True,
            return_fraglines = True)
        with open(outfile, 'w') as out:
            out.write(
                '\n'.join(
                    map(
                        lambda l:
                            '%.07f\t%s\t%s%s' %\
                                tuple(l[:3] +
                                    ['\t%s'%l[3] if len(l) >= 4 else '']),
                        lst
                    )
                )
            )
    
    def ms2_read_metabolites(self, fname, extra_fragments = True,
                             return_fraglines = False):
        """
        First part is from Toby Hodges.
        Reads metabolite fragments data and autogenerates fragment series
        like fatty acid fragments with a defined range of carbon counts and
        unsaturation.
        
        """
        
        Metabolites = []
        Hgroupfrags = {}
        rehg = re.compile(r'.*\(([\+;A-Z]+)\).*')
        rehgsep = re.compile(r'[;\+]')
        
        with open(fname, 'r') as Handle:
            lst = \
                list(
                    map(
                        lambda l:
                            list(
                                map(
                                    lambda i:
                                        i[1] if i[0] > 0 else self.to_float(i[1]),
                                    enumerate(
                                        l.strip().split('\t')
                                    )
                                )
                            ),
                        Handle.readlines()
                    )
                )
        
        if extra_fragments:
            
            if 'negative' in fname:
                
                if not self.only_marcos_fragments:
                    
                    lst += self.auto_fragment_list(fragments.FAminusH, -1, name = 'FA')
                    # fatty acid -CO2- fragments:
                    lst += self.auto_fragment_list(
                        fragments.FattyFragment, -1,
                        minus = ['CO2', 'H'], name = 'FA')
                    
                    lst += self.auto_fragment_list(fragments.FAAlkylminusH, -1)
                    
                    for hg in ('PE', 'PC', 'PS', 'PI', 'PG', 'PA'):
                        
                        lst += self.auto_fragment_list(
                            getattr(fragments, 'Lyso%s' % hg), -1
                        )
                        
                        lst += self.auto_fragment_list(
                            getattr(fragments, 'Lyso%s' % hg), -1, minus = ['H2O']
                        )
                        
                        if hg != 'PA' and hg != 'PG':
                            lst += self.auto_fragment_list(
                                getattr(fragments, 'Lyso%sAlkyl' % hg), -1
                            )
                            lst += self.auto_fragment_list(
                                getattr(fragments, 'Lyso%sAlkyl' % hg), -1, minus = ['H2O']
                            )
                        
                        if hg == 'PI':
                            lst += self.auto_fragment_list(
                                fragments.LysoPI, -1, minus = ['H', 'H2O', 'C6H10O5']
                            )
                            lst += self.auto_fragment_list(
                                fragments.LysoPI, -1, minus = ['CO2']
                            )
                        
                        if hg == 'PE':
                            lst += self.auto_fragment_list(
                                fragments.LysoPE, -1, minus = ['CO2']
                            )
                        
                        if hg == 'PC':
                            lst += self.auto_fragment_list(
                                fragments.LysoPC, -1, minus = ['CH3']
                            )
                            lst += self.auto_fragment_list(
                                fragments.LysoPC, -1, minus = ['CH3', 'H2O']
                            )
                    
                    lst += self.auto_fragment_list(fragments.CerFA, -1)
                    lst += self.auto_fragment_list(fragments.CerFAminusC, -1)
                    lst += self.auto_fragment_list(fragments.CerFAminusN, -1)
                    lst += self.auto_fragment_list(fragments.CerFAminusC2H5N, -1)
                    lst += self.auto_fragment_list(fragments.CerSphi, -1,
                                                   cmin = 14, unsatmin = 0,
                                                   cmax = 19, unsatmax = 3)
                    lst += self.auto_fragment_list(fragments.CerSphiMinusN, -1,
                                                   cmin = 14, unsatmin = 0 ,
                                                   cmax = 19, unsatmax = 3)
                    lst += self.auto_fragment_list(fragments.CerSphiMinusNO, -1,
                                                   cmin = 14, unsatmin = 0 ,
                                                   cmax = 19, unsatmax = 3)
                    
                else:
                    
                    lst += self.auto_fragment_list(fragments.FAminusH, -1)
                    lst += self.auto_fragment_list(fragments.LysoPA, -1, minus = ['H2O'])
                    lst += self.auto_fragment_list(fragments.CerFA, -1)
            
            if 'positive' in fname:
                
                if not self.only_marcos_fragments:
                    
                    lst += self.auto_fragment_list(fragments.NLFAminusH2O, 0)
                    lst += self.auto_fragment_list(fragments.NLFA, 0)
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 14, unsatmin = 0,
                                                   cmax = 19, unsatmax = 3,
                                                   minus = ['H5'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 14, unsatmin = 0,
                                                   cmax = 19, unsatmax = 3,
                                                   minus = ['H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 14, unsatmin = 0,
                                                   cmax = 19, unsatmax = 3,
                                                   minus = ['H2O', 'H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 14, unsatmin = 0,
                                                   cmax = 19, unsatmax = 3,
                                                   minus = ['C', 'H2O', 'H2O'])
                    lst += self.auto_fragment_list(fragments.NLFAplusOH, 0)
                lst += self.auto_fragment_list(fragments.FAplusGlycerol, 1)
                lst += self.auto_fragment_list(fragments.NLFAplusNH3, 0)
                lst += self.auto_fragment_list(fragments.FAminusO, 1)
                
                if self.only_marcos_fragments:
                    
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 16, unsatmin = 0,
                                                   cmax = 16, unsatmax = 3,
                                                   minus = ['H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 16, unsatmin = 0,
                                                   cmax = 16, unsatmax = 3,
                                                   minus = ['H2O', 'H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 16, unsatmin = 0,
                                                   cmax = 16, unsatmax = 3,
                                                   minus = ['C', 'H2O', 'H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 18, unsatmin = 0,
                                                   cmax = 18, unsatmax = 3,
                                                   minus = ['H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 18, unsatmin = 0,
                                                   cmax = 18, unsatmax = 3,
                                                   minus = ['H2O', 'H2O'])
                    lst += self.auto_fragment_list(fragments.SphingosineBase, 1,
                                                   cmin = 18, unsatmin = 0,
                                                   cmax = 18, unsatmax = 3,
                                                   minus = ['C', 'H2O', 'H2O'])
            if return_fraglines:
                return lst
        
        for Line in lst:
            try:
                MetabMass, MetabType, MetabCharge = Line[:3]
            except:
                sys.stdout.write('\tWrongly formatted fragment line:'\
                    '\n\t  %s\n' % str(Line))
                sys.stdout.flush()
            hgm = rehg.match(Line[1])
            
            if (len(Line) == 4 and len(Line[3])) or hgm:
                
                Hgroupfrags[MetabType] = set([])
                
                if len(Line) == 4:
                    Hgroupfrags[MetabType] = Hgroupfrags[MetabType] | \
                        set(rehgsep.split(Line[3]))
                
                if hgm:
                    Hgroupfrags[MetabType] = Hgroupfrags[MetabType] | \
                        set(rehgsep.split(hgm.groups()[0]))
            
            Metabolites.append([MetabMass, MetabType, MetabCharge])
            if '+' not in MetabCharge and '-' not in MetabCharge \
                and 'NL' not in MetabCharge:
                sys.stdout.write('WARNING: fragment %s has no '\
                    'valid charge information!\n' % \
                    MetabType)
        
        Headgroups = {}
        for frag, hgs in iteritems(Hgroupfrags):
            for hg in hgs:
                if len(hg):
                    if hg not in Headgroups:
                        Headgroups[hg] = set([])
                    Headgroups[hg].add(frag)
        
        Metabolites = \
            np.array(
                sorted(Metabolites,
                    key = lambda x: x[0]
                ),
                dtype = np.object
            )
        
        return Metabolites, Hgroupfrags, Headgroups
    
    def auto_fragment_list(self, typ, charge, cmin = 2, unsatmin = 0,
        cmax = 36, unsatmax = 6, minus = [], plus = [], **kwargs):
        """
        Generates a series of fragments of a given type, iterating
        over a range of carbon counts and unsaturated bond counts.
        In addition, certain groups can be added or deduced from the
        fragment mass, e.g. -H2O for a water loss.
        """
        series = fragments.FAFragSeries(typ, charge,
            cmin = cmin, unsatmin = unsatmin,
            cmax = cmax, unsatmax = unsatmax,
            minus = minus, plus = plus,
            **kwargs)
        return list(series.iterfraglines())

    def ms2_filenames(self):
        """
        Files are placed in 2 root directories, in specific subdirectories,
        positive and negative MS mode in separate files.
        This function collects all the file paths and returns them 
        in a dict of dicts. Keys are LTP names, and 'pos'/'neg'.
        """
        redirname = re.compile(r'(^[0-9a-zA-Z]+)[ _](pos|neg)')
        remgfname = re.compile(r'(^[0-9a-zA-Z_]+)_([a-zA-Z0-9]{3,})_(pos|neg)'\
            r'_([A-Z][0-9]{1,2})\.mgf')
        refractio = re.compile(r'.*_([A-Z][0-9]{1,2}).*')
        fnames = {}
        
        for d in self.datadirs:
            
            try:
                proteindd = os.listdir(d)
            except OSError:
                sys.stdout.write('\t:: Directories missing. '
                                 'Please mount the shared folder.\n')
                return None
            
            if self.ms2dir is not None and self.ms2dir in proteindd:
                
                proteindd = [self.ms2dir]
                ms2dir = True
            
            else:
                proteindd = [i for i in proteindd if os.path.isdir(os.path.join(d, i))]
                ms2dir = False
            
            if len(proteindd) == 0:
                sys.stdout.write('\t:: Please mount the shared folder!\n')
                return None
            
            for proteind in proteindd:
                
                dirname = redirname.findall(proteind)
                
                if len(dirname):
                    
                    proteinname, pos = dirname[0]
                    proteinname = proteinname.upper()
                    
                elif not ms2dir:
                    # other dirs are irrelevant:
                    continue
                
                if not ms2dir:
                    fpath = [proteind, 'Results']
                else:
                    fpath = [proteind]
                
                if os.path.isdir(os.path.join(d, *fpath)):
                    
                    for f in os.listdir(os.path.join(d, *fpath)):
                        
                        if f.endswith('mgf') \
                            and 'buffer' not in f \
                            and 'TLC_STD' not in f \
                            and 'MeOH' not in f:
                            
                            try:
                                fr = refractio.match(f).groups()[0]
                                fr = '%s%u' % (fr[0], int(fr[1:]))
                            except AttributeError:
                                sys.stdout.write(
                                    'could not determine '\
                                    'fraction number: %s' % f)
                            
                            if ms2dir:
                                null, proteinname, pos, fr = \
                                    remgfname.match(f).groups()
                                
                                fr = '%s%u' % (fr[0], int(fr[1:]))
                            
                            fr = 'B1' if fr == 'B01' \
                                else 'A9' if fr == 'A09' \
                                else fr
                            
                            if proteinname not in fnames:
                                fnames[proteinname] = {}
                            
                            if pos not in fnames[proteinname] \
                                or proteind.endswith('update'):
                                
                                fnames[proteinname][pos] = {}
                            
                            fnames[proteinname][pos][fr] = \
                                os.path.join(*([d] + fpath + [f]))
        
        self.ms2files = fnames
    
    def ms2_map(self, proteins = None, silent = False):
        """
        Maps offsets of the beginning of each scan in MS2 mgf files.
        """
        stRpos = 'pos'
        stRneg = 'neg'
        refrac = re.compile(r'([A-Z])([\d]{1,2})$')
        result = dict((protein.upper(), {'pos': None, 'neg': None,
                'ms2files': {'pos': {}, 'neg': {}}}) \
            for protein, d in iteritems(self.ms2files))
        
        missing_ms1 = 0
        if not hasattr(self, 'missing_ms1'):
            self.missing_ms1 = set([])
        
        if proteins is not None and len(proteins) == 1:
            silent = True
        
        if not silent:
            prg = progress.Progress(len(self.ms2files) * 2,
                                    'Indexing MS2 data', 1,
                                    percent = False)
        
        for protein, d in iteritems(self.ms2files):
            
            if proteins is not None and protein not in proteins:
                continue
            
            pFeatures = []
            nFeatures = []
            uprotein = protein.upper()
            
            ifrac = self.fraction_indices(protein)
            
            for mode, files in iteritems(d):
                
                if not silent:
                    prg.step()
                
                features = pFeatures if mode == stRpos else nFeatures
                #fractions = set([])
                for fr, fl in iteritems(files):
                    
                    fr = refrac.match(fr).groups()
                    fr = '%s%u' % (fr[0], int(fr[1]))
                    try:
                        frnum = ifrac[fr][0]
                    except KeyError:
                        missing_ms1 += 1
                        self.missing_ms1.add((protein, mode, fr))
                        continue
                    
                    mm = self.ms2_index(fl, frnum,
                                        charge = self.ms2_precursor_charge)
                    m = np.vstack(mm)
                    features.extend(mm)
                    result[uprotein]['ms2files'][mode][fr] = fl
            
            pFeatures = np.array(sorted(pFeatures, key = lambda x: x[0]),
                dtype = np.float64)
            nFeatures = np.array(sorted(nFeatures, key = lambda x: x[0]),
                dtype = np.float64)
            result[uprotein]['pos'] = pFeatures
            result[uprotein]['neg'] = nFeatures
        
        if not silent:
            prg.terminate()
        
        if missing_ms1 > 0:
            sys.stdout.write('\n\t:: Warning: for certain MS2 files '\
                'no MS1 data available. See them in `missing_ms1`.\n')
            sys.stdout.flush()
        
        self.ms2map = result

    def ms2_index(self, fl, fr, charge = 1):
        """
        Looking up offsets in one MS2 mgf file.
        
        Columns:
            -- pepmass
            -- intensity
            -- retention time
            -- scan num
            -- offset in file
            -- fraction num
        """
        stRrtinseconds = 'RTINSECONDS'
        stRtitle = 'TITLE'
        stRbe = 'BE'
        stRen = 'EN'
        stRch = 'CH'
        stRpepmass = 'PEPMASS'
        stRempty = ''
        reln = re.compile(r'^([A-Z]+).*=([\d\.]+)[\s]?([\d\.]*)["]?$')
        features = []
        offset = 0
        cap_next = False
        
        with open(fl, 'rb', 8192) as fp:
            
            for l in fp:
                
                l = l.decode('ascii')
                
                if not l[0].isdigit() and \
                    not l[:2] == stRbe and not l[:2] == stRen:
                    
                    if not l[:2] == stRch:
                        
                        try:
                            m = reln.match(l).groups()
                        except:
                            print(fl, l)
                            continue
                        
                        if m[0] == stRtitle:
                            scan = float(m[1])
                        
                        if m[0] == stRrtinseconds:
                            rtime = float(m[1]) / 60.0
                        
                        if m[0] == stRpepmass:
                            pepmass = float(m[1])
                            intensity = 0.0 if m[2] == stRempty else float(m[2])
                            if charge is None:
                                cap_next = True
                        
                    else:
                        _charge = int(l[7]) if len(l) >= 8 else None
                        if charge is None or _charge == charge:
                            cap_next = True
                
                elif cap_next:
                    
                    features.append([pepmass, intensity,
                        rtime, scan, offset, fr])
                    scan = None
                    rtime = None
                    intensity = None
                    pepmass = None
                    _charge = None
                    cap_next = False
                
                offset += len(l)
        
        return features

    def ms2_main(self, proteins = None, verbose = False, outfile = None,
                 silent = False):
        """
        For all LTPs and modes obtains the MS2 data from original files.
        """
        # ms2map columns: pepmass, intensity, rtime, scan, offset, fraction
        
        if proteins is not None and len(proteins) == 1:
            silent = True
        
        if not silent:
            prg = progress.Progress(
                (len(self.valids) if proteins is None else len(proteins)) * 2,
                'Looking up MS2 fragments', 1,
                percent = False)
        
        if verbose:
            outfile = outfile if hasattr(outfile, 'write') else \
                sys.stdout if outfile is None else open(outfile, 'w')
        
        for protein, d in iteritems(self.valids):
            
            if proteins is None or protein in proteins:
                
                for pn, tbl in iteritems(d):
                    
                    if not silent:
                        prg.step()
                    
                    # we look up the real measured MS1 m/z's in MS2,
                    # so will divide recalibrated values by the drift ratio
                    drift = 1.0 \
                        if not hasattr(self, 'proteins_drifts') \
                        or 'recalibrated' not in tbl \
                        or not tbl['recalibrated'] \
                        else self.ppm2ratio(np.nanmedian(
                                np.array(self.proteins_drifts[
                                    protein][pn].values())
                            ))
                    
                    if verbose:
                        outfile.write('\t:: Recalibrated m/z: '\
                        '%.08f; drift = %.08f; measured m/z: %.08f\n' % \
                            (tbl['mz'][0], drift, tbl['mz'][0] / drift))
                    
                    ms2matches = self.ms2_match(tbl['mz'],
                        tbl['rt'], tbl['i'],
                        protein, pn, drift = drift,
                        verbose = verbose, outfile = outfile)
                    
                    # this already returns final result, from protein containing
                    # fractions and with the relevant retention times
                    tbl['ms2'] = self.ms2_lookup(protein, pn, ms2matches,
                                                 verbose = verbose,
                                                 outfile = outfile)
                    
                    tbl['ms2r'] = self.ms2_result(tbl['ms2'])
                    #if verbose:
                        #print('\n')
                        #print('number of positive mzs:', len(posMzs))
                        #print('negative matching:', len(pos_matches))
                        #print('number of negative mzs:', len(negMzs))
                        #print('negative matching:', len(neg_matches))
        
        if hasattr(outfile, 'close') and outfile != sys.stdout:
            outfile.close()
        
        if not silent:
            prg.terminate()

    def ms2_result(self, ms2matches):
        """
        Extracts the most relevant information from the MS2
        result arrays, throwing away the rest.
        
        Columns in output arrays (6):
            # MS2 fragment name, MS2 adduct name, (0-1)
            # MS2 fragment m/z, MS2 fragment intensity (2-3)
            # fraction number, scan number (4-5)
        """
        result = dict((oi, []) for oi in ms2matches.keys())
        
        for oi, ms2s in iteritems(ms2matches):
            
            for ms2i in ms2s:
                
                result[oi].append(np.array([ms2i[7], ms2i[8],
                    ms2i[1], ms2i[2], ms2i[14], ms2i[12]], dtype = np.object))
        
        for oi, ms2s in iteritems(result):
            
            result[oi] = np.vstack(ms2s) if len(ms2s) > 0 else np.array([])
            result[oi].shape = (len(result[oi]), 6)
        
        return result

    def ms2_match(self, ms1Mzs, ms1Rts, ms1is, protein, pos,
        verbose = False, outfile = None, drift = 1.0,
        rt_tolerance = 1.0):
        """
        Looks up matching pepmasses for a list of MS1 m/z's.
        """
        opened_here = False
        if hasattr(outfile, 'write'):
            outfile = outfile
        elif outfile is None:
            outfile = sys.stdout
        else:
            open(outfile, 'w')
            opened_here = True
        
        matches = []
        ms2tbl = self.ms2map[protein][pos]
        # iterating over MS1 m/z, MS1 original index, and retention time
        for ms1Mz, ms1i, rt in zip(ms1Mzs, ms1is, ms1Rts):
            # drift is the ratio
            # we divide here to have the original measured MS1 m/z,
            # not the recalibrated one:
            if verbose:
                outfile.write('\t:: Recalibrated m/z: %.08f; drift = %.08f; m'\
                    'easured m/z: %.08f\n' % (ms1Mz, drift, ms1Mz / drift))
            ms1Mz = ms1Mz / drift
            # if error comes here, probably MS2 files are missing
            try:
                iu = ms2tbl[:,0].searchsorted(ms1Mz)
            except IndexError:
                sys.stdout.write('\nMissing MS2 files for %s-%s?\n' % \
                    (protein, pos))
                continue
            if verbose:
                outfile.write('\t:: Looking up MS1 m/z %.08f. '\
                    'Closest values found: %.08f and %.08f\n' % (
                    ms1Mz,
                    ms2tbl[iu - 1, 0] if iu > 0 else 0.0,
                    ms2tbl[iu, 0] if iu < ms2tbl.shape[0] else 0.0))
            rt_mean = np.mean(rt)
            rt = (rt_mean - rt_tolerance, rt_mean + rt_tolerance)
            u = 0
            if iu < ms2tbl.shape[0]:
                while iu + u < ms2tbl.shape[0]:
                    if ms2tbl[iu + u,0] - ms1Mz <= self.ms1_tlr:
                        if verbose:
                            outfile.write('\t -- Next value within '\
                                'range of tolerance: %.08f\n' % \
                                    ms2tbl[iu + u, 0])
                            if not self.ms2_rt_within_range:
                                outfile.write('\t -- Not checking RT\n')
                            elif ms2tbl[iu + u, 2] >= rt[0] and \
                                 ms2tbl[iu + u, 2] <= rt[1]:
                                outfile.write('\t -- Retention time OK, '\
                                    'accept this match\n')
                            else:
                                outfile.write('\t -- Retention time is %.02f'\
                                    ', should be in range %.02f-%.02f to'\
                                    ' match.\n' % \
                                    (ms2tbl[iu + u, 2], rt[0], rt[1]))
                                outfile.write('\t -- Retention time not OK, '\
                                    'drop this match\n')
                        # checking retention time
                        if not self.ms2_rt_within_range or (
                            ms2tbl[iu + u, 2] >= rt[0] and \
                            ms2tbl[iu + u, 2] <= rt[1]):
                            if verbose:
                                outfile.write('\t -- value found at index %u\n' % (iu + u))
                            matches.append((ms1Mz, iu + u, ms1i))
                        u += 1
                    else:
                        break
            l = 1
            if iu > 0:
                while iu >= l:
                    if ms1Mz - ms2tbl[iu - l,0] <= self.ms1_tlr:
                        # checking retention time
                        if verbose:
                            outfile.write('\t -- Next value within '\
                                'range of tolerance: %.08f\n' % \
                                    ms2tbl[iu - l, 0])
                            if not self.ms2_rt_within_range:
                                outfile.write('\t -- Not checking RT\n')
                            elif ms2tbl[iu - l, 2] >= rt[0] and \
                                 ms2tbl[iu - l, 2] <= rt[1]:
                                outfile.write('\t -- Retention time OK, '\
                                    'accept this match\n')
                            else:
                                outfile.write('\t -- Retention time is %.02f'\
                                    ', should be in range %.02f-%.02f to'\
                                    ' match.\n' % \
                                    (ms2tbl[iu - l, 2], rt[0], rt[1]))
                                outfile.write('\t -- Retention time not OK, '\
                                    'drop this match\n')
                        if not self.ms2_rt_within_range or (
                            ms2tbl[iu - l, 2] >= rt[0] and \
                            ms2tbl[iu - l, 2] <= rt[1]):
                            if verbose:
                                outfile.write('\t -- value found at index %u\n' % (iu - l))
                            matches.append((ms1Mz, iu - l, ms1i))
                        l += 1
                    else:
                        break
        if opened_here and type(outfile) is file and outfile != sys.stdout:
            outfile.close()
        return sorted(uniqList(matches), key = lambda x: x[0])

    def ms2_verbose(self, protein, outfile = None):
        self.ms2_main(proteins = [protein], verbose = True,
                outfile = outfile)

    def ms2_lookup(self, protein, mode, ms1matches, verbose = False, outfile = None):
        """
        For the matching MS2 m/z's given, reads and identifies
        the list of fragments.
        
        Columns in output arrays (15):
            # MS1 m/z, MS2 fragment m/z, MS2 fragment intensity, (0-2)
            # MS2 table index, direct/inverted match, fraction number, (3-5)
            # MS2 fragment m/z in annotation, MS2 fragment name, (6-7)
            # MS2 adduct name, (8)
            # MS1 pepmass, MS1 intensity, rtime, MS2 scan, (9-12)
            # MS2 file offset, fraction number (13-14)
        """
        
        fraglist = self.pFragments if mode == 'pos' else self.nFragments
        ms2map = self.ms2map[protein][mode]
        ms2files = self.ms2map[protein]['ms2files'][mode]
        # indices of fraction numbers
        ifracs = self.fraction_indices(protein)
        fracsi = dict(map(lambda fr: (fr[1][0], fr[0]), iteritems(ifracs)))
        
        sample_i = {
            9: 1, 10: 2, 11: 3, 12: 4, 1: 5
        }
        stRcharge = 'CHARGE'
        
        # opening all files to have file pointers ready
        files = dict((fr, open(fname, 'r')) \
            for fr, fname in iteritems(ms2files))
        
        # Initializing dict with original indices
        ms2matches = dict((m[2], []) for m in ms1matches)
        
        # iterating over MS1 m/z, MS2 table index, MS1 original index
        if verbose:
            outfile.write('\t :: MS2 lookup starts\n')
        
        for ms1mz, ms2i, ms1oi in ms1matches:
            
            # ms2map columns: pepmass, intensity, rtime, scan, offset, fraction
            ms2item = ms2map[ms2i,:]
            fr = int(ms2item[5])
            frlab = fracsi[fr]
            
            if verbose:
                outfile.write('\t -- Looking up %.08f in fraction %s (#%u), '\
                    'line %u\n' % (ms1mz, frlab, fr, ms2i))
            
            # only fractions with the protein
            if verbose:
                if frlab in files:
                    outfile.write('\t -- Have file for fraction %s\n' % frlab)
                else:
                    outfile.write('\t -- Do not have file for fraction %s; files: %s\n' % \
                        (frlab, str(files.keys())))
            
            if (not self.ms2_only_protein_fractions or \
                fractions[sample_i[fr]] == 1) and frlab in files:
                
                fp = files[frlab]
                
                # jumping to offset
                fp.seek(int(ms2item[4]), 0)
                if verbose:
                    outfile.write('\t -- Reading from file %s\n' % fp.name)
                
                # zero means no clue about charge
                charge = 0
                
                for l in fp:
                    
                    if l[:6] == stRcharge:
                        # one chance to obtain the charge
                        charge = int(l.strip()[-2])
                        continue
                    if not l[0].isdigit():
                        # finish at next section
                        break
                    else:
                        # reading fragment masses
                        mi = l.strip().split()
                        mass = float(mi[0])
                        
                        intensity = float(mi[1]) if len(mi) > 1 else np.nan
                        # matching fragment --- direct
                        ms2hit1 = self.ms2_identify(mass, fraglist,
                            compl = False)
                        # matching fragment --- inverted
                        ms2hit2 = self.ms2_identify(ms2item[0] - mass,
                            fraglist, compl = True)
                        # matched fragment --- direct
                        # columns (14):
                        # MS1 m/z, MS2 fragment m/z, MS2 fragment intensity,
                        # MS2 table index, direct/inverted match, 
                        # fraction number,
                        # MS2 fragment m/z in annotation, MS2 fragment name,
                        # MS2 adduct name,
                        # MS1 pepmass, MS1 intensity, rtime, MS2 scan,
                        # MS2 file offset, fraction number
                        for frag in ms2hit1:
                            ms2matches[ms1oi].append(np.concatenate(
                                (np.array([ms1mz, mass,
                                    intensity, ms2i, 0, fr]),
                                frag, ms2item)
                            ))
                        # matched fragment --- inverted
                        for frag in ms2hit2:
                            ms2matches[ms1oi].append(np.concatenate((
                                np.array([ms1mz, mass,
                                    intensity, ms2i, 1, fr]),
                                frag, ms2item)
                            ))
                        # no fragment matched --- unknown fragment
                        if not len(ms2hit1) and not len(ms2hit2):
                            thisRow = np.concatenate((
                                np.array([ms1mz, mass,
                                    intensity, ms2i, 0, fr],
                                    dtype = np.object),
                                np.array([None, 'unknown', 'unknown'],
                                    dtype = np.object),
                                ms2item)
                            )
                            thisRow.shape = (1, 15)
                            ms2matches[ms1oi].append(thisRow)
                
                if verbose:
                    outfile.write('\t -- %u lines have been read\n' % len(ms2matches[ms1oi]))
        
        # removing file pointers
        for fp in files.values():
            fp.close()
        
        for oi, ms2match in iteritems(ms2matches):
            
            if len(ms2match) > 0:
                ms2matches[oi] = np.vstack(ms2match)
            else:
                ms2matches[oi] = np.array([[]])
                ms2matches[oi].shape = (0, 15)
        
        return ms2matches

    def ms2_identify(self, mass, fraglist, compl):
        """
        Looks up one MS2 m/z value in list of known fraglist masses.
        Either with matching between MS2 m/z and fragment m/z within
        a given tolerance, or between the fragment mass and the
        residual mass after substracting MS2 m/z from MS1 m/z.
        Returns the fragment's mass, name and adduct type, or None
        in case of no match.
        """
        result = []
        i = -1
        du = None
        dl = None
        iu = fraglist[:,0].searchsorted(mass)
        if iu < len(fraglist) and \
            ((compl and 'NL' in fraglist[iu,2]) or \
            (not compl and ('+' in fraglist[iu,2] or \
            '-' in fraglist[iu,2]) and 'NL' not in fraglist[iu,2])):
            du = fraglist[iu,0] - mass
        if iu > 0 and \
            ((compl and 'NL' in fraglist[iu - 1,2]) or \
            (not compl and ('+' in fraglist[iu - 1,2] or \
            '-' in fraglist[iu - 1,2]) \
            and 'NL' not in fraglist[iu - 1,2])):
            dl = mass - fraglist[iu - 1,0]
        if du is not None and (dl is None or du < dl) and du < self.ms2_tlr:
            i = iu
            st = 1
        if dl is not None and (du is None or dl < du) and dl < self.ms2_tlr:
            i = iu - 1
            st = -1
        val = fraglist[i,0]
        while len(fraglist) > i >= 0 and fraglist[i,0] == val:
            if (compl and 'NL' in fraglist[i,2]) or \
                (not compl and \
                ('-' in fraglist[i,2] or '+' in fraglist[i,2]) \
                and 'NL' not in fraglist[i,2]):
                result.append(fraglist[i,:])
                i += st
        return result

    def ms2_collect(self, ms2matches, ms1mz, unknown = False):
        """
        Deprecated.
        """
        result = []
        fraglist = ms2matches[ms2matches[:,0] == ms1mz,:]
        for frag in uniqList(fraglist[:,7]):
            if frag != 'Unknown':
                thisFrag = fraglist[fraglist[:,7] == frag,:]
                maxInt = thisFrag[:,2].max()
                thisFragMass = thisFrag[0,1]
                result.append((frag, maxInt, thisFragMass))
        if unknown:
            unknowns = fraglist[fraglist[:,7] == 'Unknown',:]
            result += [('Unknown', l[2], l[1]) for l in unknowns]
        return result
    
    def ms2_headgroups2(self, proteins = None):
        """
        This collects the possible headgroups from the
        advanced MS2 identification (done by ms2_scans_identify()).
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mode, tbl in iteritems(d):
                
                tbl['ms2hg2'] = {}
                for oi, ms2i in iteritems(tbl['ms2i2']):
                    tbl['ms2hg2'][oi] = set(
                        map(
                            lambda hgssc:
                                hgssc[0],
                            filter(
                                lambda hgssc:
                                    hgssc[1] > 0,
                                map(
                                    lambda hgsca:
                                        (hgsca[0],
                                            sum(
                                                map(
                                                    lambda scan:
                                                        scan['score'],
                                                    hgsca[1]
                                                )
                                            )
                                        ),
                                    iteritems(tbl['ms2i2'][oi])
                                )
                            )
                        )
                    )
    
    def consensus_identity(self, proteins = None):
        """
        Attempts to come to the most accurate identification using
        results of the MS2 spectra processing and the database lookups
        whereever these are available.
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mode, tbl in iteritems(d):
                
                ids = {}
                idlevels = {}
                
                for oi, ms2i in iteritems(tbl['ms2i2']):
                    
                    this_id   = []
                    idlevel   = 'IX'
                    ms2_level = False
                    
                    for hg in sorted(ms2i.keys()):
                        
                        ms2ii = ms2i[hg]
                        fa_level_id = False
                        hg_in_ms2   = False
                        
                        for ms2iii in ms2ii:
                            
                            if ms2iii['score'] > 0:
                                
                                hg_in_ms2 = True
                                ms2_level = True
                                
                                for ms2fa in ms2iii['fattya']:
                                    
                                    sumcc = reduce(
                                        lambda cc1, cc2:
                                            (cc1[0] + cc2[0], cc1[1] + cc2[1]),
                                        map(
                                            lambda cc:
                                                (int(cc.split(':')[0]), int(cc.split(':')[1])),
                                            ms2fa.replace('O-', '').replace('d', '').split('/')
                                        )
                                    )
                                    sumccstr = '%u:%u' % sumcc
                                    if hg in tbl['ms1fa'][oi] and \
                                        sumccstr in tbl['ms1fa'][oi][hg]:
                                            this_id.append('%s(%s)' % (hg, ms2fa))
                                            fa_level_id = True
                        
                        if hg_in_ms2 and not fa_level_id:
                            if hg in tbl['ms1fa'][oi]:
                                for ms1fa in tbl['ms1fa'][oi][hg]:
                                    this_id.append('%s(%s)' % (hg, ms1fa))
                            elif hg in tbl['ms1hg'][oi]:
                                this_id.append(hg)
                    
                    if (
                        len(
                            tbl['ms2hg2'][oi] &
                            tbl['ms1hg'][oi]
                        ) == 1 or
                        len(
                            tbl['ms2hg2'][oi] &
                            tbl['ms1hg'][oi] &
                            set(['BMP', 'PG'])
                        ) == 2):
                        # MS1 and MS2 identifications
                        # confirm each other
                        idlevel = 'I'
                        
                    elif len(tbl['ms1hg'][oi]):
                        # MS1 and MS2 identification
                        # but not confirming
                        # or MS2 completely unknown
                        idlevel = 'III.2'
                        this_id = sorted(set(self.get_ms1_ids(
                            protein, mode, oi)))
                        
                    else:
                        # has no MS2
                        # MS1 either unknown or matches
                        # some species in the database
                        idlevel = 'III.5'
                        this_id = sorted(set(self.get_ms1_ids(
                            protein, mode, oi)))
                    
                    ms2f = tbl['ms2f'][oi]
                    
                    if (abs(ms2f._scans[ms2f.best_scan].deltart)
                        < self.deltart_threshold):
                        
                        ids[oi]      = sorted(set(this_id))
                        idlevels[oi] = idlevel
                
                for oi in tbl['i']:
                    
                    # no MS2
                    if oi not in idlevels:
                        
                        if len(tbl['ms1hg'][oi]) == 1:
                            # no MS2, but only one MS1
                            # identification
                            idlevels[oi] = 'II'
                            
                        elif len(tbl['ms1hg'][oi]) > 1:
                            # no MS2 and more than one
                            # MS1 identification
                            idlevels[oi] = 'IV'
                            
                        else:
                            # what remains: no MS1
                            # and no MS2
                            idlevels[oi] = 'III.6'
                        
                        ids[oi] = sorted(set(self.get_ms1_ids(
                            protein, mode, oi)))
                
                tbl['cid']     = ids
                tbl['idlevel'] = idlevels
    
    def get_ms1_ids(self, protein, mode, oi):
        """
        Returns the summary of MS1 based identifications
        for one feature based on its original index.
        """
        
        tbl = self.valids[protein][mode]
        
        ids = set()
        ids.update(
            set(
                itertools.chain(
                    *map(
                        lambda hg:
                            map(
                                lambda fa:
                                    '%s(%s)' % (hg[0], fa),
                                hg[1]
                            ),
                        iteritems(tbl['ms1fa'][oi])
                    )
                )
            )
        )
        
        ids.update(tbl['ms1hg'][oi] -
                   set(tbl['ms1fa'][oi].keys()))
        
        return ids
    
    def ms2_headgroups(self, proteins = None):
        """
        Creates dictionaries named ms2hg having the
        original IDs as keys and the sets of the
        identified possible headgroups as values.
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mode, tbl in iteritems(d):
                
                tbl['ms2hg'] = {}
                tbl['ms2fa'] = {}
                tbl['ms2fai'] = {}
                tbl['ms2fas'] = {}
                hgfrags = self.pHgfrags \
                    if mode == 'pos' else self.nHgfrags
                headgroups = self.pHeadgroups \
                    if mode == 'pos' else self.nHeadgroups
                result = {}
                for oi, ms2r in iteritems(tbl['ms2r']):
                    hgroups = self.ms2_headgroup2(ms2r, hgfrags, headgroups)
                    ms2fa, ms2fai, ms2fas = self.ms2_fattya(ms2r)
                    tbl['ms2hg'][oi] = hgroups
                    tbl['ms2fa'][oi] = ms2fa
                    tbl['ms2fai'][oi] = ms2fai
                    tbl['ms2fas'][oi] = ms2fas

    def ms2_headgroup(self, ms2r, hgfrags, headgroups):
        """
        Identifies headgroups from MS2 results for one
        feature, based on dictionary of fragments and
        the characteristic combinations of fragments
        identifying headgroups.
        """
        hgroups = None
        frags = set([])
        # collecting all possible headgroups for
        # each fragment into `hgroups` and all
        # fragments into `frags`
        for ms2item in ms2r:
            if ms2item[0] in hgfrags:
                hgroups = hgfrags[ms2item[0]] if hgroups is None \
                    else hgroups & hgfrags[ms2item[0]]
                frags.add(ms2item[0])
        # if any headgroup related fragment found
        if hgroups is not None and len(hgroups) > 0:
            missingFrags = set([])
            for hg in hgroups:
                if len(headgroups[hg] - frags) > 0:
                    missingFrags.add(hg)
            hgroups = hgroups - missingFrags if \
                len(hgroups - missingFrags) > 0 else hgroups
        return hgroups

    def ms2_headgroup2(self, ms2r, hgfrags, headgroups):
        """
        Identifies headgroups from MS2 results for one
        feature, based on dictionary of fragments and
        the characteristic combinations of fragments
        identifying headgroups.
        """
        _hgroups = []
        _hgroups2 = []
        hgroups = set([])
        frags = set([])
        # collecting all possible headgroups for
        # each fragment into `hgroups` and all
        # fragments into `frags`
        for ms2item in ms2r:
            if ms2item[0] in hgfrags:
                _hgroups.append(hgfrags[ms2item[0]])
                frags.add(ms2item[0])
        # if any headgroup related fragment found
        if len(_hgroups) > 0:
            for hgs in _hgroups:
                added = False
                for i, hgs2 in enumerate(_hgroups2):
                    if len(hgs & hgs2) > 0:
                        _hgroups2[i] = hgs & hgs2
                        added = True
                if not added:
                    _hgroups2.append(hgs)
        for hgs in _hgroups2:
            hgroups = hgroups | hgs
        return hgroups

    def headgroups_negative_positive(self, ms):
        """
        Creates dictionaries named ms1hg_pos, ms1hg_neg,
        ms2hg_pos or ms2hg_neg with the original
        IDs of the given mode as keys, with dicts as
        values having the original IDs of the other mode
        as keys and the combined set of headgroups as 
        values. The combined set is the intersection of
        those detected in the 2 modes, or the union, if
        there is no intersection.
        """
        for protein, d in iteritems(self.valids):
            d['pos']['%shg_neg'%ms] = {}
            d['neg']['%shg_pos'%ms] = {}
            for poi, nois in iteritems(d['pos']['neg']):
                if poi in d['pos']['%shg'%ms]:
                    if d['pos']['%shg'%ms][poi] is not None:
                        for noi in nois.keys():
                            if noi in d['neg']['%shg'%ms]:
                                if d['neg']['%shg'%ms][noi] is not None:
                                    if poi not in d['pos']['%shg_neg'%ms]:
                                        d['pos']['%shg_neg'%ms][poi] = {}
                                    if noi not in d['neg']['%shg_pos'%ms]:
                                        d['neg']['%shg_pos'%ms][noi] = {}
                                    poshg = d['pos']['%shg'%ms][poi]
                                    neghg = d['neg']['%shg'%ms][noi]
                                    combined = poshg & neghg \
                                        if len(poshg & neghg) > 0 \
                                        else poshg | neghg
                                    d['pos']['%shg_neg'%ms][poi][noi] = \
                                        combined
                                    d['neg']['%shg_pos'%ms][noi][poi] = \
                                        combined

    def ms2_fattya(self, ms2r, highest = 2):
        """
        Identifies the fatty acids with highest intensities.
        Returns only number of `highest` fatty acids.
        """
        recnum = re.compile(r'.*[^0-9]([0-9]+:[0-9]+).*')
        # reverse sort by intensities
        ms2fa = set([])
        ms2fai = []
        ms2fas = None
        # if we have MS2 data at all:
        if ms2r.shape[0] > 0:
            ms2r = ms2r[ms2r[:,3].argsort()[::-1],:]
            fanum = 0
            for ms2f in ms2r:
                if 'FA' in ms2f[0] or 'Lyso' in ms2f[0]:
                    fa = recnum.match(ms2f[0])
                    fa = fa.groups()[0]
                    if fa not in ms2fa:
                        # a set with fatty std formulas
                        # [carbon count]:[unsaturated count]
                        ms2fa.add(fa)
                        # a list with intensities decreasing
                        ms2fai.append((fa, ms2f[3]))
                        fanum += 1
                if fanum == highest:
                    break
            # sum formula
            if len(ms2fa) == highest:
                carbs = 0
                unsats = 0
                for fa in ms2fa:
                    carb, unsat = map(int, fa.split(':'))
                    carbs += carb
                    unsats += unsat
                ms2fas = '%u:%u' % (carbs, unsats)
        return ms2fa, ms2fai, ms2fas
    
    def ms2_cleanup(self, proteins = None):
        """
        Removes all MS2 related objects for all proteins or
        only those in list `proteins` in order to free up
        memory.
        """
        
        #to_remove = ['ms2hg', 'ms2fa', 'ms2f',
        #             'ms2fai', 'ms2r', 'ms2i',
        #             'ms2fas', 'ms2fai', 'ms2i2',
        #             'ms2hg2', 'ms2i3', 'ms2']
        
        to_remove = ['ms2f', 'ms2r', 'ms2']
        
        proteins = list(self.valids.keys()) if proteins is None else proteins
        
        for protein in proteins:
            
            del self.ms2map[protein]
            
            for tbl in self.valids[protein].values():
                
                for key in to_remove:
                    
                    if key in tbl:
                        del tbl[key]
    
    def ms2_rt(self, proteins = None):
        
        proteins = list(self.valids.keys()) if proteins is None else proteins
        
        for protein in proteins:
            
            for mode, tbl in iteritems(self.valids[protein]):
                
                rtms2  = []
                drtms2 = []
                
                for i, oi in enumerate(tbl['i']):
                    
                    if oi in tbl['ms2f']:
                        
                        ms2f = tbl['ms2f'][oi]
                        closest_ms2_rt = ms2f.scans[ms2f.best_scan][0,11]
                        rtms2.append(closest_ms2_rt)
                        drtms2.append(abs(np.mean(tbl['rt'][i,:]) -
                                          closest_ms2_rt))
                    
                    else:
                        rtms2.append(np.nan)
                        drtms2.append(np.nan)
                
                tbl['ms2rt']   = np.array(rtms2)
                tbl['deltart'] = np.array(drtms2)
    
    """
    END: MS2 functions
    """
    
    """
    BEGIN: New identification methods
    """
    
    def ms2_scans_identify(self, proteins = None,
                           silent = False, identify = True):
        
        if proteins is not None and len(proteins) == 1:
            silent = True
        
        if not silent:
            prg = progress.Progress(len(self.valids) * 2,
                'Analysing MS2 scans and identifying features', 1, percent = False)
        
        logdir = 'ms2log_%s' % self.today()
        
        if not os.path.isdir(logdir):
            os.mkdir(logdir)
        else:
            map(
                lambda f:
                    os.remove(os.path.join(logdir,f)),
                os.listdir(logdir)
            )
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mode, tbl in iteritems(d):
                
                logfile = '%s-%s.txt' % (protein, mode)
                self.ms2log = os.path.join(logdir, logfile)
                
                if not silent:
                    prg.step()
                
                tbl['ms2f'] = {}
                tbl['ms2i'] = {}
                tbl['ms2i2'] = {}
                for i, oi in enumerate(tbl['i']):
                    if oi in tbl['ms2']:
                        tbl['ms2f'][oi] = ms2.Feature(self, protein, mode, oi)
                        
                        if identify:
                            
                            tbl['ms2f'][oi].identify()
                            tbl['ms2f'][oi].identify2()
                            tbl['ms2i'][oi] = tbl['ms2f'][oi].identities
                            tbl['ms2i2'][oi] = tbl['ms2f'][oi].identities2
        
        if not silent:
            prg.terminate()
    
    def ms2_get_hgs(self):
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                tbl['ms2i3'] = {}
                for oi, ms2results in iteritems(tbl['ms2i2']):
                    tbl['ms2i3'][oi] = set([])
                    for hg, ms2res in iteritems(ms2results):
                        if sum(map(lambda i: i['score'], ms2res)) > 0:
                            if hg in tbl['ms1hg'][oi]:
                                tbl['ms2i3'][oi].add(hg)
    
    """
    END: New identification methods
    """
    
    """
    Pipeline elements
    """
    
    def write_out(self, matches, fname):
        """
        In:
        [0] pos_m/z, pos_profile_score, 
        [2] pos_control_profile_score, pos_rank_profile_boolean,
        [4] pos_ubiquity_score, pos_ubiquity_score,
        [6] pos_original_index, pos_swisslipids_ac, pos_level, 
        [9] pos_lipid_name, pos_lipid_formula, pos_adduct, pos_adduct_m/z
        [13] neg_m/z, neg_profile_score, neg_control_profile_score,
        [16] neg_rank_profile_boolean,
        [17] neg_ubiquity_score, neg_ubiquity_score, neg_original_index,
        [20] neg_swisslipids_ac, neg_level,
        [22] neg_lipid_name, neg_lipid_formula, neg_adduct, neg_adduct_m/z
        """
        with open(fname, 'w') as f:
            hdr = ['LTP',
                'Positive_m/z', 'Positive_m/z_in_SwissLipids',
                'Positive_adduct',
                'Negative_m/z', 'Negative_m/z_in_SwissLipids', 
                'Negative_adduct',
                'SwissLipids_AC', 'SwissLipids_formula', 'SwissLipids_name']
            f.write('\t'.join(hdr) + '\n')
            for ltp, tbl in iteritems(matches):
                for l in tbl:
                    f.write('\t'.join([ltp, '%08f'%l[0],
                        '%08f'%l[12], l[11], '%08f'%l[13],
                        '%08f'%l[25], l[24], l[7], l[10], str(l[9])]) + '\n')
    
    def counts_redundancy_table(self, lipids, unknowns):
        with open('unique_features_counts.csv', 'w') as f:
            f.write('\t'.join(['LTP-mode', 'unknown_features', 
                'lipid_matching_features', 'lipids']) + '\n')
            f.write('\n'.join('%s\t%u\t%u\t%u'%i \
                for i in zip(
                ['%s-%s'%(b,a) for b in unknowns.keys() \
                    for a in unknowns[b].keys()],
                [len(uniqList(list(a[:,7]))) for b in lipids.values() \
                    for a in b.values()], 
                [len(uniqList(list(a[:,7]))) for b in unknowns.values() \
                    for a in b.values()],
                [len(list(a[:,7])) for b in lipids.values() \
                    for a in b.values()])))
    
    """
    These are the highest level methods for reading all data and annotations,
    making sure everything is ready to run the analysis.
    """
    
    def init_from_scratch(self, save = False):
        """
        Does all the initial preprocessing.
        Saves intermediate data, so it can be loaded faster 
        for next sessions.
        """
        self.get_filenames()
        # at first run, after reading from saved textfile
        self.read_annotations()
        self.pp2()
        self.write_pptable()
        if 'ctrl' in self.datafiles:
            del self.datafiles['ctrl']
        if save:
            self.save()
        self.read_data()
        self.check_missing_gel()
        if save:
            self.save_data()

    def init_reinit(self, data = False):
        """
        Initializing from preprocessed and dumped data.
        Pickle file has a 2.0GB size.
        """
        self.load()
        if data:
            self.load_data()
        self.read_annotations()
        self.pp2()
    
    #
    # Reading annotations
    #
    
    def read_annotations(self):
        """
        Reads all additional annotations necessary for data processing.
        These are:
        - fractions: which fraction is protein containing,
        control, or have not been measured
        - lipid names: short notations, database keywords and most
        abundant adducts for lipid classes
        - binding properties: known binders of lipid 
        
        """
        self.read_wrong_fractions()
        self.read_fractions()
        self.upper_fractions()
        self.read_lipid_names()
        self.read_binding_properties()
    
    def read_fractions(self):
        """
        Reads from file sample/control annotations
        for each proteins.
        Also reads the highest fractions which is used if we use
        the profile evaluation method from Enric.
        """
        
        refr = re.compile(r'([A-Z])([0-9]{1,2})-?([A-Z]?)([0-9]{0,2})')
        
        def get_names(names):
            
            return \
                filter(
                    lambda n:
                        n.upper() == n,
                    map(
                        lambda n:
                            n.replace(')', '').strip(),
                        names.split('(')
                    )
                )
        
        def get_fractions(frlist):
            
            return \
                list(
                    itertools.chain(
                        *map(
                            lambda i:
                                map(
                                    lambda num:
                                        '%s%s' % (i[0], num),
                                    xrange(int(i[1]), int(i[3]) + 1)
                                ) if len(i[2]) and len(i[3]) else \
                                ['%s%s' % (i[0], i[1])],
                            refr.findall(frlist)
                        )
                    )
                )
        
        def get_hpeak(hfracs):
            
            return (
                list(
                    map(
                        lambda hf:
                            hf.split('-'),
                        hfracs.split(':')
                    )
                )
            )
        
        if not self.pcont_fracs_from_abs:
            
            data = {}
            
            with open(self.fractionsf, 'r') as fp:
                
                hdr = fp.readline().strip()
                
                col2fr = dict(
                    map(
                        lambda fr:
                            (fr[0], fr[1].upper()),
                        enumerate(hdr.split(',')[1:])
                    )
                )
                
                for l in fp:
                    
                    l = l.strip().split(',')
                    
                    data[l[0].replace('"', '').upper()] = (
                    dict(
                            map(
                                lambda x:
                                    (col2fr[x[0]], self.to_int(x[1])),
                                filter(
                                    lambda x:
                                        x[1] != '',
                                    enumerate(l[1:])
                                )
                            )
                        )
                    )
            
            self.fractions = data
            
        else:
            
            if self.fractionsf.endswith('xlsx'):
                
                tab = self.read_xls(self.fractionsf, sheet = 'status')
                
                self._fractions, self.hpeak = (
                    tuple(
                        map(
                            dict,
                            zip(
                                *itertools.chain(
                                    *map(
                                        lambda l:
                                            list(
                                                map(
                                                    lambda n:
                                                        [
                                                            (
                                                                n, # protein name
                                                                get_fractions(l[15])
                                                            ),
                                                            (
                                                                n, # protein name
                                                                get_hpeak(l[18])
                                                            )
                                                        ],
                                                    get_names(l[0])
                                                )
                                            ),
                                        filter(
                                            lambda l:
                                                len(l[15]) and \
                                                    l[15].strip() != 'NA' and (
                                                        l[15][0].upper() ==
                                                        l[15][0]) and (
                                                    l[15][1].isdigit()),
                                            tab[1:]
                                        )
                                    )
                                )
                            )
                        )
                    )
                )
                
            else:
                
                with open(self.fractionsf, 'r') as fp:
                    
                    self._fractions = \
                        dict(
                            map(
                                lambda l:
                                    (l[0], l[1:]),
                                map(
                                    lambda l:
                                        l.split(';'),
                                    filter(
                                        len,
                                        fp.read().split('\n')
                                    )
                                )
                            )
                        )
    
    def read_wrong_fractions(self):
        """
        Reads the list of potentially wrong fractions.
        """
        result = {}
        
        if not os.path.exists(self.wrongfracsf):
            
            if self.filter_wrong_fractions:
                
                sys.stdout.write('\t:: Filtering of wrong fractions enabled, '
                                'but input file missing: `%s`.\n' % (
                                self.wrongfracsf))
            
            return None
        
        with open(self.wrongfracsf, 'r') as fp:
            
            _ = fp.readline()
            
            for line in fp:
                
                if not len(line):
                    continue
                
                line = line.split(';')
                
                _ = result.setdefault(line[0], {})
                
                for i, mode in enumerate(['neg', 'pos']):
                    
                    if len(line[i + 1].strip()):
                        wfracs = line[i + 1].strip().replace('?', '').split(',')
                        result[line[0]][mode] = wfracs
        
        self.wrong_fracs = result
    
    def nonzero_in_wrong(self):
        """
        Creates a boolean attribute which is True at features which are
        nonzero in fractions marked as potentially wrong.
        """
        
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                tbl['wnz'] = np.array([False] * tbl['fe'].shape[0])
                
                if (protein not in self.wrong_fracs or
                    mode not in self.wrong_fracs[protein]):
                    continue
                
                ifracs = self.fraction_indices(protein)
                
                for wfrac in self.wrong_fracs[protein][mode]:
                    
                    col = ifracs[wfrac][0]
                    
                    tbl['wnz'] = tbl['fe'][:,col] > 0
    
    def set_measured(self):
        """
        Creates a dict of arrays with value 1 if the fracion is either
        protein containing or control (i.e. measured), or `None` otherwise.
        
        The result is stored in `measured` attribute.
        """
        self.measured = dict((
                k,
                np.array([1 if x is not None else None for x in v])
            )
            for k, v in iteritems(self.fractions))

    def upper_fractions(self):
        """
        Creates the dict ``fractions_upper`` of fractions with
        uppercase LTP names, with same content as ``fractions``.
        """
        if hasattr(self, 'fractions'):
            self.fractions_upper = \
                dict((l.upper(), s) for l, s in iteritems(self.fractions))
    
    def set_onepfraction(self):
        """
        Creates a list of those proteins present only in one fraction.
        Abundance ratio for these can not be calculated.
        
        The result stored in ``onepfraction`` attribute.
        """
        self.onepfraction = [k.upper() for k, v in iteritems(self.fractions) \
            if sum((i for i in v.values() if i is not None)) == 1]
    
    def read_lipid_names(self, add_fa = True):
        """
        Reads annotations for lipid classes:
        - full names
        - short notations
        - database keywords
        (to process long names from SwissLipids and LipidMaps)
        - most abundant adducts
        
        The input file is given by the ``lipnamesf`` attribute.
        """
        result = {}
        with open(self.lipnamesf, 'r') as f:
            nul = f.readline()
            for l in f:
                l = l.strip().split('\t')
                result[l[0]] = {
                    'full_name': l[1],
                    'swl': self.process_db_keywords(l[2]),
                    'lmp': self.process_db_keywords(l[3]),
                    'pos_adduct': l[4] if l[4] != 'ND' and self.adducts_constraints else None,
                    'neg_adduct': l[5] if l[5] != 'ND' and self.adducts_constraints else None
                }
        
        result['FA'] = {'full_name': 'Fatty acid', 'swl': [], 'lmp': [],
                        'pos_adduct': None, 'neg_adduct': None}
        
        self.lipnames = result

    def read_binding_properties(self):
        """
        Reads the known binders for each lipid class.
        The input file name is provided by `bindpropf` attribute.
        """
        result = {}
        with open(self.bindpropf, 'r') as f:
            data = \
                list(
                    map(
                        lambda l:
                            l.strip('\n').split('\t'),
                        list(
                            filter(lambda l:
                                len(l),
                                f
                            )
                        )[1:]
                    )
                )
        
        for l in data:
            if l[2] not in result:
                result[l[2]] = set([])
            try:
                for lip in l[6].split(';'):
                    if lip != 'ND' and lip != '':
                        lip = lip if lip != 'Cer1P' else 'CerP'
                        result[l[2]].add(lip)
            except IndexError:
                print(l)
        
        self.bindprop = result
    
    #
    # END: Reading annotations
    #
    
    def basic_filters(self, profile_treshold = 0.25, ubiquity_treshold = 7):
        """
        Deprecated with new data structure.
        """
        self.apply_filters()
        self.validity_filter()
        self.profile_filter()
        self.profile_filter(prfx = 'c')
        self.ubiquity_filter()
        self.val_ubi_filter(ubiquity = ubiquity_treshold)
        self.rprofile_filter()
        self.val_prf_filter(treshold = profile_treshold)
        self.val_rpr_filter()
        self.val_ubi_prf_filter(treshold = profile_treshold,
            ubiquity = ubiquity_treshold)
        self.val_ubi_prf_rprf_filter(treshold = profile_treshold,
            ubiquity = ubiquity_treshold)

    def basic_filters_with_evaluation(self,
        profile_treshold = 0.25, ubiquity_treshold = 7):
        """
        Deprecated with new data structure.
        """
        filtr_results = {}
        for f in ['quality', 'charge', 'area', 'peaksize', 'validity']:
            filtr_results[f] = self.eval_filter(f)
        filtr_results['ubiquity'] = self.eval_filter('ubiquity',
        param = {'only_valid': False},
        runtime = True, repeat = 1, number = 1,
        hit = lambda x: x < ubiquity_treshold)
        filtr_results['val_ubi'] = eval_filter('val_ubi',
            param = {'ubiquity': ubiquity_treshold})
        filtr_results['profile025'] = eval_filter('profile',
            param = {'prfx': ''},
            runtime = True, repeat = 1, number = 1, 
            hit = lambda x: x <= profile_treshold)
        filtr_results['cprofile025'] = self.eval_filter('cprofile',
            runtime = True, repeat = 1, number = 1, 
            hit = lambda x: x <= profile_treshold)
        filtr_results['rprofile'] = self.eval_filter('rprofile',
            runtime = True)
        filtr_results['val_prf'] = self.eval_filter('val_prf',
            param = {'treshold': profile_treshold})
        filtr_results['val_rpr'] = self.eval_filter('val_rpr')
        filtr_results['val_ubi_prf'] = self.eval_filter('val_ubi_prf',
            param = {'treshold': profile_treshold,
                     'ubiquity': ubiquity_treshold})
        filtr_results['val_ubi_prf_rprf'] = \
            self.eval_filter('val_ubi_prf_rprf',
            param = {'treshold': profile_treshold,
                     'ubiquity': ubiquity_treshold})
        self.filter_results = filtr_results

    def positive_negative_runtime(self, ):
        return timeit.timeit('negative_positive(lipids)', 
            setup = 'from __main__ import negative_positive, lipids',
            number = 1)

    def valid_features(self, cache = False):
        """
        Creates new dict of arrays with only valid features.
        Keys:
        - 'fe': features
        - 'mz': m/z values
        - 'i': original index
        
        """
        
        self.fractions_marco()
        
        if cache and os.path.exists(self.validscache):
            with open(self.validscache, 'rb') as fp:
                self.valids = pickle.load(fp)
            
            return None
        
        self.apply_filters()
        self.validity_filter()
        self.valids = dict((protein.upper(), {'pos': {}, 'neg': {}}) \
            for protein in self.data.keys())
        
        for protein, d in iteritems(self.data):
            
            protein = protein.upper()
            
            for pn, tbl in iteritems(d):
                
                self.valids[protein][pn]['fe'] = \
                    np.array(tbl['smp'][tbl['vld']])
                
                self.valids[protein][pn]['mz'] = \
                    np.array(tbl['ann'][tbl['vld'], 2])
                
                self.valids[protein][pn]['qua'] = \
                    np.array(tbl['ann'][tbl['vld'], 0])
                
                self.valids[protein][pn]['sig'] = \
                    np.array(tbl['ann'][tbl['vld'], 1])
                
                self.valids[protein][pn]['z'] = \
                    np.array(tbl['ann'][tbl['vld'], 5])
                
                self.valids[protein][pn]['rt'] = \
                    np.array(tbl['ann'][tbl['vld'], 3:5])
                
                self.valids[protein][pn]['aa'] = \
                    np.array(tbl['aa'][tbl['vld']])
                
                for key in ['peaksize', 'pslim02', 'pslim05',
                    'pslim10', 'pslim510', 'rtm']:
                    self.valids[protein][pn][key] = np.array(tbl[key][tbl['vld']])
                self.valids[protein][pn]['i'] = np.where(tbl['vld'])[0]
        
        self.norm_all()
        self.get_area()
        
        with open(self.validscache, 'wb') as fp:
            pickle.dump(self.valids, fp)
    
    def get_area(self):
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                try:
                    tbl['are'] = np.nanmax(tbl['fe'], 1)
                except ValueError:
                    sys.stdout.write('\t:: WARNING: Could not'\
                        ' calculate area: %s, %s\n' % (protein, mode))
    
    def data2valids(self, key):
        for protein_l, dd in iteritems(self.data):
            for mode, tbld in iteritems(dd):
                if key in tbld:
                    tbl = self.valids[protein_l.upper()][mode]
                    valids_array = []
                    for oi in tbl['i']:
                        valids_array.append(tbld[key][(oi,)])
                    tbl[key] = np.array(valids_array)
    
    def delete_array(self, key):
        """
        Deletes the data belonging to one key
        from all tables (at every protein in each modes).
        """
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                if key in tbl:
                    del tbl[key]

    def norm_all(self):
        """
        Creates table with all the profiles normalized
        Keys:
        - 'no': normalized profiles
        """
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                tbl['no'] = self.norm_profiles(tbl['fe'])

    """
    END: pipeline elements
    """

    """
    Distance metrics
    """
    
    def profiles_corr(self, metric, prfx, pprofs = ''):
        """
        Calculates custom correlation metric
        between each feature and the protein profile.
        
        This method is deprecated, should not be used.
        """
        
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        pprs = getattr(self, 'pprofs%s' % pprofs)
        for protein, d in iteritems(self.valids):
            ppr = \
                np.array(
                    map(
                        lambda s:
                            pprs[protein][frs[s[1]]],
                        filter(
                            lambda s:
                                s[0] != 0 and s[1] is not None,
                            enumerate(self.fractions_upper[protein])
                        )
                    )
                )
            ppr = self.norm_profile(ppr).astype(np.float64)
            for pn, tbl in iteritems(d):
                if 'no' not in tbl:
                    self.norm_all()
                tbl['%sv%s' % (prfx, pprofs)] = np.zeros((tbl['no'].shape[0],),
                    dtype = np.float64)
                tbl['%sp%s' % (prfx, pprofs)] = np.zeros((tbl['no'].shape[0],),
                    dtype = np.float64)
                for i, fe in enumerate(tbl['no']):
                    # if one feature is not detected
                    # in any fraction of the sample,
                    # it will have a nan value:
                    if np.any(np.isnan(\
                            fe[np.where(\
                                [fr == 1 \
                                    for fr in self.fractions_upper[protein][1:] \
                                    if fr is not None]
                            )]
                        )):
                        vp = (np.nan, 0.0)
                    else:
                        vp = metric(fe, ppr)
                    tbl['%sv%s' % (prfx, pprofs)][i] = vp[0]
                    tbl['%sp%s' % (prfx, pprofs)][i] = vp[1]

    def gkgamma(self, x, y):
        """
        Calls Goodman-Kruskal's gamma from vcdExtra
        R package.
        """
        gkg = rvcd.GKgamma(rbase.matrix(rbase.c(*(list(x) + list(y))),
            nrow = 2))
        # gamma, 0.0, C, D, sigma, CIlevel, CI
        return tuple([gkg[0][0], 0.0] + [i[0] for i in gkg[1:]])

    def roco(self, x, y):
        """
        Calls R function robust correlation coefficient
        with test from rococo R package.
        """
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        if _x.size == 0 or _y.size == 0 or \
            np.all(_x == _x[0]) or np.all(_y == _y[0]):
            return 0.0, 0.0
        _x = rbase.c(*_x)
        _y = rbase.c(*_y)
        rt = rococo.rococo_test( \
            rbase.as_vector(rstats.na_omit(_x.rx( \
                rbase.c(*list(np.where(\
                    map(lambda i: not i, list(rbase.is_na(_y))))[0] + 1))))), \
            rbase.as_vector(rstats.na_omit(_y.rx( \
                rbase.c(*list(np.where(\
                    map(lambda i: not i, list(rbase.is_na(_x))))[0] + 1))))))
        return rt.slots['sample.gamma'][0], rt.slots['p.value'][0]

    def _diff_profiles(self, x, y):
        """
        Wrapper for diff_profiles() to return a tuple.
        """
        return self.diff_profiles(x, y), 0.0

    def euclidean_dist(self, x, y):
        """
        Calculates simple euclidean distance after removing NaN values.
        """
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        return (sp.spatial.distance.euclidean(_x, _y), 0.0) \
            if len(_x) > 0 else (np.inf, 0.0)

    def euclidean_dist_norm(self, x, y):
        """
        This euclidean distance is normalized by the number of dimensions.
        """
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        return (sp.spatial.distance.euclidean(_x, _y) / len(_x), 0.0) \
            if len(_x) > 0 else (1.0, 0.0)

    def profiles_corrs(self, pprofs = ''):
        """
        Calculates an array of similarity/correlation
        metrics between each MS intensity profile and 
        the corresponding protein concentration profile.
        """
        metrics = [
            (stats.spearmanr, 'sp'),
            (stats.kendalltau, 'kt'),
            (self._diff_profiles, 'df'),
            (self.gkgamma, 'gk'),
            (self.roco, 'rc'),
            (self.euclidean_dist, 'eu'),
            (self.euclidean_dist_norm, 'en'),
            (stats.pearsonr, 'pe'),
            (self._comp_profiles, 'cp')
        ]
        for metric, prfx in metrics:
            sys.stdout.write('Calculating %s\n' % metric.__name__)
            self.profiles_corr(metric, prfx, pprofs)

    def inconsistent(self, valids, metric = 'en'):
        attr = '%sv'%metric
        target = '%si'%metric
        sort_alll(valids, attr)
        for protein, d in iteritems(valids):
            for pn, tbl in iteritems(d):
                vals = tbl[attr]
                incs = np.diff(vals)
                inco = np.array(
                    [0.0] * 3 + \
                    map(lambda i:
                        incs[i] - np.mean(incs[:i]) / np.std(incs[:i]),
                        xrange(2, len(incs))
                    )
                )
                tbl['%sde'%metric] = np.concatenate((np.array([0.0]), incs),
                    axis = 0)
                tbl['%sdde'%metric] = np.concatenate((np.array([0.0]), 
                    np.diff(tbl['%sde'%metric])), axis = 0)
                tbl[target] = inco

    """
    END: disctance metrics
    """

    """
    Functions for clustering
    """

    def distance_matrix(self, metrics = ['eu'], with_pprof = False, proteins = None):
        _metrics = {
            'eu': ('euclidean distance', self.euclidean_dist),
            'en': ('normalized euclidean distance', self.euclidean_dist_norm)
        }
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        t0 = time.time()
        for m in metrics:
            prg = progress.Progress(
                len(self.valids) * 2 if proteins is None else len(proteins) * 2,
                'Calculating %s' % _metrics[m][0],
                1, percent = False)
            for protein, d in iteritems(self.valids):
                if proteins is None or protein in proteins:
                    if with_pprof:
                        ppr = np.array([self.pprofs[protein.upper()][frs[i]] \
                            for i, fr in enumerate(self.fractions_upper[protein]) \
                                if i != 0 and fr is not None])
                        ppr = self.norm_profile(ppr).astype(np.float64)
                    for pn, tbl in iteritems(d):
                        prg.step()
                        fnum = tbl['no'].shape[0]
                        # square shape matrix of all features vs all features
                        if with_pprof:
                            tbl['_%sd'%m] = np.empty((fnum + 1, fnum + 1))
                        else:
                            tbl['_%sd'%m] = np.empty((fnum, fnum))
                        # to keep track the order accross sorting
                        tbl['_%so'%m] = np.copy(tbl['i'])
                        for i in xrange(fnum):
                            for j in xrange(fnum):
                                tbl['_%sd'%m][i,j] = \
                                    _metrics[m][1](tbl['no'][i,:],
                                        tbl['no'][j,:])[0]
                        if with_pprof:
                            for i in xrange(fnum):
                                ppr_dist = _metrics[m][1](tbl['no'][i,:], ppr)[0]
                                tbl['_%sd'%m][i,-1] = ppr_dist
                                tbl['_%sd'%m][-1,i] = ppr_dist
                            tbl['_%sd'%m][-1,-1] = _metrics[m][1](ppr, ppr)[0]
        prg.terminate()
        sys.stdout.write('\t:: Time elapsed: %us\n'%(time.time() - t0))
        sys.stdout.flush()

    def features_clustering(self, dist = 'en', method = 'ward', proteins = None):
        """
        Using the distance matrices calculated by
        `distance_matrix()`, builds clusters using
        the linkage method given by `method` arg.
        """
        prg = progress.Progress(len(self.valids)*2 \
            if proteins is None else len(proteins)*2,
            'Calculating clusters', 1, percent = False)
        for protein, d in iteritems(self.valids):
            if proteins is None or protein in proteins:
                for pn, tbl in iteritems(d):
                    prg.step()
                    tbl['_%sc'%dist] = fastcluster.linkage(tbl['_%sd'%dist],
                        method = method, metric = 'euclidean',
                            preserve_input = True)
        prg.terminate()

    def distance_corr(self, valids, dist = 'en'):
        for protein, d in iteritems(valids):
            for pn, tbl in iteritems(d):
                tbl['_%sdc'%dist] = np.array([
                        sp.stats.pearsonr(tbl['_%sd'%dist][:,i],
                            tbl['_%sd'%dist][:,-1])[0] \
                    for i in xrange(tbl['_%sd'%dist].shape[1])])

    def fcluster_with_protein(self, lin, t):
        fc = sp.cluster.hierarchy.fcluster(lin, t, criterion = 'distance')
        return np.where(fc == fc[-1])[0]

    def nodes_in_cluster(self, lin, i):
        n = lin.shape[0]
        nodes = map(int, lin[i,:2])
        singletons = set(filter(lambda x: x <= n, nodes))
        upper_levels = set(nodes) - singletons
        for u in upper_levels:
            singletons = singletons | self.nodes_in_cluster(lin, u - n - 1)
        return singletons

    def _get_link_colors_dist(self, tbl, dist, threshold,
        highlight_color, base_color):
        protein_fc = set(self.fcluster_with_protein(tbl['_%sc' % dist], threshold))
        return dict(zip(
            xrange(tbl['_%sc'%dist].shape[0] + 1,
                   tbl['_%sc'%dist].shape[0] * 2 + 2),
            map(lambda x: 
                highlight_color \
                    if set(self.nodes_in_cluster(tbl['_%sc' % dist], x)) <= \
                        protein_fc \
                    else base_color,
            xrange(tbl['_%sc'%dist].shape[0])) + \
                [highlight_color \
                    if len(protein_fc) == tbl['_%sd'%dist].shape[0] \
                    else base_color]
        ))

    def _get_link_colors_corr(self, tbl, dist, cmap, threshold,
        highlight_color, base_color):
        return dict(zip(
            xrange(tbl['_%sc'%dist].shape[0] + 1,
                tbl['_%sc'%dist].shape[0]*2 + 2), # 180
            map(lambda col:
                '#%s' % ''.join(map(lambda cc: '%02X'%(cc*255), col[:3])) \
                    if type(col) is tuple else col,
                map(lambda c:
                    # color from correlation value:
                    #'#%s%s00' % ('%02X'%(c*255), '%02X'%(c*255)) \
                    cmap(c) if c > 0.0 else base_color, # '#FEFE00'
                    map(lambda x:
                        # minimum of these correlations
                        # for each link:
                        min(map(lambda xx:
                            # correlation of distances for one link:
                            tbl['_%sdc'%dist][xx], # 0.9999928, 0.9999871
                            # nodes for each link:
                            list(nodes_in_cluster(tbl['_%sc' % dist], x))
                        )),
                        # all links (rows in linkage matrix):
                        xrange(tbl['_%sc'%dist].shape[0]) # 45
                    ) + \
                    # this is for the root:
                    [min(tbl['_%sdc'%dist])]
                )
            )
        ))

    def _get_dendrogram_cmap(self, cmap, threshold,
        highlight_color, base_color):
        if cmap is None:
            if thershold is None:
                _cmap = 'inferno'
            else:
                _cmap = lambda x: \
                    highlight_color if x >= threshold else base_color
        elif type(cmap) is str and hasattr(mpl.cm, cmap):
                _cmap = mpl.cm.get_cmap(cmap)
        return _cmap

    def _cluster_size_threshold(self, tbl, dist, threshold):
        dist_values = tbl['_%sc'%dist][:,2]
        dist_values = dist_values[np.where(dist_values > 0.0)]
        if threshold is None:
            threshold = len(dist_values) * 0.1
        _max = np.max(dist_values)
        _min = np.min(dist_values)
        _lower = _min
        _upper = _max
        _conv = 9999
        _dconvs = []
        while True:
            _threshold = (_lower + _upper) / 2.0
            cls = sp.cluster.hierarchy.fcluster(tbl['_%sc'%dist], _threshold)
            mean_size = np.mean(collections.Counter(cls).values())
            _over = mean_size > threshold
            if mean_size > threshold:
                _upper = max(_upper - (_upper - _lower) / 2.0, 0.0)
            elif mean_size < threshold:
                _lower = max(_lower + (_upper - _lower) / 2.0, 0.0)
            if _lower >= _upper:
                _upper += (_upper - _lower) / 4.0
            _dconvs.append(_conv - abs(mean_size - threshold))
            _conv = abs(threshold - mean_size)
            if len(_dconvs) > 10 and \
                abs(np.mean(_dconvs[-10:]) - _dconvs[-1]) < 0.005:
                return _threshold

    def _inconsistency_threshold(self, tbl, dist, threshold):
        n = tbl['_%sd'%dist].shape[0]
        if threshold is None:
            threshold = n/20
        clustering = tbl['_%sc'%dist]
        incons = sp.cluster.hierarchy.inconsistent(clustering)
        maxincons = sp.cluster.hierarchy.maxinconsts(clustering, incons)
        fc = sp.cluster.hierarchy.fcluster(clustering, threshold, 
            criterion = 'maxclust_monocrit', monocrit = maxincons)
        nodes = set(np.where(fc == fc[-1])[0])
        print('nodes in same cluster as protein: %u' % len(nodes))
        _threshold = max(
            map(lambda i: 
                clustering[i,2],
                filter(lambda i: 
                    nodes_in_cluster(clustering, i) <= nodes,
                    xrange(n - 1)
                )
            )
        )
        return _threshold

    def _dendrogram_get_threshold(self, tbl, dist, threshold, threshold_type):
        dist_values = tbl['_%sc'%dist][:,2]
        if threshold_type == 'percent':
            # _threshold = dist_values.max() * threshold / tbl['%sc'%dist].shape[0]
            _threshold = 10**(np.log10(dist_values.max()) * threshold)
        elif threshold_type == 'quantile':
            _threshold = 10**(np.percentile(
                np.log10(dist_values)[np.where(dist_values > 0.0)], threshold))
        elif threshold_type == 'clsize':
            _threshold = _cluster_size_threshold(tbl, dist, threshold)
        elif threshold_type == 'incons':
            _threshold = _inconsistency_threshold(tbl, dist, threshold)
        return _threshold

    def plot_heatmaps_dendrograms_gradient(self, *args, **kwargs):
        pass

    def plot_heatmaps_dendrograms(self, dist = 'en', 
        fname = None, proteins = None, cmap = None,
        highlight_color = '#FFAA00', base_color = '#000000',
        coloring = 'corr', threshold = None,
        threshold_type = 'percent',
        save_selection = None, pca = False):
        """
        For each protein plots heatmaps and dendrograms.
        Thanks to http://stackoverflow.com/a/3011894/854988
        """
        all_hgs = set()
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                for ids in tbl['identity'].values():
                    for hg in ids.keys():
                        all_hgs.add(hg)
        hg_cols = dict(map(lambda hg: 
            (hg[1], self.colors[hg[0]]),
            enumerate(sorted(list(all_hgs)))
        ))
        t0 = time.time()
        fname = 'features_clustering-%s%s%s.pdf' % \
                (coloring, 
                '-%s'%threshold_type if threshold is not None else '',
                '-%02f'%threshold if threshold is not None else '') \
            if fname is None else fname
        if coloring == 'corr':
            _cmap = self._get_dendrogram_cmap(cmap, threshold,
                highlight_color, base_color)
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        with mpl.backends.backend_pdf.PdfPages(fname) as pdf:
            prg = progress.Progress(len(self.valids) * 2 \
                if proteins is None else len(proteins)*2,
                'Plotting heatmaps with dendrograms', 1, percent = False)
            for protein, d in iteritems(self.valids):
                if proteins is None or protein in proteins:
                    ppr = np.array([self.pprofs[protein.upper()][frs[i]] \
                        for i, fr in enumerate(self.fractions_upper[protein]) \
                            if i != 0 and fr is not None])
                    ppr = self.norm_profile(ppr).astype(np.float64)
                    for pn, tbl in iteritems(d):
                        prg.step()
                        if coloring == 'dist':
                            _threshold = self._dendrogram_get_threshold(tbl, dist,
                                threshold, threshold_type)
                        labels = ['%u'%(f) for f in tbl['_%so'%dist]] + [protein]
                        names = map(lambda oi:
                            ', '.join(sorted(map(lambda hg:
                                hg[0],
                                filter(lambda hg:
                                    hg[1]['ms1_%s'%pn] and hg[1]['ms2_%s'%pn],
                                    iteritems(tbl['identity'][int(oi)])
                                )
                            ))) or oi,
                            labels[:-1]
                        ) + [protein]
                        
                        
                        if coloring == 'corr':
                            _link_colors = \
                                self._get_link_colors_corr(tbl, dist, _cmap,
                                threshold, highlight_color, base_color)
                        elif coloring == 'dist':
                            _link_colors = self._get_link_colors_dist(tbl, dist,
                                _threshold, highlight_color, base_color)
                        protein_fc = \
                            set(self.fcluster_with_protein(tbl['_%sc' % dist],
                            _threshold))
                        if save_selection is not None:
                            oi2i = \
                                dict(zip(labels[:-1], xrange(len(labels) - 1)))
                            tbl[save_selection] = np.array(
                                map(lambda oi:
                                    oi2i['%u'%oi] in protein_fc,
                                    tbl['i']
                                )
                            )
                        
                        mpl.rcParams['lines.linewidth'] = 0.1
                        mpl.rcParams['font.family'] = 'Helvetica Neue LT Std'
                        fig = mpl.figure.Figure(figsize = (8, 8))
                        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                        gs = mpl.gridspec.GridSpec(2, 2, 
                            height_ratios=[2, 8], width_ratios = [2, 8])
                        
                        # First dendrogram
                        ax1 = fig.add_subplot(gs[1,0])
                        Z1 = hc.dendrogram(tbl['_%sc'%dist],
                            orientation = 'left',
                            labels = names,
                            leaf_rotation = 0, ax = ax1,
                            link_color_func = lambda i: _link_colors[i])
                        ax1.yaxis.grid(False)
                        ax1.set_xscale('symlog')
                        null = [tl.set_fontsize(1.5) \
                            for tl in ax1.get_yticklabels()]
                        null = [(tl.set_color(hg_cols[tl._text]),
                                tl.set_fontweight('bold')) \
                            for tl in ax1.get_yticklabels() \
                                if tl._text in hg_cols]
                        null = [(tl.set_color(highlight_color),
                                tl.set_fontweight('bold'),
                                tl.set_fontsize(4)) \
                            for tl in ax1.get_yticklabels() if tl._text == protein]
                        if _threshold is not None:
                            ax1.axvline(x = _threshold,
                                c = '#FFAA00', alpha = 0.5)
                        
                        # Compute and plot second dendrogram.
                        ax2 = fig.add_subplot(gs[0,1])
                        Z2 = hc.dendrogram(tbl['_%sc'%dist], 
                            labels = names,
                            leaf_rotation = 90, ax = ax2,
                            #color_threshold = _threshold)
                            link_color_func = lambda i: _link_colors[i])
                        ax2.xaxis.grid(False)
                        ax2.set_yscale('symlog')
                        null = [tl.set_fontsize(1.5) \
                            for tl in ax2.get_xticklabels()]
                        null = [(tl.set_color(hg_cols[tl._text]),
                                tl.set_fontweight('bold')) \
                            for tl in ax2.get_xticklabels() \
                                if tl._text in hg_cols]
                        null = [(tl.set_color(highlight_color), 
                                tl.set_fontweight('bold'),
                                tl.set_fontsize(4)) \
                            for tl in ax2.get_xticklabels() if tl._text == protein]
                        if _threshold is not None:
                            ax2.axhline(y = _threshold, c = '#FFAA00',
                                alpha = 0.5)
                        
                        # Plot distance matrix.
                        ax3 = fig.add_subplot(gs[1,1])
                        idx1 = Z1['leaves']
                        idx2 = Z2['leaves']
                        D = tbl['_%sd'%dist][idx1,:]
                        D = D[:,idx2]
                        im = ax3.matshow(D, aspect = 'auto', origin = 'lower',
                            cmap = mpl.cm.get_cmap('Blues'))
                        ax3.xaxis.grid(False)
                        ax3.yaxis.grid(False)
                        ax3.set_xticklabels([])
                        ax3.set_yticklabels([])
                        
                        fig.suptitle('%s :: %s mode\n'\
                            'clustering valid features; '\
                            'features in highlighted cluster: %u' % \
                            (protein, pn, len(protein_fc) - 1),
                            color = '#AA0000' if protein in self.onepfraction else '#000000')
                        
                        cvs.draw()
                        fig.tight_layout()
                        fig.subplots_adjust(top = 0.90)
                        
                        cvs.print_figure(pdf)
                        fig.clf()
                        
                        if pca:
                            oi2i = dict(zip(tbl['i'],
                                xrange(tbl['no'].shape[0])))
                            pca = sklearn.decomposition.PCA(n_components = 2)
                            fe = np.vstack((tbl['no'][map(lambda oi:
                                    oi2i[int(oi)],
                                    labels[:-1]
                                ),:], ppr))
                            col = map(lambda i:
                                highlight_color \
                                    if i in protein_fc else base_color,
                                xrange(len(labels) - 1)
                            ) + ['#3383BE']
                            fe[np.where(np.isnan(fe))] = 0.0
                            pca = pca.fit(fe)
                            coo = pca.transform(fe)
                            fig = mpl.figure.Figure(figsize = (8, 8))
                            cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                            ax = fig.gca()
                            ax.scatter(coo[:,0], coo[:,1], c = col, 
                                linewidth = 0.0, alpha = 0.7)
                            ax.set_title('%s :: %s mode :: PCA' % (protein, pn),
                                color = '#AA0000' \
                                    if protein in onepfraction else '#000000')
                            cvs.draw()
                            fig.tight_layout()
                            cvs.print_figure(pdf)
                            fig.clf()
            
            pdfinf = pdf.infodict()
            pdfinf['Title'] = 'Features clustering'
            pdfinf['Author'] = 'Dénes Türei'.decode('utf-8')
            pdfinf['Subject'] = 'Clustering MS1 features '\
                'based on Euclidean distances'
            pdfinf['Keywords'] = 'lipid transfer protein, protein,'\
                ' lipidomics, mass spectrometry'
            pdfinf['CreationDate'] = datetime.datetime(2016, 2, 22)
            pdfinf['ModDate'] = datetime.datetime.today()
        
        prg.terminate()
        sys.stdout.write('\t:: Time elapsed: %us\n'%(time.time() - t0))
        sys.stdout.write('\t:: Plots saved to %s\n'%fname)
        sys.stdout.flush()

    """
    END: Clustering
    """

    def plot_increment(self, valids, onepfraction, metric = 'en',
        fname = 'increments.pdf'):
        with mpl.backends.backend_pdf.PdfPages(fname) as pdf:
            for protein, d in iteritems(valids):
                for pn, tbl in iteritems(d):
                    fig = mpl.figure.Figure(figsize = (8,8))
                    cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                    ax = fig.gca()
                    ax.plot(xrange(tbl['%si'%metric].shape[0]),
                        tbl['%si'%metric], c = '#FCCC06',
                        label = 'Inconsistency')
                    ax.plot(xrange(tbl['%sde'%metric].shape[0]),
                        tbl['%sde'%metric] * 100, c = '#007B7F',
                        label = '1st derivate',
                        lw = 2, alpha = 0.5)
                    ax.plot(xrange(tbl['%sde'%metric].shape[0]),
                        tbl['%sdde'%metric] * 100, c = '#6EA945',
                        label = '2nd derivate',
                        lw = 2, alpha = 0.5)
                    handles, labels = ax.get_legend_handles_labels()
                    ax.legend(handles, labels)
                    ax.set_title('%s :: %s mode :: distance increments' % \
                        (protein, pn),
                        color = '#AA0000' if protein in onepfraction else '#000000')
                    ax.set_xlim([3.0, 50.0])
                    cvs.print_figure(pdf)
                    fig.clf()
            
            pdfinf = pdf.infodict()
            pdfinf['Title'] = 'Features distance increment'
            pdfinf['Author'] = 'Dénes Türei'.decode('utf-8')
            pdfinf['Subject'] = 'Features distance increment'
            pdfinf['Keywords'] = 'lipid transfer protein, protein,'\
                ' lipidomics, mass spectrometry'
            pdfinf['CreationDate'] = datetime.datetime(2016, 2, 22)
            pdfinf['ModDate'] = datetime.datetime.today()

    def kmeans(self, valids, pprofs, fractions):
        cfracs = ['c0'] + self.fracs
        prg = progress.Progress(len(valids) * 2,
            'Calculating k-means', 1, percent = False)
        for protein, d in iteritems(valids):
            ppr = np.array([pprofs[protein.upper()][cfracs[i]] \
                for i, fr in enumerate(fractions[protein]) if fr == 1 and i != 0])
            ppr = norm_profile(ppr).astype(np.float64)
            for pn, tbl in iteritems(d):
                prg.step()
                with_protein = np.vstack(tbl['no'], ppr)
                whitened = sp.cluster.vq.whiten(with_protein)
                code_book = sp.cluster.vq.kmeans(whitened, 2)
                # tbl['km'] = 
        prg.terminate()

    def read_positives(self, basedir, fname = 'manual_positive.csv'):
        """
        Reads manually annotated positive hits
        from file.
        """
        with open(os.path.join(basedir, fname), 'r') as f:
            _result = [[c.strip() for c in l.split('\t')] \
                for l in f.read().split('\n')]
        result = dict((protein, []) for protein in uniqList([x[0] \
            for x in _result if x[0] != '']))
        for l in _result:
            if l[0] != '':
                result[l[0]].append(l)
        return result

    def spec_sens(self, protein, pos, metric, asc):
        """
        Calculates specificity, sensitivity, precision,
        false discovery rate as a function of critical
        values of a score, for one LTP for one mode.
        Returns dict of lists.
        """
        result = {'spec': [], 'sens': [], 'prec': [],
            'fdr': [], 'cutoff': [], 'n': 0}
        tbl = self.valids[protein][pos]
        ioffset = 0 if pos == 'pos' else 6
        if len(tbl['_std']) > 0:
            self._sort_all(tbl, metric, asc)
            p = len(tbl['_std'])
            for i, cutoff in enumerate(tbl[metric]):
                tp = sum([1 for oi in tbl['i'][:i + 1] if oi in tbl['_std']])
                fp = sum([1 for oi in tbl['i'][:i + 1] \
                    if oi not in tbl['_std']])
                fn = sum([1 for oi in tbl['i'][i + 1:] if oi in tbl['_std']])
                tn = sum([1 for oi in tbl['i'][i + 1:] \
                    if oi not in tbl['_std']])
                result['cutoff'].append(cutoff)
                result['spec'].append(tn / float(tn + fp))
                result['sens'].append(tp / float(tp + fn))
                result['prec'].append(tp / float(tp + fp))
                result['fdr'].append(fp / float(tp + fp))
            result['n'] = p
            self._sort_all(tbl, 'mz')
        return result

    def evaluate_scores(self, tasks = None, stdltps = None):
        """
        Calculates specificity, sensitivity, precision,
        false discovery rate as a function of critical
        values of each scores, for all LTPs with manual
        positive annotations, for all modes.
        Returns 4x embedded dicts of
        LTPs/modes/metrics/performance metrics.
        E.g. result['STARD10']['pos']['ktv']['sens']
        """
        std_proteins = self.known_binders_detected if std_proteins is None else std_proteins
        result = dict((protein, {'pos': {}, 'neg': {}}) for protein in std_proteins)
        metrics = \
            [
                ('Kendall\'s tau', 'ktv', False),
                ('Spearman corr.', 'spv', False),
                ('Pearson corr.', 'pev', False),
                ('Euclidean dist.', 'euv', True),
                ('Robust corr.', 'rcv', False),
                ('Goodman-Kruskal\'s gamma', 'gkv', False),
                ('Difference', 'dfv', True)
            ] if tasks == 'old' else \
            [
                ('Euclidean, +0 mcl', 'env', True),
                ('Euclidean, +15 mcl', 'env15', True),
                ('Euclidean, +45 mcl', 'env45', True),
                ('Peak ratio', 'prs', True)
            ]
        for protein in std_proteins:
            for pos in ['pos', 'neg']:
                tbl = self.valids[protein][pos]
                for m in metrics:
                    result[protein][pos][m[1]] = \
                        self.spec_sens(protein, pos, m[1], m[2])
        self.scores_eval = result
    
    def known_binders_as_standard(self):
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                if 'known_binder' in tbl:
                    try:
                        tbl['_std'] = set(tbl['i'][np.where(tbl['known_binder'])])
                    except TypeError:
                        print(protein, mode)

    def count_threshold_filter(self, score, threshold, count = 10,
        threshold_type = 'fix', asc = True):
        """
        Builds a boolean array whether the values of a score fall below
        or above certain critical value. The critical value can be defined
        as a fix value,
        or as a fraction of the minimum or maximum value of the score.
        Threshold type is either `fix`, `relative` or `best_fraction`.
        
        - fix: values below or equal this number if lower is the better, otherwise
            values above or equal will be selected
        
        - relative: threshold will be set the minimum value (if ordered ascending)
            or maximum value multiplied by the threshold
        
        - best_fraction: opposite way as at the `relative`, at ascending order
            the maximum, at descending the minimum will be multiplied by the
            threshold
        
        - count: the absolute maximum number of selected instances
        """
        sort_alll(self.valids, score, asc = True)
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                limScore = np.nanmin(tbl[score]) \
                    if asc and threshold_type =='relative' or \
                        not asc and threshold_type == 'best_fraction' \
                    else np.nanmean(tbl[score]) \
                    if threshold_type == 'mean_relative' \
                    else np.nanmedian(tbl[score]) \
                    if threshold_type == 'median_relative' \
                    else np.nanmax(tbl[score])
                _threshold = limScore * threshold \
                    if threshold_type == 'relative' or \
                        threshold_type == 'best_fraction' or \
                        threshold_type == 'mean_relative' or \
                        threshold_type == 'median_relative' \
                    else threshold
                cnt = 0
                boolArray = []
                for i in (xrange(len(tbl[score])) if asc \
                    else xrange(len(tbl[score])-1,-1,-1)):
                    if np.isnan(tbl[score][i]) or np.isinf(tbl[score][i]):
                        boolArray.append(False)
                        continue
                    if cnt > count or asc and tbl[score][i] > _threshold or \
                        not asc and tbl[score][i] < _threshold:
                        boolArray.append(False)
                    else:
                        cnt += 1
                        boolArray.append(True)
                tbl['bool_%s'%score] = np.array(boolArray)

    def scores_plot(self, score = 'env', asc = True, 
        score_name = 'Euclidean distance', pdfname = None, onepfraction = None,
        hlines = [3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 10.0],
        derivates = True):
        pdfname = 'scores_%s.pdf' % score if pdfname is None else pdfname
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        sort_alll(valids, score, asc = asc)
        fig, axs = plt.subplots(8, 8, figsize = (20, 20))
        proteins = sorted(valids.keys())
        bool_score = 'bool_%s' % score
        prg = progress.Progress(len(proteins), 'Plotting scores', 1,
            percent = False)
        for i in xrange(len(proteins)):
            prg.step()
            ax = axs[i / 8, i % 8]
            proteinname = proteins[i]
            for pn in ['pos', 'neg']:
                col = '#CC0000' if pn == 'pos' else '#0000CC'
                scoreMin = np.nanmin(self.valids[proteinname][pn][score]) if asc \
                    else np.nanmax(self.valids[proteinname][pn][score])
                ax.plot(np.arange(len(self.valids[proteinname][pn][score])),
                    self.valids[proteinname][pn][score], color = col,
                    alpha = 0.7, ls = '-',
                    linewidth = 0.3)
                if derivates:
                    xd = np.arange(1, len(self.valids[proteinname][pn][score]))
                    yd = np.array([(self.valids[proteinname][pn][score][j] - \
                            self.valids[proteinname][pn][score][j-1]) for j in xd])
                    yd = yd / np.nanmax(yd)
                    if not asc:
                        yd = -1 * yd
                    ax.plot(xd, yd, color = col, alpha = 0.2, ls = '-',
                        linewidth = 0.3)
                _xlim = ax.get_xlim()
                _ylim = ax.get_ylim()
                plt.setp(ax.xaxis.get_majorticklabels(), rotation = 90)
                best_ones = [j for j, b in \
                    enumerate(self.valids[proteinname][pn][bool_score]) if b]
                xbreak = np.nanmax(best_ones) if len(best_ones) > 0 else 0.0
                ybreak = valids[proteinname][pn][score][int(xbreak)]
                ax.plot([xbreak], [ybreak], marker = 'o', markersize = 2.0,
                    markerfacecolor = col, alpha = 0.7)
                ax.annotate(
                        '%u'%(xbreak), 
                        xy = (xbreak, ybreak), 
                        #xytext = (xbreak + (20.0 if pn == 'pos' else 0.0), 
                        #ybreak + (0.5 if pn == 'neg' else 0.0)), 
                        #xycoords = 'data',
                        #textcoords = 'offset points',
                        #ha = 'center', va = 'bottom', 
                        color = col,
                        alpha = 0.7,
                        fontsize = 'xx-small',
                        arrowprops = dict(arrowstyle = '-',
                            connectionstyle = 'arc,rad=.0',
                            color = col, edgecolor = col, alpha = 0.7, 
                            visible = True, linewidth = 0.2), 
                    )
                if type(hlines) is list:
                    for c in hlines:
                        ax.axhline(y = scoreMin*c, 
                            linewidth = 0.2, alpha = 0.5, color = col)
                ax.set_xlim(_xlim)
                ax.set_ylim(_ylim)
            ax.set_title(proteinname, color = '#CC0000' \
                if onepfraction is not None and proteinname in onepfraction else '#000000')
            ax.set_ylabel(score_name)
            ax.set_xlabel('Features (ordered %s)' % \
                'ascending' if asc else 'descending')
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()

    def fractions_barplot(self, features = False,
        highlight = False, highlight2 = False,
        all_features = True, pdfname = None):
        
        if pdfname is None:
            pdfname = 'pp%s.pdf' % \
                ('' if features is None else '_features')
        
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        n = int(np.ceil(np.sqrt(len(self.pprofs))))
        fig, axs = plt.subplots(n, n, figsize = (12 + n, 12 + n))
        proteins = sorted(self.fractions.keys())
        prg = progress.Progress(len(proteins), 'Plotting profiles', 1,
            percent = False)
        
        for i in xrange(len(proteins)):
            
            prg.step()
            ax = axs[i // n, i % n]
            protein_name = proteins[i]
            fracs = self.all_fractions(protein_name)
            #sorted(self.pprofs[protein_name].keys(),
            #               key = lambda i: (i[0], int(i[1:])))
            ppr = np.array([self.pprofs[protein_name][fr] for fr in fracs])
            if hasattr(self, 'pprofs_original'):
                ppr_o = np.array([self.pprofs_original[protein_name][fr] \
                    for fr in fracs])
            else:
                ppr_o = ppr
            if features:
                ppmax = np.nanmax(ppr_o)
                ppmin = np.nanmin(ppr_o)
                ppr = (ppr - ppmin) / (ppmax - ppmin)
                ppr[ppr < 0.0] = 0.0
                ppr_o = (ppr_o - ppmin) / (ppmax - ppmin)
            #B6B7B9 gray (not measured)
            #6EA945 mantis (protein sample)
            #007B7F teal (control/void)
            col = ['#6EA945' if self.fractions[protein_name][fr] == 1 else \
                    '#B6B7B9' if self.fractions[protein_name][fr] is None else \
                    '#007B7F' \
                for fr in fracs]
            ax.bar(np.arange(len(ppr)), ppr_o, color = col, alpha = 0.1,
                edgecolor = 'none')
            ax.bar(np.arange(len(ppr)), ppr, color = col, edgecolor = 'none')
            if features and self.valids is not None:
                for pn in ['pos', 'neg']:
                    for fi, fe in enumerate(self.valids[protein_name][pn]['no']):
                        alpha = 0.20
                        lwd = 0.1
                        lst = '-'
                        plot_this = False
                        color = '#FFCCCC' if pn == 'pos' else '#CCCCFF'
                        try:
                            def _feg(fe):
                                for fei in fe:
                                    yield fei
                            feg = _feg(fe)
                            if highlight2 and \
                                self.valids[protein_name][pn][highlight2][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.15
                                lwd = 0.4
                                lst = ':'
                                plot_this = True
                            if highlight and \
                                self.valids[protein_name][pn][highlight][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.75
                                lwd = 0.4
                                lst = '-'
                                plot_this = True
                            if all_features or plot_this:
                                ax.plot(np.arange(len(ppr)) + 0.4,
                                    np.array([feg.next() \
                                        if s is not None else 0.0 \
                                        for s in self.all_fractions(protein_name)]),
                                    linewidth = lwd, markersize = 0.07,
                                    linestyle = lst, 
                                    color = color, alpha = alpha, marker = 'o')
                        except ValueError:
                            print('Unequal length dimensions: %s, %s' % \
                                (protein_name, pn))
            ax.set_xticks(np.arange(len(ppr)) + 0.4)
            ax.set_xticklabels(fracs, rotation = 90)
            ax.set_title('%s protein conc.'%protein_name)
        
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()
    
    def fractions_barplot2(self, features = False,
        highlight = False, highlight2 = False,
        all_features = True, pdfname = None,
        ignore_offsets = False):
        
        if pdfname is None:
            pdfname = 'pp%s3.pdf' % \
                ('' if not features else '_features')
        
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, axs = plt.subplots(8, 8, figsize = (20, 20))
        proteins = sorted(self.valids.keys())
        prg = progress.Progress(len(proteins), 'Plotting profiles', 1,
            percent = False)
        width = (
            0.9 if ignore_offsets else
            0.3 if len(self.fr_offsets) == 2 else
            0.45
        )
        
        for i in xrange(len(proteins)):
            
            prg.step()
            ax = axs[i // 8, i % 8]
            protein = proteins[i].upper()
            
            for (w, offset), label in \
                zip(enumerate([0.0, self.fr_offsets[0],
                                    self.fr_offsets[-1]]),
                   ['', 'L', 'U']):
                
                if len(label) and ignore_offsets:
                    continue
                
                if label == 'U':
                    if not hasattr(self, 'pprofs_originalU'):
                        continue
                
                ppr = np.array([getattr(self, 'pprofs%s' % label)\
                    [protein][fr] for fr in self.all_fractions(protein)])
                ppr_o = np.array([getattr(self, 'pprofs_original%s' % label)\
                    [protein][fr] for fr in self.all_fractions(protein)])
                
                if features:
                    ppmax = np.nanmax(ppr_o)
                    ppmin = np.nanmin(ppr_o)
                    ppr = (ppr - ppmin) / (ppmax - ppmin)
                    ppr[ppr < 0.0] = 0.0
                    ppr_o = (ppr_o - ppmin) / (ppmax - ppmin)
                #B6B7B9 gray (not measured)
                #6EA945 mantis (protein sample)
                #007B7F teal (control/void)
                col = ['#6EA945' 
                       if self.fractions[protein][fr]
                       else '#007B7F'
                    for fr in self.all_fractions(protein)]
                ax.bar(np.arange(len(ppr)) + width * w, ppr_o, width,
                    color = col, alpha = 0.1, edgecolor = 'none')
                ax.bar(np.arange(len(ppr)) + width * w, ppr, width,
                    color = col, edgecolor = 'none')
            if features and self.valids is not None:
                
                ifracs = self.fraction_indices(protein)
                
                for pn in ['pos', 'neg']:
                    
                    for fi, fe in enumerate(self.valids[protein][pn]['no']):
                        
                        alpha = 0.20
                        lwd = 0.1
                        lst = '-'
                        plot_this = False
                        color = '#FFCCCC' if pn == 'pos' else '#CCCCFF'
                        
                        try:
                            if highlight2 and \
                                self.valids[protein][pn][highlight2][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.15
                                lwd = 0.4
                                lst = ':'
                                plot_this = True
                            if highlight and \
                                self.valids[protein][pn][highlight][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.75
                                lwd = 0.4
                                lst = '-'
                                plot_this = True
                            if all_features or plot_this:
                                ax.plot(np.arange(len(ppr)) + 0.4,
                                    np.array([fe[ifracs[fr][0]]
                                        for fr in self.sort_fracs_str(ifracs.keys())]),
                                    linewidth = lwd, markersize = 0.07,
                                    linestyle = lst, 
                                    color = color, alpha = alpha, marker = 'o')
                            
                        except ValueError:
                            print('Unequal length dimensions: %s, %s' % \
                                (protein, pn))
            
            ax.set_xticks(np.arange(len(ppr)) + 0.4)
            ax.set_xticklabels(self.all_fractions(protein))
            ax.set_title('%s protein conc.'%protein)
        
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()
    
    def peak_ratios_histo(self, lower = 0.5, pdfname = None):
        upper = 1.0 / lower
        if pdfname is None:
            pdfname = 'peak_ratios.pdf'
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, axs = plt.subplots(8, 8, figsize = (20, 20))
        proteins = sorted(self.fractions_upper.keys())
        prg = progress.Progress(len(proteins), 'Plotting peak ratios', 1,
            percent = False)
        width = 0.3
        for i in xrange(len(proteins)):
            prg.step()
            ax = axs[i / 8, i % 8]
            protein_name = proteins[i].upper()
            if self.valids[protein_name]['pos']['ipr'].shape[1] > 0:
                ppr = np.array(list(self.ppratios[protein_name]\
                    [sorted(self.ppratios[protein_name].keys())[0]]))
                x1 = self.valids[protein_name]['pos']['ipr'][:,0]
                x1 = x1[np.isfinite(x1)]
                x2 = self.valids[protein_name]['neg']['ipr'][:,0]
                x2 = x2[np.isfinite(x2)]
                bins = np.linspace(ppr[0] * 0.25, ppr[1] * 1.50, 15)
                ax.hist(x1, bins, color = '#6EA945', alpha = 0.2, edgecolor = 'none')
                ax.hist(x2, bins, color = '#007B7F', alpha = 0.2, edgecolor = 'none')
                ax.scatter(ppr, np.array([0.0, 0.0]), c = '#DA0025',
                    alpha = 0.5, edgecolor = 'none')
                ax.axvline(x = ppr[0] * lower, c = '#DA0025', linewidth = 0.2, alpha = 0.5, linestyle = '-.')
                ax.axvline(x = ppr[1] * upper, c = '#DA0025', linewidth = 0.2, alpha = 0.5, linestyle = '-.')
                ax.set_xlim((ppr[0] * 0.25, ppr[1] * 2.50))
                x1a = x1[np.where((x1 > ppr[0] * lower) & (x1 < ppr[1] * upper))]
                m1 = np.mean(x1a)
                std1 = np.std(x1a)
                x2a = x2[np.where((x2 > ppr[0] * lower) & (x2 < ppr[1] * upper))]
                m2 = np.mean(x2a)
                std2 = np.std(x2a)
                x3 = list(x1a)
                x3.extend(list(x2a))
                m3 = np.mean(x3)
                std3 = np.std(x3)
                
                # positive: mean, 1SD, 2SD
                ax.axvline(x = m1, c = '#6EA945', linewidth = 0.2, alpha = 0.5)
                ax.axvline(x = m1 + std1, c = '#6EA945', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                ax.axvline(x = m1 - std1, c = '#6EA945', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                #ax.axvline(x = m1 + 2 * std1, c = '#6EA945', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                #ax.axvline(x = m1 - 2 * std1, c = '#6EA945', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                
                # negative: mean, 1SD, 2SD
                ax.axvline(x = m2, c = '#007B7F', linewidth = 0.2, alpha = 0.5)
                ax.axvline(x = m2 + std2, c = '#007B7F', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                ax.axvline(x = m2 - std2, c = '#007B7F', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                #ax.axvline(x = m2 + 2 * std2, c = '#007B7F', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                #ax.axvline(x = m2 - 2 * std2, c = '#007B7F', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                
                # both: mean, 1SD, 2SD
                ax.axvline(x = m3, c = '#996A44', linewidth = 0.2, alpha = 0.5)
                ax.axvline(x = m3 + std3, c = '#996A44', linewidth = 0.2, alpha = 0.5, linestyle = ':')
                ax.axvline(x = m3 - std3, c = '#996A44', linewidth = 0.2, alpha = 0.5, linestyle = ':')
            ax.set_title('%s int. peak ratios'%protein_name)
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()
    
    def lookup_nans(self):
        """
        Looks up those features having too many NaNs or zeros
        among the protein containing fractions.
        Stores result under key `na`.
        """
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                pfe = self.pcont_columns(protein, mode, 'fe')
                
                if self.permit_profile_end_nan:
                    
                    hasnans = np.logical_or(
                        np.logical_and(
                            np.logical_or(
                                np.sum(np.isnan(pfe[:,:-1]),
                                    axis = 1, dtype = np.bool),
                                np.sum(pfe[:,:-1] == 0.0,
                                    axis = 1, dtype = np.bool)
                            ),
                            np.logical_or(
                                np.isnan(pfe[:,-1]),
                                pfe[:,-1] == 0.0
                            )
                        ),
                        np.logical_and(
                            np.logical_or(
                                np.sum(np.isnan(pfe[:,1:]),
                                    axis = 1, dtype = np.bool),
                                np.sum(pfe[:,1:] == 0.0,
                                    axis = 1, dtype = np.bool)
                            ),
                            np.logical_or(
                                np.isnan(pfe[:,0]),
                                pfe[:,0] == 0.0
                            )
                        ),
                    )
                    
                else:
                    
                    hasnans = np.logical_or(
                                np.sum(np.isnan(pfe),
                                    axis = 1, dtype = np.bool),
                                np.sum(pfe == 0.0,
                                    axis = 1, dtype = np.bool)
                            )
                
                tbl['na'] = hasnans
    
    @staticmethod
    def ratio_limits(lower):
        """
        Returns a symmetric range of tolerance for ratio-like
        values. This means lower / expected = expected / upper
        """
        return (lower, 1.0 / lower)
    
    def peak_ratio_score(self, exclude_wrong = None):
        """
        Calculates the difference between the mean intensity peak ratio
        expressed in number of standard deviations.
        The mean and the SD calculated in the range of the expected
        ratio minus lower plus upper.
        """
        
        sys.stdout.write('\n\t=== Peak ratio score messages: ======\n\n')
        
        lower, upper = self.ratio_limits(self.peak_ratio_score_max_bandwidth)
        proteins = sorted(self.valids.keys())
        w = self.filter_wrong_fractions
        if exclude_wrong is not None:
            w = exclude_wrong
        
        def set_bandwith(ipr, ppr):
            """
            Attempts to set the bandwith of the peak ratio filter.
            `ipr` is the array of intensity ratios, `ppr` is
            the corresponding protein ratio.
            """
            count = self.peak_ratio_score_optimal_population
            minct = self.peak_ratio_score_minimal_population
            sgdct = self.peak_ratio_score_still_good_population
            lower = self.peak_ratio_score_max_bandwidth
            lowe2 = self.peak_ratio_score_optimal_bandwidth
            rdiff = ipr / ppr
            rdiff[np.where(rdiff > 1.0)] = 1.0 / rdiff[np.where(rdiff > 1.0)]
            rdiff.sort()
            rdiff = rdiff[::-1]
            for cnt in xrange(count, minct, -1):
                bw = rdiff[cnt - 1]
                if bw >= lowe2:
                    return self.ratio_limits(bw)
                if cnt < sgdct and bw >= lower:
                    return self.ratio_limits(bw)
            
            return (None, None)
        
        def get_score(protein, col, fkey, lower, upper,
                      modes = ['neg', 'pos']):
            
            if not modes:
                # these won't be used anyways
                return None, None
            
            with np.errstate(divide = 'ignore', invalid = 'ignore'):
                
                with warnings.catch_warnings():
                    
                    warnings.simplefilter('ignore', RuntimeWarning)
                    # the expected value of the peak ratio:
                    ppr = self.ppratios[protein][fkey]
                    # all intensity peak ratios:
                    p   = self.valids[protein]['pos']['ipr'][:,col]
                    pa  = p[np.isfinite(p)]
                    n   = self.valids[protein]['neg']['ipr'][:,col]
                    na  = n[np.isfinite(n)]
                    # all intensity peak ratios:
                    a   = np.array([])
                    if 'neg' in modes:
                        a = np.concatenate((a, na))
                    if 'pos' in modes:
                        a = np.concatenate((a, pa))
                    # range considering fraction offsets, having tolerance
                    # -/+ upper/lower
                    if self.adaptive_peak_ratio_score:
                        
                        lim = set_bandwith(a, np.mean(ppr))
                        
                        if lim[0] is None:
                            sys.stdout.write(
                                '\t:: Could not calculate peak ratio '\
                                'score at protein `%s`\n\t   between '\
                                'fractions %s and %s: no features in '\
                                'range of tolerance.\n' % (
                                    protein, fkey[0], fkey[1]
                                ))
                            return None, None
                        
                        a   = a[np.where(
                            (a >= lim[0]) &
                            (a <= lim[1])
                        )]
                        
                    else:
                        
                        a   = a[np.where(
                            (a > ppr[0]  * lower) &
                            (a < ppr[-1] * upper)
                        )]
                    
                    m   = np.nanmean(a)
                    s   = np.nanstd(a)
                    # scores: difference from mean in SD
                    scp = np.abs(p - m) / s
                    scn = np.abs(n - m) / s
                    return scp, scn
        
        def scores_array(protein, scores, mode):
            
            try:
                return (
                    np.column_stack(
                        tuple(
                            map(
                                lambda fk:
                                    scores[fk],
                                map(
                                    lambda fk:
                                        fk[0],
                                    sorted(
                                        iteritems(
                                            self.valids[protein][mode]['iprsa']
                                        ),
                                        key = lambda fk: fk[1]
                                    )
                                )
                            )
                        )
                    )
                )
            except:
                sys.stdout.write(protein)
        
        for protein in self.valids.keys():
            
            if self.valids[protein]['pos']['ipr'].shape[1] > 0:
                
                fracs = self.protein_containing_fractions(protein)
                if w:
                    wfracs = {
                        'neg': self.wrong_fractions(protein, 'neg'),
                        'pos': self.wrong_fractions(protein, 'pos')
                    }
                    wfracsall  = set().union(*wfracs.values())
                
                scoresp = {}
                scoresn = {}
                
                iprsa    = {'neg': {}, 'pos': {}}
                prsa_col = {'neg': 0,  'pos': 0}
                
                for i in xrange(len(fracs) - 1):
                    
                    for j in xrange(i + 1, len(fracs)):
                        
                        modes = set([])
                        col   = {}
                        frac1, frac2 = fracs[i], fracs[j]
                        fkey = (frac1, frac2)
                        
                        for mode in ['neg', 'pos']:
                            
                            if (
                                not w or (
                                    frac1 not in wfracs[mode] and
                                    frac2 not in wfracs[mode]
                                )
                            ):
                                
                                iprsa[mode][fkey] = prsa_col[mode]
                                prsa_col[mode] += 1
                                modes.add(mode)
                        
                        col = self.valids[protein][mode]['ipri'][fkey]
                        
                        pscores, nscores = get_score(
                            protein, col, fkey,
                            lower, upper,
                            modes = modes
                        )
                        
                        if pscores is None:
                            modes = modes.difference('pos')
                            _ = iprsa['pos'].pop(fkey, None)
                        
                        if nscores is None:
                            modes = modes.difference('neg')
                            _ = iprsa['neg'].pop(fkey, None)
                        
                        if 'neg' in modes:
                            scoresn[fkey] = nscores
                        if 'pos' in modes:
                            scoresp[fkey] = pscores
                
                self.valids[protein]['pos']['iprsa'] = iprsa['pos']
                self.valids[protein]['neg']['iprsa'] = iprsa['neg']
                
                self.valids[protein]['pos']['prsa'] = \
                    scores_array(protein, scoresp, 'pos')
                self.valids[protein]['neg']['prsa'] = \
                    scores_array(protein, scoresn, 'neg')
            
            else:
                self.valids[protein]['pos']['prsa'] = \
                    np.zeros(
                        (self.valids[protein]['pos']['ipr'].shape[0],),
                        dtype = np.float
                    )
                self.valids[protein]['neg']['prsa'] = \
                    np.zeros(
                        (self.valids[protein]['neg']['ipr'].shape[0],),
                        dtype = np.float
                    )
            
            for mode in ['neg', 'pos']:
                
                if w and wfracs[mode]:
                    
                    sys.stdout.write('\t:: At protein `%s` in mode `%s` fractions %s ignored.\n'
                        '\t   Calculated scores for %u ratios '
                        'instead of %u.\n' % (
                            protein, mode, ', '.join(wfracs[mode]),
                            self.valids[protein][mode]['prsa'].shape[1],
                            len(self.valids[protein][mode]['ipri'])
                        )
                    )
    
    def combine_peak_ratio_scores(self):
        """
        Peak ratio scores calculated for each pairs of protein containing
        fractions. Here we take their mean and select the scores for the
        first and highest fraction pairs.
        """
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                fkeyf = self.first_ratio[protein]
                fkeyh = self.last_ratio[protein]
                
                tbl['prs'] = np.array(
                    list(
                        map(
                            lambda a:
                                (
                                    np.mean(a[np.isfinite(a)])
                                    if sum(np.isfinite(a)) > 0
                                    else np.inf
                                ),
                            tbl['prsa']
                        )
                    )
                ) # mean ignoring infinites and NaNs
                #tbl['prs'] = np.nanmean(tbl['prsa'], axis = 1) # mean score
                tbl['prs'][np.where(tbl['na'])] = np.inf
                if fkeyf in tbl['iprsa']:
                    tbl['prsf'] = tbl['prsa'][:,tbl['iprsa'][fkeyf]] # score for first
                if fkeyh in tbl['iprsa']:
                    tbl['prsh'] = tbl['prsa'][:,tbl['iprsa'][fkeyh]] # highest diff.
                # taking the scores for fraction pairs
                # around the peak (like in slope filter)
                if 'sloi0' in tbl:
                    # ignore warnings, as nan's produced in any invalid case
                    # and that is fine
                    with np.errstate(divide = 'ignore', invalid = 'ignore'):
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore', RuntimeWarning)
                            tbl['prss'] = np.array(
                                list(
                                    map(
                                        lambda i:
                                            np.nanmean(
                                                list(
                                                    filter(
                                                        lambda x:
                                                            np.isfinite(x),
                                                        map(
                                                            lambda frs:
                                                                tbl['prsa'][i,tbl[
                                                                    'iprsa'][frs]],
                                                            filter(
                                                                lambda frs:
                                                                    frs in tbl['iprsa'],
                                                                tbl['sloi0'] + (
                                                                    tbl['sloi1']
                                                                    if 'sloi1' in tbl
                                                                    else []
                                                            )
                                                            )
                                                        )
                                                    )
                                                )
                                            ),
                                        xrange(tbl['prs'].shape[0])
                                    )
                                )
                            )
    
    def export_scores(self, fname = 'scores.csv'):
        """
        Exports a table with all scores for plotting with ggplot.
        """
        hdr = ['protein', 'mode', 'boolppr', 'boolprs', 'boolenr',
               'boolprss', 'prs', 'prsf', 'prsh']
        
        with open(fname, 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr))
            
            for protein, d in iteritems(self.valids):
                
                for mode, tbl in iteritems(d):
                    
                    for i in xrange(tbl['prs'].shape[0]):
                        
                        fp.write('%s\n' % '\t'.join([
                            protein,
                            mode,
                            str(tbl['prr'][i]),
                            str(tbl['prs'][i] < 1.0),
                            str(tbl['slobb'][i]),
                            str(tbl['prss'][i] < 1.0),
                            '%.08f' % tbl['prs'][i],
                            (
                                '%.08f' % tbl['prsf'][i]
                                if 'prsf' in tbl else 'NA'
                            ),
                            (
                                '%.08f' % tbl['prsh'][i]
                                if 'prsh' in tbl else 'NA'
                            )
                        ]))
    
    def peak_ratio_score_hist(self, pdfname = None):
        """
        Plots a histogram for peak ratio scores.
        """
        if pdfname is None:
            pdfname = 'peak_ratios_hist.pdf'
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        gridsize = int(np.ceil(np.sqrt(len(self.valids))))
        fig, axs = plt.subplots(gridsize, gridsize,
                                figsize = (gridsize + 12, gridsize + 12))
        proteins = sorted(self.fractions_upper.keys())
        prg = progress.Progress(len(proteins), 'Plotting peak ratio scores', 1,
            percent = False)
        width = 0.3
        
        for i in xrange(len(proteins)):
            
            prg.step()
            ax = axs[i / gridsize, i % gridsize]
            protein_name = proteins[i].upper()
            p = self.valids[protein_name]['pos']['prs']
            n = self.valids[protein_name]['neg']['prs']
            p = p[np.isfinite(p)]
            n = n[np.isfinite(n)]
            bins = np.linspace(0.0, 5.0, 25)
            ax.hist(p, bins, color = '#6EA945', alpha = 0.2, edgecolor = 'none')
            ax.hist(n, bins, color = '#007B7F', alpha = 0.2, edgecolor = 'none')
            ax.set_xlim((0.0, 5.0))
            ax.set_title('%s p/r scores' % protein_name)
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()
    
    def peak_ratio_score_bool(self, exclude_na = True):
        """
        Creates a boolean attribute from peak ratio scores
        along a certain cutoff.
        """
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                tbl['prs1'] = tbl['prs'] <= self.peak_ratio_score_threshold
        
        if exclude_na:
            self.logical_not('na', 'not_na')
            self.logical_and(one = 'prs1', two = 'not_na', result = 'prs1')
    
    def peak_ratio_score_best(self, best = 10, exclude_na = True):
        """
        Creates a boolean attribute selecting the n features with lowest
        peak ratio scores.
        """
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                b = min([
                    best,
                    tbl['prs'].shape[0],
                    np.isfinite(tbl['prs']).sum()
                ])
                
                threshold    = tbl['prs'][tbl['prs'].argsort()][b - 1]
                
                tbl['prsb'] = tbl['prs'] <= threshold
        
        if exclude_na:
            self.logical_not('na', 'not_na')
            self.logical_and('prsb', 'not_na', 'prsb')
    
    def slope_filter(self):
        """
        Implements the feature selection according to Enric's proposal.
        This method considers only the direction of the slope in the
        vicinity of the highest fractions.
        """
        
        if not self.slope_profile_selection:
            return None
        
        sys.stdout.write('\n\t=== Slope filter messages: ======\n\n')
        
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                hpeak  = self.hpeak[protein]
                result = []
                
                for i, hfracs in enumerate(hpeak):
                    
                    self._slope_filter(protein, mode, tbl, hfracs, i)
                    result.append(tbl['slob%u' % i])
                
                # this is guaranteed to be only one vector
                # for each table of features
                tbl['slobb'] = (
                    np.any(
                        np.hstack(
                            result
                        ),
                        axis = 1
                    )
                )
    
    def _slope_filter(self, protein, mode, tbl, hfracs, i):
        """
        Calculates a simple slope filter for one protein
        and one set of highest fractions.
        """
        
        w = self.filter_wrong_fractions
        
        wfracs = self.wrong_fractions(protein, mode)
        
        if wfracs and w:
            sys.stdout.write('\t:: Attempting to ignore fractions `%s` at '\
                'protein `%s` in mode `%s`.\n' % (
                    ', '.join(wfracs), protein, mode
                )
            )
        
        if w and not set(hfracs) - wfracs:
            sys.stdout.write('\t:: At protein `%s` mode `%s` all highest '\
                'fractions (%s) marked as potentially wrong.\n'\
                '\t   Falling back to use these wrong fractions.\n' % (
                    protein, mode, ', '.join(hfracs)))
            wfracs = set([])
            w = False
        
        # removing `wrong fractions` from the peak
        if w:
            hfracs     = list(filter(lambda fr: fr not in wfracs, hfracs))
        
        hfracs     = tuple(hfracs)
        ratio      = self.slope_equal_fractions_ratio
        has_ratio1 = True
        has_ratio2 = True
        
        nfracs = (
            dict(
                enumerate(
                    self.protein_containing_fractions(
                        protein,
                        except_wrong = mode if w else False
                    )
                )
            )
        )
        
        pfracs = (
            dict(
                map(
                    lambda fr:
                        (fr[1], fr[0]),
                    iteritems(nfracs)
                )
            )
        )
        
        ifracs = self.fraction_indices(protein,
                                       except_wrong = mode if w else False)
        pprofs = self.pprofs[protein]
        fe     = tbl['fe']
        shape  = [fe.shape[0], 1]
        
        # this is the default, we select the one just before
        # the highest
        i_minus1 = pfracs[hfracs[0]]  - 1 # h-1
        i_plus1  = pfracs[hfracs[-1]] + 1 # h+1
        
        if len(nfracs) > i_plus1:
            ratio2_descending = True
            fr21  = hfracs[-1]                # always the last of highest h(-1)
            fr22  = nfracs[i_plus1]           # the one after that (h+1)
        else:
            # except if we have no more fractions:
            if i_minus1 - 1 < 0:
                has_ratio2 = False
            else:
                ratio2_descending = False
                fr21  = nfracs[i_minus1 - 1]
                fr22  = nfracs[i_minus1]
        
        if i_minus1 < 0:
            # in few cases we don't have the fraction -1
            # then we select the fraction by 2 after the highest
            only_descending = True
            i_minus1 = pfracs[hfracs[-1]]  + 2
            if i_minus1 < len(nfracs):
                fr11 = fr21             # the h+1 one
                fr12 = nfracs[i_minus1] # the h+2 one
            else:
                has_ratio1 = False
        else:
            only_descending = False
            fr11 = nfracs[i_minus1] # the h-1 one
            fr12 = hfracs[0]        # the h(1)
        
        fr31 = hfracs[0]  # only matters if we have 2 highest
        fr32 = hfracs[-1] #
        
        tbl['sloi%u' % i] = []
        
        with np.errstate(divide = 'ignore', invalid = 'ignore'):
            # as we control all arrays to be greater than zero
            # otherwise result will be False
            # we can safely ignore zero division or
            # invalid values
            
            if has_ratio1:
                
                pratio1 = pprofs[fr11] / pprofs[fr12] # either -1 or +2
                
                iratio1 = (fe[:,ifracs[fr11][0]] / # numeric vector
                        fe[:,ifracs[fr12][0]])  # of ratios
                
                if only_descending:
                    # descending, but not more than certain threshold
                    iratio1 = np.logical_and(
                                iratio1 < pratio1,
                                iratio1 > (
                                    pratio1 *
                                    self.slope_descending_slope_diff_tolerance
                                )
                            )
                else:
                    iratio1 = iratio1 < 1.0 # ascending
                
                iratio1   = np.logical_and(
                                np.logical_and(
                                    fe[:,ifracs[fr11][0]] > 0.0,
                                    fe[:,ifracs[fr12][0]] > 0.0
                                ),
                                iratio1 # this is bool vector
                            )
                
                tbl['slo%u%s%s' % (i, fr11, fr12)] = iratio1.reshape(shape)
                tbl['sloi%u' % i].append((fr11, fr12))
            
            if has_ratio2:
                
                pratio2 = pprofs[fr21] / pprofs[fr22] # always +1
                
                iratio2 = fe[:,ifracs[fr21][0]] / fe[:,ifracs[fr22][0]]
                
                iratio2   = np.logical_and(
                                np.logical_and(
                                    fe[:,ifracs[fr21][0]] > 0.0,
                                    fe[:,ifracs[fr22][0]] > 0.0
                                ),
                                (iratio2 > 1.0
                                if ratio2_descending
                                else iratio2 < 1.0)
                            )
                
                tbl['slo%u%s%s' % (i, fr21, fr22)] = iratio2.reshape(shape)
                tbl['sloi%u' % i].append((fr21, fr22))
            
            if fr31 != fr32:
                # to not calculate it twice
                iratio3 = (fe[:,ifracs[fr31][0]] /
                           fe[:,ifracs[fr32][0]])
                
                iratio3 = np.logical_and(
                            np.logical_and(
                                fe[:,ifracs[fr31][0]] > 0.0,
                                fe[:,ifracs[fr32][0]] > 0.0
                            ),
                            np.logical_and(
                                iratio3 > ratio,
                                iratio3 < 1.0 / ratio
                            )
                          )
                
                tbl['slo%u%s%s' % (i, fr31, fr32)] = iratio3.reshape(shape)
                tbl['sloi%u' % i].append((fr31, fr32))
        
        tbl['slob%u' % i] = (
                                np.all(
                                    np.hstack(
                                        list(
                                            map(
                                                lambda v:
                                                    tbl['slo%u%s%s' % (
                                                        i, v[0], v[1]
                                                    )],
                                                tbl['sloi%u' % i]
                                            )
                                        )
                                    ),
                                    axis = 1
                                ).reshape(shape) # one bool vector
                            )
        
        if len(tbl['sloi%u' % i]) < 2:
            sys.stdout.write('\t:: Warning: at protein `%s` '
                'in mode `%s` we could'
                ' use only %u ratio.\n' % (
                    protein, mode, len(tbl['sloi%u' % i])
                )
            )
    
    def plot_score_performance(self, perf):
        metrics = [
            ('Kendall\'s tau', 'ktv', False),
            ('Spearman corr.', 'spv', False),
            ('Pearson corr.', 'pev', False),
            ('Euclidean dist.', 'euv', True),
            ('Robust corr.', 'rcv', False),
            ('Goodman-Kruskal\'s gamma', 'gkv', False),
            ('Difference', 'dfv', True)
        ]
        perfmetrics = [
            ('sens', '#6EA945', 'Sensitivity'),
            ('spec', '#FCCC06', 'Specificity'),
            ('prec', '#DA0025', 'Precision'),
            ('fdr', '#007B7F', 'FDR')
        ]
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        # plt.gca().set_color_cycle(['red', 'green', 'blue', 'yellow'])
        for protein, d1 in iteritems(perf):
            for pos, d2 in iteritems(d1):
                fig, axs = plt.subplots(len(metrics),
                    figsize = (10, 20), sharex = False)
                plt.suptitle('Performance of scores :: %s, %s' % \
                    (protein, pos), size = 16)
                for i, m in enumerate(metrics):
                    for pm in perfmetrics:
                        axs[i].plot(d2[m[1]]['cutoff'], d2[m[1]][pm[0]],
                            '-', linewidth = 0.33,
                            color = pm[1], label = pm[2])
                    leg = axs[i].legend()
                    axs[i].set_ylabel('Performance\nmetrics')
                    axs[i].set_xlabel('%s value'%m[0])
                    axs[i].set_title(m[0])
                    if m[2]:
                        axs[i].set_xlim(axs[i].get_xlim()[::-1])
                fig.tight_layout()
                plt.subplots_adjust(top = 0.95)
                fig.savefig('score_performance_%s-%s.pdf'%(protein, pos))
                plt.close(fig)
    
    def plot_roc(self, perf):
        metrics = [
            ('Kendall\'s tau', 'ktv', False),
            ('Spearman corr.', 'spv', False),
            ('Pearson corr.', 'pev', False),
            ('Euclidean dist.', 'euv', True),
            ('Robust corr.', 'rcv', False),
            ('Goodman-Kruskal\'s gamma', 'gkv', False),
            ('Difference', 'dfv', True)
        ]
        def colors():
            _colors = ['#6EA945', '#FCCC06', '#DA0025', 
                '#007B7F', '#454447', '#996A44']
            for c in _colors:
                yield c
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, axs = plt.subplots((len(metrics) + len(metrics) % 2) / 2, 2,
            figsize = (10, 20), sharex = False, sharey = False)
        plt.suptitle('Reciever operating characteristics', size = 16)
        for i, m in enumerate(metrics):
            c = colors()
            ax = axs[i/2][i%2]
            for protein, d1 in iteritems(perf):
                for pos, d2 in iteritems(d1):
                    ax.plot(1 - np.array(d2[m[1]]['spec']),
                        np.array(d2[m[1]]['sens']),
                        '-', linewidth = 0.33, color = c.next(),
                        label = '%s-%s n = %u' % (protein, pos, d2[m[1]]['n']))
            ax.plot([0,1], [0,1], '--', linewidth = 0.33, color = '#777777')
            leg = ax.legend(loc = 4)
            ax.set_ylabel('Sensitivity')
            ax.set_xlabel('1 - Specificity')
            ax.set_title('ROC :: %s'%m[0])
        fig.tight_layout()
        plt.subplots_adjust(top = 0.95)
        fig.savefig('roc.pdf')
        plt.close(fig)
    
    def plot_roc2(self):
        metrics = [
            ('Euclidean, +0 mcl', 'env', True),
            ('Euclidean, +15 mcl', 'env15', True),
            ('Euclidean, +45 mcl', 'env45', True),
            ('Peak ratio', 'prs', True)
        ]
        def colors():
            _colors = ['#6EA945', '#FCCC06', '#DA0025', 
                '#007B7F', '#454447', '#996A44']
            for c in _colors:
                yield c
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        nplots = sum(
            map(
                lambda protein:
                    len(
                        filter(
                            lambda mode:
                                len(self.valids[protein][mode]['_std']),
                            ['pos', 'neg']
                        )
                    ),
                self.known_binders_detected
            )
        )
        nrows = int(np.ceil(np.sqrt(nplots)))
        fig, axs = plt.subplots(nrows, nrows,
            figsize = (nrows * 5, nrows * 5),
            sharex = False, sharey = False)
        plt.suptitle('Reciever operating characteristics', size = 16)
        i = 0
        for protein in self.known_binders_detected:
            for mode in ['pos', 'neg']:
                n = len(self.valids[protein][mode]['_std'])
                if n:
                    #
                    x1, y1 = self.get_cutoff_coo(
                        self.scores_eval[protein][mode]['prs'], 1.0)
                    #
                    c = colors()
                    ax = axs[i/nrows][i%nrows]
                    for m in metrics:
                        d2 = self.scores_eval[protein][mode]
                        ax.plot(1 - np.array(d2[m[1]]['spec']),
                            np.array(d2[m[1]]['sens']),
                            '-', linewidth = 0.33, color = c.next(),
                            label = m[0])
                    ax.scatter([x1], [y1], c = '#007B7F', alpha = 0.5,
                        label = 'Peak ratio score = 1.0', edgecolors = 'none')
                    ax.plot([0,1], [0,1], '--', linewidth = 0.33, color = '#777777')
                    leg = ax.legend(loc = 4)
                    ax.set_ylabel('Sensitivity')
                    ax.set_xlabel('1 - Specificity')
                    ax.set_title('ROC :: %s :: %s :: n = %u' % (protein, mode, n))
                    i += 1
        fig.tight_layout()
        plt.subplots_adjust(top = 0.95)
        fig.savefig('roc2.pdf')
        plt.close(fig)
    
    def get_cutoff_coo(self, d, value):
        for i, val in enumerate(d['cutoff']):
            if val >= value:
                return 1 - d['spec'][i], d['sens'][i]
        return 1.0, 1.0
    
    def best_combined(self, valids, scores, best = 10, ubiquity_treshold = 5):
        for protein, d in iteritems(valids):
            for pn, tbl in iteritems(d):
                sorted_scores = []
                for sc in scores:
                    oi_sorted = np.copy(tbl['i'][tbl[sc[0]].argsort()])
                    if sc[1]:
                        sorted_scores.append(oi_sorted)
                    else:
                        sorted_scores.append(oi_sorted[::-1])
                all_ranks = np.empty((len(scores), len(tbl['i'])))
                for j, oi in enumerate(tbl['i']):
                    for i, scs in enumerate(sorted_scores):
                        for si, sc in enumerate(scs):
                            if oi == sc:
                                if tbl['cpv'][j] < 1.0 or \
                                    np.isnan(
                                        np.sum([tbl[s[0]][j] for s in scores])
                                    ) or (\
                                        'ubi' in tbl and \
                                        tbl['ubi'][j] > ubiquity_treshold
                                    ):
                                    all_ranks[i,j] = np.inf
                                else:
                                    all_ranks[i,j] = si
                combined_ranks = np.sum(all_ranks, axis = 0)
                best_oi = tbl['i'][np.argsort(combined_ranks)][:best]
                tbl['best%u'%best] = best_oi

    def best_gk_eu(self, valids, best = 10):
        best_combined(valids, [('gkv', False), ('euv', True)], best = best)

    def best_table(self, valids, fname, pos, best = 10):
        hdr = ['protein', 'm/z', 'Database_ID', 'Level', 'Full_name',
               'Formula', 'Adduct', 'Database_m/z', 'Adduct_m/z']
        sort_alll(valids, 'mz')
        with open(fname, 'w') as f:
            f.write('\t'.join(hdr) + '\n')
            for protein, d in iteritems(valids):
                tbl = d[pos]
                for oi in tbl['best%u'%best]:
                    i = np.where(tbl['i'] == oi)[0]
                    mz = tbl['mz'][i]
                    if tbl['lip'][oi] is not None:
                        for lip in tbl['lip'][oi]:
                            f.write('\t'.join([protein, '%.08f'%mz] + \
                                [str(x) for x in list(lip)]) + '\n')
                    else:
                        f.write('\t'.join([protein, '%.08f'%mz] + \
                            ['unknown']*6) + '\n')

    def true_positives(self, valids, stdpos, protein, pos = 'pos',
        tolerance = 0.01):
        """
        For one protein having known binders looks up these
        among the valid features either in positive OR
        negative mode. (So you need to run twice, once
        for +, once for -.)
        """
        ioffset = 0 if pos == 'pos' else 6
        valids[protein][pos]['std'] = {}
        for fe in stdpos[protein]:
            if fe[4 + ioffset] != '':
                mz = float(fe[ioffset + 4])
                iu = valids[protein][pos]['mz'].searchsorted(mz)
                idx = []
                if iu < len(valids[protein][pos]['mz']):
                    u = 0
                    while valids[protein][pos]['mz'][iu + u] - mz <= tolerance:
                        idx.append(iu + u)
                        u += 1
                if iu > 0:
                    l = 1
                    while mz - valids[protein][pos]['mz'][iu -l] <= tolerance:
                        idx.append(iu - l)
                        l += 1
                for i in idx:
                    identify_feature(valids[protein][pos], fe, i, ioffset)

    def identify_feature(self, tbl, fe, i, ioffset):
        """
        For one feature looks up if the lipid database ID
        and the adduct do match.
        Adds the matching record from the gold standard
        to lists in the `std` dict. There can be accessed
        by the original index.
        """
        result = []
        oi = tbl['i'][i]
        if len(np.logical_and(tbl['lip'][oi][:,0] == fe[1 + ioffset], \
            tbl['lip'][oi][:,4] == fe[5 + ioffset])) > 0:
            if oi not in tbl['std']:
                tbl['std'][oi] = []
            tbl['std'][oi].append(fe)
        else:
            sys.stdout.write('Feature %.05f matched, lipid '\
                'ID hasn\'t been found: %s' % \
                (fe[4 + ioffset], fe[1 + ioffset]))
            sys.stdout.flush()

    def sortby_score(self, tbl, attr, asc = True):
        """
        Returns sorted views of normalized features,
        m/z's and the attribute the sorting carried
        out by.
        """
        sorted_no = tbl['no'][np.argsort(tbl[attr]),:]
        sorted_mz = tbl['mz'][np.argsort(tbl[attr])]
        sorted_attr = np.sort(tbl[attr])
        if asc:
            return sorted_no, sorted_mz, sorted_attr
        else:
            return sorted_no[::-1], sorted_mz[::-1], sorted_attr[::-1]

    def _sort_all(self, tbl, attr, asc = True):
        """
        Sorts all arrays in one table by one specified 
        attribute.
        Either ascending or descending.
        """
        ind = tbl[attr].argsort()
        dim = tbl[attr].shape[0]
        for k, a in iteritems(tbl):
            if k != attr and type(tbl[k]) == np.ndarray and \
                tbl[k].shape[0] == dim and not k.startswith('_'):
                if len(a.shape) == 1:
                    if asc:
                        tbl[k] = tbl[k][ind]
                    else:
                        tbl[k] = tbl[k][ind][::-1]
                else:
                    if asc:
                        tbl[k] = tbl[k][ind,:]
                    else:
                        tbl[k] = tbl[k][ind,:][::-1]
        tbl[attr].sort()
        if not asc:
            tbl[attr] = tbl[attr][::-1]

    def sort_alll(self, attr, asc = True):
        """
        Sorts all arrays in all tables by one specified 
        attribute.
        Either ascending or descending.
        """
        for d in self.valids.values():
            for tbl in d.values():
                self._sort_all(tbl, attr, asc)
    
    def get_scored_hits(self, data):
        hits = val_ubi_prf_rpr_hits(data, ubiquity = 70, profile_best = 50000)
        [v[0].shape[0] if v is not None else None \
            for vv in hits.values() for v in vv.values()]
        hits_upper = dict((l.upper(), {'pos': None, 'neg': None}) \
            for l in hits.keys())
        for l, d in iteritems(hits):
            for pn, tbl in iteritems(d):
                if hits_upper[l.upper()][pn] is None:
                    hits_upper[l.upper()][pn] = tbl
        return hits_upper

    """
    MS1 lipid identification
    """

    def lipid_lookup(self, stage0, runtime = None):
        """
        Obtains full SwissLipids data
        """
        pAdducts, nAdducts = get_swisslipids(
            adducts = ['[M+H]+', '[M+NH4]+', '[M-H]-'],
            formiate = True)
        lipids = find_lipids(stage0, pAdducts, nAdducts)
        if runtime:
            runtime = timeit.timeit(
                'find_lipids(final_hits, pAdducts, nAdducts)',
                setup = 'from __main__ import find_lipids, final_hits,'\
                'pAdducts, nAdducts', number = 1)
        return pAdducts, nAdducts, lipids, runtime

    def lipid_lookup_exact(self, verbose = False, outfile = None, charge = 1):
        """
        Fetches data from SwissLipids and LipidMaps
        if not given.
        Looks up lipids based on exact masses (but
        taking into account the adducts).
        Writes hits into arrays in `lip` dict, having
        original indices as keys.
        Returns an array of lipid databases with exact
        masses.
        """
        if self.exacts is None:
            self.get_swisslipids_exact()
            self.lipidmaps_exact()
            self.add_nonidet()
        self.find_lipids_exact(verbose = verbose,
            outfile = outfile, charge = charge)

    def evaluate_results(self, stage0, stage2, lipids, fractions_upper,
        letter = 'e'):
        """
        Tracks the number of features/lipids along stages.
        Does some plotting.
        """
        stage3 = [(
                l.upper(), 
                [len(i['neg']), len(i['pos'])], 
                len(stage2[l.upper()]), 
                [len(stage0[l]['neg'][0]), len(stage0[l]['pos'][0])], 
                np.nansum([x for x in fractions_upper[l] if x is not None])
            ) if len(i.values()) > 0 \
            else (
                l.upper(), 0, 0, 0,
                np.nansum([x for x in fractions_upper[l] if x is not None])
            ) \
            for l, i in iteritems(lipids)]
        #
        [sys.stdout.write(str(i) + '\n') for i in stage3]
        sum([i[2] > 0 for i in stage3])
        #
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[1][0] for i in stage3 if i[-1] == f]
            x2 = [i[1][1] for i in stage3 if i[-1] == f]
            y = [i[2] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, yy in zip(x1, x2, y):
                p = plt.plot([xa, xb], [yy, yy], lw = .3, c = '#CCCCCC')
        plt.xlabel('Lipids (square: --; dot: +)')
        plt.ylabel('Lipids matching (square: -; dot: +)')
        fig.tight_layout()
        fig.savefig('hits-1-1000-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[3][0] for i in stage3 if i[-1] == f]
            y1 = [i[1][0] for i in stage3 if i[-1] == f]
            x2 = [i[3][1] for i in stage3 if i[-1] == f]
            y2 = [i[1][1] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y1, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y2, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, ya, yb in zip(x1, x2, y1, y2):
                p = plt.plot([xa, xb], [ya, yb], lw = .3, c = '#CCCCCC')
        #
        plt.xlabel('Num of m/z`s (green: --; yellow: +)')
        plt.ylabel('Lipids (green: --; yellow: +)')
        fig.tight_layout()
        fig.savefig('hits-2-1000-%s.pdf'%letter)
        plt.close()
        # 
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[3][0] for i in stage3 if i[-1] == f]
            x2 = [i[3][1] for i in stage3 if i[-1] == f]
            y = [i[2] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, yy in zip(x1, x2, y):
                p = plt.plot([xa, xb], [yy, yy], lw = .3, c = '#CCCCCC')
        plt.xlabel('Num of m/z`s (square: --; dot: +)')
        plt.ylabel('Lipids matching (square: -; dot: +)')
        fig.tight_layout()
        fig.savefig('hits-5-1000-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        y = flatList([i[3] for i in stage3 if type(i[3]) is not int])
        x = flatList([[i[4]]*2 for i in stage3 if type(i[3]) is not int])
        p = plt.scatter(x, y, c = '#6ea945', edgecolors = 'none')
        plt.xlabel('Fractions in sample')
        plt.ylabel('M/z min(pos, neg)')
        fig.tight_layout()
        fig.savefig('hits-3-150-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        y = [i[2] for i in stage3]
        x = [i[4] for i in stage3]
        p = plt.scatter(x, y, c = '#6ea945', edgecolors = 'none', alpha = .2)
        plt.xlabel('Fractions in sample')
        plt.ylabel('Lipids matched (pos, neg)')
        fig.tight_layout()
        fig.savefig('hits-4-1000-%s.pdf' % letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        p = plt.bar(xrange(len(stage3)),
            [i[2] for i in sorted(stage3,
                key = lambda x: x[2], reverse = True)],
            0.5, color = '#6EA945', edgecolor = 'none')
        plt.xlabel('Lipid transfer protein')
        plt.ylabel('Number of valid lipid matches\n'\
            '(with no filtering based on profiles)')
        ax.set_xticks(np.arange(len(stage3)) + .25)
        ax.set_xticklabels([i[0] for i in \
                sorted(stage3, key = lambda x: x[2], reverse = True)],
            rotation = 90, fontsize = 5)
        fig.tight_layout()
        fig.savefig('lipid-matches-3000-%s.pdf'%letter)
        plt.close()

    def ms1_headgroups(self, verbose = False):
        """
        Identifies headgroups by keywords and fatty acids
        from database record names.
        """
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                tbl['ms1hg'] = {}
                tbl['ms1fa'] = {}
                for oi, lips in iteritems(tbl['lip']):
                    tbl['ms1hg'][oi] = set([])
                    tbl['ms1fa'][oi] = {}
                    if lips is not None:
                        for lip in lips:
                            if lip[7] is not None:
                                if verbose:
                                    sys.stdout.write('\t:: %s-%s: '\
                                        'found %s\n' % \
                                        (protein, pn, lip[7]))
                                    sys.stdout.flush()
                                posAdd = self.lipnames[lip[7]]['pos_adduct']
                                negAdd = self.lipnames[lip[7]]['neg_adduct']
                                thisModeAdd = \
                                    self.lipnames[lip[7]]['%s_adduct'%pn]
                                hg = lip[7]
                                fa = '%u:%u' % (lip[8][0], lip[8][1]) \
                                    if lip[8] is not None else None
                                if (posAdd is None and negAdd is None) or \
                                    thisModeAdd == lip[4]:
                                    if verbose:
                                        sys.stdout.write('\t\taccepting '\
                                            '%s-%s for'\
                                            ' %s-%s\n' % \
                                            (hg, lip[4], protein, pn))
                                        sys.stdout.flush()
                                    tbl['ms1hg'][oi].add(hg)
                                    if fa is not None:
                                        if hg not in tbl['ms1fa'][oi]:
                                            tbl['ms1fa'][oi][hg] = set([])
                                        tbl['ms1fa'][oi][hg].add(fa)
                                else:
                                    if verbose:
                                        sys.stdout.write(
                                            '\t\tdiscarding %s-%s'\
                                            ' for %s, %s\n'\
                                            ' in %s mode' % (
                                                lip[7],
                                                lip[4],
                                                protein,
                                                    '%s is the main adduct' % \
                                                    thisModeAdd \
                                                    if thisModeAdd \
                                                        is not None \
                                                    else \
                                                        '%s does not'\
                                                        ' ionize' % \
                                                        lip[4],
                                                pn
                                            )
                                        )
                                        sys.stdout.flush()

    def headgroups_by_fattya(self, proteins = None, verbose = False):
        """
        Limits the number of possible headgroups based on detected
        MS2 fatty acids.
        Creates dict `hgfa`.
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mod, tbl in iteritems(d):
                
                tbl['hgfa'] = {}
                for oi in tbl['i']:
                    tbl['hgfa'][oi] = set([])
                    if oi in tbl['ms2fas'] and tbl['ms2fas'][oi] is not None:
                        ms2fa = tbl['ms2fas'][oi]
                        for hg, ms1fa in iteritems(tbl['ms1fa'][oi]):
                            if ms2fa in ms1fa:
                                tbl['hgfa'][oi].add(hg)
                                if verbose:
                                    sys.stdout.write('\t%s is among '\
                                        'possible fat'\
                                        'ty acids (%s) for %s based on MS1, '\
                                        'for feature %u at %s-%s \n' % \
                                        (ms2fa, ', '.join(list(ms1fa)), 
                                            hg, oi, protein, mod))
                            else:
                                if verbose:
                                    sys.stdout.write('\t%s not among '\
                                        'possible fat'\
                                        'ty acids (%s) for %s based on MS1, '\
                                        'for feature %u at %s-%s \n' % \
                                        (ms2fa, ', '.join(list(ms1fa)), 
                                            hg, oi, protein, mod))

    def identity_combined(self, proteins = None):
        """
        Combined identification based on MS1 database lookup,
        MS2 headgroup fragments and MS2 fatty acids.
        Creates dicts `combined_hg` and `combined_fa`.
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mod, tbl in iteritems(d):
                
                tbl['combined_hg'] = {}
                tbl['combined_fa'] = {}
                for oi in tbl['i']:
                    hgs = set([])
                    tbl['combined_fa'][oi] = {}
                    if oi not in tbl['ms2hg'] or tbl['ms2hg'][oi] is None:
                        hgs = tbl['ms1hg'][oi]
                    else:
                        if len(tbl['ms1hg'][oi] & tbl['ms2hg'][oi]):
                            hgs = tbl['ms1hg'][oi] & tbl['ms2hg'][oi]
                        else:
                            hgs = tbl['ms1hg'][oi] | tbl['ms2hg'][oi]
                    if len(tbl['hgfa'][oi]) and len(hgs & tbl['hgfa'][oi]):
                        hgs = hgs & tbl['hgfa'][oi]
                    else:
                        hgs = tbl['hgfa'][oi]
                    tbl['combined_hg'][oi] = hgs
                    for hg in hgs:
                        if hg in tbl['ms1fa'][oi] and \
                            tbl['ms2fas'][oi] is not None and \
                            tbl['ms2fas'][oi] in tbl['ms1fa'][oi][hg]:
                            if hg not in tbl['combined_fa'][oi]:
                                tbl['combined_fa'][oi][hg] = set([])
                            # currently only the most certain cases:
                            tbl['combined_fa'][oi][hg].add(tbl['ms2fas'][oi])
                            # maybe later we need those with less evidence

    def identity_combined_ms2(self):
        """
        Combined identification based on MS1 database lookup,
        MS2 headgroup fragments and MS2 fatty acids.
        Creates dicts `combined_hg` and `combined_fa`.
        """
        for protein, d in iteritems(self.valids):
            for mod, tbl in iteritems(d):
                tbl['combined_hg_ms2'] = {}
                for oi in tbl['i']:
                    tbl['combined_hg_ms2'][oi] = set([])
                    if oi in tbl['ms2hg']:
                        hgs = tbl['ms1hg'][oi] & tbl['ms2hg'][oi]
                    if len(tbl['hgfa'][oi]) and len(hgs & tbl['hgfa'][oi]):
                        hgs = hgs & tbl['hgfa'][oi]
                        tbl['combined_hg_ms2'][oi] = hgs

    def ms1_table(self, include = 'cl70pct'):
        proteins = sorted(self.valids.keys())
        # collecting primary and secondary column names
        # and cell values
        result = dict((protein, {}) for protein in proteins)
        colnames = {}
        for protein in proteins:
            for pn, tbl in iteritems(self.valids[protein]):
                _np = 'pos' if pn == 'neg' else 'neg'
                nptbl = tbl[_np]
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        if tbl['lip'][oi] is not None:
                            for lip in tbl['lip'][oi]:
                                hg = lip[7]
                                if hg is not None:
                                    fa = '%u:%u' % tuple(lip[8]) \
                                        if lip[8] is not None else 'unknown'
                                    # colnames
                                    if hg not in colnames:
                                        colnames[hg] = set([])
                                    colnames[hg].add(fa)
                                    # cells
                                    if hg not in result[protein]:
                                        result[protein][hg] = {}
                                    if fa not in result[protein][hg]:
                                        # 3 bool values: +, -,
                                        # both + & - at same exact mass
                                        result[protein][hg][fa] = \
                                            [False, False, False]
                                    # cell values
                                    if pn == 'pos':
                                        result[protein][hg][fa][0] = True
                                    elif pn == 'neg':
                                        result[protein][hg][fa][1] = True
                                    for _oi, pnlips in \
                                        iteritems(tbl['%s_lip'%_np][oi]):
                                        if not result[protein][hg][fa][2] and \
                                            pnlips is not None:
                                            for pnlip in pnlips:
                                                pfa = '%u:%u' % \
                                                    tuple(pnlip[12]) \
                                                    if pnlip[12] is not None \
                                                    else 'unknown'
                                                nfa = '%u:%u' % \
                                                    tuple(pnlip[13]) \
                                                    if pnlip[13] is not None \
                                                    else 'unknown'
                                                if pfa == fa and nfa == fa and \
                                                    pnlip[10] == hg and \
                                                    pnlip[11] == hg:
                                                    result[protein][hg][fa][2] = \
                                                        True
                                                    break
        return colnames, result
    
    def today(self):
        return datetime.date.today().strftime(r'%Y-%m-%d')
    
    def ms1_table_html(self, filename = None, include = 'cl70pct'):
        filename = 'results_%s_ms1.html' % self.today() \
            if filename is None \
            else filename
        colnames, ms1tab = self.ms1_table(include = include)
        title = 'Binding specificities of LTPs detected in MS1'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tableshcell = '\t\t\t<th colspan="%u">\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % ''
        th2 = tablehcell % 'LTP'
        for hg in sorted(colnames.keys()):
            th1 += tableshcell % (len(colnames[hg]), hg)
            for fa in sorted(colnames[hg]):
                th2 += tablehcell % fa
        table += tablerow % th1
        table += tablerow % th2
        for protein in sorted(self.valids.keys()):
            row = tablecell % ('rowname', protein, protein)
            for hg in sorted(colnames.keys()):
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[protein] and fa in ms1tab[protein][hg]:
                        if ms1tab[protein][hg][fa][2]:
                            row += tablecell % ('matching', 
                                '%s (%s) detected in '\
                                'Positive & Negative modes at %s' % \
                                    (hg, fa, protein),
                                '')
                        elif ms1tab[protein][hg][fa][0] and \
                            ms1tab[protein][hg][fa][1]:
                            row += tablecell % ('both', '%s (%s) detected in '\
                                'Positive & Negative modes at %s' % \
                                    (hg, fa, protein),
                                '')
                        elif ms1tab[protein][hg][fa][0]:
                            row += tablecell % ('positive',
                                '%s (%s) detected '\
                                'in Positive mode at %s' % (hg, fa, protein), '')
                        elif ms1tab[protein][hg][fa][1]:
                            row += tablecell % ('negative',
                                '%s (%s) detected '\
                                'in Negative mode at %s' % (hg, fa, protein), '')
                    else:
                        row += tablecell % ('empty',
                            '%s (%s) not detected at %s' % (hg, fa, protein), '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))

    def ms1_table_html_simple(self, filename = None, include = 'cl70pct'):
        """
        Outputs a HTML table proteins vs lipid classes (headgroups)
        based on MS1 identifications.
        """
        filename = 'results_%s_ms1hg.html' % self.today() \
            if filename is None \
            else filename
        colnames, ms1tab = self.ms1_table(include = include)
        title = 'Binding specificities of proteins by headgroups detected in MS1'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'LTP'
        for hg in sorted(colnames.keys()):
            th1 += tablehcell % hg
        table += tablerow % th1
        for protein in sorted(self.valids.keys()):
            row = tablecell % ('rowname', protein, protein)
            for hg in sorted(colnames.keys()):
                pos_neg = False
                pos = False
                neg = False
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[protein] and fa in ms1tab[protein][hg]:
                        if ms1tab[protein][hg][fa][2]:
                            pos_neg = True
                        elif ms1tab[protein][hg][fa][0] and \
                            ms1tab[protein][hg][fa][1]:
                            pos_neg = True
                        elif ms1tab[protein][hg][fa][0]:
                            pos = True
                        elif ms1tab[protein][hg][fa][1]:
                            neg = True
                if pos_neg:
                    row += tablecell % ('matching', 
                        '%s detected in Positive & '\
                        'Negative modes' % hg, '')
                elif pos:
                    row += tablecell % ('positive', '%s detected in Positive '\
                        'mode' % hg, '')
                elif neg:
                    row += tablecell % ('negative', '%s detected in Negative '\
                        'mode' % hg, '')
                else:
                    row += tablecell % ('empty', '%s not detected' % hg, '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))
    
    def ms1_table_latex_simple(self, filename = 'ms1headgroups.tex', include = 'cl70pct', break_half = True):
        """
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS1 identifications.
        """
        colnames, ms1tab = self.ms1_table(include = include)
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        tablerow = '\t%s\\\\\n'
        hdr = '& '
        hdr += ' & '.join(
            sorted(
                map(
                    lambda hg:
                        '\\rot{%s}' % hg,
                    colnames.keys()
                )
            )
        )
        table += tablerow % hdr
        for i, protein in enumerate(sorted(self.valids.keys())):
            row = [protein]
            for hg in sorted(colnames.keys()):
                pos_neg = False
                pos = False
                neg = False
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[protein] and fa in ms1tab[protein][hg]:
                        if ms1tab[protein][hg][fa][2]:
                            pos_neg = True
                        elif ms1tab[protein][hg][fa][0] and \
                            ms1tab[protein][hg][fa][1]:
                            pos_neg = True
                        elif ms1tab[protein][hg][fa][0]:
                            pos = True
                        elif ms1tab[protein][hg][fa][1]:
                            neg = True
                if pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)

    """
    END: MS1 lipid identification
    """

    def ms2_table_html_simple(self, filename = None, include = 'cl70pct'):
        """
        Outputs a HTML table LTPs vs lipid classes (headgroups)
        based on MS2 identifications.
        """
        filename = 'results_%s_ms2hg.html' % self.today() \
            if filename is None \
            else filename
        title = 'Binding specificities of LTPs by headgroups detected in MS2'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'protein'
        colnames = set([])
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                for hg in tbl['ms2hg'].values():
                    if hg is not None:
                        colnames = colnames | hg
        colnames = sorted(list(colnames))
        for hg in colnames:
            th1 += tablehcell % hg
        table += tablerow % th1
        for protein in sorted(self.valids.keys()):
            row = tablecell % ('rowname', protein, protein)
            for hg in colnames:
                pos_neg_same = False
                pos_neg = False
                pos = False
                neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                for pn, tbl in iteritems(self.valids[protein]):
                    for ms2hg in tbl['ms2hg'].values():
                        if ms2hg is not None and hg in ms2hg:
                            if pn == 'pos':
                                pos = True
                                if len(ms2hg) == 1:
                                    pos_unambig = True
                            elif pn == 'neg':
                                neg = True
                                if len(ms2hg) == 1:
                                    neg_unambig = True
                if pos and neg:
                    pos_neg = True
                if pos_unambig or neg_unambig:
                    pos_neg_unambig = True
                for ms2hgs in self.valids[protein]['pos']['ms2hg_neg'].values():
                    for ms2hg in ms2hgs.values():
                        if hg in ms2hg:
                            pos_neg_same = True
                            if len(ms2hg) == 1:
                                pos_neg_unambig = True
                if pos_neg_same:
                    unambig = 'UA' if pos_neg_same_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_same_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('matching', 'Detected in Positive &'\
                        ' Negative modes,\nat same exact mass%s'%\
                        unambig2, unambig)
                elif pos_neg:
                    unambig = 'UA' if pos_neg_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('both', 'Detected in Positive &'\
                        ' Negative modes,\nat different exact mass%s'%\
                        unambig2, unambig)
                elif pos:
                    unambig2 = '\nUnambiguous at least once' if pos_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if pos_unambig else 'A'
                    row += tablecell % ('positive',
                        'Detected in Positive mode%s' % \
                        unambig2, unambig)
                elif neg:
                    unambig2 = '\nUnambiguous at least once' if neg_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if neg_unambig else 'A'
                    row += tablecell % ('negative',
                        'Detected in Negative mode%s'%\
                        unambig2, unambig)
                else:
                    row += tablecell % ('empty', 'Not detected', '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))

    def ms2_table_latex_simple(self, filename = 'ms2headgroups.tex', include = 'cl70pct', break_half = True):
        """
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS2 identifications.
        """
        colnames = set([])
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                for hg in tbl['ms2hg'].values():
                    if hg is not None:
                        colnames = colnames | hg
        tablerow = '\t%s\\\\\n'
        colnames.remove('')
        colnames = sorted(list(colnames))
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        for i, protein in enumerate(sorted(self.valids.keys())):
            row = [protein]
            for hg in colnames:
                pos_neg_same = False
                pos_neg = False
                pos = False
                neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                for pn, tbl in iteritems(self.valids[protein]):
                    for ms2hg in tbl['ms2hg'].values():
                        if ms2hg is not None and hg in ms2hg:
                            if pn == 'pos':
                                pos = True
                                if len(ms2hg) == 1:
                                    pos_unambig = True
                            elif pn == 'neg':
                                neg = True
                                if len(ms2hg) == 1:
                                    neg_unambig = True
                if pos and neg:
                    pos_neg = True
                if pos_unambig or neg_unambig:
                    pos_neg_unambig = True
                for ms2hgs in self.valids[protein]['pos']['ms2hg_neg'].values():
                    for ms2hg in ms2hgs.values():
                        if hg in ms2hg:
                            pos_neg_same = True
                            if len(ms2hg) == 1:
                                pos_neg_unambig = True
                if pos_neg_same or pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)

    def ms1_ms2_table_html_simple(self, filename = None, include = 'cl70pct'):
        """
        Outputs a HTML table LTPs vs lipid classes (headgroups)
        based on MS1 and MS2 identifications.
        """
        filename = 'results_%s_ms1ms2hg.html' % self.today() \
            if filename is None \
            else filename
        title = 'Binding specificities of LTPs by headgroups'\
            ' detected in MS1 and MS2'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'LTP'
        colnames = set([])
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                for hgs in tbl['identity'].values():
                    for hg, ids in iteritems(hgs):
                        if hg is not None and \
                            (ids['ms2_pos'] or ids['ms2_neg']):
                            colnames.add(hg)
        colnames = sorted(list(colnames))
        # header row (lipid species)
        for hg in colnames:
            th1 += tablehcell % hg
        table += tablerow % th1
        # rows by protein
        for protein in sorted(self.valids.keys()):
            row = tablecell % ('rowname', protein, protein)
            for hg in colnames:
                pos = False
                neg = False
                pos_neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                pos_neg_same = False
                tbl = self.valids[protein]['pos']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                            pos = True
                            if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                                pos_neg_same = True
                                for noi in tbl['neg'][oi].keys():
                                    if hg in valids[protein]['neg']\
                                        ['identity'][noi]:
                                        if self.valids[protein]['neg']['identity']\
                                            [noi][hg]['ms1_neg'] \
                                                and \
                                            self.valids[protein]['neg']['identity']\
                                            [noi][hg]['ms2_neg']:
                                            pos_neg = True
                                            if sum(map(lambda _hg:
                                                    _hg[0] != hg \
                                                            and \
                                                        _hg[1]['ms1_neg'] \
                                                            and \
                                                        _hg[1]['ms2_neg'],
                                                    iteritems(
                                                        valids[protein]['neg']\
                                                        ['identity']\
                                                        [noi])
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda _hg:
                                    _hg[0] != hg and _hg[1]['ms1_pos'] \
                                        and _hg[1]['ms2_pos'],
                                    iteritems(tbl['identity'][oi])
                                )) == 0:
                                pos_unambig = True
                tbl = self.valids[protein]['neg']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                            neg = True
                            if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                                pos_neg_same = True
                                for noi in tbl['pos'][oi].keys():
                                    if hg in self.valids[protein]['pos']\
                                        ['identity'][noi]:
                                        if valids[protein]['pos']['identity']\
                                            [noi][hg]['ms1_pos'] and \
                                            self.valids[protein]['pos']['identity']\
                                                [noi][hg]['ms2_pos']:
                                            pos_neg = True
                                            if sum(map(lambda _hg:
                                                    _hg[0] != hg and \
                                                        _hg[1]['ms1_pos'] \
                                                            and \
                                                        _hg[1]['ms2_pos'],
                                                    iteritems(
                                                        valids[protein]['pos']\
                                                        ['identity']\
                                                        [noi])
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda _hg:
                                    _hg[0] != hg and _hg[1]['ms1_neg'] \
                                        and _hg[1]['ms2_neg'],
                                    iteritems(tbl['identity'][oi])
                                )) == 0:
                                neg_unambig = True
                if pos_unambig and neg_unambig:
                    pos_neg_unambig = True
                if pos_neg_same:
                    unambig = 'UA' if pos_neg_same_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_same_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('matching', '%s detected in Positive &'\
                        ' Negative modes,\nat same exact mass%s'%\
                        (hg, unambig2), unambig)
                elif pos_neg:
                    unambig = 'UA' if pos_neg_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('both', '%s detected in Positive &'\
                        ' Negative modes,\nat different exact mass%s' % \
                        (hg, unambig2), unambig)
                elif pos:
                    unambig2 = '\nUnambiguous at least once' if pos_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if pos_unambig else 'A'
                    row += tablecell % ('positive',
                        '%s detected in Positive mode%s' % \
                        (hg, unambig2), unambig)
                elif neg:
                    unambig2 = '\nUnambiguous at least once' if neg_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if neg_unambig else 'A'
                    row += tablecell % ('negative',
                        '%s detected in Negative mode%s' % \
                        (hg, unambig2), unambig)
                else:
                    row += tablecell % ('empty', '%s not detected' % hg, '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))
    
    def ms1_ms2_table_latex_simple(self, filename = 'ms1ms2headgroups.tex',
        include = 'cl70pct', break_half = True):
        """
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS1 and MS2 identifications.
        """
        colnames = set([])
        for ltp, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                for hgs in tbl['identity'].values():
                    for hg, ids in iteritems(hgs):
                        if hg is not None and \
                            (ids['ms2_pos'] or ids['ms2_neg']):
                            colnames.add(hg)
        colnames.remove('')
        colnames = sorted(list(colnames))
        tablerow = '\t%s\\\\\n'
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        # rows by LTP
        for i, protein in enumerate(sorted(self.valids.keys())):
            row = [protein]
            for hg in colnames:
                pos = False
                neg = False
                pos_neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                pos_neg_same = False
                tbl = self.valids[protein]['pos']
                
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    
                    if hg in tbl['identity'][oi]:
                        
                        this_hg = tbl['identity'][oi][hg]
                        
                        if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                            
                            pos = True
                            
                            if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                                
                                pos_neg_same = True
                                for noi in tbl['neg'][oi].keys():
                                    
                                    if hg in valids[protein]['neg']\
                                        ['identity'][noi]:
                                        
                                        if self.valids[protein]['neg']['identity']\
                                            [noi][hg]['ms1_neg'] \
                                                and \
                                            self.valids[protein]['neg']['identity']\
                                            [noi][hg]['ms2_neg']:
                                            
                                            pos_neg = True
                                            
                                            if sum(map(lambda _hg:
                                                    _hg[0] != hg \
                                                            and \
                                                        _hg[1]['ms1_neg'] \
                                                            and \
                                                        _hg[1]['ms2_neg'],
                                                    iteritems(
                                                        valids[protein]['neg']\
                                                        ['identity']\
                                                        [noi]
                                                    )
                                                )) == 0:
                                                
                                                pos_neg_same_unambig = True
                            
                            if sum(map(lambda _hg:
                                    _hg[0] != hg and _hg[1]['ms1_pos'] \
                                        and _hg[1]['ms2_pos'],
                                    iteritems(tbl['identity'][oi])
                                )) == 0:
                                pos_unambig = True
                
                tbl = self.valids[protein]['neg']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    
                    if hg in tbl['identity'][oi]:
                        
                        this_hg = tbl['identity'][oi][hg]
                        
                        if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                            
                            neg = True
                            if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                                
                                pos_neg_same = True
                                
                                for noi in tbl['pos'][oi].keys():
                                    
                                    if hg in self.valids[protein]['pos']\
                                        ['identity'][noi]:
                                        
                                        if valids[protein]['pos']['identity']\
                                            [noi][hg]['ms1_pos'] and \
                                            self.valids[protein]['pos']['identity']\
                                                [noi][hg]['ms2_pos']:
                                            
                                            pos_neg = True
                                            
                                            if sum(map(lambda _hg:
                                                    _hg[0] != hg and \
                                                        _hg[1]['ms1_pos'] \
                                                            and \
                                                        _hg[1]['ms2_pos'],
                                                    iteritems(
                                                        valids[protein]['pos']\
                                                        ['identity']\
                                                        [noi]
                                                    )
                                                )) == 0:
                                                
                                                pos_neg_same_unambig = True
                            
                            if sum(map(lambda _hg:
                                    _hg[0] != hg and _hg[1]['ms1_neg'] \
                                        and _hg[1]['ms2_neg'],
                                    iteritems(tbl['identity'][oi])
                                )) == 0:
                                neg_unambig = True
                
                if pos_unambig and neg_unambig:
                    pos_neg_unambig = True
                
                if pos_neg_unambig:
                    row.append('\\cellcolor{emblyellow!75}')
                
                elif pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                
                else:
                    row.append('')
            
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)
    
    def identities_latex_table(self, filename = 'ms1ms2hcheadgroups2.tex', break_half = True, include = 'cl70pct'):
        allhgs = set([])
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                for oi, ms2results in iteritems(tbl['ms2i2']):
                    for hg, ms2result in iteritems(ms2results):
                        if sum(map(lambda i: i['score'], ms2result)) > 0:
                            allhgs.add(hg)
        colnames = sorted(list(allhgs))
        tablerow = '\t%s\\\\\n'
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        for l, protein in enumerate(sorted(self.valids.keys())):
            row = [protein]
            for hg in colnames:
                detected = {'neg': False, 'pos': False}
                for mode in ['neg', 'pos']:
                    for i, oi in enumerate(self.valids[protein][mode]['i']):
                        if oi in self.valids[protein][mode]['ms2i3']:
                            ms2results = self.valids[protein][mode]['ms2i3'][oi]
                            if not include or self.valids[protein][mode][include][i]:
                                if hg in ms2results:
                                    detected[mode] = True
                if detected['pos'] and \
                    detected['neg']:
                    row.append('\\cellcolor{emblyellow!75}')
                elif detected['neg']:
                    row.append('\\cellcolor{emblpetrol!75}')
                elif detected['pos']:
                    row.append('\\cellcolor{emblgreen!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and l == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)
    
    def feature_identity_table(self):
        """
        Creates dictionaries named `identity`, having
        original IDs as keys and 4 element dictionaries
        as values with keys ms1_pos, ms2_pos, ms1_neg, ms2_neg,
        each having a boolean value.
        """
        self.sort_alll('mz')
        for protein, d in iteritems(self.valids):
            for pn, tbl in iteritems(d):
                opp_mode = 'neg' if pn == 'pos' else 'pos'
                tbl['identity'] = {}
                for oi in tbl['i']:
                    tbl['identity'][oi] = {}
                    ms1 = set([]) if tbl['ms1hg'][oi] is None \
                        else tbl['ms1hg'][oi]
                    ms2 = set([]) \
                        if oi not in tbl['ms2hg'] or tbl['ms2hg'][oi] is None \
                        else tbl['ms2hg'][oi]
                    ms1_opp = set([])
                    ms2_opp = set([])
                    for opp_oi in tbl[opp_mode][oi].keys():
                        if opp_oi in d[opp_mode]['ms1hg']:
                            ms1_opp = ms1_opp | d[opp_mode]['ms1hg'][opp_oi]
                        if opp_oi in d[opp_mode]['ms2hg'] and \
                            d[opp_mode]['ms2hg'][opp_oi] is not None:
                            ms2_opp = ms2_opp | d[opp_mode]['ms2hg'][opp_oi]
                    hg_all = ms1 | ms2 | ms1_opp | ms2_opp
                    for hg in hg_all:
                        tbl['identity'][oi][hg] = {
                            'ms1_%s'%pn: hg in ms1,
                            'ms2_%s'%pn: hg in ms2,
                            'ms1_%s'%opp_mode: hg in ms1_opp,
                            'ms2_%s'%opp_mode: hg in ms2_opp
                        }

    def feature_identity_html_table(self, valids, bindprop,
        fits_pprop = 'cl5pct',
        outf = 'identities.html'):
        hdr = ['LTP', 'm/z', 'mode', 'fits protein +', 'fits protein -', 
            'headgroup', 'MS1+', 'MS2+', 'MS1-', 'MS2-']
        title = 'Identities for all features'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        hrow = ''
        for coln in hdr:
            hrow += tablehcell % coln
        hrow = tablerow % hrow
        table = hrow
        empty_ids = {'': {'ms1_pos': False, 'ms2_pos': False, 
            'ms1_neg': False, 'ms2_neg': False}}
        for protein, d in iteritems(valids):
            for pn, tbl in iteritems(d):
                _np = 'pos' if pn == 'neg' else 'neg'
                for i, oi in enumerate(tbl['i']):
                    for hg, ids in iteritems((tbl['identity'][oi]) \
                        if len(tbl['identity'][oi]) > 0 \
                        else iteritems(empty_ids)):
                        this_row = ''
                        this_row += tablecell % ('rowname', '', protein)
                        this_row += tablecell % \
                            ('nothing', '', '%.05f'%tbl['mz'][i])
                        this_row += tablecell % ('nothing', '', 
                            '+' if pn == 'pos' else '-')
                        fits_pos = tbl[fits_pprop][i] if pn == 'pos' else \
                            bool(sum(map(lambda _i: _i[1],
                                    filter(lambda _i: 
                                        d['pos']['i'][_i[0]] in tbl['pos'][oi], 
                                        enumerate(d['pos'][fits_pprop])
                                    )
                                )
                            ))
                        fits_neg = tbl[fits_pprop][i] if pn == 'neg' else \
                            bool(sum(map(lambda _i: _i[1],
                                    filter(lambda _i: 
                                        d['neg']['i'][_i[0]] in tbl['neg'][oi], 
                                        enumerate(d['neg'][fits_pprop])
                                    )
                                )
                            ))
                        this_row += tablecell % (
                            'positive' if fits_pos else 'nothing',
                            'Positive mode fits well on protein profile' \
                                if fits_pos \
                                else 'Positive mode does not fit well on'\
                                    ' protein profile or not available',
                            '')
                        this_row += tablecell % (
                            'positive' if fits_neg else 'nothing',
                            'Negative mode fits well on protein profile' \
                                if fits_neg \
                                else 'Negative mode does not fit well on'\
                                    ' protein profile or not available',
                            '')
                        this_row += tablecell % (
                            'positive' if hg in bindprop[protein] else 'nothing',
                            '%s is known binder of %s' % (hg, protein) \
                                if hg in bindprop[protein] \
                                else 'No literature data '\
                                    'about %s binding %s' % \
                                    (protein, hg),
                            hg)
                        this_row += tablecell % (
                            'positive' if ids['ms1_pos'] else 'nothing',
                            'Identified in MS1 positive mode' \
                                if ids['ms1_pos'] \
                                else 'Not identified in MS1 positive mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms2_pos'] else 'nothing',
                            'Identified in MS2 positive mode' \
                                if ids['ms2_pos'] \
                                else 'Not identified in MS2 positive mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms1_neg'] else 'nothing',
                            'Identified in MS1 negative mode' \
                                if ids['ms1_neg'] \
                                else 'Not identified in MS1 negative mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms2_neg'] else 'nothing',
                            'Identified in MS2 negative mode' \
                                if ids['ms2_neg'] \
                                else 'Not identified in MS2 negative mode',
                            '')
                        this_row = tablerow % this_row
                        table += this_row
        with open(outf, 'w') as f:
            f.write(html_table_template % (title, title, table))
    
    def known_binders_enrichment(self, valids, bindprop, classif = 'cl5pct'):
        for protein, d in iteritems(valids):
            for pn, tbl in iteritems(d):
                if 'slo' not in tbl:
                    tbl['slo'] = {}
                if classif not in tbl['slo']:
                    tbl['slo'][classif] = {}
                tbl['slo'][classif]['tp'] = len(
                    filter(lambda ioi:
                        tbl[classif][ioi[0]] and \
                            len(bindprop[protein] & \
                                set(tbl['identity'][ioi[1]].keys())) > 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['slo'][classif]['fp'] = len(
                    filter(lambda ioi:
                        tbl[classif][ioi[0]] and \
                            len(bindprop[protein] & \
                                set(tbl['identity'][ioi[1]].keys())) == 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['slo'][classif]['tn'] = len(
                    filter(lambda ioi:
                        not tbl[classif][ioi[0]] and \
                            len(bindprop[protein] & \
                                set(tbl['identity'][ioi[1]].keys())) == 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['slo'][classif]['fn'] = len(
                    filter(lambda ioi:
                        not tbl[classif][ioi[0]] and \
                            len(bindprop[protein] & \
                                set(tbl['identity'][ioi[1]].keys())) > 0,
                        enumerate(tbl['i'])
                    )
                )
                try:
                    tbl['slo'][classif]['or'] = \
                        float(tbl['slo'][classif]['tp'] * \
                        tbl['slo'][classif]['tn']) / \
                        float(tbl['slo'][classif]['fp'] * \
                        tbl['slo'][classif]['fn'])
                except ZeroDivisionError:
                    tbl['slo'][classif]['or'] = np.inf
                tbl['slo'][classif]['contab'] = np.array([
                    [tbl['slo'][classif]['tp'], tbl['slo'][classif]['fp']],
                    [tbl['slo'][classif]['fn'], tbl['slo'][classif]['tn']]])
                tbl['slo'][classif]['fisher'] = \
                    sp.stats.fisher_exact(tbl['slo'][classif]['contab'])
    
    def enrichment_barplot(self, valids, classif = 'cl5pct', 
        outf = 'known_binders_enrichment.pdf'):
        w = 0.45
        labels = sorted(filter(lambda protein:
            valids[protein]['pos']['slo'][classif]['tp'] + \
                valids[protein]['pos']['slo'][classif]['fn'] > 0 or \
            valids[protein]['neg']['slo'][classif]['tp'] + \
                valids[protein]['neg']['slo'][classif]['fn'] > 0,
            valids.keys()
        ))
        ppvals = map(lambda protein:
            '%.03f' % valids[protein]['pos']['slo'][classif]['fisher'][1],
            labels
        )
        npvals = map(lambda protein:
            '%.03f' % valids[protein]['neg']['slo'][classif]['fisher'][1],
            labels
        )
        pors = map(lambda protein:
            '%.03f' % valids[protein]['pos']['slo'][classif]['fisher'][0],
            labels
        )
        nors = map(lambda protein:
            '%.03f' % valids[protein]['neg']['slo'][classif]['fisher'][0],
            labels
        )
        fig = mpl.figure.Figure(figsize = (8, 4))
        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
        ax = fig.gca()
        ax.bar(np.arange(len(labels)), pors, w, color = '#97BE73', lw = 0.0)
        ax.bar(np.arange(len(labels)) + w, nors, w, color = '#49969A', lw = 0.0)
        ax.set_yscale('symlog')
        ax.set_xticks(np.arange(len(labels)) + w)
        ax.set_xticklabels(labels, rotation = 90)
        ax.set_xlabel('LTPs')
        ax.set_ylabel('Enrichment (odds ratio)')
        cvs.draw()
        fig.tight_layout()
        cvs.print_figure(outf)
        fig.clf()

    def identification_levels(self, protein = 'STARD10',
        hg = 'PC', classif = None):
        visited = {'neg': set([]), 'pos': set([])}
        labels = ['MS1 & MS2 both +/-', 
            'MS1 & MS2 only +', 'MS1 & MS2 only -',
            'MS1 both +/-, no MS2', 'MS2 both +/-, no MS1',
            'MS1 + and MS2 -', 'MS1 - and MS2 +',
            'Only MS1 +', 'Only MS1 -', 'Only MS2 +', 'Only MS2 -',
            'MS1 both +/-, MS2 +', 'MS1 both +/-, MS2 -',
            'MS1 +, MS2 both +/-', 'MS1 -, MS2 both +/-',
            'Nothing', 'Non %s' % hg,
            'Total', '+/- Total', 'Only + Total', 'Only - Total']
        values = dict(zip(labels, [0] * len(labels)))
        modes = {'pos': '+', 'neg': '-'}
        keys = ['ms1_pos', 'ms2_pos', 'ms1_neg', 'ms2_neg']
        combinations = {
            (True, True, True, True): 'MS1 & MS2 both +/-',
            (True, True, False, False): 'MS1 & MS2 only +',
            (False, False, True, True): 'MS1 & MS2 only -',
            (True, False, True, False): 'MS1 both +/-, no MS2',
            (False, True, False, True): 'MS2 both +/-, no MS1',
            (True, False, False, True): 'MS1 + and MS2 -',
            (False, True, True, False): 'MS1 - and MS2 +',
            (True, False, False, False): 'Only MS1 +',
            (False, False, True, False): 'Only MS1 -',
            (False, True, False, False): 'Only MS2 +',
            (False, False, False, True): 'Only MS2 -',
            (True, True, True, False): 'MS1 both +/-, MS2 +',
            (True, False, True, True): 'MS1 both +/-, MS2 -',
            (True, True, False, True): 'MS1 +, MS2 both +/-',
            (False, True, True, True): 'MS1 -, MS2 both +/-'
        }
        for mode, opp_mod in [('pos', 'neg'), ('neg', 'pos')]:
            tbl = valids[protein][mode]
            opp_tbl = self.valids[protein][opp_mod]
            for i, oi in enumerate(tbl['i']):
                if oi not in visited[mode] and \
                    (classif is None or tbl[classif][i]):
                    visited[mode].add(oi)
                    values['Total'] += 1
                    opp_indices = map(lambda opp_oi:
                        np.where(opp_tbl['i'] == opp_oi)[0][0],
                        tbl[opp_mod][oi].keys()
                    )
                    opp_class = classif is None or len(opp_indices) == 0 or \
                        sum(opp_tbl[classif][opp_indices]) > 0
                    if len(tbl[opp_mod][oi]) > 0 and opp_class:
                        values['+/- Total'] += 1
                        for opp_oi in tbl[opp_mod][oi].keys():
                            if classif is None or \
                            any(opp_tbl[classif][opp_indices]):
                                visited[opp_mod].add(opp_oi)
                    else:
                        values['Only %s Total'%modes[mode]] += 1
                    if hg not in tbl['identity'][oi] or (\
                        hg in tbl['identity'][oi] and \
                        not tbl['identity'][oi][hg]['ms1_%s'%mode] and \
                        not tbl['identity'][oi][hg]['ms2_%s'%mode]):
                        if len(set(tbl['identity'][oi]) - set([hg])) > 0:
                            values['Non %s'%hg] += 1
                        else:
                            values['Nothing'] += 1
                    else:
                        comb = map(lambda k: 
                            tbl['identity'][oi][hg][k], 
                            keys
                        )
                        if not opp_class:
                            if opp_mod == 'neg':
                                comb[2] = False
                                comb[3] = False
                            else:
                                comb[0] = False
                                comb[1] = False
                        comb = tuple(comb)
                        values[combinations[comb]] += 1
        return values
    
    """
    {'MS1 both +/-, MS2 +': 7, 'Only MS2 -': 0, 'MS1 - and MS2 +': 2,
     'MS1 both +/-, MS2 -': 1, 'MS1 +, MS2 both +/-': 0, '+/- Total': 46,
     'MS2 both +/-, no MS1': 0, 'Non PC': 79, 'Nothing': 221,
     'MS1 & MS2 only +': 12, 'MS1 & MS2 both +/-': 8,
     'MS1 & MS2 only -': 0, 'Only + Total': 298, 'Only MS1 -': 9,
     'MS1 + and MS2 -': 0, 'MS1 both +/-, no MS2': 17,
     'MS1 -, MS2 both +/-': 1, 'Total': 448, 'Only - Total': 104,
     'Only MS1 +': 68, 'Only MS2 +': 23}
    """
    
    def plot_identification_levels(self, idlevels, protein, hg,
        fname = '%s-%s-classes.pdf'):
        fname = fname % (protein, hg)
        w = 0.8 / len(idlevels)
        cols = ['#97BE73', '#49969A', '#996A44', '#FDD73F']
        labels = ['MS1 & MS2 both +/-', 
            'MS1 & MS2 only +', 'MS1 & MS2 only -',
            'MS1 both +/-, no MS2', 'MS2 both +/-, no MS1', 
            'MS1 + and MS2 -', 'MS1 - and MS2 +',
            'Only MS1 +', 'Only MS1 -', 'Only MS2 +', 'Only MS2 -', 
            'MS1 both +/-, MS2 +', 'MS1 both +/-, MS2 -',
            'MS1 +, MS2 both +/-', 'MS1 -, MS2 both +/-',
            'Nothing', 'Non %s'%hg,
            'Total', '+/- Total', 'Only + Total', 'Only - Total']
        fig = mpl.figure.Figure(figsize = (8, 4))
        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
        ax = fig.gca()
        i = 0
        
        for lab in sorted(idlevels.keys()):
            ax.bar(np.arange(len(labels)) + i * w, 
                map(lambda l: idlevels[lab][l], labels), 
                w, color = cols[i], lw = 0.0)
            i += 1
        
        lhandles = map(lambda ilab:
            mpl.patches.Patch(color = cols[ilab[0]], label = ilab[1]),
            enumerate(sorted(idlevels.keys())))
        
        leg = ax.legend(handles = lhandles)
        ax.set_xticks(np.arange(len(labels)) + w * i / 2.0)
        ax.set_xticklabels(labels, rotation = 90)
        ax.set_xlabel('Identification levels')
        ax.set_ylabel('Number of features')
        cvs.draw()
        fig.tight_layout()
        cvs.print_figure(fname)
        fig.clf()
    
    def find_known_binders(self):
        classes = set(['PA', 'PC', 'PE', 'PG', 'PS', 'PI', 'SM', 'Cer'])
        having_binders = filter(
            lambda protein:
                len(self.bindprop[protein] & classes),
            self.valids.keys()
        )
        self.having_known_binders = having_binders
        self.known_binders_detected = set([])
        for protein in self.having_known_binders:
            d = self.valids[protein]
            for mode, tbl in iteritems(d):
                is_known_binder = []
                for i, oi in enumerate(tbl['i']):
                    if oi not in tbl['ms2i2']:
                        is_known_binder.append(False)
                    else:
                        binders = \
                            set(
                                map(
                                    lambda _hg:
                                        _hg[0],
                                    filter(
                                        lambda _hg:
                                            len(filter(
                                                lambda ms2is:
                                                    ms2is['score'] > 0,
                                                _hg[1]
                                            )),
                                        iteritems(tbl['ms2i2'][oi])
                                    )
                                )
                            )
                        is_known_binder.append(bool(len(binders & self.bindprop[protein])))
                tbl['known_binder'] = np.array(is_known_binder)
                if sum(is_known_binder) > 0:
                    self.known_binders_detected.add(protein)
    
    def original_id(self, protein, mode, mz):
        """
        Looks up an m/z and returns its original (stable) id.
        """
        
        self.sort_alll('mz')
        tbl = self.valids[protein][mode]
        ui = tbl['mz'].searchsorted(mz)
        i = (
            ui if ui == 0 or
            tbl['mz'][ui] - mz < mz - tbl['mz'][ui - 1]
            else ui - 1
        )
        if abs(tbl['mz'][i] - mz) > self.ms1_tlr:
            return None
        else:
            return tbl['i'][i]
    
    def print_ms2_identifications(self, protein, mode, mz):
        """
        Prints all MS2 identifications for an m/z.
        """
        
        oi = self.original_id(protein, mode, mz)
        
        if oi is None:
            return None
        
        tbl = self.valids[protein][mode]
        ms2f = tbl['ms2f'][oi]
        
        for sid, sc in iteritems(ms2f._scans):
            
            sc.print_identities()
    
    def mz_report(self, protein, mode, mz):
        """
        Looks up an m/z value and prints a report about the closest one.
        """
        
        self.sort_alll('mz')
        tbl = self.valids[protein][mode]
        ui = tbl['mz'].searchsorted(mz)
        i = (
            ui if ui == 0 or
            tbl['mz'][ui] - mz < mz - tbl['mz'][ui - 1]
            else ui - 1
        )
        oi = tbl['i'][i]
        ifracs = sorted(iteritems(self.fraction_indices(protein)),
                        key = lambda i: i[1][0])
        
        sys.stdout.write(
            
            '\n\t:: Looking up %.08f\n'\
            '\t %s8\n'\
            '\t -- The closest m/z found is %.08f\n'\
            '\t -- Current index is %u\n'\
            '\t -- Original index is %u\n'\
            '\t -- MS1 headgroups identified: %s\n'\
            '\t -- MS2 headgroups identified: %s\n'\
            '\t -- MS1 fatty acids identified: %s\n'\
            '\t -- MS2 fatty acids identified: %s\n'\
            '\t -- Headgroups based on fatty acids: %s\n'\
            '\t -- Combined identity: %s\n'\
            '\t %s8\n'\
            '\t -- RT range: %.02f--%.02f\n'\
            '\t -- Peak size: %.02f\n'\
            '\t -- Avg. area: %u\n'\
            '\t -- Protein containing fractions intensities:\n'\
            '\t\t%s\n'\
            '\t\t%s\n'\
            '\t -- Control fractions intensities:\n'\
            '\t\t%s\n'\
            '\t\t%s\n'\
            '\n' % (
             mz,
             '=' * 70,
             tbl['mz'][i],
             i,
             oi,
            ', '.join(sorted(list(tbl['ms1hg'][oi]))) \
                if 'ms1hg' in tbl else '',
            ', '.join(sorted(list(tbl['ms2hg'][oi]))) \
                if 'ms2hg' in tbl and \
                    oi in tbl['ms2hg'] and \
                    type(tbl['ms2hg'][oi]) is set else '',
            ', '.join(map(lambda hgfa: 
                    '%s: %s' % (hgfa[0], ', '.join(list(hgfa[1]))), 
                    iteritems(tbl['ms1fa'][oi])
                )
            ) if 'ms1fa' in tbl else '',
            ', '.join(list(tbl['ms2fa'][oi])) \
                if 'ms2fa' in tbl and oi in tbl['ms2fa'] else '',
            ', '.join(list(tbl['hgfa'][oi])) \
                if 'hgfa' in tbl else '',
            ', '.join(list(tbl['combined_hg'][oi])) \
                if 'combined_hg' in tbl else '',
            '=' * 70,
            tbl['rt'][i][0],
            tbl['rt'][i][1],
            tbl['peaksize'][i],
            tbl['aa'][i],
            '\t'.join(map(lambda fr: fr[0].rjust(7), filter(lambda fr: fr[1][1],
                                                 ifracs))),
            '\t'.join(map(lambda fr: ('%u' % int(tbl['fe'][i,fr[1][0]])).rjust(7),
                          filter(lambda fr: fr[1][1], ifracs))),
            '\t'.join(map(lambda fr: fr[0].rjust(7), filter(lambda fr: not fr[1][1],
                                                 ifracs))),
            '\t'.join(map(lambda fr: ('%u' % int(tbl['fe'][i,fr[1][0]])).rjust(7),
                          filter(lambda fr: not fr[1][1], ifracs))),
            )
        )
        sys.stdout.flush()
    
    def protein_name_upper2mixed(self, protein):
        protein_original = {}
        for protein_name in self.data.keys():
            if protein_name.upper() == protein:
                for mode in ['neg', 'pos']:
                    if mode in self.data[protein_name]:
                        protein_original[mode] = protein_name
        return protein_original
    
    def mz_report_raw(self, protein, mode, mz):
        protein = self.protein_name_upper2mixed(protein)
        protein = protein[mode]
        tbl = self.data[protein][mode]
        ui = tbl['ann'][:,2].searchsorted(mz)
        du = 999.0 if ui == tbl['ann'].shape[0] else tbl['ann'][ui,2] - mz
        dl = 999.0 if ui == 0 else mz - tbl['ann'][ui - 1,2]
        i = ui if du < dl else ui - 1
        sys.stdout.write('\n\t:: Looking up %.08f\n'\
            '\t -- The closest m/z found is %.08f\n'\
            '\t -- Original index is %u\n'\
            '\t -- Retention time range is %.02f--%.02f\n'\
            '\t -- Protein containing fractions intensities:\n'\
            '\t\t%s\n'\
            '\t -- Control fractions intensities:\n'\
            '\t\t%s\n'\
            '\t -- Peak size:\t%.02f\t(%s)\n'\
            '\t -- Avg. area:\t%u\t(%s)\n'\
            '\t -- Quality:\t%.02f\t(%s)\n'\
            '\t -- Charge:\t%u\t(%s)\n\n' % \
            (
                mz,
                tbl['ann'][i,2],
                i,
                tbl['ann'][i,3],
                tbl['ann'][i,4],
                str(list(tbl['lip'][i,:])),
                str(list(tbl['ctr'][i,:])),
                tbl['peaksize'][i],
                tbl['pks'][i],
                tbl['aa'][i],
                int(tbl['are'][i]),
                tbl['ann'][i,0],
                tbl['qly'][i],
                tbl['ann'][i,5],
                tbl['crg'][i]
            )
        )
        sys.stdout.flush()
    
    def _database_details_list(self, lip):
        if lip is None: lip = []
        return '\n'.join(map(lambda l:
            '⚫ %s\t%.04f\t%s' % (l[4], l[5], l[2]),
            lip
        ))
    
    def _fragment_details_list(self, ms2r, ms2files, path):
        fractions = {9: 'A09', 10: 'A10', 11: 'A11', 12: 'A12', 1: 'B01'}
        if ms2r is not None:
            try:
                sort = ms2r[:,3].argsort()[::-1]
            except:
                print(ms2r.shape)
        else:
            sort = []
        return '\n'.join(map(lambda i:
            '⚫ %s (%.02f)\n    @ file: %s\n    @ scan no. %u' % \
                (ms2r[i,0], ms2r[i,3],
                    ms2files[fractions[int(ms2r[i,4])]].replace(path, '')[1:],
                    int(ms2r[i,5])),
            filter(lambda i:
                ms2r[i,0] != 'unknown',
                sort
            )
        ))
    
    def _features_table_row(self, protein, mod, tbl, oi, i, fits_profile, drift):
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        tableccell = '\t\t\t<td class="%s" title="%s"'\
            ' onclick="showTooltip(event);">'\
            '\n\t\t\t\t%s\n\t\t\t</td>\n'
        row = []
        original_mz = tbl['mz'][i] / drift
        # m/z
        marco = 'marco' in tbl and tbl['marco'][i]
        denes = (tbl['aaa'][i] > self.aaa_threshold[mod] or (
                    oi in tbl['ms1hg'] and oi in tbl['ms2hg2'] and \
                    len(tbl['ms1hg'][oi] & tbl['ms2hg2'][oi])
                )) and tbl['prs'][i] <= 1.0 and tbl['peaksize'][i] >= 5.0
        row.append(tablecell % \
            (   'both' \
                if marco and denes \
                else 'positive' if denes \
                else 'negative' if marco \
                else 'nothing',
                'm/z measured; %s, %s mode; raw: %.07f, '\
                'recalibrated: %.07f; original index: %u; %s%s %s' % \
                (
                    protein,
                    'negative' if mod == 'neg' else 'positive',
                    original_mz,
                    tbl['mz'][i],
                    oi,
                    'This feature has %s intensity (%u) than'\
                        ' the threshold in %s mode (%u), ' % \
                        (
                            'larger' if tbl['aaa'][i] > \
                                self.aaa_threshold[mod] else 'lower',
                            tbl['aaa'][i],
                            'positive' if mod == 'pos' else 'negative',
                            self.aaa_threshold[mod]
                        ),
                    '%s %s MS2 confirmed identification.' % \
                        ('but' if tbl['aaa'][i] <= self.aaa_threshold[mod] \
                            else 'and',
                        'has' \
                            if oi in tbl['ms1hg'] and oi in tbl['ms2hg2'] and \
                                len(tbl['ms1hg'][oi] & tbl['ms2hg2'][oi]) \
                            else 'does not have'
                        ),
                    '%sn Marco\'s standards.' % \
                        ('I' if marco else 'Not i')
                ),
                '%.04f (%.04f)' % (tbl['mz'][i], original_mz)))
        # RT
        row.append(tablecell % (
                'nothing',
                'Retention time range (mean): %.02f-%.02f (%.02f)' % \
                    (tbl['rt'][i,0], tbl['rt'][i,1], np.mean(tbl['rt'][i,])),
                '%.02f' % np.mean(tbl['rt'][i,])
            )
        )
        # intensity
        row.append(tablecell % (
                'nothing' \
                    if tbl['aaa'][i] < self.aaa_threshold[mod] \
                    else 'positive',
                'Average intensity: %u' % int(tbl['aaa'][i]),
                '%u' % int(tbl['aaa'][i] / 1000.0)
            )
        )
        row.append(tableccell % (
            ('nothing clickable',
                self._database_details_list(tbl['lip'][oi]),
                'see DB records') \
                    if tbl['lip'][oi] is not None and \
                        len(tbl['lip'][oi]) > 0 \
                    else \
            ('nothing',
                'Lookup of this m/z resulted no records',
                'No DB records')
        ))
        row.append(tablecell % \
            ('nothing', 'Possible headgroups based on database records',
            ', '.join(tbl['ms1hg'][oi])))
        row.append(tablecell % \
            (('nothing ms2cell" ms2spectra="%s' % tbl['ms2f'][oi].html_table_b64(),
                    'Click to open MS2 spectra in pop-up viewer',
                    'see MS2 frags.') \
                if oi in tbl['ms2r'] else \
            ('nothing', 'No MS2 results for this feature', 'No MS2'))
        )
        row.append(tablecell % \
            ('nothing', 'Possible MS2 headgroups based on fragmets lookup',
            ', '.join(list(tbl['ms2hg2'][oi])) \
                if oi in tbl['ms2hg2'] and tbl['ms2hg2'][oi] is not None else '')
        )
        row.append(tablecell % \
            ('nothing', 'MS2 headroups and fatty acids identified based on '\
                'MS1 and MS2 new rule sets.',
            '' if oi not in self.valids[protein][mod]['ms2f'] \
                else str(self.valids[protein][mod]['ms2f'][oi])
        ))
        row.append(tablecell % \
            ('nothing',
            'Possible fatty acids based on database records',
            '; '.join(map(lambda hgfa:
                '%s: %s' % (hgfa[0], ', '.join(list(hgfa[1]))),
                iteritems(tbl['ms1fa'][oi])
            )) if len(tbl['ms1fa'][oi]) else '')
        )
        row.append(tablecell % \
            ('nothing',
            'Fatty acids identified in MS2',
            ', '.join(list(tbl['ms2fa'][oi])) \
                if oi in tbl['ms2fa'] and len(tbl['ms2fa'][oi]) else '')
        )
        row.append(tablecell % \
            ('nothing',
            'Combined identity (MS1, MS2, fatty acids)',
            ', '.join(list(tbl['combined_hg'][oi])) \
                if len(tbl['combined_hg'][oi]) else '')
        )
        peaks = [
            'Peaksize = %.02f' % tbl['peaksize'][i],
            'Peaksize > 2.0--5.0: %s' % (tbl['peaksize'][i] > tbl['pslim05'][i]),
            'Peaksize > 2.0--10.0: %s' % (tbl['peaksize'][i] > tbl['pslim10'][i]),
            'Peaksize > 5.0--10.0: %s' % (tbl['peaksize'][i] > tbl['pslim510'][i])
        ]
        row.append(tablecell % \
            ('nothing clickable',
                '\n'.join(peaks),
                '!' if (
                    not (tbl['peaksize'][i] > tbl['pslim05'][i]) or \
                    not (tbl['peaksize'][i] > tbl['pslim10'][i]) or \
                    not (tbl['peaksize'][i] > tbl['pslim510'][i])
                ) else ''
            )
        )
        if tbl['prr'] is None:
            row.append(tablecell % (
                'nothing',
                'Only one protein containing fraction, '\
                    'could not calculate peak ratio.',
                '--'
            ))
        else:
            row.append(tablecell % (
                'positive' if tbl[fits_profile][i] else 'nothing',
                '%s protein profile. Peak size ratio score: %.04f. '\
                'Peak ratio%s in expected range.' % \
                    ('Fits' if tbl[fits_profile][i] else 'Does not fit',
                        tbl['prs'][i], '' if tbl['prr'][i] else ' not'),
                'Yes%s' % ('*' if tbl['prr'][i] else '') \
                    if tbl[fits_profile][i] \
                    else 'No%s' % ('*' if tbl['prr'][i] else '')
            ))
        return row

    def features_table(self, filename = None,
        fits_profile = 'prs1'):
        
        prg = progress.Progress(len(self.valids),
                                'Generating HTML tables', 1, percent = False)
        
        hdr = ['+m/z', 'RT', 'Int', '+Database', '+MS1 HGs',
            '+MS2 frags', '+MS2 HGs', '+ID',
            '+MS1 FAs', '+MS2 FAs',
            '+MS1&2',
            '+Psize',
            '+Fits protein',
            '-m/z', 'RT', 'Int', '-Database', '-MS1 HGs',
            '-MS2 frags', '-MS2 HGs', '-ID',
            '-MS1 FAs', '-MS2 FAs',
            '-MS1&2',
            '-Psize',
            '-Fits protein']
        
        filename = 'results_%s_identities_details' % self.today() \
            if filename is None \
            else filename
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        
        # navigation html
        if not os.path.isdir(filename):
            os.mkdir(filename)
        navigation = '%s.html' % filename
        title = 'Identification details of all features'
        navhtml = ''
        navcnum = 7
        navrnum = len(self.valids) / navcnum + 1
        ltps = sorted(self.valids.keys())
        for i in xrange(navrnum):
            thisRow = []
            for j in xrange(navcnum):
                thisRow.append(tablecell % \
                    (('nothing',
                    'See identifications for all features of %s' % \
                        (proteins[j*navrnum + i]),
                    '<a href="%s/%s_%s.html">%s</a>' % \
                        (filename, filename, 
                        proteins[j*navrnum + i], proteins[j*navrnum + i])
                    ) if j*navrnum + i < len(proteins) else ('nothing', '', ''))
                )
            navhtml += tablerow % '\n'.join(thisRow)
        with open(navigation, 'w') as f:
            f.write(self.html_table_template % (title, title, navhtml))
        
        for protein, d in iteritems(self.valids):
            prg.step()
            title = '%s: identities for all features, detailed' % protein
            thisFilename = '%s/%s_%s.html' % (filename, filename, protein)
            table = ''
            hrow = ''
            for coln in hdr:
                hrow += tablehcell % coln
            hrow = tablerow % hrow
            table = hrow
            visited = {'pos': set([]), 'neg': set([])}
            self.sort_alll('aaa', asc = False)
            for mod, tbl in iteritems(d):
                opp_mod = 'neg' if mod == 'pos' else 'pos'
                drift = 1.0 if not self.proteins_drifts \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                        np.array(self.proteins_drifts[protein][mod].values())
                    ))
                opp_drift = 1.0 if not self.proteins_drifts \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                        np.array(self.proteins_drifts[protein][opp_mod].values())
                    ))
                for i, oi in enumerate(tbl['i']):
                    if oi not in visited[mod]:
                        thisRow = {'pos': [], 'neg': []}
                        visited[mod].add(oi)
                        thisRow[mod].append(
                            self._features_table_row(protein, mod, tbl,
                                oi, i, fits_profile, drift))
                        try:
                            if len(tbl[opp_mod][oi]) == 0:
                                thisRow[opp_mod] = [map(lambda i:
                                    tablecell % ('nothing', '', ''),
                                    xrange(len(hdr) / 2)
                                )]
                            else:
                                for opp_oi in tbl[opp_mod][oi].keys():
                                    opp_row = []
                                    visited[opp_mod].add(opp_oi)
                                    opp_i = np.where(d[opp_mod]['i'] == \
                                        opp_oi)[0][0]
                                    thisRow[opp_mod].append(
                                        self._features_table_row(
                                            protein, opp_mod, d[opp_mod],
                                            opp_oi, opp_i, fits_profile,
                                            opp_drift
                                        )
                                    )
                        except KeyError:
                            print(protein, mod, opp_mod)
                        for pos_row in thisRow['pos']:
                            for neg_row in thisRow['neg']:
                                #try:
                                table += tablerow % ('\n%s\n%s\n' % \
                                    ('\n'.join(pos_row), '\n'.join(neg_row)))
                                #except TypeError:
                                    #print(pos_row)
                                    #print(neg_row)
            with open(thisFilename, 'w') as f:
                f.write(self.html_table_template % (title, title, table))
        prg.terminate()

    def combined_table(self, valids, filename = 'headgroups_combined.html',
        include = 'cl70pct', identity = 'combined_hg'):
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        tableccell = '\t\t\t<td class="%s" title="%s"'\
            ' onclick="showTooltip(event);">'\
            '\n\t\t\t\t%s\n\t\t\t</td>\n'
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        title = 'Headgroups, combined identification'
        sort_alll(valids, 'mz')
        all_hgs = set([])
        for protein, d in iteritems(valids):
            for mod, tbl in iteritems(d):
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        all_hgs = all_hgs | tbl[identity][oi]
        all_hgs = sorted(list(all_hgs))
        data = dict(map(lambda protein:
                (protein, dict(map(lambda hg:
                    (hg, []), 
                    all_hgs
                ))),
                valids.keys()
            ))
        for protein, d in iteritems(valids):
            for mod, tbl in iteritems(d):
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        for hg in tbl[identity][oi]:
                            adds = ';'.join(uniqList(\
                                list(tbl['lip'][oi][\
                                    np.where(tbl['lip'][oi][:,7] == hg)[0], 4\
                                ])\
                            ))
                            fa = ','.join(list(tbl['combined_fa'][oi][hg])) \
                                if hg in tbl['combined_fa'][oi] else \
                                tbl['ms2fas'][oi] if oi in tbl['ms2fas'] else \
                                ','.join(list(tbl['ms1fa'][oi][hg]))
                            data[protein][hg].append(
                                ['%.04f' % tbl['mz'][i], adds, fa]
                            )
        hdr = map(lambda hg:
            tablehcell % hg,
            all_hgs
        )
        table = tablerow % '\n'.join([tablehcell % ('')] + hdr)
        for protein in sorted(valids.keys()):
            thisRow = [tablecell % ('rowname', '', protein)]
            for hg in all_hgs:
                c = data[protein][hg]
                if not c.__len__():
                    thisRow.append(tablecell % ('nothing', '', ''))
                else:
                    thisRow.append(tableccell % (
                        'positive clickable' \
                            if 'neg' not in map(lambda i: i[1], c) else \
                        'negative clickable' \
                            if 'pos' not in map(lambda i: i[1], c) else \
                        'both clickable',
                        '%s\n%s' % (hg, '\n'.join(map(lambda l:
                            '⚫ %s' % '     '.join(l),
                            c
                        ))),
                        ''
                    ))
            table += tablerow % '\n'.join(thisRow)
        with open(filename, 'w') as f:
            f.write(html_table_template % (title, title, table))
    
    def i2oi(self, protein, mode, i):
        return self.valids[protein][mode]['i'][i] \
            if (0 <= i < len(self.valids[protein][mode]['i'])) \
            else None
    
    def oi2i(self, protein, mode, oi):
        i = np.where(self.valids[protein][mode]['i'] == oi)[0]
        return i[0] if len(i) > 0 else None
    
    def ppratios_table(self):
        with open('protein_peak_ratios.txt', 'w') as f:
            f.write('%s\n' % '\t'.join([
                'Protein',
                'Fractions',
                'Lower',
                'Upper'
            ]))
            for protein, pr in iteritems(self.ppratios):
                for fracs, prs in iteritems(pr):
                    f.write('%s\n' % '\t'.join([
                        protein,
                        '%s:%s' % fracs,
                        '%.04f' % prs[0],
                        '%.04f' % prs[1]
                    ]))
    
    def pprofs_table(self, original = True):
        original = '_original' if original else ''
        renondigit = re.compile(r'[^\d]+')
        with open(self.pptable_file, 'w') as f:
            f.write('%s\n' % '\t'.join([
                'Protein',
                'Fraction',
                'Cc_%.03f' % self.fr_offsets[0],
                'Cc_%.03f' % self.fr_offsets[1]
            ]))
            for protein, prof in getattr(self, iteritems('pprofs%sL'%original)):
                for frac in sorted(prof.keys(), key = lambda x: (x[0], int(renondigit.sub('', x)))):
                    f.write('%s\n' % '\t'.join([
                        protein,
                        frac,
                        '%.04f' % getattr(self, 'pprofs%sL'%original)[protein][frac],
                        '%.04f' % getattr(self, 'pprofs%sU'%original)[protein][frac]
                    ]))
    
    def ms2identities_summary(self, string = False):
        result = \
            dict(
                map(
                    lambda protein:
                        (protein[0],
                        dict(
                            map(
                                lambda _tbl:
                                    (_tbl[0],
                                    filter(
                                        lambda i:
                                            i is not None,
                                        map(
                                            lambda f:
                                                None \
                                                if len(f.identities) == 0 \
                                                else reduce(
                                                    lambda ids:
                                                        ids[0] | ids[1],
                                                    f.identities
                                                ),
                                            _tbl[1]['ms2f'].values()
                                        ),
                                    )
                                    ),
                                iteritems(protein[1])
                            )
                            )
                        ),
                    iteritems(self.valids)
                )
            )
        
        if not string:
            return result
        
        result_str = \
            ''.join(
                map(
                    lambda protein:
                        'Protein: %s\n%s\n' % (
                            protein[0],
                            ''.join(
                                map(
                                    lambda res:
                                        '\tMode: %s\n\t   %s\n' % (
                                            res[0],
                                            '\n\t   '.join(
                                                map(
                                                    lambda hg:
                                                        '%s (detected %u '\
                                                            'times)' % \
                                                                (hg,
                                                                collections.\
                                                                Counter(res[1])\
                                                                    [hg]
                                                                ),
                                                    uniqList(res[1])
                                                ),
                                            )
                                        ),
                                    iteritems(protein[1])
                                )
                            )
                        ),
                    iteritems(result)
                )
            )
        return result_str
    
    def bindprop_latex_table(self, only_with_ms_data = False, outfile = 'binding_properties_table.tex'):
        allhgs = \
        sorted(
            list(
                reduce(
                    lambda llips:
                        llips[0] | llips[1],
                    map(
                        lambda lips:
                            lips[1],
                        filter(
                            lambda lips:
                                not only_with_ms_data or lips[0] in self.fractions_upper,
                            iteritems(self.bindprop)
                        )
                    )
                )
            )
        )
        tbl = r"""\begin{tabular}{l%s}
            & %s \\
            %s \\
        \end{tabular}
        """
        colalign = 'l' * len(allhgs)
        rownames = \
        sorted(
            filter(
                lambda protein:
                    not only_with_ms_data or protein in self.fractions_upper,
                self.bindprop.keys()
            )
        )
        hdr = ' & '.join(
            map(
                lambda lip:
                    '\\rot{%s}' % lip,
                allhgs
            )
        )
        cells = \
        '\\\\\n'.join(
            map(
                lambda protein:
                    '%s & %s' % (
                        protein,
                        ' & '.join(
                            map(
                                lambda lip:
                                    r'\cellcolor{emblgreen}' \
                                    if lip in self.bindprop[protein] \
                                    else '',
                                allhgs
                            )
                        )
                    ),
                rownames
            )
        )
        with open(outfile, 'w') as f:
            f.write(tbl % (colalign, hdr, cells))
    
    def hg_fa_list(self, filename = 'headgroups_fattya.tex', include = 'cl70pct'):
        out = '\\begin{itemize}'
        allhgs = set([])
        for protein, d in iteritems(self.valids):
            for mode, tbl in iteritems(d):
                for oi, ms2results in iteritems(tbl['ms2i2']):
                    for hg, ms2result in iteritems(ms2results):
                        if sum(map(lambda i: i['score'], ms2result)) > 0:
                            allhgs.add(hg)
        allhgs = sorted(list(allhgs))
        for protein in sorted(self.valids.keys()):
            d = self.valids[protein]
            thisProtein = '\n  \\item %s' % protein
            rows = []
            for hg in allhgs:
                thisRow = set([])
                for mode, tbl in iteritems(d):
                    for i, oi in enumerate(tbl['i']):
                        if oi in tbl['ms2i2']:
                            for res in tbl['ms2i2'][oi][hg]:
                                if res['score'] > 0:
                                    thisRow = thisRow | res['fattya']
                                    if not len(res['fattya']) and hg in tbl['ms1fa'][oi]:
                                        thisRow = thisRow | tbl['ms1fa'][oi][hg]
                if len(thisRow):
                    thisRow = '    \\item %s: %s' % (hg, ', '.join(sorted(list(thisRow))))
                    rows.append(thisRow)
            if len(rows):
                thisProtein += '\n  \\begin{itemize}\n%s\n  \\end{itemize}' % '\n'.join(rows)
            out += thisProtein
        out += '\n\\end{itemize}'
        with open(filename, 'w') as f:
            f.write(out)
    
    @staticmethod
    def colw_scale(w):
        """
        Scales column widths for xlsxwriter output.
        Below are empirical values to adjust the column widths.
        """
        return (w - 0.1705710377055909) / 0.24899187310525259
    
    def add_sheet(self, xls, tbl, name, colws = None):
        """
        Converts a table to xls worksheet, with formatting.
        
        :param xls: xlsxwriter workbook
        :param list tbl: table as list of lists
        :param str name: worksheet name
        :param list colws: column widths, list of floats
        """
        sheet  = xls.add_worksheet(name)
        plain  = xls.add_format({})
        bold   = xls.add_format({'bold': True,
                                #'rotation': 90
                                })
        green  = xls.add_format({'bg_color':   '#A9C98B'})
        violet = xls.add_format({'font_color': '#CA00FA'})
        red    = xls.add_format({'bg_color':   '#FFD1D7',
                                 'font_color': '#FA0021'})
        #sheet.set_row(row = 0, height = 115.0)
        sheet.freeze_panes(1, 0)
        for i, content in enumerate(tbl[0]):
            sheet.write(0, i, content, bold)
        
        for j, row in enumerate(tbl[1:]):
            for i, content in enumerate(row):
                if type(content) is tuple:
                    if content[1] in locals():
                        style = locals()[content[1]]
                    content = content[0]
                else:
                    style = plain
                try:
                    sheet.write(j + 1, i, content, style)
                except:
                    traceback.print_exc()
                    sys.stdout.write('\n\t:: Could not write row to xlsx:\n')
                    sys.stdout.write('\t\t%s\n' % row)
        if colws is not None:
            for coln, colw in enumerate(colws):
                sheet.set_column(coln, coln, self.colw_scale(colw))
        #sheet.set_row(row = 0, height = 115.0)
    
    def good_features(self, check_deltart = False, proteins = None):
        """
        Creates a boolean vector expressing whether the features satisfy all
        critera to be classified as good.
        """
        
        for protein, d in iteritems(self.valids):
            
            if proteins is not None and protein not in proteins:
                continue
            
            for mode, tbl in iteritems(d):
                
                good = []
                
                for i, oi in enumerate(tbl['i']):
                    
                    aaa = (
                        tbl['aa'][i]
                        if self.use_original_average_area
                        else tbl['aaa'][i]
                    )
                    
                    if self.slope_profile_selection:
                        profile_condition = tbl['slobb'][i]
                    else:
                        profile_condition = (
                            tbl['prr'] is None or
                            tbl['prr'][i]
                        )
                    
                    _good = (
                        tbl['peaksize'][i] >= 5.0 and
                        profile_condition and (
                            aaa >= self.aa_threshold[mode] or (
                                not np.isnan(tbl['ms2rt'][i]) and
                                tbl['deltart'][i]  <= 1.0
                            ) or
                            not check_deltart
                        ) and
                        not tbl['na'][i]
                    )
                    
                    good.append(_good)
                
                tbl['good'] = np.array(good)
    
    def std_layout_tables_xlsx(self, check_deltart = False, one_table = False,
                               proteins = None, silent = False):
        
        def make_xls(protein, xls):
            method = self.std_layout_table_all if protein == 'all' else \
                     self.std_layout_table
            _argsp = ['pos'] if protein == 'all' else [protein, 'pos']
            _argsn = ['neg'] if protein == 'all' else [protein, 'neg']
            _kwargs = {'colws': True, 'check_deltart': check_deltart}
            tbl, colw = method(*_argsp, **_kwargs)
            self.add_sheet(xls, tbl, '%s_positive' % protein, colw)
            tbl, colw = method(*_argsn, **_kwargs)
            self.add_sheet(xls, tbl, '%s_negative' % protein, colw)
            _kwargs['only_best'] = True
            tbl, colw = method(*_argsp, **_kwargs)
            self.add_sheet(xls, tbl, '%s_positive_best' % protein, colw)
            tbl, colw = method(*_argsn, **_kwargs)
            self.add_sheet(xls, tbl, '%s_negative_best' % protein, colw)
        
        if proteins is not None and len(proteins) == 1:
            silent = True
        
        xlsdir = 'top_features'
        if not os.path.exists(xlsdir):
            os.mkdir(xlsdir)
        
        if proteins is None:
            for fname in os.listdir(xlsdir):
                os.remove(os.path.join(xlsdir, fname))
        
        if self.use_original_average_area:
            self.sort_alll('aa', asc = False)
        else:
            self.sort_alll('aaa', asc = False)
        
        if one_table:
            
            sys.stdout.write('\t:: Exporting xlsx table...')
            sys.stdout.flush()
            
            xlsname = os.path.join(xlsdir, 'all_top_features.xlsx')
            
            if os.path.exists(xlsname):
                os.remove(xlsname)
            
            xls = xlsxwriter.Workbook(xlsname, {'constant_memory': True,
                                                'nan_inf_to_errors': True})
            make_xls('all', xls)
            xls.close()
            
            sys.stdout.write(' Ready.\n')
            sys.stdout.flush()
            
        else:
            
            if not silent:
                prg = progress.Progress(len(self.valids),
                                        'Exporting xlsx tables',
                                        1, percent = False)
            
            for protein in self.valids.keys():
                
                if not silent:
                    prg.step()
                
                if proteins is not None and protein not in proteins:
                    continue
                
                xlsname = os.path.join(xlsdir,
                                       '%s_top_features.xlsx' % protein)
                xls = xlsxwriter.Workbook(xlsname, {'constant_memory': True,
                                                    'nan_inf_to_errors': True})
                make_xls(protein, xls)
                xls.close()
            
            if not silent:
                prg.terminate()
    
    def std_layout_table_all(self, mode, **kwargs):
        
        allrows = []
        hdr = None
        
        for protein in self.valids.keys():
            rows = self.std_layout_table(protein, mode, **kwargs)
            
            if 'colws' in kwargs and kwargs['colws']:
                colw = rows[1]
                colw = [3.0] + colw
                rows = rows[0]
            
            rows = list(map(lambda r: [protein] + r, rows))
            
            if hdr is None:
                hdr = rows[0]
                hdr[0] = 'Protein'
                hdr = [hdr]
            
            rows = rows[1:]
            allrows.extend(rows)
        
        allrows.sort(key = lambda r: r[3])
        allrows = hdr + allrows
        
        if 'colws' in kwargs and kwargs['colws']:
            return allrows, colw
        else:
            return allrows
    
    def std_layout_table(self, protein, mode,
                         only_best = False, colws = False,
                         check_deltart = False):
        
        rows = []
        hdr = [
            ('Quality', 1.8),
            ('Significance', 1.8),
            ('m/z', 1.8),
            ('Has MS2', 1.8),
            ('RT range', 1.8),
            ('RT mean', 1.8),
            ('RT (MS2 closest)', 1.8),
            ('dRT', 1.8),
            ('Protein Ratio', 1.8),
            ('[M+H]+ Lipids' if mode == 'pos' else '[M-H]- Lipids', 6.4),
            ('[M+NH4]+ Lipids' if mode == 'pos' else '[M+HCOO]- Lipids', 6.4),
            ('[M+Na]+ Lipids' if mode == 'pos' else 'Nothing', 6.4),
            ('Avg. Area', 2.83),
            ('m/z corrected', 2.83),
            ('SwissLipids Identification', 10.77),
            ('Class', 2.83),
            ('Confirmed by MS2', 2.83),
            ('Comment', 2.83),
            ('MS2 precursor mass', 1.86),
            ('1 MS2 Ion Mass (intensity)', 4.29),
            ('1 MS2 Fragment (mass)', 8.60),
            ('2 MS2 Ion Mass (intensity)', 4.29),
            ('2 MS2 Fragment (mass)', 8.60),
            ('3 MS2 Ion Mass (intensity)', 4.29),
            ('3 MS2 Fragment (mass)', 8.60),
            ('4 MS2 Ion Mass (intensity)', 4.29),
            ('4 MS2 Fragment (mass)', 8.60),
            ('5 MS2 Ion Mass (intensity)', 4.29),
            ('5 MS2 Fragment (mass)', 8.60),
            ('z', 0.48),
            ('MS2 File', 2.12),
            ('Scan num.', 2.83),
            ('MS2 All Fragments', 4.29),
            ('Protein Ratio OK', 2.12),
            ('Lipid Intensity Ratio', 2.12),
            ('Protein Ratio Limit %.03f' % self.fr_offsets[0], 2.12),
            ('Protein Ratio Limit %.03f' % self.fr_offsets[-1], 2.12),
            ('Protein Ratio from Fractions', 2.12),
            ('Protein Ratio Score', 2.12),
            ('Peaksize', 2.12),
            ('MS1 Headgroups (automatic identification)', 1.48),
            ('MS2 Headgroups (automatic identification)', 1.48),
            ('Identity (automatically assigned)', 3.62),
            ('[M+H]+ Lipids (LipidMaps)' if mode == 'pos' \
                else '[M-H]- Lipids (LipidMaps)', 3.31),
            ('[M+NH4]+ Lipids (LipidMaps)' if mode == 'pos' \
                else '[M+HCOO]- Lipids (LipidMaps)', 3.31),
            ('[M+Na]+ Lipids (LipidMaps)' if mode == 'pos' \
                else 'Nothing', 3.31),
            ('Present in all protein containing fractions', 2.12),
            
        ]
        
        if self.slope_profile_selection:
            
            hdr.extend([
                ('Slope filter final', 2.12),
                ('Fractions 1a', 2.12),
                ('Fractions 2a', 2.12),
                ('Fractions 3a', 2.12),
                ('Filter 1a', 2.12),
                ('Filter 2a', 2.12),
                ('Filter 3a', 2.12),
                ('Filter a', 2.12),
                ('Fractions 1b', 2.12),
                ('Fractions 2b', 2.12),
                ('Fractions 3b', 2.12),
                ('Filter 1b', 2.12),
                ('Filter 2b', 2.12),
                ('Filter 3b', 2.12),
                ('Filter b', 2.12)
            ])
        
        colw = list(map(lambda f: f[1], hdr))
        
        rows.append(list(map(lambda f: f[0], hdr)))
        
        tbl = self.valids[protein][mode]
        drift = 1.0 if not self.proteins_drifts \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                        np.array(self.proteins_drifts[protein][mode].values())
                    ))
        
        def lipid_name(lip):
            names = lip[2].split('|')
            for name in names:
                if name is not None:
                    return name
        
        def get_lipids(lips, add, db = None):
            return '' if lips is None else \
                '; '.join(
                    uniqList(
                        map(
                            lambda r:
                                lipid_name(r) \
                                    if self.marco_lipnames_from_db else \
                                '%s(%s%u:%u)' % (r[7], 'O-' if r[9] else '',
                                               r[8][0], r[8][1]) \
                                    if r[8] is not None else '%s' % r[7],
                            filter(
                                lambda r:
                                    (
                                        (not self.marco_lipnames_from_db and \
                                        r[7] is not None)
                                            or \
                                        (self.marco_lipnames_from_db and \
                                        (db is None or r[0][0] == db))
                                    ) \
                                    and r[4] == add,
                                lips
                            )
                        )
                    )
                )
        
        min_ms2_mz = 70.0 if mode == 'neg' else 135.0
        
        for i, oi in enumerate(tbl['i']):
            
            ms2_best = None if oi not in tbl['ms2f'] or \
                    tbl['ms2f'][oi].best_scan is None else \
                    tbl['ms2f'][oi].best_scan
            
            ms2_rt = 'NA' \
                if ms2_best is None else \
                tbl['ms2f'][oi].scans[ms2_best][0,11]
            delta_rt = 'NA' \
                if ms2_best is None else \
                tbl['ms2f'][oi]._scans[ms2_best].deltart
            lips1 = get_lipids(tbl['lip'][oi], '[M+H]+', 'S') if mode == 'pos' \
                else get_lipids(tbl['lip'][oi], '[M-H]-', 'S')
            lips2 = get_lipids(tbl['lip'][oi], '[M+NH4]+', 'S') if mode == 'pos' \
                else get_lipids(tbl['lip'][oi], '[M+HCOO]-', 'S')
            lips3 = get_lipids(tbl['lip'][oi], '[M+Na]+', 'S') if mode == 'pos' \
                else ''
            
            lips1l = get_lipids(tbl['lip'][oi], '[M+H]+', 'L') if mode == 'pos' \
                else get_lipids(tbl['lip'][oi], '[M-H]-', 'L')
            lips2l = get_lipids(tbl['lip'][oi], '[M+NH4]+', 'L') if mode == 'pos' \
                else get_lipids(tbl['lip'][oi], '[M+HCOO]-', 'L')
            lips3l = get_lipids(tbl['lip'][oi], '[M+Na]+', 'L') if mode == 'pos' \
                else ''
            
            ms2_mz = 'NA' if ms2_best is None else \
                tbl['ms2f'][oi].scans[ms2_best][0,9]
            
            ms2i1, ms2f1 = \
                tbl['ms2f'][oi]._scans[ms2_best].get_by_rank(1, min_ms2_mz) \
                if ms2_best is not None else ('', '')
            ms2i2, ms2f2 = \
                tbl['ms2f'][oi]._scans[ms2_best].get_by_rank(2, min_ms2_mz) \
                if ms2_best is not None else ('', '')
            ms2i3, ms2f3 = \
                tbl['ms2f'][oi]._scans[ms2_best].get_by_rank(3, min_ms2_mz) \
                if ms2_best is not None else ('', '')
            ms2i4, ms2f4 = \
                tbl['ms2f'][oi]._scans[ms2_best].get_by_rank(4, min_ms2_mz) \
                if ms2_best is not None else ('', '')
            ms2i5, ms2f5 = \
                tbl['ms2f'][oi]._scans[ms2_best].get_by_rank(5, min_ms2_mz) \
                if ms2_best is not None else ('', '')
            
            ms2_file = '' if ms2_best is None else \
                tbl['ms2f'][oi]._scans[ms2_best].ms2_file
            
            ms2_scan = '' if ms2_best is None else \
                tbl['ms2f'][oi]._scans[ms2_best].scan_id[0]
            
            ms2_full = '' if ms2_best is None else \
                tbl['ms2f'][oi]._scans[ms2_best].full_list_str()
            
            mz_original = tbl['mz'][i] / drift
            
            ms2_style = (
                'green'
                if (
                    len(ms2_full) and
                    (
                        delta_rt == 'NA' or abs(delta_rt) <
                        self.deltart_threshold
                    )
                ) else
                'plain'
            )
            
            oi = tbl['i'][i]
            
            aaa = tbl['aa'][i] if self.use_original_average_area else tbl['aaa'][i]
            
            if self.slope_profile_selection:
                profile_condition = tbl['slobb'][i]
            else:
                profile_condition = tbl['prr'] is None or tbl['prr'][i]
            
            good = tbl['peaksize'][i] >= 5.0 and \
                profile_condition and \
                (aaa >= self.aa_threshold[mode] or \
                (oi in tbl['ms2f'] and len(tbl['ms2f'][oi].deltart) and \
                    (min(map(abs, tbl['ms2f'][oi].deltart.values())) <= 1.0 or \
                        not check_deltart))) and not tbl['na'][i]
            
            _good = tbl['peaksize'][i] >= 5.0 and \
                profile_condition and \
                ((aaa >= self.aa_threshold[mode] and \
                    (any(map(len, [lips1, lips2, lips3, lips1l, lips2l, lips3l])))) or \
                (oi in tbl['ms1hg'] and oi in tbl['ms2hg2'] and \
                    len(tbl['ms1hg'][oi] & tbl['ms2hg2'][oi]))) and not tbl['na'][i]
            
            if not only_best or good:
                
                this_row = [
                    tbl['qua'][i],
                    tbl['sig'][i],
                    (mz_original, 'green' if _good else 'plain'),
                    
                    ('yes' if len(ms2_full) else 'no', ms2_style),
                    '%.02f - %.02f' % (tbl['rt'][i][0], tbl['rt'][i][1]),
                    tbl['rtm'][i],
                    # '%u:%u' % (int(tbl['rtm'][i]) / 60, tbl['rtm'][i] % 60),
                    ms2_rt,
                    ((delta_rt, 'plain')
                        if delta_rt == 'NA' or
                        abs(delta_rt) < self.deltart_threshold else
                    (delta_rt, 'red')),
                    'Inf' if np.isinf(tbl['iprf'][i]) else \
                        'NA' if np.isnan(tbl['iprf'][i]) or \
                            tbl['iprf'] is None else \
                        tbl['iprf'][i],
                    (lips1, 'green' if len(lips1) else 'plain'),
                    (lips2, 'green' if len(lips2) else 'plain'),
                    (lips3, 'green' if len(lips3) else 'plain'),
                    (aaa,
                    'green' if aaa >= self.aa_threshold[mode] \
                        else 'plain'
                    ),
                    tbl['mz'][i],
                    (('%s: %s /// %s: %s%s' % (
                        '[M+H]+' if mode == 'pos' else '[M-H]-',
                        lips1 if len(lips1) else 'nothing',
                        '[M+NH4]+' if mode == 'pos' else '[M+HCOO]-',
                        lips2 if len(lips2) else 'nothing',
                        '' if mode == 'neg' else ' /// (%s: %s' % \
                            ('[M+Na]+', lips3) if len(lips3) else 'nothing')) \
                        if len(lips1) or len(lips2) or len(lips3) \
                            else 'unknown',
                    'green' if len(lips1) or len(lips2) or len(lips3) else 'plain'),
                    # 3 empty cols to be filled manually
                    '' if oi not in tbl['cid'] or not len(tbl['cid'][oi]) else \
                        (', '.join(sorted(set(tbl['cid'][oi]))), 'violet'), # Class
                    '', # Confirmed by MS2
                    (tbl['idlevel'][oi], 'violet'), # Comment
                    (ms2_mz, ms2_style),
                    (ms2i1, ms2_style),
                    (ms2f1, ms2_style),
                    (ms2i2, ms2_style),
                    (ms2f2, ms2_style),
                    (ms2i3, ms2_style),
                    (ms2f3, ms2_style),
                    (ms2i4, ms2_style),
                    (ms2f4, ms2_style),
                    (ms2i5, ms2_style),
                    (ms2f5, ms2_style),
                    tbl['z'][i],
                    (ms2_file.split('/')[-1], ms2_style),
                    (ms2_scan, ms2_style),
                    (ms2_full, ms2_style),
                    ('NA' if tbl['prr'] is None else \
                        'OK' if tbl['prr'][i] else 'NOT_OK',
                    'plain' if tbl['prr'] is None or not tbl['prr'][i] \
                        else 'green'
                    ),
                    'NA' if self.first_ratio[protein] is None \
                        or np.isnan(tbl['iprf'][i]) \
                        else 'Inf' if np.isinf(tbl['iprf'][i]) \
                        else tbl['iprf'][i],
                    'NA' if self.first_ratio[protein] is None else \
                        'Inf' if np.isinf(self.ppratios[protein][
                            self.first_ratio[protein]][0]) else \
                        self.ppratios[protein][self.first_ratio[protein]][0],
                    'NA' if self.first_ratio[protein] is None else \
                        'Inf' if np.isinf(self.ppratios[protein][
                            self.first_ratio[protein]][-1]) else \
                        self.ppratios[protein][self.first_ratio[protein]][-1],
                    'NA' if self.first_ratio[protein] is None \
                        else '%s:%s' % self.first_ratio[protein],
                    ('NA' if self.first_ratio[protein] is None \
                        or np.isnan(tbl['prs'][i]) \
                        else 'Inf' if np.isinf(tbl['prs'][i]) \
                        else tbl['prs'][i],
                    'plain' if self.first_ratio[protein] is None \
                        or tbl['prs'][i] > 1.0 \
                        else 'green'
                    ),
                    (tbl['peaksize'][i], 'green' \
                        if tbl['peaksize'][i] >= 5.0 else 'plain'),
                    '' if oi not in tbl['ms1hg'] or not len(tbl['ms1hg'][oi]) else \
                        (', '.join(sorted(list(tbl['ms1hg'][oi]))), 'green'),
                    '' if oi not in tbl['ms2hg2'] or not len(tbl['ms2hg2']) else \
                        (', '.join(sorted(list(tbl['ms2hg2'][oi]))), 'green'),
                    '' if oi not in tbl['cid'] or not len(tbl['cid'][oi]) else \
                        (', '.join(sorted(tbl['cid'][oi])), 'green'),
                    (lips1l, 'green' if len(lips1l) else 'plain'),
                    (lips2l, 'green' if len(lips2l) else 'plain'),
                    (lips3l, 'green' if len(lips3l) else 'plain'),
                    ('missing values' if tbl['na'][i] else 'no missing values',
                    'plain' if tbl['na'][i] else 'green')
                ]
                
                if self.slope_profile_selection:
                    
                    # 1st columnt is the final outcome
                    this_row.append(
                        ('OK'    if tbl['slobb'][i] else 'NOT_OK',
                         'green' if tbl['slobb'][i] else 'plain'),
                    )
                    
                    # now the 1 sometimes 2 ``highest`` i.e. peak
                    for hi, hfracs in enumerate(self.hpeak[protein]):
                        
                        hfracs = tuple(hfracs)
                        
                        # the fraction pairs in this peak
                        # 3 cols: the compared fractions
                        for fri in tbl['sloi%u' % hi]:
                            
                            if fri[0] != fri[1]:
                                this_row.append(
                                    ('%s:%s' % fri,
                                     (
                                        'green'
                                        if tbl['slo%u%s%s' % (
                                            hi, fri[0], fri[1]
                                        )][i]
                                        else 'plain'
                                     )
                                    )
                                )
                            else:
                                this_row.append(('None', 'plain'))
                        
                        # if we use only 2 ratios
                        if len(tbl['sloi%u' % hi]) == 2:
                            this_row.append(('None', 'plain'))
                        
                        # 3 cols: OK or NOT_OK
                        for fri in tbl['sloi%u' % hi]:
                            
                            if fri[0] != fri[1]:
                                this_row.append(
                                    (
                                        'OK'
                                        if tbl['slo%u%s%s' % (
                                            hi, fri[0], fri[1]
                                        )][i]
                                        else 'NOT_OK',
                                        'green'
                                        if tbl['slo%u%s%s' % (
                                            hi, fri[0], fri[1]
                                        )][i]
                                        else 'plain'
                                    )
                                )
                            else:
                                this_row.append(('None', 'plain'))
                        
                        # if we use only 2 ratios
                        if len(tbl['sloi%u' % hi]) == 2:
                            this_row.append(('None', 'plain'))
                        
                        # 1 col: this peak overall is OK or NOT_OK
                        this_row.append(
                            ('OK'    if tbl['slob%u' % hi][i] else 'NOT_OK',
                             'green' if tbl['slob%u' % hi][i] else 'plain')
                        )
                    
                    # if we have only 1 peak, fill the place of the 2nd
                    if len(hfracs) == 1:
                        this_row.extend([('None', 'plain')] * 7)
                
                rows.append(this_row)
                
        if colws:
            return rows, colw
        else:
            return rows
    
    def std_layout_table_stripped(self, protein, mode, only_best = False):
        tbl = self.std_layout_table(protein, mode, only_best = only_best)
        tbl = \
            list(
                map(
                    lambda l:
                        list(
                            map(
                                lambda c:
                                    c[0] if type(c) is tuple else c,
                                l
                            )
                        ),
                    tbl
                )
            )
        return tbl
    
    def marco_standards(self):
        result = {}
        files = filter(
            lambda fname:
                fname[-4:] == 'xlsx',
            os.listdir(self.marco_dir)
        )
        def tbl2mzs(tbl):
            return filter(
                lambda x:
                    type(x) is float,
                map(
                    lambda row:
                        self.to_float(row[2]),
                    tbl
                )
            )
        def mzs2ois(protein, mode, mzs):
            ois = set([])
            for mz in mzs:
                oi = self.mz2oi(protein, mode, mz)
                if oi is None:
                    sys.stdout.write('\t:: m/z not found: %.05f (%s, %s)\n' % \
                        (mz, protein, mode))
                    sys.stdout.flush()
                else:
                    ois.add(oi)
            return ois
        def ois2attr(protein, mode, ois):
            tbl = self.valids[protein][mode]
            tbl['marco'] = \
                np.array(
                    map(
                        lambda i:
                            i in ois,
                        tbl['i']
                    )
                )
        for fname in files:
            protein = fname.split('_')[0]
            result[protein] = {'pos': [], 'neg': []}
            pos_table = self.read_xls(
                os.path.join(self.marco_dir, fname), sheet = 0)
            neg_table = self.read_xls(
                os.path.join(self.marco_dir, fname), sheet = 1)
            pos_mzs = tbl2mzs(pos_table[1:])
            neg_mzs = tbl2mzs(neg_table[1:])
            pos_ois = mzs2ois(protein, 'pos', pos_mzs)
            neg_ois = mzs2ois(protein, 'neg', neg_mzs)
            ois2attr(protein, 'pos', pos_ois)
            ois2attr(protein, 'neg', neg_ois)
    
    def read_marco_standards(self):
        result = {}
        xlsfiles = filter(
            lambda fname:
                fname[-4:] == 'xlsx',
            os.listdir(self.marco_dir)
        )
        
        csvfiles = filter(
            lambda fname:
                fname[-3:] == 'csv',
            os.listdir(self.marco_dir)
        )
        
        for fname in xlsfiles:
            protein = fname.split('_')[0]
            mode = fname.split('_')[1]
            xls = openpyxl.load_workbook(os.path.join(self.marco_dir, fname),
                                         read_only = True)
            if protein not in result:
                result[protein] = {}
            if len(xls.worksheets) == 1:
                result[protein] = {mode: self.read_xls(os.path.join(self.marco_dir, fname), sheet = 0)}
            else:
                result[protein]['pos'] = self.read_xls(
                    os.path.join(self.marco_dir, fname), sheet = 0)
                result[protein]['neg'] = self.read_xls(
                    os.path.join(self.marco_dir, fname), sheet = 1)
        
        for fname in csvfiles:
            protein = fname.split('_')[0]
            mode = fname.split('_')[1]
            with open(os.path.join(self.marco_dir, fname), 'r') as f:
                tbl = list(map(lambda l: list(map(lambda c: c.replace('"', ''), l.strip().split(','))), f))
            if not tbl[0][0] == '' and not tbl[0][1] == '':
                if protein not in result:
                    result[protein] = {}
                result[protein][mode] = tbl
        
        self.marco_std_data = result
    
    def standards_crosscheck(self, logfile = 'crosscheck.log', only_best = True):
        def log(msg):
            logf.write('%s\n' % msg)
        
        def lookup_column(tbl, name):
            for i, colname in enumerate(tbl[0]):
                if colname.strip() == name:
                    return i
        
        def get_mmzs(tbl):
            mzcol = lookup_column(tbl, 'm.z')
            mzs = list(filter(lambda mz: mz[1] is not None,
                                map(lambda l: (l[0], self.to_float(l[1][mzcol])),
                            enumerate(tbl))))
            return np.array(list(map(list, sorted(mzs, key = lambda mz: mz[1]))))
        
        def get_dmzs(tbl):
            return np.array(list(map(list, sorted(
                    list(map(lambda l: (l[0], l[1][2]), enumerate(tbl))),
                    key = lambda mz: mz[1]
                ))))
        
        def mz_lookup(mz, lst):
            ui = lst[:,1].searchsorted(mz)
            du = None
            dl = None
            if ui < lst.shape[0]:
                if lst[ui,1] - mz < 0.01:
                    du = lst[ui,1] - mz
            if ui > 0:
                if mz - lst[ui - 1,1] < 0.01:
                    dl = mz - lst[ui - 1,1]
            if du is not None and dl is not None:
                if du < dl:
                    return ui
                else:
                    return ui - 1
            elif du is not None:
                return ui
            elif dl is not None:
                return ui - 1
            else:
                return None
        
        def compare(m, d, colname, dcol, name, fun,
                    mfun = lambda x: x, dfun = lambda x: x):
            mcol = lookup_column(mtbl, colname)
            if mcol is None:
                log('\t\t[ !! ] Column `%s` could not be found at Marco' % colname)
            else:
                print(m[mcol], d[dcol])
                mval = mfun(m[mcol])
                dval = dfun(d[dcol])
                print(mval, dval, type(mval), type(dval))
                ok = fun((mval, dval))
                if ok:
                    log('\t\t[ OK ] %s is the same at Marco & Denes' % name)
                else:
                    log('\t\t[ !! ] %s mismatch: %s at Marco, %s at Denes' % \
                        (name, str(mval), str(dval)))
        
        logf = open(logfile, 'w')
        
        for protein, d in iteritems(self.marco_std_data):
            for mode, mtbl in iteritems(d):
                log('===[ %s, %s mode ]=======================================8' % \
                    (protein, 'positive' if mode == 'pos' else 'negative'))
                dtbl = self.std_layout_table_stripped(protein, mode, only_best = only_best)[1:]
                
                reccol = lookup_column(mtbl, 'm.z_corrected')
                
                mmzs = get_mmzs(mtbl)
                dmzs = get_dmzs(dtbl)
                allmzs = []
                if len(mmzs):
                    allmzs.extend(list(mmzs[:,1]))
                if len(dmzs):
                    allmzs.extend(list(dmzs[:,1]))
                allmzs = sorted(list(set(allmzs)))
                for mz in allmzs:
                    di = mz_lookup(mz, dmzs) if len(dmzs) else None
                    mi = mz_lookup(mz, mmzs) if len(mmzs) else None
                    if mi is None:
                        log('\t[ !! ] m/z %.04f present only at Denes' % mz)
                        dii = dmzs[di][0]
                        # something to tell about features missing from Marco
                        continue
                    if di is None:
                        log('\t[ !! ] m/z %.04f present only at Marco' % mz)
                        continue
                    else:
                        log('\t[ OK ] m/z %.04f present at both Marco & Denes' % mz)
                    mii = int(mmzs[mi][0])
                    dii = int(dmzs[di][0])
                    compare(mtbl[mii], dtbl[dii], 'm.z_corrected', 13, 'recalibrated m/z',
                            lambda v: abs(v[0] - v[1]) < 0.005,
                            mfun = lambda x: self.to_float(x))
                    
                    compare(mtbl[mii], dtbl[dii], 'RT.mean', 5, 'RT mean',
                            lambda v: abs(v[0] - v[1]) < 0.1,
                            mfun = lambda x: self.to_float(x))
                    compare(mtbl[mii], dtbl[dii], 'protein_peak_ratio', 8,
                            'Lipid intensity ratio',
                            lambda v: abs(v[0] - v[1]) < 0.1,
                            mfun = lambda x: 1.0 if x == 'NA' else self.to_float(x),
                            dfun = lambda x: 1.0 if x == 'NA' else self.to_float(x))
                    compare(mtbl[mii], dtbl[dii], 'Avg..Area', 12, 'Average area',
                            lambda v: abs(v[0] - v[1]) < max(v[0], v[1]) * 0.1,
                            mfun = lambda x: self.to_float(x),
                            dfun = lambda x: self.to_float(x))
                    compare(mtbl[mii], dtbl[dii],
                            'lipid_.M.H.' if mode == 'pos' else 'lipid_.M.H..',
                            9, '[M+H]+ lipids' if mode == 'pos' else '[M-H]- lipids',
                            lambda v: len(v[0] ^ v[1]) == 0,
                            mfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';')))),
                            dfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';'))))
                        )
                    
                    compare(mtbl[mii], dtbl[dii],
                            'lipid_.M.NH4.' if mode == 'pos' else 'lipid_.M.COOH..',
                            10, '[M+NH4]+ lipids' if mode == 'pos' else '[M+HCOO]- lipids',
                            lambda v: len(v[0] ^ v[1]) == 0,
                            mfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';')))),
                            dfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';'))))
                        )
                    
                    if mode == 'pos':
                        compare(mtbl[mii], dtbl[dii],
                            'lipid_.M.Na.',
                            11, '[M+Na]+ lipids',
                            lambda v: len(v[0] ^ v[1]) == 0,
                            mfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';')))),
                            dfun = lambda x: set(list(map(lambda i: i.strip(), x.split(';'))))
                        )
                    
                    mms2 = self.to_float(mtbl[mii][lookup_column(mtbl, 'Scan')]) is not None
                    dms2 = type(dtbl[dii][31]) is int
                    
                    if mms2 and not dms2:
                        log('\t\t[ !! ] MS2 missing from Denes. Scan num at Marco is %s' % \
                            mtbl[mii][lookup_column(mtbl, 'Scan')])
                    if dms2 and not mms2:
                        log('\t\t[ !! ] MS2 missing from Marco. Scan num at Denes is %u' % \
                            dtbl[dii][31])
                    
                    if mms2 and dms2:
                        log('\t\t[ OK ] MS2 data both at Marco (scan %u) & Denes (scan %u)' % \
                            (int(float(mtbl[mii][lookup_column(mtbl, 'Scan')])), dtbl[dii][31]))
                        compare(mtbl[mii], dtbl[dii], 'Peptide.Mass', 18, 'MS2 precursor mass',
                                lambda v: abs(v[0] - v[1]) < 0.0005,
                                mfun = lambda x: self.to_float(x))
                        compare(mtbl[mii], dtbl[dii], 'Retention.Time..mins.secs.', 6,
                                'closest MS2 scan RT',
                                lambda v: abs(v[0] - v[1]) < 0.1,
                                mfun = lambda x: self.to_float(x))
                        compare(mtbl[mii], dtbl[dii], 'delta_RT', 7, 'delta RT',
                                lambda v: abs(v[0] - v[1]) < 0.1,
                                mfun = lambda x: self.to_float(x))
                        compare(mtbl[mii], dtbl[dii], 'Scan', 31, 'Best scan No.',
                                lambda v: v[0] == v[1],
                                mfun = lambda x: int(self.to_float(x)))
                        for i in xrange(5):
                            compare(mtbl[mii], dtbl[dii], 'MS2.Ion.%u.Mass.Intensity.' % (i + 1), 19 + i * 2,
                                    'MS2 ion #%u mass' % (i + 1),
                                    lambda v: abs(v[0] - v[1]) < 0.01,
                                    mfun = lambda x: self.to_float(x.split('(')[0]) if len(x) else 0.0,
                                    dfun = lambda x: self.to_float(x.split('(')[0]) if len(x) else 0.0)
                            compare(mtbl[mii], dtbl[dii], 'MS2.Ion.%u.Mass.Intensity.' % (i + 1), 19 + i * 2,
                                    'MS2 ion #%u intensity' % (i + 1),
                                    lambda v: abs(v[0] - v[1]) < 1,
                                    mfun = lambda x: self.to_int(x.split('(')[1][:-1]) if len(x) else 0,
                                    dfun = lambda x: self.to_int(x.split('(')[1][:-1]) if len(x) else 0)
        
        logf.close()
    
    #
    # Processing manually curated results
    #
    
    def read_manual(self):
        """
        Reads adequate columns from manually annotated tables.
        Provides data in `Screening().manual` attribute.
        """
        
        reclass = re.compile(r'(^[IV]*\.?[0-9]?).*')
        
        def read_line(l, protein, mode):
            if len(l[17]) and len(l[12]) and len(l[13]):
                return \
                    [
                        float(l[13]), # m/z corrected
                        reclass.match(l[17]).groups()[0], # result class
                        l[14], # SwissLipids name
                        l[15], # main headgroup class
                        int(float(l[12])), # intensity
                        float(l[2]), # m/z original
                        protein, # protein name
                        mode, # ion mode
                        float(l[5]), # RT mean
                        float(l[6]) if l[6] != 'NA' else np.nan, # RT MS2 closest
                        float(l[4].split('-')[0].strip()), # RT lower
                        float(l[4].split('-')[1].strip())  # RT greater
                    ]
        
        def read_table(tbl, protein, mode):
            return \
                list(
                    filter(
                        lambda l:
                            l is not None,
                        map(
                            lambda l:
                                read_line(l, protein, mode),
                            tbl
                        ),
                    )
                )
        
        data = {}
        
        fnames = \
            list(
                filter(
                    lambda f:
                        f.endswith(self.manualend),
                    os.listdir(self.manualdir)
                )
            )
        
        for f in fnames:
            protein = f.split('_')[0]
            xlsname = os.path.join(self.manualdir, f)
            tblneg = self.read_xls(xlsname,
                                   sheet = '%s_negative_best' % protein)
            tblpos = self.read_xls(xlsname,
                                   sheet = '%s_positive_best' % protein)
            data[protein] = {}
            data[protein]['neg'] = read_table(tblneg[1:], protein, 'neg')
            data[protein]['pos'] = read_table(tblpos[1:], protein, 'pos')
        
        self.manual = data
    
    def read_manual2(self, fname = 'All_results_v04.xlsx'):
        """
        Reads manually annotated results from Marco's final table.
        """
        result = {}
        reclass = re.compile(r'(^[IV]*\.?[0-9]?).*')
        
        tbl = self.read_xls(fname, sheet = 'final data')
        
        for l in tbl:
            
            protein, mode = l[1].split('_')
            mode = mode[:3]
            
            if protein not in result:
                result[protein] = {}
            
            if mode not in result[protein]:
                result[protein][mode] = []
            
            if l[2].strip()[:4] == 'Qual':
                continue
            
            result[protein][mode].append([
                float(l[15]), # m/z corrected
                reclass.match(l[19]).groups()[0], # result class
                l[16].replace(u'−', '-'), # SwissLipids name
                l[17].replace(u'−', '-'), # main headgroup class
                int(float(l[14])), # intensity
                float(l[4]), # m/z original
                protein, # protein name
                mode, # ion mode
                float(l[7]), # RT mean
                float(l[8]) if l[8] != 'NA' else np.nan, # RT MS2 closest
                float(l[6].split('-')[0].strip()), # RT lower
                float(l[6].split('-')[1].strip())  # RT greater
            ])
        
        self.manual = result
    
    def manual_df(self, screen = 'A', only_swl_col = False):
        """
        Creates a pandas dataframe from manual results.
        """
        
        shgs = {
            'Monoalkylmonoacylglycerol-O': 'DAG-O',
            'hydroquinone?': 'HQ',
            'Ganglioside GM3': 'GM3',
            'alpha-tocopherol metabolite': 'VE',
            'alpha-tocopherol': 'VE',
            r'Retinol {calculated as -H2O adduct '\
                r'is not in applied database}': 'VA',
            'docosapentaenoate': 'PUFA',
            'octacosatetraenoate': 'PUFA',
            'octacosapentaenoate': 'PUFA',
            'octadecatetraenoate': 'PUFA',
            'octatriacontatetraenoate': 'PUFA',
            'tetracosapentaenoate': 'PUFA',
            'hexacosatetraenoate': 'PUFA',
            'hexacosanoate': 'PUFA',
            'dotriacontapentaenoate': 'PUFA',
            'Sterol ester': 'SE',
            'nothing': 'NA',
            'unknown': 'NA',
            'Monoalkylglycerol-O': 'MAG-O',
            'Monoalkyldiacylglycerol-O': 'TAG-O',
            'Dihexosyldiacylglycerol': 'HexDAG',
            'Monohexosyldiacylglycerol': 'HexDAG',
            'Monoalkylmonoacylglycerol-O': 'DAG-O',
            'Monoalkyldiacylglycerol-O': 'TAG-O',
            'Monoalkylmonoacylglycerol': 'DAG-O',
            '24-Hydroxy-19-norgeminivitamin D3': 'VD',
            'NP40': 'P40',
            'Cer1P': 'CerP',
            'Phosphatidylcholine': 'PC',
            'Phosphatidylethanolamine': 'PE',
            'Phosphatidylcholine-O': 'PC-O',
            'Phosphatidylethanolamine-O': 'PE-O',
            'Dihexosyl ceramide': 'Hex2Cer',
            'Sulfodihexosyl ceramide': 'SHex2Cer',
            'BMP / PG': 'PG/BMP',
            'BMP/PG': 'PG/BMP',
            'Lyso-O-PE': 'LysoPE-O',
            'Lyso-O-PC': 'LysoPC-O',
            'Lyso-O-PG': 'LysoPG-O',
            'Lyso-O-PS': 'LysoPS-O',
            'LysoO-PE': 'LysoPE-O',
            'LysoO-PG': 'LysoPG-O',
            'LysoO-PC': 'LysoPC-O',
            'LysoO-PS': 'LysoPS-O',
            'Hex-Cer': 'HexCer'
        }
        
        shgs2 = {
            'Ganglioside': 'GM',
            'Vit.A1': 'VA',
            'Vit. E metabolite': 'VE',
            'SulfohexCer': 'SHexCer',
            'SulfoHexCer': 'SHexCer',
            'SulfodihexCer': 'SHex2Cer',
            'DiHexCer-OH': 'Hex2CerOH',
            'DiHexCer': 'Hex2Cer',
            'PI2xP': 'PIP2',
            'MAMAG': 'DAG-O'
        }
        
        uhgs = {
            'Hex2Cer': 'Hex2Cer',
            'Hex2Cer-OH': 'Hex2CerOH',
            'HexCer-OH': 'HexCerOH',
            'GM3': 'GM',
            'Detergent': 'P40',
            'Hex-Cer': 'HexCer'
        }
        
        def get_names(swl_names, manual_names):
            """
            Extracts the lipid names and carbon counts
            from the SwissLipids IDs field.
            """
            
            something   = []
            nothing     = []
            
            # fixing typos and inconsequent naming:
            manual_name_1 = manual_names.strip().split('(')[0]
            manual_name_1_l = manual_name_1.lower().strip()
            
            if (manual_name_1_l == 'ambiguous' or
                manual_name_1_l == 'ambigous'):
                
                manual_name_1 = 'ambiguous'
            
            if manual_name_1_l == 'unknown' or manual_name_1_l == 'unkown':
                manual_name_1 = 'NA'
            
            if not len(manual_name_1_l):
                manual_name_1 = 'NA'
            
            # :done
            
            # testing if there is PG/BMP ambiguity
            pg_bmp = not bool(set(['BMP', 'PG']) -
                              set(x.split('(')[0].strip()
                                  for x in manual_names.split(',')))
            
            if manual_names == 'BMP(18:1/19:1), PG(18:1/19:1)':
                print(pg_bmp)
            
            for lips in (
                swl_names.split(r'///')
                if (
                    only_swl_col or
                    manual_name_1 == 'NA' or
                    manual_name_1 == 'ambiguous' or
                    manual_name_1 == 'adduct'
                ) else # at Enric we use the manual names column:
                manual_names.split(',')):
                
                lips = lips.strip()
                
                if pg_bmp:
                    if lips[:2] == 'PG':
                        # at PG we replace with PG/BMP
                        lips = lips.replace('PG', 'PG/BMP')
                        manual_name_1 = 'PG/BMP'
                    elif lips[:3] == 'BMP':
                        # at BMP we skip
                        continue
                
                if pg_bmp:
                    print(lips)
                
                # matching the adduct type
                add = self.readd.match(lips)
                
                if add is not None:
                    add = add.groups()[0]
                else:
                    add = ''
                
                for lip in lips.split(';'):
                    
                    lip = lip.strip()
                    
                    if lip and lip[0] == '(':
                        lip = lip[1:]
                    
                    if 'lyso' in lip.lower():
                        lyso = 'Lyso'
                        manual_name_1 = manual_name_1.replace('yso-', 'yso')
                    else:
                        lyso = ''
                    
                    swl_parsed = self.headgroup_from_lipid_name(['S', None, lip])[0]
                    
                    if swl_parsed is None:
                        swl_parsed = lip
                        if ']' in swl_parsed:
                            swl_parsed = swl_parsed.split(']')[1]
                        swl_parsed = swl_parsed.split('(')[0].strip()
                        if ':' in swl_parsed:
                            swl_parsed = swl_parsed.split(':')[1].strip()
                    
                    if pg_bmp:
                        print('--', swl_parsed)
                    
                    if 'nothing' in swl_parsed:
                        swl_parsed = 'NA'
                    
                    # regex finds the total carbon count
                    cc1 = self.recount1.findall(lip)
                    
                    # special case if the fatty acid
                    # name is greek name
                    if swl_parsed in fa_greek:
                        cc1 = [('', fa_greek[swl_parsed][0],
                                    fa_greek[swl_parsed][1])]
                        swl_parsed = 'FA'
                    
                    # a full headgroup name:
                    fullhg = '%s%s%s' % (
                        lyso if not swl_parsed.startswith('Lyso') else '',
                        swl_parsed,
                        '%s' % ('-O' if len(cc1) and cc1[0][0] == 'O' else '')
                    )
                    
                    # regex finds 2-3 fatty acids
                    cc2s = self.recount3.findall(lip)
                    
                    # the total carbon count
                    ccpart = (
                        [cc1[0][0], int(cc1[0][1]), int(cc1[0][2])]
                        if len(cc1) else
                        ['', np.nan, np.nan]
                    )
                    
                    # carbon counts of fatty acids
                    if len(cc2s) and (
                        any(map(lambda cc2: cc2[4], cc2s)) or
                        swl_parsed == 'FA' or
                        lyso
                    ):
                        
                        faccparts = (
                            list(
                                map(
                                    lambda cc2:
                                        [
                                            # FA1
                                            cc2[0],
                                            int(cc2[1]),
                                            int(cc2[2]),
                                            # FA2
                                            cc2[3],
                                            int(cc2[4]) if cc2[4] else np.nan,
                                            int(cc2[5]) if cc2[5] else np.nan,
                                            # FA3
                                            cc2[6],
                                            int(cc2[7]) if cc2[7] else np.nan,
                                            int(cc2[8]) if cc2[8] else np.nan
                                        ],
                                    filter(
                                        lambda cc2:
                                            # if this is a Lyso species
                                            # or single fatty acid
                                            # we have only one cc:unsat
                                            # otherwise we must have at least 2
                                            cc2[4] or swl_parsed == 'FA' or lyso,
                                        cc2s
                                    )
                                )
                            )
                        )
                        
                    else:
                        faccparts = [
                            [
                                '', np.nan, np.nan,
                                '', np.nan, np.nan,
                                '', np.nan, np.nan
                            ]
                        ]
                    
                    for faccpart in faccparts:
                        
                        if cc2s and not cc1:
                            ccpart = [
                                faccpart[0],
                                np.nansum([faccpart[1], faccpart[4], faccpart[7]]),
                                np.nansum([faccpart[2], faccpart[5], faccpart[8]])
                            ]
                        
                        res = []
                        res.append(swl_parsed)
                        res.append(lyso)
                        res.extend(ccpart)
                        res.extend(faccpart)
                        res.append(fullhg)
                        res.append(manual_name_1)
                        
                        if res[14] == 'NA':
                            nothing.append(res)
                        else:
                            something.append(res)
                    
                    if pg_bmp:
                        print(swl_parsed, fullhg, manual_name_1)
            
            return something or nothing
        
        if not hasattr(self, 'manual') or self.manual is None:
            self.read_manual2()
        
        if not hasattr(self, 'lipnames') or self.lipnames is None:
            self.read_lipid_names()
        
        result = []
        
        def get_uhg(cnt, counts):
            if cnt[-1] not in ['ambiguous', 'adduct']:
                uhg = cnt[-1]
            else:
                uhg = cnt[-2]
            
            if uhg in uhgs:
                uhg = uhgs[uhg]
            
            uhg = uhg.replace('-O-', '-O')
            
            if uhg == 'NA' and len(counts) == 1:
                uhg = cnt[14]
            
            return uhg
        
        for protein, d in iteritems(self.manual):
            
            for mode, tbl in iteritems(d):
                
                for i, l in enumerate(tbl):
                    
                    counts = get_names(l[2], l[3])
                    
                    if l[3].strip() in shgs2:
                        l[3] = shgs2[l[3].strip()]
                    
                    res = [protein, mode, i, l[0], l[5], l[4], l[1], l[3]] + \
                        l[8:12] # 12 cols: protein -- rtup
                    
                    #this_feature_hgs = set([get_uhg(cnt, counts) for cnt in counts])
                    #if not (set(['PG', 'BMP']) - this_feature_hgs):
                    #    pass
                    
                    for cnt in counts:
                        res1 = res[:]
                        
                        if cnt[1] == 'O':
                            cnt[0] = '%s-O' % cnt[0]
                        
                        if cnt[0].strip() in shgs:
                            cnt[0] = shgs[cnt[0].strip()]
                        
                        cnt[-1] = cnt[-1].strip()
                        
                        for hgi in [-1, -2]:
                            
                            if cnt[hgi] in shgs:
                                cnt[hgi] = shgs[cnt[hgi]]
                            if cnt[hgi] in shgs2:
                                cnt[hgi] = shgs2[cnt[hgi]]
                        
                        uhg = get_uhg(cnt, counts)
                        
                        cnt.append(uhg)
                        
                        if not np.isnan(cnt[3]) and not np.isnan(cnt[4]):
                            cnt.append('%s(%u:%u)' % (
                                cnt[16] if cnt[16] != 'NA' else cnt[14],
                                cnt[3],
                                cnt[4]
                            ))
                            cnt.append('%u:%u' % (cnt[3], cnt[4]))
                        else:
                            cnt.append('NA')
                            cnt.append('NA')
                        
                        facc = []
                        for i in [6, 9, 12]:
                            if not np.isnan(cnt[i]) and not np.isnan(cnt[i+1]):
                                facc.append((cnt[i], cnt[i+1]))
                        
                        facc = '/'.join(map(lambda cc: '%u:%u' % cc,
                                            sorted(facc)))
                        
                        if len(facc):
                            cnt.append('%s(%s)' % (
                                cnt[16] if cnt[16] != 'NA' else cnt[14],
                                facc)
                            )
                            cnt.append(facc)
                        else:
                            cnt.append('NA')
                            cnt.append('NA')
                        
                        
                        res1.extend(cnt)
                        res1.append(screen)
                        
                        result.append(res1)
        
        self.pmanual = pd.DataFrame(result, columns = self.df_header)
    
    def auto_df(self, screen_name = 'E'):
        """
        Compiles a data frame in the same format as `manual_df`
        just from the programmatic results.
        """
        
        result = []
        
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                ii = 0
                
                for i, incl in enumerate(tbl['good']):
                    
                    if not incl:
                        
                        continue
                    
                    mz = tbl['mz'][i]
                    oi = tbl['i'][i]
                    idlevel = tbl['idlevel'][oi]
                    intensity = round(tbl['aaa'][i])
                    rt = tbl['rt'][i,:]
                    rtmean = np.mean(tbl['rt'][i,:])
                    rtms2 = tbl['ms2rt'][i]
                    
                    cids = tbl['cid'][oi]
                    
                    clm = None
                    if not cids:
                        clm = 'unknown'
                        cids = ['unknown']
                    elif len(cids) > 1:
                        clm = 'ambiguous'
                    
                    for lip in cids:
                        
                        res = []
                        
                        res.append(protein)
                        res.append(mode)
                        res.append(ii)
                        res.append(mz)
                        res.append(mz)
                        res.append(intensity)
                        res.append(idlevel)
                        
                        cl = lip.split('(')[0] if clm is None else clm
                        hg = 'NA' if clm == 'unknown' else lip.split('(')[0]
                        lyso = 'Lyso' if 'lyso' in lip.lower() else ''
                        hg = 'NA' if clm == 'unknown' else lip.split('(')[0]
                        pref = 'O' if '-O' in hg else ''
                        if 'Cer' in hg:
                            pref = 'd' if 'CerOH' not in hg else 't'
                        
                        if pref =='O':
                            hg = '%s-O' % hg
                        
                        hg0 = hg.replace(
                            '-O', '').replace(
                            'CerOH', 'Cer').replace(
                            'Cer1P', 'CerP')
                        
                        res.append(hg)
                        res.append(rtmean)
                        res.append(rtms2)
                        res.append(rt[0])
                        res.append(rt[1])
                        res.append(hg0)
                        res.append(lyso)
                        # res.append(pref)
                        
                        cc  = self.recount3.findall(lip)
                        
                        if cc:
                            
                            sumcc  = sum(map(lambda cci:
                                            int(cc[0][cci]) if cc[0][cci] else 0,
                                        [1, 4, 7]))
                            sumuns = sum(map(lambda uni:
                                            int(cc[0][uni]) if cc[0][uni] else 0,
                                        [2, 5, 8]))
                            
                            res.extend([pref, sumcc, sumuns])
                            
                            hgcc = '%s(%u:%u)' % (hg, sumcc, sumuns)
                            sumccuns = '%u:%u' % (sumcc, sumuns)
                            fa = '%s:%s' % (cc[0][1], cc[0][2])
                            if cc[0][4]:
                                fa = '%s/%s:%s' % (fa, cc[0][4], cc[0][5])
                            if cc[0][7]:
                                fa = '%s/%s:%s' % (fa, cc[0][7], cc[0][8])
                            
                        else:
                            res.extend(['', np.nan, np.nan])
                            
                            hgcc = 'NA'
                            sumccuns = 'NA'
                            fa = 'NA'
                        
                        cc2 = self.recount2.findall(lip)
                        
                        if (
                            cc and
                            cc[0][1] and
                            cc[0][2] and (
                                cc[0][4] or hg == 'FA' or lyso
                            )):
                            
                            res.extend([cc[0][0],
                                int(cc[0][1]),
                                int(cc[0][2]),
                                cc[0][3],
                                int(cc[0][4]) if cc[0][4] else np.nan,
                                int(cc[0][5]) if cc[0][5] else np.nan,
                                cc[0][6],
                                int(cc[0][7]) if cc[0][7] else np.nan,
                                int(cc[0][8]) if cc[0][8] else np.nan
                            ])
                            
                        else:
                            
                            res.extend(['', np.nan, np.nan,
                                        '', np.nan, np.nan,
                                        '', np.nan, np.nan])
                        
                        res.append(hg)
                        res.append(hg if clm is None else clm)
                        res.append('NA' if clm == 'unknown' else hg)
                        res.append(hgcc)
                        res.append(sumccuns)
                        res.append('%s(%s)' % (hg, fa))
                        res.append(fa)
                        
                        res.append(screen_name)
                        
                        result.append(res)
                        
                        ii += 1
        
        self.pauto = pd.DataFrame(result, columns = self.df_header)
    
    def headgroups_cross_screening(self, label1 = 'Screen1',
                                   label2 = 'Screen2',
                                   idlevels = set(['I']),
                                   outfile = 'headgroups_%s_%s.tab'):
        """
        Does a quick comparison at headgroup/protein level
        between 2 screenings.
        """
        
        self.manual_df()
        
        result = []
        
        for protein, d in iteritems(self.valids):
            
            for mode, tbl in iteritems(d):
                
                s1_hg = set(
                    map(
                        lambda hg:
                            hg.replace('-O', ''),
                        self.pmanual[
                            np.logical_and(
                                np.logical_and(
                                    self.pmanual.protein == protein,
                                    self.pmanual.ionm    == mode
                                ),
                                self.pmanual.cls.isin(idlevels)
                            )
                        ]['uhgroup']
                    )
                )
                
                s2_hg = (
                    set(
                        reduce(
                            lambda h1, h2:
                                h1 | h2,
                            map(
                                lambda i:
                                    set(
                                        map(
                                            lambda hgfa:
                                                hgfa.split('(')[0],
                                            tbl['cid'][i[1]]
                                        )
                                    ),
                                filter(
                                    lambda i:
                                        (
                                            tbl['slobb'][i[0]] and
                                            tbl['idlevel'][i[1]] in idlevels
                                        ),
                                    enumerate(tbl['i'])
                                )
                            ),
                            set([])
                        )
                    )
                )
                
                for hg12 in (s2_hg & s1_hg):
                    result.append(['%s-%s' % (protein, mode), hg12,
                                   '%s_%s' % (label1, label2)])
                
                for hg1 in (s1_hg - s2_hg):
                    result.append(['%s-%s' % (protein, mode), hg1, label1])
                
                for hg2 in (s2_hg - s1_hg):
                    result.append(['%s-%s' % (protein, mode), hg2, label2])
        
        outfile = outfile % (label1, label2)
        
        hdr = ['protein_mode', 'hg', 'found_in']
        
        self.cross_screen_hg = result
        
        with open(outfile, 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr))
            
            fp.write('\n'.join(map(lambda row: '\t'.join(row), result)))
    
    def bubble_altair(self,
                      classes = ['I', 'II'],
                      subtitle = '',
                      main_title = ''):
        
        smodes = {'pos': '+', 'neg': '-'}
        # select the classes
        data = self.pmanual[self.pmanual.cls.isin(classes)]
        
        nrows = 0
        ncols = len(data.ionm.unique())
        subplot_titles = []
        
        allhgs = sorted(data.headgroup.unique())
        unsat = np.arange(min(data.unsat), max(data.unsat) + 1)
        carb = np.arange(min(data.carb), max(data.carb) + 1)
        inte = (min(data.intensity), max(data.intensity))
        traces = []
        
        xlim = [min(unsat), max(unsat)]
        ylim = [min(carb), max(carb)]
        
        data = data.sort_values(by = ['protein', 'ionm'])
        
        a = altair.Chart(data).mark_point().encode(
            row = 'protein',
            column = 'ionm',
            size = 'Intensity:average(intensity)',
            x = altair.X('unsat', axis = altair.Axis(title = 'Unsaturated count')),
            y = altair.Y('carb', axis = altair.Axis(title = 'Carbon count'))
        )
        
        return a
    
    @staticmethod
    def export_df(df, fname, **kwargs):
        """
        Exports the results from a `pandas.DataFrame` to csv.
        """
        
        if 'sep' not in kwargs:
            kwargs['sep'] = '\t'
        if 'na_rep' not in kwargs:
            kwargs['na_rep'] = 'NaN'
        if 'index' not in kwargs:
            kwargs['index'] = False
        
        df.to_csv(fname, **kwargs)
    
    def bubble_plotly(self,
                     classes = ['I', 'II'],
                     subtitle = '',
                     main_title = ''):
        
        smodes = {'pos': '+', 'neg': '-'}
        # select the classes
        data = self.pmanual[self.pmanual.cls.isin(classes)]
        
        nrows = 0
        ncols = len(data.ionm.unique())
        subplot_titles = []
        
        allhgs = sorted(data.headgroup.unique())
        unsat = np.arange(min(data.unsat), max(data.unsat) + 1)
        carb = np.arange(min(data.carb), max(data.carb) + 1)
        inte = (min(data.intensity), max(data.intensity))
        traces = []
        
        xlim = [min(unsat), max(unsat)]
        ylim = [min(carb), max(carb)]
        
        for protein in sorted(data.protein.unique()):
            nrows += 1
            
            for mode in sorted(data.ionm.unique()):
                
                subplot_titles.append('%s%s%s' % (
                    protein,
                     smodes[mode],
                     ', %s' % subtitle if len(subtitle) else '')
                )
                
                this_data = \
                    data[(data.protein == protein) & (data.ionm == mode)]
                
                vals = this_data.groupby(['carb', 'unsat'])['intensity'].sum()
                #print(list(iteritems(vals)))
                x = list(map(lambda i: i[0][1], iteritems(vals)))
                y = list(map(lambda i: i[0][0], iteritems(vals)))
                s = list(map(lambda i: i[1] / float(inte[1]), iteritems(vals)))
                
                #print(protein, x, y, s)
                
                traces.append(go.Scatter(x = x, y = y,
                                         mode = 'markers',
                                         marker = dict(size = s, sizemode = 'area', sizeref = 0.0001),
                                         name = '%s%s' % (protein, smodes[mode]), fill = '#333333', showlegend = False,
                                         xaxis = dict(range = xlim),
                                         yaxis = dict(range = ylim))
                                    )
        
        fig = plotly.tools.make_subplots(rows=nrows,
                                         cols=ncols,
                                         print_grid = False,
                                         subplot_titles=subplot_titles
                                        )
        
        for i, trace in enumerate(traces):
            fig.append_trace(trace, row = int(np.floor(i / ncols) + 1), col = (i % ncols) + 1)
        
        fig['layout'].update(height = nrows * 500, width = 600, title = main_title,
                             xaxis = dict(range = xlim), yaxis = dict(range = ylim))
        
        pl.iplot(fig, show_link = False)
    
    def piecharts_plotly(self, by_class = True, main_title = 'Lipid classes by protein', result_classes = {'I'}):
        """
        Plots piecharts of detected lipids for each protein based on manually
        annotated results.
        Uses plotly, output accessible in Jupyter notebook.
        """
        
        def get_names(r, by_class = True):
            
            counts = []
            
            for lips in r[2].split('///'):
                
                for lip in lips.split(';'):
                    
                    if 'nothing' in lip or not len(lip.strip()):
                        continue
                    
                    cl = self.headgroup_from_lipid_name(['S', None, lip])[0]
                    
                    if cl is None:
                        cl = lip.split('(')[0].strip()
                        if ':' in cl:
                            cl = cl.split(':')[1].strip()
                    
                    if by_class:
                        cc = ''
                    else:
                        cc = self.recount2.findall(lip)
                        
                        if not len(cc):
                            cc = self.recount1.findall(lip)
                        
                        cc = cc[0] if len(cc) else '?'
                    
                    counts.append('%s(%s)' % (cl, cc) if len(cc) else cl)
            
            return counts
        
        if not hasattr(self, 'manual') or self.manual is None:
            self.read_manual()
        
        if not hasattr(self, 'lipnames') or self.lipnames is None:
            self.read_lipid_names()
        
        main_title = '%s (class %s)' % (main_title, ', '.join(sorted(list(result_classes))))
        
        modes = {'pos': 'positive', 'neg': 'negative'}
        smodes = {'pos': '+', 'neg': '-'}
        nrows = int(np.ceil(len(self.manual) / 2.0))
        height = 500 * nrows
        param = {
            'data': [],
            'layout': {
                'title': main_title,
                'annotations': [],
                'autosize': False,
                'width': 600,
                'height': height
            }
        }
        
        traces = []
        #fig = plotly.tools.make_subplots(rows=nrows, cols=2, print_grid = False)
                          #subplot_titles=('First Subplot','Second Subplot', 'Third Subplot'))
        
        n = 0
        for protein in sorted(self.manual.keys()):
            
            for mode in ['neg', 'pos']:
                
                
                this_data = {}
                this_anno = {'font': {'size': 10}, 'showarrow': False}
                lab_val = {}
                
                for r in self.manual[protein][mode]:
                    
                    if r[1].strip() in result_classes:
                        
                        label = '/'.join(get_names(r, by_class = by_class))
                        
                        if label not in lab_val:
                            lab_val[label] = 0.0
                        
                        lab_val[label] += r[4]
                    
                this_data['labels'], this_data['values'] = \
                    zip(*sorted(lab_val.items(), key = lambda i: i[0])) \
                        if len(lab_val) else (['None'], [1])
                this_data['name'] = '%s %s, \nsum of intensities' % (protein, modes[mode])
                this_data['type'] = 'pie'
                this_data['hole'] = 0.4
                this_data['hoverinfo'] = 'label+percent+name'
                this_data['domain'] = {
                    'x': [
                        n % 2 / 2.0,
                        n % 2 / 2.0 + 0.5
                    ],
                    'y': [
                        1.0 - (0.48 / nrows * np.floor(n / 2) + 0.003),
                        1.0 - (0.48 / nrows * (np.floor(n / 2) + 1) - 0.003)
                    ]
                }
                this_pie = go.Pie(**this_data)
                traces.append(this_pie)
                
                # print('%s: n = %u,   %s' % (protein, n, str(this_data['domain'])))
                
                # fig.append_trace(this_pie, int(n % 2 + 1), int(np.floor(n / 2.0) + 1))
                this_anno['text'] = '%s [%s]' % (protein, smodes[mode])
                this_anno['x'] = n % 2 / 2.0 + 0.25
                this_anno['y'] = 1.0 - (0.48 / nrows * np.floor(n / 2) + 0.48 / nrows / 2.0)
                this_anno['xanchor'] = 'center'
                this_anno['yanchor'] = 'middle'
                
                param['data'].append(this_data)
                param['layout']['annotations'].append(this_anno)
                
                n += 1
        
        layout = go.Layout(annotations = param['layout']['annotations'],
                        height = height, title = main_title,
                        #width = 600, autosize = False
                        )
        fig = go.Figure(data = traces, layout = layout)
        # print(param)
        #fig['layout'].update(showlegend = True, title = 'Lipid classes by protein')
        pl.iplot(fig, show_link = False)
    
    #
    # Methods for preparing a diff between 2 sets of xls outputs
    #
    
    def features_xls_diff(self, dir1, dir2,
                          outdir = 'top_features_diff', e = 0.0005):
        
        def get_xls(d, f):
            return openpyxl.load_workbook(os.path.join(d, f), read_only = True)
        
        def get_header(xls, sheet_name = 0):
            """
            Returns values in header row of one sheet
            """
            sheet = xls[sheet_name]
            return list(map(lambda c: c.value, next(sheet.iter_rows())))
        
        def index_sheet(sheet):
            
            def table_cell_attr(sheet, fun, strip_header = True):
                tbl = \
                    list(
                        map(
                            lambda r:
                                list(map(lambda c: fun(c), r)),
                            sheet.iter_rows()
                        )
                    )
                if strip_header:
                    tbl = tbl[1:]
                return tbl
            
            content = table_cell_attr(sheet, lambda c: c.value)
            colors = table_cell_attr(sheet, lambda c: str(c.fill.bgColor.rgb)[:6])
            
            index = \
                np.array(
                    list(
                        map(
                            lambda i:
                                # [n, n, n] from (n, (n, n))
                                [i[0], i[1][0], i[1][1]],
                            enumerate(
                                map(
                                    lambda r:
                                        # m/z and intensity
                                        (float(r[2]), float(r[12])),
                                    content
                                )
                            )
                        )
                    )
                ).reshape(len(content), 3)
            
            index = index[index[:,1].argsort()]
            
            return content, colors, index
        
        def diff(xls1, xls2):
            
            result = {}
            
            for s in sorted(xls1.sheetnames):
                
                if s not in xls2.sheetnames:
                    sys.stdout.write('\tSheet `%s` missing!\n' % s)
                
                sheet1 = xls1[s]
                sheet2 = xls2[s]
                
                try:
                    result[s] = sdiff(sheet1, sheet2)
                except:
                    print(s)
                
            return result
        
        def sdiff(sheet1, sheet2):
            """
            Returns rows in 1 missing from 2.
            """
            dat1, col1, i1 = index_sheet(sheet1)
            dat2, col2, i2 = index_sheet(sheet2)
            missing = []
            
            for i in i1:
                ui = i2[:,1].searchsorted(i[1])
                if ui < i2.shape[0]:
                    if abs(i2[ui,1] - i[1]) < e:
                        continue
                if ui > 0:
                    if abs(i2[ui - 1,1] - i[1]) < e:
                        continue
                missing.append([i[0], i[2]])
            
            nmissg = len(missing)
            missing = np.array(missing).reshape(nmissg, 2)
            
            # ordering by intensities desc
            
            missing = missing[missing[:,1].argsort()[::-1]]
            
            # ordering all outputs the same way
            return \
                list(map(lambda i: dat1[i],
                         missing[:,0].astype(np.int, copy = False))), \
                list(map(lambda i: col1[i],
                         missing[:,0].astype(np.int, copy = False))), \
                      i1[missing[:,0].astype(np.int, copy = False),]
        
        # starting with empty outdir
        if not os.path.exists(outdir):
            os.mkdir(outdir)
        
        for f in os.listdir(outdir):
            os.remove(os.path.join(outdir, f))
        
        # column widths as set in the original method:
        colws = [
            1.8, 1.8, 1.8, 1.8, 1.8, 1.8, 1.8, 1.8, 1.8, 6.4, 6.4, 6.4, 2.83,
            2.83, 10.77, 2.83, 2.83, 2.83, 1.86, 4.29, 8.60, 4.29, 8.60, 4.29,
            8.60, 4.29, 8.60, 4.29, 8.60, 0.48, 2.12, 2.83, 4.29, 2.12, 2.12,
            2.12, 2.12, 2.12, 2.12, 2.12, 1.48, 1.48, 3.62, 3.31, 3.31, 3.31
        ]
        
        
        
        fnames1 = \
            list(
                filter(
                    lambda f:
                        f.endswith('_top_features.xlsx'),
                    os.listdir(dir1)
                )
            )
        
        prg = progress.Progress(len(fnames1),
                                'Compiling diffs, generating xls',
                                1, percent = False)
        
        # old +, new +, + new, + false, 
        # old best +, new best +, ...
        stats = {}
        stats['hdr'] = ['', 'old +', 'new +', '+ new', '+ false',
                'old best +', 'new best +', 'best + new', 'best + false',
                'old -', 'new -', '- new', '- false',
                'old best -', 'new best -', 'best - new', 'best - false'
            ]
        
        for f in fnames1:
            
            prg.step()
            
            protein = f.split('_')[0]
            xls1 = get_xls(dir1, f)
            xls2 = get_xls(dir2, f)
            nhdr = get_header(xls1, '%s_negative' % protein)
            phdr = get_header(xls1, '%s_positive' % protein)
            nums = [protein]
            # new records in recent tables:
            new = diff(xls2, xls1)
            # false records in old tables:
            old = diff(xls1, xls2)
            
            for mode in ['positive', 'negative']:
                for sel in ['', '_best']:
                    
                    sheet_name = '%s_%s%s' % (protein, mode, sel)
                    nums.extend([
                        '%u' % (xls1[sheet_name].max_row - 1),
                        '%u' % (xls2[sheet_name].max_row - 1),
                        '%u' % len(new[sheet_name][0]),
                        '%u' % len(old[sheet_name][0])
                    ])
            
            stats[protein] = nums
            
            difffname = '%s_diff.xlsx' % protein
            difffname = os.path.join(outdir, difffname)
            
            xls = xlsxwriter.Workbook(difffname, {'constant_memory': True})
            
            for typ, d in [('new', new), ('false', old)]:
                
                for sh in sorted(d.keys()):
                    
                    tbl = \
                        list(
                            map(
                                lambda r:
                                    list(
                                        map(
                                            lambda c:
                                                (c[0], 'plain') \
                                                    if c[1] == '000000' else \
                                                (c[0], 'green'),
                                            zip(r[0], r[1])
                                        )
                                    ),
                                zip(d[sh][0], d[sh][1])
                            )
                        )
                    
                    sheet_name = '%s_%s' % (sh, typ)
                    hdr = nhdr if 'negative' in sheet_name else phdr
                    self.add_sheet(xls, [hdr] + tbl, sheet_name, colws)
            
            xls.close()
        
        prg.terminate()
        
        with open('diff_stats.txt', 'w') as f:
            f.write('%s\n' % '\t'.join(stats['hdr']))
            del stats['hdr']
            for protein in sorted(stats.keys()):
                f.write('%s\n' % '\t'.join(stats[protein]))
    
    def to_uniprot(self, protein):
        """
        Gets the UniProt from LTP name.
        """
        self.read_uniprots()
        
        return self.uniprots[protein]['uniprot'] \
            if protein in self.uniprots else None
    
    def get_family(self, protein):
        """
        Returns the lipid binding domain family from LTP name.
        """
        self.read_uniprots()
        
        return self.uniprots[protein]['family'] \
            if protein in self.uniprots else None
    
    def get_name(self, uniprot):
        """
        Returns the LTP name from its UniProt.
        """
        self.read_uniprots()
        
        return self.names[uniprot] \
            if uniprot in self.names else None
    
    def read_uniprots(self, reread = True):
        """
        Reads the UniProt IDs and domain families of LTPs.
        Result stored in `uniprots` attribute.
        """
        if self.uniprots is None or reread:
            with open(self.ltplistf, 'r') as f:
                self.uniprots = dict(
                    map(
                        lambda l:
                            (l[1].strip(), {'uniprot': l[2].strip(),
                                            'family': l[0].strip()}),
                        map(
                            lambda l:
                                l.strip().split('\t'),
                            f
                        )
                    )
                )
            
            self.names = \
                dict(
                    map(
                        lambda i:
                            (i[1]['uniprot'], i[0]),
                        iteritems(self.uniprots)
                    )
                )
    
    def get_comppi_localizations(self, minor = False):
        """
        Downloads localization data from ComPPI.
        Results stored in `ulocs` and `locs` attributes.
        """
        url = self.comppi_url
        post = {
            'fDlSet': 'protnloc',
            'fDlSpec': '0',
            'fDlMloc': 'all',
            'fDlSubmit': 'Download'
        }
        
        c = _curl.Curl(url = url, post = post,
                        large = True, silent = False, compr = 'gz')
        
        self.read_uniprots()
        
        ultps = set(map(lambda p: p['uniprot'], self.uniprots.values()))
        
        self.ulocs = \
            dict(
                map(
                    lambda l:
                        (
                            l[0],
                            dict(
                                map(
                                    lambda loc:
                                        (
                                            loc[0],
                                            float(loc[1])
                                        ),
                                    map(
                                        lambda loc:
                                            loc.split(':'),
                                        l[3].split('|')
                                    )
                                )
                            )
                        ),
                    map(
                        lambda l:
                            l.strip().split('\t'),
                        filter(
                            lambda l:
                                l[:6] in ultps or l[:10] in ultps,
                            c.result
                        )
                    )
                )
            )
        
        self.locs = dict(map(lambda i: (self.names[i[0]], i[1]),
                             iteritems(self.ulocs)))
    
    def read_membrane_constitutions(self):
        """
        Reads membrane lipid constitutions from Charlotte.
        """
        pass
    
    def read_manual_localizations(self):
        """
        Reads localization data from our manual collection.
        """
        abbrev = {
            'G': 'golgi',
            'ER': 'ER',
            'LE/LY': 'late endosome, lysosome',
            'Cy': 'cytosol',
            
        }
        tbl = self.read_xls(self.localizationf, sheet = 1)
        return tbl
    
    def get_go_goa(self, organism='human'):
        """
        Downloads GO annotation from UniProt GOA.
        """
        
        def add_annot(a, result):
            if a[1] not in result[a[8]]:
                result[a[8]][a[1]] = []
            result[a[8]][a[1]].append(a[4])
        
        result = {'P': {}, 'C': {}, 'F': {}}
        
        url = self.goa_url % (organism.upper(), organism)
        c = _curl.Curl(url, silent=False, large = True)
        
        _ = \
            list(
                map(
                    lambda l:
                        add_annot(l.strip().split('\t'), result),
                    filter(
                        lambda l:
                            l[0] != '!' and len(l),
                        map(
                            lambda l:
                                l.decode('ascii'),
                            c.result
                        )
                    )
                )
            )
        
        return result

    def get_go_quick(self, organism=9606, slim=False, names_only=False):
        """
        Loads GO terms and annotations from QuickGO.
        Returns 2 dicts: `names` are GO terms by their IDs,
        `terms` are proteins GO IDs by UniProt IDs.
        """
        def add_term(a, terms, names, names_only):
            if not names_only:
                if a[0] not in terms[a[3][0]]:
                    terms[a[3][0]][a[0]] = set([])
                terms[a[3][0]][a[0]].add(a[1])
            names[a[1]] = a[2]
        
        termuse = 'slim' if slim or slim is None else 'ancestor'
        goslim = '' if termuse == 'ancestor' \
            else '&goid=%s' % ','.join(get_goslim(url=slim))
        terms = {'C': {}, 'F': {}, 'P': {}}
        names = {}
        url = self.quickgo_url % (goslim, termuse, organism)
        c = _curl.Curl(url, silent=False, large=True)
        _ = c.result.readline()
        
        _ = \
            list(
                map(
                    lambda l:
                        add_term(l, terms, names, names_only),
                    map(
                        lambda l:
                            l.decode('ascii').strip().split('\t'),
                        c.result
                    )
                )
            )
        
        return {'terms': terms, 'names': names}
    
    def count_nonzero_features(self, outf = 'nonzero_counts.tab'):
        """
        Counts the number of nonzero features for each fraction.
        """
        result = {}
        hdr = ['protein', 'mode', 'fr', 'set', 'peak', 'cnt']
        f = open(outf, 'w')
        f.write('%s\n' % '\t'.join(hdr))
        
        for protein, d in iteritems(self.valids):
            
            ifracs = self.fraction_indices(protein)
            
            for mode, tbl in iteritems(d):
                
                for attr, key in [('valids', 'fe'), ('data', 'raw')]:
                    
                    k = (protein, mode, attr)
                    if k not in result:
                        result[k] = {}
                    
                    for fr, fi in iteritems(ifracs):
                        
                        try:
                            cnt = np.sum(getattr(self, attr)[
                                protein][mode][key][:,fi[0]] > 0)
                        except:
                            print(protein, mode, fr)
                            cnt = 0
                        
                        result[k][fr] = cnt
                        
                        f.write('%s\n' % \
                            '\t'.join([protein, mode, fr,
                                       attr,
                                       'peak' if fi[1] else 'non-peak',
                                       '%u' % cnt]
                            )
                        )
        
        f.close()
        
        self.nonzero_features = result
    
    def count_features(self, attr, fname = 'feature_count.tab'):
        """
        Exports a table with number of features selected by a filter.
        """
        
        hdr = ['protein', 'mode', 'cat', 'num']
        
        with open(fname, 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr))
            
            for protein, d in iteritems(self.valids):
                
                for mode, tbl in iteritems(d):
                    
                    allr = [protein, mode, 'all', '%u' % tbl['fe'].shape[0]]
                    
                    goodr = [protein, mode, 'good', '%u' % sum(tbl[attr])]
                    
                    fp.write('%s\n' % '\t'.join(allr))
                    fp.write('%s\n' % '\t'.join(goodr))
    
    def get_scan(self, protein, mode, mz):
        """
        Returns the best MS2 scan for a feature.
        """
        
        oi  = self.original_id(protein, mode, mz)
        ms2 = self.valids[protein][mode]['ms2f'][oi]
        
        return ms2._scans[ms2.best_scan]
    
    def read_prior_knowledge(self, fname = '../ltp/170224_LTP_master_table.xlsx'):
        """
        Reads literature knowledge and Charlotte's results from LiMA assays.
        """
        
        rehptlc = re.compile(r'[A-z]{2,}')
        repinum = re.compile(r'(PI)\(?(\d),?(\d?),?(\d?)\)?(P\d?)')
        recer   = re.compile(r'C\d+(Cer1?P?)')
        relima  = re.compile(r'''(?x)
                ([A-z\d][ ,A-z\d]*?[A-z\d])
                [ ,]*\(
                ([A-z][ ,A-z]*?[A-z])
                \)''')
        reccloc = re.compile(r'''(?x)
                (?:[ A-z0-9:]*: ?)?[ ]
                ([A-Z][ A-z]*[A-z])\.?[ ]
                (?:[ \.\{\}\|A-z\d=:;]+?\.)''')
        rehpalo = re.compile(r'([ A-z]+) \(.*?\)')
        renothi = re.compile(r'unknown|n\.d\.|ND|NA')
        resep   = re.compile(r'[ ]?[;,/][ ]?')
        
        nothing = set(['unknown', 'n.d.', 'ND', 'NA'])
        
        shortnames = {
            'photo-25-hydroxycholesterol': 'HCH',
            '25-hydroxycholesterol': 'HCH',
            '25-hydroxylcholesterol': 'HCH',
            '22(r)-hydroxycholesterol': 'HCH',
            'cholesterol-3-o-sulfate': 'CHS',
            'chenodeoxycholic acid': 'CCA',
            'cholic acid': 'CA',
            'cholesterol': 'CH',
            'photo-cholesterol': 'CH',
            'dehydroergosterol': 'DES',
            'cholestatrienol': 'CHT',
            'lipid iva': 'IVA',
            'alpha-tocopherol': 'VE',
            '11-cis-retinaldehyde': 'VA',
            '11-cis-retinol': 'VA',
            '11-cis-retinal': 'VA',
            'squalene': 'SQ',
            'retinoic acid': 'VA',
            'synthetic retinoid': 'VA',
            'bilirubin': 'BR',
            'cholesterol-3-O-sulfate': 'CHS',
            'tetra-acylated lipid iva': 'IVA',
            'rrr-alpha-tocopherylquinone': 'VE',
            'biotinylated alpha-tocopherol': 'VE',
            'sterols': 'CH',
            '7-ketocholesterol': 'KCH',
            'cholestatrienol': 'CH',
            'arachidonate': 'FA',
            'pips': 'PIP',
            'cerebroside': 'HexCer',
            'retinal': 'VA',
            'retinol': 'VA',
            'c12-bodipy-glucosylceramide': 'HexCer',
            'palmitate': 'FA',
            '16-doxylstearic acid': 'FA',
            'long-chain fa acyl-coa esters': 'FA',
            'fa acyl-fa esters': 'FA',
            'very-long-chain fa acyl-coa esters': 'FA',
            'biliverdin': 'BR',
            'bile acids (taurocholate': 'CA',
            'ursodeoxycholate': 'CA',
            'taurochenodeoxycholate)': 'CA',
            'LPA': 'LysoPA',
            'oleate': 'FA',
            'cholate': 'CA',
            'chenodeoxycholate': 'CA',
            'bile acids (various)': 'CA',
            'hydroxyeicosatetraenoic acids': 'PUFA',
            'hydroperoxyeicosatetraenoic acids': 'PUFA',
            'oleoyl-coa ester': 'FA',
            'phytanic acid': 'FA',
            'linoleate': 'FA',
            'stearate': 'FA',
            'palmitate': 'FA',
            'elaidate': 'FA',
            'laurate': 'FA',
            'fa acyl-coa esters': 'FA',
            'pips (pi34p2': 'PI34P2',
            'docosahexaenate)': 'PUFA',
            'progesterone': 'PR',
            'fatty acids (oleate': 'FA',
            'triacylglycerol': 'TAG',
            'galcer': 'HexCer',
            'ceramide': 'Cer',
            'bodypi-glccer': 'HexCer',
            'ceramide-1-phosphate': 'CerP',
            'Cer1P': 'CerP',
            'laccer': 'HexCer',
            'PAF, LPAF': 'PAF',
            'PHCH': 'HCH',
            'PCH': 'CH'
        }
        
        toremove = set([
            '3',
            "3'",
            'α-pinene',
            'and vanillin)',
            'citralva',
            'Odorants (IBMP',
            'Imatinib',
            'Melatonin',
            "5'-triiodo-L-thyronine",
            '5-triiodo-L-thyronine',
            'L-thyroxine',
            "'compound 19'",
            '(lyso)platelet activating factor',
            'Iron (FeDHBA',
            'FeENT)',
            'Platelet activating factor'
        ])
        
        locnames = {
            'Mainly cytosolic': 'Cplasm',
            'Cytoplasm': 'Cplasm',
            'Cytosol': 'Cplasm',
            
            'Rough endoplasmic reticulum': 'ER',
            'Endoplasmic reticulum membrane': 'ER',
            'Endoplasmic reticulum': 'ER',
            
            'Golgi stacks in interphase cells': 'Golgi',
            'Golgi stack membrane': 'Golgi',
            'Golgi apparatus': 'Golgi',
            'TGN exit sites': 'TGN',
            
            'Late endosome membrane': 'LE',
            'Late endosome': 'LE',
            'Early endosome membrane': 'EE',
            
            'Cytoplasmic granule membrane': 'Gran',
            'Granule': 'Gran',
            
            'Nucleus membrane': 'Nuc',
            'Nuclear membrane': 'Nuc',
            'Nucleus envelope': 'Nuc',
            'Nuclear speckles': 'Nuc',
            'Nucleus outer membrane': 'Nuc',
            'Nucleus': 'Nuc',
            'Nucleoplasm': 'Nuc',
            'Nucleoli': 'Nuc',
            
            'Cell membrane': 'PM',
            'Cleavage furrow': 'PM',
            'Plasma membrane': 'PM',
            
            'Mitochondrion envelope': 'Mit',
            'Mitochondrion': 'Mit',
            'Mitochondria': 'Mit',
            
            'Also': None,
            'Localizes to the daughter centriole during mitosis': 'MTOC',
            'Microtubule organizing center': 'MTOC',
            'Centrosome': 'MTOC',
            'Centriole': 'MTOC',
            
            'III membrane protein': None,
            'Isoform': None,
            'Peripheral membrane protein': None,
            'Peripheral membrane': None,
            'Membrane': None,
            
            'Lipid droplet': 'Lipdrop',
            'Secreted': 'Secr',
            'Peroxisome': 'Perox',
            'Peroxisomes': 'Perox',
            'Lysosome': 'Lyso',
            'Microtubules': 'Microt',
            'Actin filaments': 'Actin',
            'Vesicles': 'Vesi',
            'Cytoplasmic bodies': 'Cybody'
            
        }
        
        raw = self.read_xls(fname)
        
        hptlc = {} # HPTLC binders
        hpalo = {} # compartment localization from HPA
        ccloc = {} # compartment localization form GO
        litlo = {} # compartment localization from literature
        litli = {} # literature ligands from Charlotte
        limal = {} # lima assay lipids
        limam = {} # lima assay artificial membranes
        limcm = {} # lima assay complex membranes
        litme = {} # membrane lipid affinities from literature
        litpr = {} # membrane related protein interactions from literature
        famly = {} # domain families
        synon = {}
        
        for r in raw[1:]:
            
            protein = r[2].strip()
            
            famly[protein] = r[1].strip().replace('STAR', 'START')
            synon[protein] = r[3].split()
            
            # known ligands from literature
            litli[protein] = set([])
            
            for lip in resep.split(repinum.sub(r'\1\2\3\4\5', r[14])):
                
                lip = lip.strip()
                if lip.lower() in shortnames:
                    litli[protein].add(shortnames[lip.lower()])
                elif lip.lower() in fa_greek:
                    litli[protein].add('FA')
                else:
                    litli[protein].add(lip)
            
            litli[protein].difference_update(nothing)
            litli[protein].difference_update(toremove)
            
            # HPTLC results
            hptlc[protein] = set(map(lambda lip: (
                                shortnames[lip.lower()]
                                if lip.lower() in shortnames
                                else lip),
                                rehptlc.findall(r[15])
                             ))
            
            hptlc[protein].difference_update(nothing)
            
            # membrane lipid affinities from literature
            litme[protein] = set(map(lambda x: x.strip(), r[19].split(',')))
            litme[protein].discard('')
            
            # membrane protein interactions from literature
            litpr[protein] = set(map(lambda x: x.strip(), r[20].split(',')))
            
            # lima assay results
            lip_mem = recer.sub(r'\1', repinum.sub(r'\1\2\3\4\5', r[21]))
            lip_mem = lip_mem.replace('S1P', 'SphP')
            lip_mem = lip_mem.replace('Cer1P', 'CerP')
            
            limal[protein] = set([])
            limam[protein] = set([])
            limcm[protein] = set([])
            
            for lip, mem in relima.findall(lip_mem):
                
                limal[protein].update(set(map(lambda x: x.strip(),
                                          lip.split(','))))
                limam[protein].update(set(filter(lambda x: x[1:4] == 'OPC',
                                          map(lambda x: x.strip(),
                                          mem.split(',')))))
                limcm[protein].update(set(filter(lambda x: x[1:4] != 'OPC',
                                          map(lambda x: x.strip(),
                                          mem.split(',')))))
            
            limal[protein].discard('not expressed')
            limal[protein].discard('IMM')
            limcm[protein].discard('stop codon')
            
            if 'IMM' in lip_mem:
                limcm[protein].add('IMM')
            if 'OMM' in lip_mem:
                limcm[protein].add('OMM')
            
            litpr[protein] = set(map(lambda x: x.strip().split('-')[0],
                                     resep.split(r[24])))
            
            ccloc[protein] = set(map(lambda x: (
                                     locnames[x]
                                     if x in locnames
                                     else x),
                                     reccloc.findall(r[25])))
            
            if 'Endomembrane system' in ccloc[protein]:
                ccloc[protein].update(set(['ER', 'Golgi']))
                ccloc[protein].remove('Endomembrane system')
            if 'Endosome membrane' in ccloc[protein]:
                ccloc[protein].update(set(['EE', 'LE']))
                ccloc[protein].remove('Endosome membrane')
            
            ccloc[protein].discard(None)
            
            hpalo[protein] = set(map(lambda x: (
                                     locnames[x]
                                     if x in locnames
                                     else x),
                                     rehpalo.findall(r[26])))
        
        # build a complete synonym dictionary
        # from synonyms to our preferred standard names
        stdnm = {}
        for st, syns in iteritems(synon):
            
            stdnm[st] = st
            
            for syn in syns:
                
                stdnm[syn] = st
        
        # typos...
        stdnm['ARGHAP1'] = 'ARHGAP1'
        
        self.prior = {
            'hptlc': hptlc, # HPTLC binders
            'hpalo': hpalo, # compartment localization from HPA
            'ccloc': ccloc, # compartment localization form GO
            'litlo': litlo, # compartment localization from literature
            'litli': litli, # literature ligands from Charlotte
            'limal': limal, # lima assay lipids
            'limam': limam, # lima assay artificial membranes
            'limcm': limcm, # lima assay complex membranes
            'litme': litme, # membrane lipid affinities from literature
            'litpr': litpr, # memb
            'famly': famly, # domain families
            'synon': synon, # synonyms
            'stdnm': stdnm  # standard names
        }
    
    def uniprot_genesymbol(self):
        """
        Build dictionaries for mapping between UniProts and GeneSymbols.
        """
        
        self.u2g = {}
        self.g2u = {}
        
        url = 'ftp://ftp.uniprot.org/pub/databases/uniprot/'\
            'current_release/knowledgebase/reference_proteomes/'\
            'Eukaryota/UP000005640_9606.idmapping.gz'
        
        c = _curl.Curl(url, large = True, silent = False)
        
        for l in c.result:
            
            l = l.decode('utf-8').strip().split('\t')
            
            if l[1] == 'Gene_Name':
                if l[0] not in self.u2g:
                    self.u2g[l[0]] = set([])
                self.u2g[l[0]].add(l[2])
                if l[2] not in self.g2u:
                    self.g2u[l[2]] = set([])
                self.g2u[l[2]].add(l[0])
        
        c.result.close()
        c.close()
    
    def all_uniprots(self):
        """
        Obtains a list of all human SwissProt IDs.
        """
        
        url = 'http://www.uniprot.org/uniprot/'
        post = {
            'query': 'organism:9606 AND reviewed:yes',
            'format': 'tab',
            'columns': 'id'
        }
        c = _curl.Curl(url, post = post, silent = False)
        data = c.result
        self.all_up = list(
            filter(lambda x: len(x) > 0,
                map(lambda l: l.strip(), data.split('\n')[1:])))
    
    def combined_network(self,
                         fname = 'combined_network_%s.csv',
                         classes = set(['I', 'II'])):
        """
        Builds a combined network from prior knowledge,
        LiMA screen from Charlotte and MS screens from Antonella
        and Enric.
        """
        
        hdr_e = ['source', 'target', 'type', 'datasource', 'category']
        hdr_v = ['name', 'type', 'domain', 'uniprot']
        hdr_d = ['protein', 'other', 'ptype', 'otype', 'uniprot', 'domain',
                 'etype', 'datasource', 'category']
        hdr_p = ['charge', 'netcharge', 'fa_chains',
                 'complexity', 'haccept', 'hdonor', 'heavyatoms',
                 'tpsa', 'xlogp', 'rotatablebonds']
        
        all_groups = sorted(set.union(*[i['hg_groups'] for i in  self.lipiprop.values()]))
        
        def get_prop_line(lip):
            
            def get_field(name):
                return (
                    self.lipiprop[lip][name]
                    if name in self.lipiprop[lip] else 'NA')
            
            if lip not in self.lipiprop:
                
                return ['NA'] * len(hdr_p)
            
            def get_groups():
                
                return ['%u' % int(gr in self.lipiprop[lip]['hg_groups'])
                        for gr in all_groups]
            
            return [
                get_field('charge'),
                '%u' % (get_field('charge').count('+') -
                        get_field('charge').count('-')),
                get_field('fa_chains'),
                get_field('Complexity'),
                get_field('HBondAcceptorCount'),
                get_field('HBondDonorCount'),
                get_field('HeavyAtomCount'),
                get_field('TPSA'),
                get_field('XLogP'),
                get_field('RotatableBondCount'),
            ] + get_groups()
        
        def get_family(protein):
            
            if protein in self.uniprots:
                return self.uniprots[protein]['family']
            elif protein in self.prior['famly']:
                return self.prior['famly'][protein]
            elif self.prior['stdnm'][protein] in self.prior['famly']:
                return self.prior['famly'][self.prior['stdnm'][protein]]
            else:
                return 'NA'
        
        def get_uniprot(protein):
            
            stdnm = None
            
            if protein in self.uniprots:
                return self.uniprots[protein]['uniprot']
            
            if protein in self.prior['stdnm']:
                stdnm = self.prior['stdnm'][protein]
                
                if stdnm in self.uniprots:
                    return self.uniprots[stdnm]['uniprot']
                
            if protein in self.g2u or stdnm in self.g2u:
                
                sw = list(filter(lambda u: u in self.all_up,
                                 (self.g2u[protein]
                                  if protein in self.g2u else
                                  self.g2u[stdnm])
                                ))
                
                if not sw:
                    sw = self.g2u[protein]
                if len(sw) > 1:
                    sys.stdout.write('\t:: Warning: more than one UniProt IDs'
                                     'for %s: %s' % (
                                         protein,
                                         ', '.join(sorted(sw))
                                     ))
                
                return sorted(sw)[0]
            else:
                return 'NA'
        
        def get_genesymbol(uniprot):
            
            if uniprot in self.u2g:
                if len(self.u2g[uniprot]) > 1:
                    sys.stdout.write('\t:: Warning: more than one GeneSymbols'
                                     'for %s: %s' % (
                                         uniprot,
                                         ', '.join(sorted(self.u2g[uniprot]))
                                     ))
                return sorted(self.u2g[uniprot])[0]
                
            else:
                return 'NA'
        
        def is_protein(typ):
            return typ in ['protein', 'ltp']
        
        def get_std_name(name):
            return (self.prior['stdnm'][name]
                    if name in self.prior['stdnm']
                    else name)
        
        def add_edge(vertices, edges, df, source, target,
                     source_type, target_type,
                     edge_type, edge_source, own):
            
            if source == 'NA' or target == 'NA':
                return None
            
            psrc = is_protein(source_type)
            nsrc = get_std_name(source)
            ntgt = get_std_name(target)
            usrc = get_uniprot(source) if is_protein(source_type) else 'NA'
            utgt = get_uniprot(target) if is_protein(target_type) else 'NA'
            fsrc = get_family(source) if source_type == 'ltp' else 'NA'
            ftgt = get_family(target) if target_type == 'ltp' else 'NA'
            
            vertices.add((nsrc, source_type, fsrc, usrc))
            vertices.add((ntgt, target_type, ftgt, utgt))
            edges.add((source, target, edge_type, edge_source, own))
            df_row = [
                nsrc if psrc else ntgt,
                ntgt if psrc else nsrc,
                source_type if psrc else target_type,
                target_type if psrc else source_type,
                usrc if psrc else utgt,
                fsrc if psrc else ftgt,
                edge_type,
                edge_source,
                own
            ]
            df_row.extend(get_prop_line(nsrc if source_type == 'lipid' else ntgt))
            df.append(df_row)
        
        def add_set(data, vertices, edges, df,
                    source_type, target_type,
                    edge_type, edge_source, own,
                    rev = False):
            
            for a, bs in iteritems(data):
                
                for b in bs:
                    
                    add_edge(vertices, edges, df,
                             b if rev else a,
                             a if rev else b,
                             source_type, target_type,
                             edge_type, edge_source, own)
        
        # obtaining various data:
        prg = progress.Progress(9, 'Reading and processing data',
                                1, percent = False)
        prg.step()
        self.read_uniprots()
        prg.step()
        self.read_binding_properties()
        prg.step()
        self.read_prior_knowledge()
        prg.step()
        self.manualdir = 'revised_20170425'
        self.manualend = 'revised.xlsx'
        self.read_manual()
        prg.step()
        self.manual_df(screen = 'E')
        prg.step()
        msenric = self.pmanual
        self.read_manual2()
        prg.step()
        self.manual_df(screen = 'A', only_swl_col = True)
        msantonella = self.pmanual
        prg.step()
        self.all_uniprots()
        prg.step()
        self.uniprot_genesymbol()
        prg.terminate()
        
        self.read_lipid_properties()
        self.load_pubchem()
        
        vertices = set([])
        edges = set([])
        df = []
        
        for r in itertools.chain(msenric.itertuples(),
                                 msantonella.itertuples()):
            
            if (r.cls in classes and
                r.uhgroup != 'Adduct' and
                r.uhgroup != 'unconfirmed ID' and
                r.uhgroup != 'P40'):
                
                add_edge(vertices, edges, df, r.protein, r.uhgroup,
                         'ltp', 'lipid', 'cargo_binding',
                         'MS%s' % r.screen, 'own_experiment')
        
        bindprop_all = {}
        for p, lips in itertools.chain(iteritems(self.prior['litli']),
                                       iteritems(self.bindprop)):
            
            pstd = get_std_name(p)
            
            if pstd not in bindprop_all:
                bindprop_all[pstd] = set([])
            
            bindprop_all[pstd].update(lips)
        
        add_set(bindprop_all,
                vertices, edges, df,
                'ltp', 'lipid',
                'cargo_binding', 'LITB', 'literature')
        add_set(self.prior['hptlc'],
                vertices, edges, df,
                'ltp', 'lipid',
                'cargo_binding', 'HPTLC', 'own_experiment')
        add_set(self.prior['hpalo'],
                vertices, edges, df,
                'ltp', 'location',
                'localization', 'HPALOC', 'database')
        add_set(self.prior['ccloc'],
                vertices, edges, df,
                'ltp', 'location',
                'localization', 'CCLOC', 'database')
        add_set(self.prior['litlo'],
                vertices, edges, df,
                'ltp', 'location',
                'localization', 'LITLOC', 'literature')
        add_set(self.prior['limal'],
                vertices, edges, df,
                'lipid', 'ltp',
                'membrane_affinity', 'LIMAL', 'own_experiment',
                rev = True)
        add_set(self.prior['limam'],
                vertices, edges, df,
                'location', 'ltp',
                'membrane_affinity', 'LIMAM', 'own_experiment',
                rev = True)
        add_set(self.prior['limcm'],
                vertices, edges, df,
                'location', 'ltp',
                'membrane_affinity', 'LIMCM', 'own_experiment',
                rev = True)
        add_set(self.prior['litme'],
                vertices, edges, df,
                'lipid', 'ltp',
                'membrane_affinity', 'LITME', 'literature',
                rev = True)
        
        with open(fname % 'v', 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr_v))
            
            for v in vertices:
                fp.write('%s\n' % '\t'.join(v))
        
        with open(fname % 'e', 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr_e))
            
            for e in edges:
                fp.write('%s\n' % '\t'.join(e))
        
        with open(fname % 'df', 'w') as fp:
            
            fp.write('%s\n' % '\t'.join(hdr_d + hdr_p + all_groups))
            
            for l in df:
                fp.write('%s\n' % '\t'.join(l))
    
    def read_lipid_properties(self):
        """
        Reads lipid properties from file.
        These include charge, chemical groups and database IDs.
        """
        
        with open(self.lipipropf, 'r') as fp:
            
            hdr = fp.readline().strip().split('\t')
            
            self.lipiprop = (
                dict(
                    (
                        ll[0],
                        dict(
                            (lab, (
                                set([]) if len(ll) <= i else
                                set(ll[i].split(';')) if i > 3 else
                                ll[i]
                            ))
                            for i, lab in enumerate(hdr)
                        )
                    )
                    for ll in (l.strip().split('\t') for l in fp)
                )
            )
    
    def export_lipid_properties(self,
                                fname = 'lipid_properties_phys-chem.csv'):
        
        def ffield(content):
            
            return (';'.join(content)
                if type(content) is set or type(content) is list
                else '{}'.format(content))
        
        main_hdr = ['lipid', 'charge', 'hg_formula', 'fa_chains',
               'hg_groups', 'databas_ids', 'database_ids_C18',
               'database_ids_C16', 'database_ids_C14']
        
        pubchem_hdr = ['CID', 'Complexity', 'HBondAcceptorCount', 'HBondDonorCount',
               'TPSA', 'XLogP', 'HeavyAtomCount', 'RotatableBondCount']
        
        with open(fname, 'w') as fp:
            
            fp.write('%s\t' % '\t'.join(main_hdr))
            fp.write('%s\n' % '\t'.join(pubchem_hdr))
            
            for k, lip in iteritems(self.lipiprop):
                
                row = []
                row.extend(ffield(lip[i]) if i in lip else ''
                           for i in main_hdr)
                row.extend(ffield(lip[i]) if i in lip else ''
                           for i in pubchem_hdr)
                
                fp.write('%s\n' % '\t'.join(row))
    
    def get_pubchem(self, cid):
        """
        Fetches PubChem data for one compound ID.
        """
        
        url = self.pubchem_url % str(cid)
        c = _curl.Curl(url, silent = True, large = False)
        d = xmltodict.parse(c.result)['PropertyTable']['Properties']
        return d
    
    def load_pubchem(self):
        """
        Downloads PubChem molecule properties
        and adds them to the `lipiprop` dict.
        """
        
        def has_pubchem(ids):
            
            for i in ids:
                if i.startswith('PubChemCID'):
                    return i.split(':')[1]
        
        prg = progress.Progress(len(self.lipiprop),
                                'Downloading PubChem data',
                                1, percent = False)
        
        for k, lip in iteritems(self.lipiprop):
            
            prg.step()
            
            for idcol in ['', '_C18', '_C16', '_C14']:
                
                cid = has_pubchem(lip['database_ids%s' % idcol])
                
                if cid:
                    
                    d = self.get_pubchem(cid)
                    
                    lip.update(d)
                    lip['pubchem'] = cid
                    
                    break
        
        prg.terminate()
    
    def hmdb_pubchem_mapping(self):
        """
        Creates dict of sets mapping tables between PubChem and HMDB.
        """
        
        url = 'ftp://ftp.ebi.ac.uk/pub/databases/chembl/UniChem/data/'\
            'wholeSourceMapping/src_id18/src18src22.txt.gz'
        
        c = _curl.Curl(url, silent = False, large = True)
        
        hmdb_pubchem = {}
        pubchem_hmdb = {}
        
        _ = c.result.readline()
        
        for l in c.result:
            
            l = l.decode('ascii').strip().split('\t')
            
            if l[0] not in hmdb_pubchem:
                hmdb_pubchem[l[0]] = set([])
            
            if l[1] not in pubchem_hmdb:
                pubchem_hmdb[l[1]] = set([])
            
            hmdb_pubchem[l[0]].add(l[1])
            pubchem_hmdb[l[1]].add(l[0])
        
        c.close()
        
        self.hmdb_pubchem = hmdb_pubchem
        self.pubchem_hmdb = pubchem_hmdb
    
    def hmdb_data(self, cleanup_period = 4000):
        """
        Looks up properties of PubChem compounds in HMDB data.
        """
        
        self.hmdb_result = {}
        used = []
        
        pubchems = set(lip['pubchem']
                       for lip in self.lipiprop.values()
                       if 'pubchem' in lip)
        
        hmdbs = set.union(*[self.pubchem_hmdb[p]
                          for p in pubchems
                          if p in self.pubchem_hmdb])
        
        pref = '{http://www.hmdb.ca}'
        metab = '%smetabolite' % pref
        acces = '%saccession' % pref
        saccs = '%ssecondary_accessions' % pref
        prprp = '%spredicted_properties' % pref
        exprp = '%sexperimental_properties' % pref
        pccid = '%spubchem_compound_id' % pref
        
        print('looking up: %s' % ', '.join(hmdbs))
        
        def elem_callback(elem):
            
            this_result = {}
            
            this_acc = set([])
            
            a = elem.find(acces)
            
            if a is not None:
                this_acc.add(a.text)
            
            sa = elem.find(saccs)
            if sa is not None:
                sas = sa.find(acces)
                if sas is not None:
                    this_acc.update(set(_sa.text for _sa in sas))
            
            h_needed = this_acc & hmdbs
            
            pc = elem.find(pccid)
            
            if pc is not None:
                
                pcid = pc.text
                
                if pc in pubchems:
                    
                    h_needed = this_acc
            
            if not h_needed:
                # print('-- none of these needed: %s' % ', '.join(this_acc))
                return None
            
            print('-- collecting information about: %s' % ', '.join(h_needed))
            
            pp = elem.find(prprp)
            
            if pp is not None:
                
                print('  -- parsing predicted properties')
                
                this_result['pprop'] = xmltodict.parse(lxml.etree.tostring(pp))
            
            else:
                print('  -- no predicted properties')
            
            ep = elem.find(exprp)
            
            if ep is not None:
                
                this_result['eprop'] = xmltodict.parse(lxml.etree.tostring(ep))
            
            for h in h_needed:
                if h not in self.hmdb_result:
                    self.hmdb_result[h] = []
                self.hmdb_result[h].append(this_result)
        
        url = 'http://www.hmdb.ca/system/downloads/'\
            'current/hmdb_metabolites.zip'
        
        c = _curl.Curl(url, large = True, silent = False,
                       files_needed = ['hmdb_metabolites.xml'])
        
        fp = c.result['hmdb_metabolites.xml']
        
        e = lxml.etree.iterparse(fp, events=('start', 'end'))
        
        for ev, elem in e:
            
            used.append(elem)
            
            if ev == 'end' and elem.tag == metab:
                
                elem_callback(elem)
            
            if len(used) > cleanup_period:
                
                for i in xrange(int(cleanup_period / 2)):
                    
                    u = used.pop()
                    u.clear()
