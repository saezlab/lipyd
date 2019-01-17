#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
#  This file is part of the `lipyd` python module
#
#  Copyright (c) 2015-2019 - EMBL
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#
#  This code is not for public use.
#  Please do not redistribute.
#  For permission please contact me.
#
#  Website: http://denes.omnipathdb.org/
#


from past.builtins import xrange


import os
import sys
import time
# workaround to avoid warnings from xlrd
# this issue later will be fixed in the
# xlrd module
if hasattr(time, 'process_time'):
    time.clock = time.process_time

import xlrd
import openpyxl

import lipyd.reader.common as common


def read_xls(xls_file, sheet = 0, use_openpyxl = False):
    """Generic function to read MS Excel XLS file, and convert one sheet
    to CSV, or return as a list of lists

    Parameters
    ----------
    xls_file :
        
    sheet :
         (Default value = 0)
    use_openpyxl :
         (Default value = False)

    Returns
    -------

    """
    
    if not use_openpyxl:
        
        try:
            
            book = xlrd.open_workbook(xls_file, on_demand = True)
            
            try:
                if type(sheet) is int:
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)
            
            for i in xrange(sheet.nrows):
                
                yield [common.basestring(c.value) for c in sheet.row(i)]
            
            use_openpyxl = False
            
        except IOError:
            
            raise FileNotFoundError(xls_file)
            
        except:
            
            use_openpyxl = True
    
    if use_openpyxl:
        
        try:
            
            book = openpyxl.load_workbook(
                filename = xls_file,
                read_only = True
            )
            
        except:
            
            raise ValueError('Could not open xls: %s' % xls_file)
            
            if not os.path.exists(xls_file):
                
                raise FileNotFoundError(xls_file)
        
        try:
            
            if type(sheet) is int:
                sheet = book.worksheets[sheet]
            else:
                sheet = book[sheet]
            
        except:
            
            sheet = book.worksheets[0]
        
        cells = sheet.get_squared_range(
            1, 1, sheet.max_column, sheet.max_row
        )
        
        for row in cells:
            
            yield [common.basestring(c.value) if c.value else '' for c in row]
    
    if 'book' in locals() and hasattr(book, 'release_resources'):
        book.release_resources()
